import streamlit as st
import streamlit.components.v1 as components
import numpy as np
import matplotlib.pyplot as plt
import random
import json
import os
import math
import base64
from io import BytesIO
from datetime import datetime, timedelta, timezone as _tz_mod

_GMT5 = _tz_mod(timedelta(hours=-5))
def _now_gmt5(): return datetime.now(_GMT5)

import database as db
from constants import *
from calculations import (
    normalize_disc_scores,
    calculate_disc_results,
    calculate_behavioral_styles,
    get_disc_temperament,
    generate_disc_mega_summary,
    calculate_valanti_results,
    calculate_wpi_results,
    load_eri_questions,
    calculate_eri_results,
    load_talent_map_questions,
    calculate_talent_map_results,
    calculate_desempeno_results,
    calculate_desempeno_lider_results,
    calculate_periodo_prueba_results,
)
from analysis import (
    analyze_disc_aptitude,
    analyze_valanti_aptitude,
    analyze_wpi_aptitude,
    analyze_eri_aptitude,
    analyze_talent_map_match,
)
from charts import (
    create_disc_plot, create_behavioral_styles_chart,
    create_valanti_radar, create_valanti_bars,
    create_wpi_radar, create_wpi_bars,
    create_eri_radar, create_eri_bars,
    create_talent_map_radar, create_talent_map_bars,
    create_talent_map_comparison,
    create_desempeno_radar, create_desempeno_bars,
)
from pdfs import (
    generate_disc_pdf, generate_valanti_pdf, generate_wpi_pdf,
    generate_eri_pdf, generate_talent_map_pdf, generate_desempeno_pdf,
    generate_desempeno_lider_pdf, generate_periodo_prueba_pdf,
)
from utils import load_disc_questions, load_disc_descriptions, load_wpi_questions, nav
from pages.desempeno import (
    show_desempeno_results_admin,
    show_desempeno_lider_results_admin,
    show_periodo_prueba_results_admin,
)
from pages.talent_map import show_talent_map_results_admin
from auth import (
    _restore_admin_session, _touch_admin_session,
    _logout_admin, _start_admin_session,
    _create_result_view_token, _parse_result_view_token,
)


def _build_secure_result_url(session_id, test_type):
    token = _create_result_view_token(session_id, test_type)
    return f"?page=shared_result&rv={token}"


def page_shared_result_view():
    st.markdown("## 🔗 Visualización Segura de Resultados")

    rv_token = None
    try:
        rv_token = st.query_params.get("rv")
    except Exception:
        rv_token = None
    if not rv_token:
        rv_token = st.session_state.get("shared_result_token")

    if not rv_token:
        st.error("❌ Enlace inválido: falta token de visualización.")
        if st.button("🏠 Ir al inicio", use_container_width=True):
            nav("home")
        return

    token_data = _parse_result_view_token(rv_token)
    if not token_data:
        st.error("❌ Enlace inválido o vencido.")
        if st.button("🏠 Ir al inicio", use_container_width=True):
            nav("home")
        return

    session_id = token_data["session_id"]
    expected_test_type = token_data["test_type"]
    session = db.get_session_by_id(session_id)

    if not session:
        st.error("❌ No se encontró la evaluación asociada al enlace.")
        return
    if session.get("test_type") != expected_test_type:
        st.error("❌ Enlace no válido para este tipo de evaluación.")
        return
    if session.get("status") != "completed":
        st.warning("⚠️ Esta evaluación aún no está completada.")
        return

    candidate = db.get_candidate_by_id(session.get("candidate_id"))
    results = db.get_results(session_id)

    if not candidate or not results:
        st.warning("⚠️ No hay datos suficientes para mostrar esta evaluación.")
        return

    st.caption(f"Sesión: {session_id} | Tipo: {session.get('test_type', '').upper()}")

    test_type = session["test_type"]
    if test_type == "disc":
        show_disc_results_admin(results, candidate, session)
    elif test_type == "valanti":
        show_valanti_results_admin(results, candidate, session)
    elif test_type == "wpi":
        show_wpi_results_admin(results, candidate, session)
    elif test_type == "eri":
        show_eri_results_admin(results, candidate, session)
    elif test_type == "talent_map":
        show_talent_map_results_admin(results, candidate, session)
    elif test_type == "desempeno":
        show_desempeno_results_admin(results, candidate, session)
    elif test_type == "desempeno_lider":
        show_desempeno_lider_results_admin(results, candidate, session)
    elif test_type == "periodo_prueba":
        show_periodo_prueba_results_admin(results, candidate, session)
    else:
        st.info("Tipo de evaluación no soportado para esta vista.")

    b1, b2 = st.columns(2)
    with b1:
        if st.button("🏠 Ir al inicio", use_container_width=True, key="shared_home"):
            try:
                st.query_params.pop("rv", None)
                st.query_params.pop("page", None)
            except Exception:
                pass
            nav("home")
    with b2:
        if st.session_state.get("admin"):
            if st.button("🛡️ Volver al dashboard", use_container_width=True, key="shared_admin"):
                try:
                    st.query_params.pop("rv", None)
                    st.query_params.pop("page", None)
                except Exception:
                    pass
                nav("admin_dashboard")

def page_admin_login():
    st.markdown("## 🔒 Acceso Administrador RH")
    if st.button("⬅️ Volver al inicio"):
        nav("home")
        st.rerun()

    with st.form("admin_login_form"):
        username = st.text_input("Usuario", key="admin_user")
        password = st.text_input("Contraseña", type="password", key="admin_pass")
        submitted = st.form_submit_button("Iniciar Sesión")
        
        if submitted:
            if not username or not password:
                st.error("❌ Por favor completa todos los campos.")
            else:
                username = username.strip()
                password = password.strip()
                
                admin = db.verify_admin(username, password)
                if admin:
                    _start_admin_session(admin)
                    nav("admin_dashboard")
                    st.rerun()
                else:
                    st.error("❌ Credenciales incorrectas. Verifica usuario y contraseña.")


# -------------------------------------------------------------------------
# ADMIN: DASHBOARD
# -------------------------------------------------------------------------
def page_admin_dashboard():
    _restore_admin_session()
    admin = st.session_state.get("admin")
    if not admin:
        nav("admin_login")
        st.rerun()
        return

    st.markdown(f"## 🛡️ Panel de Administración")
    st.caption(f"Bienvenido, {admin['name']}")

    st.markdown(
        """
        <style>
        div[class*="st-key-"][class*="_pending_edit_btn_"] button {
            border: 1px solid #cbd5e1;
        }
        div[class*="st-key-"][class*="_pending_delete_btn_"] button {
            border: 1px solid #dc2626;
            color: #b91c1c;
            background: #fff;
        }
        div[class*="st-key-"][class*="_pending_delete_btn_"] button:hover {
            border-color: #b91c1c;
            background: #fee2e2;
            color: #7f1d1d;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    # ----- Filtros compartidos (Resultados / Pruebas Pendientes) -----
    _all_raw = db.get_all_sessions()
    _all_candidates = db.get_all_candidates()
    _candidate_name_by_cedula = {c["cedula"]: c.get("name", "N/A") for c in _all_candidates}
    _sessions_by_cedula = {}
    for _s in _all_raw:
        _sessions_by_cedula.setdefault(_s["cedula"], []).append(_s)
    _cand_names = sorted(set(s["candidate_name"] for s in _all_raw)) if _all_raw else []
    _FILTER_LABELS = {
        "Todos": "📋 Todos", "disc": "🎯 DISC", "valanti": "🧭 VALANTI",
        "wpi": "💼 WPI", "eri": "🔐 ERI", "talent_map": "🌟 Talent Map",
        "desempeno": "📊 Desempeño Operativo",
        "desempeno_lider": "📊 Desempeño Líderes",
        "periodo_prueba": "📋 Período de Prueba",
    }
    _SECTION_OPTIONS = {
        "dashboard": "📈 Dashboard Gerencial",
        "create": "📋 Crear Evaluación",
        "new_candidate": "➕ Nuevo Candidato",
        "candidates": "👥 Candidatos",
        "results": "📊 Resultados",
        "pending": "⏳ Pruebas Pendientes",
        "bulk": "📤 Cargue Masivo",
        "settings": "⚙️ Configuración",
    }

    with st.sidebar:
        st.markdown("### 🛡️ Panel Admin")
        st.caption(f"Sesión: {admin['name']}")
        if st.button("🚪 Cerrar Sesión", use_container_width=True):
            _logout_admin()
            nav("home")
            st.rerun()

        st.markdown("---")
        _active_section = st.radio(
            "Ir a",
            options=list(_SECTION_OPTIONS.keys()),
            key="admin_sidebar_section",
            format_func=lambda value: _SECTION_OPTIONS[value],
        )

        _filter_type = "Todos"
        _filter_cand = "Todos"
        _sort_opt = "Fecha (reciente)"
        if _active_section in ("results", "pending"):
            st.markdown("---")
            st.markdown("### 🔎 Filtros")
            _filter_type = st.selectbox(
                "Filtrar por tipo",
                list(_FILTER_LABELS.keys()),
                key="filter_type",
                format_func=lambda x: _FILTER_LABELS.get(x, x),
            )
            _filter_cand = st.selectbox(
                "Filtrar por candidato",
                ["Todos"] + _cand_names,
                key="filter_candidate",
            )
            _sort_opt = st.selectbox(
                "Ordenar por",
                ["Fecha (reciente)", "Fecha (antigua)", "Candidato A-Z", "Candidato Z-A", "Tipo prueba"],
                key="sort_option",
            )

    _ft = _filter_type if _filter_type != "Todos" else None
    _EDITABLE_TEST_TYPES = [k for k in _FILTER_LABELS.keys() if k != "Todos"]

    def _get_sort_date(s):
        date_str = s.get("completed_at") or s.get("started_at") or s.get("created_at") or ""
        try:
            return datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
        except:
            return datetime.min

    def _sort_sessions(lst):
        if _sort_opt == "Fecha (reciente)":
            lst.sort(key=_get_sort_date, reverse=True)
        elif _sort_opt == "Fecha (antigua)":
            lst.sort(key=_get_sort_date)
        elif _sort_opt == "Candidato A-Z":
            lst.sort(key=lambda s: s["candidate_name"].lower())
        elif _sort_opt == "Candidato Z-A":
            lst.sort(key=lambda s: s["candidate_name"].lower(), reverse=True)
        elif _sort_opt == "Tipo prueba":
            lst.sort(key=lambda s: s["test_type"])

    def _render_sessions_list(sessions, tab_key):
        if not sessions:
            st.info("No hay evaluaciones que coincidan con los filtros.")
            return
        _TEST_NAME_MAP = {
            "disc": "DISC", "valanti": "VALANTI", "wpi": "WPI", "eri": "ERI",
            "talent_map": "TALENT MAP", "desempeno": "DESEMPEÑO OP.",
            "desempeno_lider": "DESEMPEÑO LÍD.", "periodo_prueba": "PER. PRUEBA",
        }
        _TEST_EMOJI_MAP = {"disc": "🎯", "valanti": "🧭", "wpi": "💼", "eri": "🔐"}
        _STATUS_EMOJI = {"pending": "⏳", "in_progress": "▶️", "completed": "✅", "expired": "⏰", "employee_done": "📝"}
        # Estado de ordenamiento por tabla (independiente por tab_key)
        _sc_key = f"_rsort_col_{tab_key}"
        _sd_key = f"_rsort_dir_{tab_key}"
        if _sc_key not in st.session_state:
            st.session_state[_sc_key] = "FECHA"
            st.session_state[_sd_key] = "desc"
        _curr_col = st.session_state[_sc_key]
        _curr_dir = st.session_state[_sd_key]

        def _col_val_sort(s, col):
            if col == "CANDIDATO": return s["candidate_name"].lower()
            if col == "PRUEBA": return s["test_type"]
            if col == "CÉDULA": return str(s["cedula"])
            if col == "ID": return s.get("id", 0)
            if col == "FECHA": return _get_sort_date(s)
            return ""

        sessions = sorted(sessions, key=lambda s: _col_val_sort(s, _curr_col), reverse=(_curr_dir == "desc"))

        _W = [0.35, 2.4, 1.6, 1.5, 1.05, 1.65, 0.45]
        st.caption(f"📊 {len(sessions)} evaluación(es) encontrada(s)")
        # Cabecera con ordenamiento por columna
        _hdr = st.columns(_W)
        _hdr[0].markdown("<div style='height:30px'></div>", unsafe_allow_html=True)
        _hdr[6].markdown("<div style='height:30px'></div>", unsafe_allow_html=True)
        for _hcol_idx, _hcol_name in [(1, "CANDIDATO"), (2, "PRUEBA"), (3, "CÉDULA"), (4, "ID"), (5, "FECHA")]:
            _arrow = " ↓" if (_curr_col == _hcol_name and _curr_dir == "desc") else (
                     " ↑" if _curr_col == _hcol_name else " ↕")
            if _hdr[_hcol_idx].button(
                f"{_hcol_name}{_arrow}",
                key=f"hdr_{tab_key}_{_hcol_name}",
                use_container_width=True,
            ):
                if _curr_col == _hcol_name:
                    st.session_state[_sd_key] = "asc" if _curr_dir == "desc" else "desc"
                else:
                    st.session_state[_sc_key] = _hcol_name
                    st.session_state[_sd_key] = "desc"
                st.rerun()
        for sess in sessions:
            status_emoji = _STATUS_EMOJI.get(sess["status"], "❓")
            test_emoji = _TEST_EMOJI_MAP.get(sess["test_type"], "📝")
            test_name = _TEST_NAME_MAP.get(sess["test_type"], sess["test_type"].upper())
            evaluador_ced = sess.get("evaluador_cedula")
            evaluador_nombre = sess.get("evaluador_nombre")
            _date_ref = sess.get("completed_at") or sess.get("started_at") or sess.get("created_at")
            fecha_str = "—"
            if _date_ref:
                try:
                    fecha_str = datetime.strptime(_date_ref, "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y %H:%M")
                except:
                    fecha_str = str(_date_ref)
            toggle_key = f"show_detail_{tab_key}_{sess['id']}"
            is_open = st.session_state.get(toggle_key, False)
            # Fila de la tabla
            rc = st.columns(_W)
            rc[0].markdown(f"<div style='font-size:17px;padding-top:3px'>{status_emoji}</div>", unsafe_allow_html=True)
            rc[1].markdown(f"<div style='font-weight:600;padding-top:5px;font-size:13px'>{sess['candidate_name']}</div>", unsafe_allow_html=True)
            rc[2].markdown(f"<div style='font-size:12px;padding-top:6px'>{test_emoji} {test_name}</div>", unsafe_allow_html=True)
            rc[3].markdown(f"<div style='font-size:12px;font-family:monospace;padding-top:6px'>{sess['cedula']}</div>", unsafe_allow_html=True)
            rc[4].markdown(f"<div style='font-size:12px;font-family:monospace;padding-top:6px'>{sess['id']}</div>", unsafe_allow_html=True)
            rc[5].markdown(f"<div style='font-size:12px;padding-top:6px'>{fecha_str}</div>", unsafe_allow_html=True)
            if rc[6].button("▲" if is_open else "▼", key=f"tog_{tab_key}_{sess['id']}", use_container_width=True):
                st.session_state[toggle_key] = not is_open
                st.rerun()
            if is_open:
                with st.container(border=True):
                    ec1, ec2, ec3, ec4 = st.columns(4)
                    ec1.metric("Estado", sess["status"].upper())
                    ec2.metric("Tiempo Límite", f"{sess['time_limit_minutes']} min")
                    _s_at = sess.get("started_at")
                    _c_at = sess.get("completed_at")
                    try:
                        started_str = datetime.strptime(_s_at, "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y %H:%M") if _s_at else "N/A"
                    except:
                        started_str = _s_at or "N/A"
                    try:
                        completed_str = datetime.strptime(_c_at, "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y %H:%M") if _c_at else "N/A"
                    except:
                        completed_str = _c_at or "N/A"
                    ec3.metric("Iniciado", started_str)
                    ec4.metric("Completado", completed_str)
                    if evaluador_ced:
                        jefe_nombre = evaluador_nombre or _candidate_name_by_cedula.get(evaluador_ced)
                        if jefe_nombre:
                            st.info(f"👔 **Jefe asignado:** {jefe_nombre} (Cédula `{evaluador_ced}`)")
                        else:
                            st.info(f"👔 **Jefe asignado:** Cédula `{evaluador_ced}`")
                    if sess["status"] == "completed":
                        details_key = f"{tab_key}_show_details_{sess['id']}"
                        if not st.session_state.get(details_key, False):
                            st.caption("Resultados detallados ocultos para acelerar carga.")
                            _dc1, _dc2 = st.columns(2)
                            with _dc1:
                                _secure_url = _build_secure_result_url(sess["id"], sess["test_type"])
                                st.link_button("🔗 Visualizar", url=_secure_url, use_container_width=True)
                            with _dc2:
                                if st.button("📥 Descargar PDF", key=f"{details_key}_pdf_btn", use_container_width=True):
                                    st.session_state[details_key] = True
                                    st.session_state[f"{details_key}_focus_pdf"] = True
                                    st.rerun()
                        elif st.session_state.get(f"{details_key}_focus_pdf", False):
                            # Modo descarga directa: solo genera y ofrece el PDF
                            _res_pdf = db.get_results(sess["id"])
                            _cand_pdf = db.get_candidate_by_cedula(sess["cedula"])
                            if _res_pdf and _cand_pdf:
                                try:
                                    if sess["test_type"] == "desempeno":
                                        _radar = create_desempeno_radar(_res_pdf.get("potencial_scores", {}))
                                        _bars = create_desempeno_bars(_res_pdf.get("rendimiento_scores", {}))
                                        _pdf_buf = generate_desempeno_pdf(
                                            candidate=_cand_pdf,
                                            rendimiento_scores=_res_pdf.get("rendimiento_scores", {}),
                                            potencial_scores=_res_pdf.get("potencial_scores", {}),
                                            radar_fig=_radar,
                                            bars_fig=_bars,
                                            session_id=sess["id"],
                                            completed_at=sess.get("completed_at"),
                                            analysis=_res_pdf.get("analysis", {}),
                                            evaluador_nombre=_res_pdf.get("evaluador"),
                                            iniciativas=_res_pdf.get("iniciativas", []),
                                        )
                                        st.download_button(
                                            "📄 Descargar PDF — Desempeño Operativo",
                                            data=_pdf_buf,
                                            file_name=f"evaluacion_desempeno_{_cand_pdf['cedula']}_{sess['id']}.pdf",
                                            mime="application/pdf",
                                            key=f"dl_pdf_direct_{sess['id']}",
                                            use_container_width=True,
                                        )
                                    elif sess["test_type"] == "desempeno_lider":
                                        _pdf_buf = generate_desempeno_lider_pdf(
                                            candidate=_cand_pdf,
                                            competencias_scores={int(k): v for k, v in _res_pdf.get("competencias_scores", {}).items()},
                                            rendimiento_scores=_res_pdf.get("rendimiento_scores", {}),
                                            potencial_scores=_res_pdf.get("potencial_scores", {}),
                                            session_id=sess["id"],
                                            completed_at=sess.get("completed_at"),
                                            analysis=_res_pdf.get("analysis", {}),
                                            evaluador_nombre=_res_pdf.get("evaluador"),
                                            nivel_cargo=_res_pdf.get("nivel_cargo"),
                                            iniciativas=_res_pdf.get("iniciativas", []),
                                        )
                                        st.download_button(
                                            "📄 Descargar PDF — Desempeño Líderes",
                                            data=_pdf_buf,
                                            file_name=f"desempeno_lider_{_cand_pdf['cedula']}_{sess['id']}.pdf",
                                            mime="application/pdf",
                                            key=f"dl_pdf_direct_{sess['id']}",
                                            use_container_width=True,
                                        )
                                    else:
                                        st.info("La descarga directa solo está disponible para evaluaciones de desempeño.")
                                        st.session_state[details_key] = True
                                        st.session_state.pop(f"{details_key}_focus_pdf", None)
                                        st.rerun()
                                except Exception as _epdf:
                                    st.error(f"Error generando PDF: {_epdf}")
                            if st.button("← Volver", key=f"{details_key}_back", use_container_width=True):
                                st.session_state.pop(f"{details_key}_focus_pdf", None)
                                st.rerun()
                        else:
                            results = db.get_results(sess["id"])
                            candidate = db.get_candidate_by_cedula(sess["cedula"])
                            if results:
                                if sess["test_type"] == "disc":
                                    show_disc_results_admin(results, candidate, sess)
                                elif sess["test_type"] == "valanti":
                                    show_valanti_results_admin(results, candidate, sess)
                                elif sess["test_type"] == "wpi":
                                    show_wpi_results_admin(results, candidate, sess)
                                elif sess["test_type"] == "eri":
                                    show_eri_results_admin(results, candidate, sess)
                                elif sess["test_type"] == "talent_map":
                                    show_talent_map_results_admin(results, candidate, sess)
                                elif sess["test_type"] == "desempeno":
                                    show_desempeno_results_admin(results, candidate, sess)
                                elif sess["test_type"] == "desempeno_lider":
                                    show_desempeno_lider_results_admin(results, candidate, sess)
                                elif sess["test_type"] == "periodo_prueba":
                                    show_periodo_prueba_results_admin(results, candidate, sess)
                            else:
                                st.warning("Resultados no disponibles.")
                    elif sess["status"] == "pending" and sess["test_type"] == "desempeno":
                        st.info("⏳ Esta evaluación está pendiente de ser completada por un evaluador.")
                        if st.button("✏️ Evaluar Ahora", key=f"{tab_key}_eval_desemp_{sess['id']}"):
                            st.session_state["desempeno_session_id"] = sess["id"]
                            nav("desempeno_eval")
                            st.rerun()
                    elif sess["status"] == "pending" and sess["test_type"] == "desempeno_lider":
                        st.info("⏳ Pendiente. El empleado debe completar su auto-evaluación primero.")
                    elif sess["status"] == "employee_done" and sess["test_type"] == "desempeno_lider":
                        st.success("📝 El empleado completó su parte. El jefe asignado debe completar la evaluación.")
                        if st.button("✏️ Completar como Admin", key=f"{tab_key}_eval_lider_admin_{sess['id']}"):
                            st.session_state["desempeno_lider_session_id"] = sess["id"]
                            nav("desempeno_lider_jefe_eval")
                            st.rerun()
                    elif sess["status"] == "pending" and sess["test_type"] == "periodo_prueba":
                        st.info("⏳ Pendiente. El empleado debe completar su auto-evaluación primero.")
                    elif sess["status"] == "employee_done" and sess["test_type"] == "periodo_prueba":
                        st.success("📝 El empleado completó su parte. El jefe asignado debe completar la evaluación.")
                        if st.button("✏️ Completar como Admin", key=f"{tab_key}_eval_pp_admin_{sess['id']}"):
                            st.session_state["periodo_prueba_session_id"] = sess["id"]
                            nav("periodo_prueba_jefe_eval")
                            st.rerun()
                    if sess["status"] == "expired":
                        st.markdown("---")
                        _reactivar_key = f"show_reactivar_{tab_key}_{sess['id']}"
                        if st.button(
                            "🔄 Reactivar prueba expirada",
                            key=f"{tab_key}_reactivar_btn_{sess['id']}",
                            use_container_width=True,
                        ):
                            st.session_state[_reactivar_key] = not st.session_state.get(_reactivar_key, False)
                        if st.session_state.get(_reactivar_key, False):
                            with st.container(border=True):
                                st.warning("⚠️ Se restablecerá la prueba a estado **pendiente**. El candidato podrá iniciarla de nuevo desde cero.")
                                with st.form(f"reactivar_{tab_key}_{sess['id']}"):
                                    _tl_opts = [15, 20, 30, 45, 60, 90]
                                    _curr_tl = sess.get("time_limit_minutes", 45)
                                    _new_tl = st.selectbox(
                                        "Nuevo tiempo límite",
                                        options=_tl_opts,
                                        index=_tl_opts.index(_curr_tl) if _curr_tl in _tl_opts else 3,
                                        format_func=lambda x: f"{x} minutos",
                                        key=f"reactivar_tl_{tab_key}_{sess['id']}",
                                    )
                                    _rc1, _rc2 = st.columns(2)
                                    _confirmar = _rc1.form_submit_button("✅ Confirmar reactivación", use_container_width=True)
                                    _cancelar = _rc2.form_submit_button("✖ Cancelar", use_container_width=True)
                                    if _confirmar:
                                        _ok_reactivate = db.reactivate_test_session(sess["id"], new_time_limit_minutes=_new_tl)
                                        st.session_state.pop(_reactivar_key, None)
                                        if _ok_reactivate:
                                            st.success(f"✅ Prueba reactivada con {_new_tl} min. El candidato puede iniciarla de nuevo.")
                                            st.rerun()
                                        st.error("❌ No fue posible reactivar la sesión (debe estar expirada).")
                                    if _cancelar:
                                        st.session_state.pop(_reactivar_key, None)
                                        st.rerun()
                    if sess["status"] == "pending":
                        st.markdown("---")
                        pending_actions_col, pending_delete_col = st.columns([3, 1])
                        edit_toggle_key = f"show_edit_pending_{tab_key}_{sess['id']}"
                        with pending_actions_col:
                            if st.button(
                                "✏️ Editar prueba pendiente",
                                key=f"{tab_key}_pending_edit_btn_{sess['id']}",
                                use_container_width=True,
                            ):
                                st.session_state[edit_toggle_key] = not st.session_state.get(edit_toggle_key, False)
                        with pending_delete_col:
                            if st.button(
                                "🗑️ Eliminar",
                                key=f"{tab_key}_pending_delete_btn_{sess['id']}",
                                use_container_width=True,
                            ):
                                st.session_state[f"confirm_del_{sess['id']}"] = True
                        if st.session_state.get(edit_toggle_key, False):
                            with st.container(border=True):
                                with st.form(f"edit_pending_{tab_key}_{sess['id']}"):
                                    cedit1, cedit2 = st.columns(2)
                                    with cedit1:
                                        edit_test_type = st.selectbox(
                                            "Tipo de prueba",
                                            options=_EDITABLE_TEST_TYPES,
                                            index=_EDITABLE_TEST_TYPES.index(sess["test_type"]) if sess["test_type"] in _EDITABLE_TEST_TYPES else 0,
                                            format_func=lambda x: _FILTER_LABELS.get(x, x),
                                            key=f"edit_type_{tab_key}_{sess['id']}",
                                        )
                                    with cedit2:
                                        _tl_opts = [15, 20, 30, 45, 60, 90]
                                        _curr_tl = sess.get("time_limit_minutes", 45)
                                        edit_time_limit = st.selectbox(
                                            "Tiempo límite",
                                            options=_tl_opts,
                                            index=_tl_opts.index(_curr_tl) if _curr_tl in _tl_opts else 3,
                                            format_func=lambda x: f"{x} minutos",
                                            key=f"edit_tl_{tab_key}_{sess['id']}",
                                        )
                                    edit_jefe_ced = st.text_input(
                                        "Cédula del jefe/evaluador",
                                        value=(sess.get("evaluador_cedula") or ""),
                                        key=f"edit_jefe_ced_{tab_key}_{sess['id']}",
                                    )
                                    edit_jefe_nom = st.text_input(
                                        "Nombre del jefe/evaluador",
                                        value=(sess.get("evaluador_nombre") or ""),
                                        key=f"edit_jefe_nom_{tab_key}_{sess['id']}",
                                    )
                                    save_edit = st.form_submit_button("💾 Guardar cambios")
                                    if save_edit:
                                        if edit_test_type in ("desempeno", "desempeno_lider", "periodo_prueba") and not edit_jefe_ced.strip():
                                            st.error("❌ La cédula del jefe/evaluador es obligatoria para este tipo de prueba.")
                                        else:
                                            ok, err = db.update_pending_session(
                                                sess["id"],
                                            edit_test_type,
                                            edit_time_limit,
                                            evaluador_cedula=edit_jefe_ced.strip() or None,
                                            evaluador_nombre=edit_jefe_nom.strip() or None,
                                        )
                                        if not ok:
                                            st.error(f"❌ {err}")
                                        else:
                                            st.success("✅ Prueba actualizada correctamente.")
                                            st.session_state[edit_toggle_key] = False
                                            st.rerun()

                    if st.session_state.get(f"confirm_del_{sess['id']}", False):
                        st.warning(f"⚠️ ¿Eliminar prueba **{sess['id']}** de **{sess['candidate_name']}**? Esta acción no se puede deshacer.")
                        col_yes, col_no, _ = st.columns([1, 1, 2])
                        with col_yes:
                            if st.button("✅ Sí, eliminar", key=f"{tab_key}_confirm_yes_{sess['id']}"):
                                db.delete_test_session(sess['id'])
                                st.session_state.pop(f"confirm_del_{sess['id']}", None)
                                st.success("✅ Prueba pendiente eliminada.")
                                st.rerun()
                        with col_no:
                            if st.button("❌ Cancelar", key=f"{tab_key}_confirm_no_{sess['id']}"):
                                st.session_state.pop(f"confirm_del_{sess['id']}", None)
                                st.rerun()

    # ----- SECCIÓN: Crear Evaluación (solo candidatos existentes) -----
    if _active_section == "create":
        st.markdown("### 📋 Crear Evaluación")

        _TEST_LABELS_SHARED = {
            "disc": "🎯 DISC", "valanti": "🧭 VALANTI", "wpi": "💼 WPI",
            "eri": "🔐 ERI", "talent_map": "🌟 Talent Map",
            "desempeno": "📊 Desempeño Operativo",
            "desempeno_lider": "📊 Desempeño Líderes",
            "periodo_prueba": "📋 Período de Prueba",
        }
        _TIPOS_CON_EVALUADOR = ("desempeno", "desempeno_lider", "periodo_prueba")

        _candidates_all = db.get_all_candidates()
        if not _candidates_all:
            st.warning("No hay candidatos registrados. Ve a **➕ Nuevo Candidato** en el menú para agregar uno.")
        else:
            _cand_opts = {f"{c['cedula']} — {c['name']}": c for c in _candidates_all}
            _PLACEHOLDER_CAND = "— Escribe o selecciona un candidato —"
            _cand_sel = st.selectbox(
                "Candidato *",
                [_PLACEHOLDER_CAND] + list(_cand_opts.keys()),
                key="create_cand_sel",
                placeholder="Busca por cédula o nombre...",
            )
            if _cand_sel == _PLACEHOLDER_CAND:
                st.info("Selecciona un candidato para continuar.")
                _cand = None
            else:
                _cand = _cand_opts[_cand_sel]
            if _cand:
                _empresa_str = _cand.get("empresa_codigo") or ""
                _regional_str = _cand.get("regional") or ""
                _cargo_str = _cand.get("position") or "N/A"
                st.caption(f"Cargo: {_cargo_str}  |  Empresa: {_empresa_str}  |  Regional: {_regional_str}")
            # Mostrar pruebas activas del candidato seleccionado
            _cand_active = []
            _pending_types = set()
            if _cand:
                _cand_active = [
                    s for s in _all_raw
                    if s["cedula"] == _cand["cedula"] and s["status"] in ("pending", "in_progress", "employee_done")
                ]
                _pending_types = {s["test_type"] for s in _cand_active}
            if _cand_active:
                _pa_lines = "  \n".join(
                    f"• ⏳ **{_TEST_LABELS_SHARED.get(s['test_type'], s['test_type'].upper())}** — ID: `{s['id']}` — Estado: *{s['status']}*"
                    for s in _cand_active
                )
                st.warning(
                    f"⚠️ Este candidato ya tiene {len(_cand_active)} prueba(s) activa(s). "
                    f"No podrás asignar el mismo tipo nuevamente hasta que finalice:\n\n{_pa_lines}"
                )

            st.markdown("---")
            st.markdown("**👤 Jefe / Evaluador** _(requerido para Desempeño y Período de Prueba)_")

            # Lookup automático por cédula — fuera del formulario para reactividad inmediata
            def _on_change_eval_ced_create():
                _ced_v = st.session_state.get("create_eval_ced", "").strip()
                _found_v = db.get_candidate_by_cedula(_ced_v) if _ced_v else None
                if _found_v:
                    st.session_state["create_eval_nom"] = _found_v["name"]
                else:
                    st.session_state["create_eval_nom"] = ""

            _eval_ced = st.text_input(
                "Cédula del Jefe / Evaluador",
                key="create_eval_ced",
                placeholder="Ingresa la cédula para buscar automáticamente",
                on_change=_on_change_eval_ced_create,
            )
            _jefe_found = db.get_candidate_by_cedula(_eval_ced.strip()) if _eval_ced.strip() else None
            _jefe_ok = True
            if _jefe_found:
                st.success(f"✅ {_jefe_found['name']} | Cargo: {_jefe_found.get('position', 'N/A')}")
                _eval_nom = _jefe_found["name"]
            elif _eval_ced.strip():
                st.error("❌ Cédula no registrada. Regístralo primero en la sección **👥 Candidatos** antes de asignarlo como evaluador.")
                _jefe_ok = False
                _eval_nom = ""
            else:
                _eval_nom = ""

            st.markdown("---")
            with st.form("create_eval_form"):
                _fcol1, _fcol2 = st.columns(2)
                with _fcol1:
                    _test_type = st.selectbox(
                        "Tipo de Evaluación",
                        list(_TEST_LABELS_SHARED.keys()),
                        format_func=lambda x: _TEST_LABELS_SHARED.get(x, x),
                        key="create_test_type",
                    )
                with _fcol2:
                    _time_limit = st.selectbox(
                        "Tiempo Límite",
                        [15, 20, 30, 45, 60, 90],
                        index=3,
                        format_func=lambda x: f"{x} minutos",
                        key="create_time_limit",
                    )
                _create_submitted = st.form_submit_button("✅ Crear Evaluación", type="primary")
                if _create_submitted:
                    _ec_val = st.session_state.get("create_eval_ced", "").strip()
                    if _test_type in _TIPOS_CON_EVALUADOR and not _ec_val:
                        st.error("❌ La cédula del Jefe/Evaluador es obligatoria para este tipo de evaluación.")
                    elif _test_type in _TIPOS_CON_EVALUADOR and not _jefe_ok:
                        st.error("❌ Debes registrar al jefe como candidato antes de continuar.")
                    elif _test_type in _pending_types:
                        st.error(
                            f"❌ El candidato ya tiene una prueba **{_TEST_LABELS_SHARED.get(_test_type, _test_type.upper())}** pendiente o en curso. "
                            f"Completa o elimina la evaluación activa antes de crear una nueva."
                        )
                    else:
                        _ec_save = _ec_val if _test_type in _TIPOS_CON_EVALUADOR else None
                        _en_save = _eval_nom if _eval_nom else None
                        _sid, _serr = db.create_test_session(
                            _cand["id"], _test_type, _time_limit, admin["id"],
                            evaluador_cedula=_ec_save, evaluador_nombre=_en_save,
                        )
                        if _serr:
                            st.warning(f"⚠️ {_serr}")
                        else:
                            _extra = f" | **Evaluador:** {_ec_save}" if _ec_save else ""
                            st.success(f"✅ Evaluación creada!\n\n**ID:** `{_sid}` | **Tipo:** {_test_type.upper()}{_extra}")

    # ----- SECCIÓN: Nuevo Candidato -----
    elif _active_section == "new_candidate":
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        def _nc_border():
            s = Side(border_style="thin", color="BFBFBF")
            return Border(left=s, right=s, top=s, bottom=s)

        def _nc_hdr(ws, r, c, v, w=None):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _nc_border()
            if w:
                ws.column_dimensions[get_column_letter(c)].width = w

        _CIUDADES_CO = [
            "Bogotá", "Medellín", "Cali", "Barranquilla", "Cartagena", "Cúcuta",
            "Bucaramanga", "Pereira", "Santa Marta", "Ibagué", "Pasto", "Manizales",
            "Neiva", "Villavicencio", "Armenia", "Valledupar", "Montería", "Sincelejo",
            "Riohacha", "Tunja", "Popayán", "Florencia", "Quibdó", "Mocoa",
            "Yopal", "Arauca", "Palmira", "Buenaventura", "Bello", "Soledad",
            "Itagüí", "Soacha", "Barrancabermeja", "Dos Quebradas", "Envigado",
            "Floridablanca", "Girón", "Piedecuesta", "Turbo", "Magangué",
            "Maicao", "Pitalito", "Sogamoso", "Duitama", "Zipaquirá", "Chía",
            "Fusagasugá", "Facatativá", "Girardot", "Espinal", "Honda",
            "Cartago", "Tuluá", "Buga", "Ipiales", "Tumaco", "Ocaña", "Pamplona",
            "San José del Guaviare", "Mitú", "Puerto Carreño", "Leticia", "Inírida",
            "Apartadó", "Caucasia", "Rionegro", "Bello", "Sabaneta", "La Estrella",
            "Caldas", "Copacabana", "Girardota", "Barbosa", "Jamundí", "Yumbo",
            "Mosquera", "Madrid", "Funza", "Cajicá", "Sopó", "La Calera",
            "Otra",
        ]
        _EMPRESAS_LIST = ["HESEGO"]

        def _nc_gen_plantilla():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BD empleados"
            ws.sheet_view.showGridLines = False
            ws.freeze_panes = "A2"
            # Todas las columnas son obligatorias
            cols = [
                ("CEDULA", 18), ("EMPRESA", 12), ("REGIONAL", 20),
                ("APELLIDOS Y NOMBRES", 30), ("CORREO", 28),
                ("CARGO", 22), ("INVITAR", 10),
                ("JEFE INMEDIATO", 25), ("NIVEL DE CARGO", 18),
            ]
            ws.row_dimensions[1].height = 32
            for ci, (nm, wd) in enumerate(cols, 1):
                _nc_hdr(ws, 1, ci, nm, wd)
                for r in range(2, 502):
                    cell = ws.cell(row=r, column=ci)
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")  # todo obligatorio
                    cell.border = _nc_border()
            ej = ["1020304050", "HESEGO", "Bogotá", "García López Juan Carlos",
                  "juan@hesego.com", "Analista de Calidad", "SI", "María Rodríguez", "Operativo"]
            for ci, v in enumerate(ej, 1):
                c = ws.cell(row=2, column=ci, value=v)
                c.fill = PatternFill("solid", fgColor="DDEEFF")
                c.font = Font(italic=True, size=10, color="2F5496")
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = _nc_border()
            # Dropdowns
            dv1 = DataValidation(type="list", formula1='"HESEGO"', allow_blank=False, showDropDown=False)
            dv1.sqref = "B2:B501"
            ws.add_data_validation(dv1)
            dv2 = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False, showDropDown=False)
            dv2.sqref = "G2:G501"
            ws.add_data_validation(dv2)
            dv3 = DataValidation(type="list", formula1='"Operativo,Líder,Directivo,Administrativo"', allow_blank=False, showDropDown=False)
            dv3.sqref = "I2:I501"
            ws.add_data_validation(dv3)
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf.read()

        st.markdown("### ➕ Nuevo Candidato")

        # ── Registro uno a uno ────────────────────────────────────────────
        _nc_cedula = st.text_input("Cédula *", placeholder="Número de identificación", key="nc_cedula_lookup")
        _nc_existing = db.get_candidate_by_cedula(_nc_cedula.strip()) if _nc_cedula.strip() else None
        if _nc_existing:
            st.warning(f"⚠️ La cédula **{_nc_cedula.strip()}** ya está registrada como **{_nc_existing['name']}**. Usa 'Crear Evaluación' para asignarle una prueba.")
        else:
            with st.form("new_candidate_form"):
                _ncf1, _ncf2 = st.columns(2)
                with _ncf1:
                    _nc_name = st.text_input("Apellidos y Nombres *", placeholder="García López Juan Carlos")
                    _nc_empresa = st.selectbox("Empresa *", _EMPRESAS_LIST)
                    _nc_regional = st.selectbox("Regional *", _CIUDADES_CO)
                    _nc_cargo = st.text_input("Cargo *", placeholder="Analista de Calidad")
                    _nc_nivel = st.selectbox("Nivel de Cargo *", ["Operativo", "Líder", "Directivo", "Administrativo"])
                with _ncf2:
                    _nc_correo = st.text_input("Correo *", placeholder="juan@empresa.com")
                    _nc_jefe = st.text_input("Jefe Inmediato *", placeholder="Nombre completo del jefe")
                    _nc_invitar = st.selectbox("Invitar a evaluación *", ["SI", "NO"])
                    _nc_sex = st.selectbox("Sexo", ["", "Masculino", "Femenino", "Otro"],
                                          format_func=lambda x: x if x else "— Opcional —")
                    _nc_edu = st.text_input("Nivel Educativo", placeholder="Universitario")
                    _nc_age = st.number_input("Edad", min_value=0, max_value=100, value=0)
                _nc_submit = st.form_submit_button("💾 Guardar Candidato", type="primary")
                if _nc_submit:
                    _nc_errs = []
                    if not _nc_cedula.strip():
                        _nc_errs.append("Cédula")
                    if not _nc_name.strip():
                        _nc_errs.append("Apellidos y Nombres")
                    if not _nc_cargo.strip():
                        _nc_errs.append("Cargo")
                    if not _nc_correo.strip():
                        _nc_errs.append("Correo")
                    if not _nc_jefe.strip():
                        _nc_errs.append("Jefe Inmediato")
                    if _nc_errs:
                        st.error(f"❌ Campos obligatorios faltantes: {', '.join(_nc_errs)}.")
                    else:
                        _nc_result = db.create_empleado(
                            cedula=_nc_cedula.strip(),
                            name=_nc_name.strip(),
                            empresa_codigo=_nc_empresa,
                            regional=_nc_regional,
                            correo=_nc_correo.strip(),
                            position=_nc_cargo.strip(),
                            jefe_inmediato=_nc_jefe.strip(),
                            nivel_cargo=_nc_nivel,
                            invitar=_nc_invitar,
                            age=_nc_age if _nc_age > 0 else None,
                            sex=_nc_sex if _nc_sex else None,
                            education=_nc_edu.strip() if _nc_edu.strip() else None,
                        )
                        if _nc_result:
                            st.success(f"✅ Candidato **{_nc_name.strip()}** (CC: {_nc_cedula.strip()}) registrado correctamente.")
                        else:
                            st.error("❌ Error al guardar. La cédula puede estar duplicada.")

    # ----- SECCIÓN: Candidatos -----
    elif _active_section == "candidates":
        st.markdown("### Candidatos Registrados")
        _cands2 = db.get_all_candidates()
        if not _cands2:
            st.info("No hay candidatos registrados.")
        else:
            _sc1, _sc2 = st.columns([3, 1])
            with _sc1:
                _cand_search = st.text_input(
                    "🔍 Buscar por nombre o cédula",
                    placeholder="Escribe para filtrar...",
                    key="cand_search_filter",
                )
            with _sc2:
                st.markdown("<div style='margin-top:28px'></div>", unsafe_allow_html=True)
                if st.button("✖ Limpiar", key="cand_search_clear", use_container_width=True):
                    st.session_state["cand_search_filter"] = ""
                    st.rerun()
            if _cand_search.strip():
                _q = _cand_search.strip().lower()
                _cands2 = [
                    c for c in _cands2
                    if _q in c.get("name", "").lower() or _q in str(c.get("cedula", "")).lower()
                ]
            st.caption(f"📋 {len(_cands2)} candidato(s) encontrado(s)")
            if not _cands2:
                st.info("No se encontraron candidatos con ese criterio.")
            else:
                # ── Ordenamiento de tabla ──────────────────────────────────
                _csort_col_key = "_csort_col"
                _csort_dir_key = "_csort_dir"
                if _csort_col_key not in st.session_state:
                    st.session_state[_csort_col_key] = "NOMBRE"
                    st.session_state[_csort_dir_key] = "asc"
                _csort_col = st.session_state[_csort_col_key]
                _csort_dir = st.session_state[_csort_dir_key]

                def _cand_sort_val(c, col):
                    if col == "NOMBRE":    return (c.get("name") or "").lower()
                    if col == "CÉDULA":    return str(c.get("cedula") or "")
                    if col == "CARGO":     return (c.get("position") or "").lower()
                    if col == "NIVEL":     return (c.get("nivel_cargo") or "").lower()
                    if col == "REGIONAL":  return (c.get("regional") or "").lower()
                    if col == "EMPRESA":   return (c.get("empresa_codigo") or "").lower()
                    return ""

                _cands2 = sorted(_cands2, key=lambda c: _cand_sort_val(c, _csort_col), reverse=(_csort_dir == "desc"))

                # ── Cabecera de tabla ──────────────────────────────────────
                # Proporciones: nombre, cédula, cargo, nivel, regional, empresa, acciones
                _CW = [2.5, 1.5, 2.0, 1.4, 1.5, 1.2, 0.45]
                _ch = st.columns(_CW)
                for _cidx, _cname in [(0, "NOMBRE"), (1, "CÉDULA"), (2, "CARGO"), (3, "NIVEL"), (4, "REGIONAL"), (5, "EMPRESA")]:
                    _arrow = " ↓" if (_csort_col == _cname and _csort_dir == "desc") else (
                             " ↑" if _csort_col == _cname else " ↕")
                    if _ch[_cidx].button(f"{_cname}{_arrow}", key=f"chdr_{_cname}", use_container_width=True):
                        if _csort_col == _cname:
                            st.session_state[_csort_dir_key] = "asc" if _csort_dir == "desc" else "desc"
                        else:
                            st.session_state[_csort_col_key] = _cname
                            st.session_state[_csort_dir_key] = "asc"
                        st.rerun()
                _ch[6].markdown("<div style='height:30px'></div>", unsafe_allow_html=True)

                st.markdown("---")

                # ── Filas de tabla ─────────────────────────────────────────
                for c in _cands2:
                    _toggle_key = f"cand_open_{c['id']}"
                    _is_open = st.session_state.get(_toggle_key, False)
                    _csessions = _sessions_by_cedula.get(c["cedula"], [])
                    _n_evals = len(_csessions)

                    rc = st.columns(_CW)
                    rc[0].markdown(
                        f"<div style='font-weight:600;padding-top:5px;font-size:13px'>👤 {c['name']}</div>",
                        unsafe_allow_html=True,
                    )
                    rc[1].markdown(
                        f"<div style='font-size:12px;font-family:monospace;padding-top:6px'>{c.get('cedula','—')}</div>",
                        unsafe_allow_html=True,
                    )
                    rc[2].markdown(
                        f"<div style='font-size:12px;padding-top:6px'>{c.get('position','—')}</div>",
                        unsafe_allow_html=True,
                    )
                    rc[3].markdown(
                        f"<div style='font-size:12px;padding-top:6px'>{c.get('nivel_cargo','—')}</div>",
                        unsafe_allow_html=True,
                    )
                    rc[4].markdown(
                        f"<div style='font-size:12px;padding-top:6px'>{c.get('regional','—')}</div>",
                        unsafe_allow_html=True,
                    )
                    rc[5].markdown(
                        f"<div style='font-size:12px;padding-top:6px'>{c.get('empresa_codigo','—')}</div>",
                        unsafe_allow_html=True,
                    )
                    if rc[6].button("▲" if _is_open else "▼", key=f"ctog_{c['id']}", use_container_width=True):
                        st.session_state[_toggle_key] = not _is_open
                        st.rerun()

                    if _is_open:
                        with st.container(border=True):
                            _di1, _di2, _di3, _di4 = st.columns(4)
                            _di1.metric("Edad", c.get("age") or "—")
                            _di2.metric("Sexo", c.get("sex") or "—")
                            _di3.metric("Educación", c.get("education") or "—")
                            _di4.metric("Evaluaciones", _n_evals)
                            if c.get("correo"):    st.caption(f"📧 {c['correo']}")
                            if c.get("jefe_inmediato"): st.caption(f"👔 Jefe: {c['jefe_inmediato']}")
                            if _csessions:
                                st.markdown("**Evaluaciones:**")
                                for s in _csessions:
                                    _se = {"pending": "⏳", "in_progress": "▶️", "completed": "✅", "expired": "⏰", "employee_done": "📝"}.get(s["status"], "❓")
                                    _fp = ""
                                    _dr = s.get("completed_at") or s.get("started_at")
                                    if _dr:
                                        try: _fp = f" — {datetime.strptime(_dr, '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y %H:%M')}"
                                        except: _fp = f" — {_dr}"
                                    st.markdown(f"  - {_se} {s['test_type'].upper()} (ID: `{s['id']}`) — {s['status']}{_fp}")

                            # Botón editar
                            _edit_key = f"edit_cand_{c['id']}"
                            if st.button("✏️ Editar datos", key=f"edit_btn_{c['id']}", use_container_width=False):
                                st.session_state[_edit_key] = not st.session_state.get(_edit_key, False)
                                st.rerun()

                            if st.session_state.get(_edit_key, False):
                                st.markdown("---")
                                st.markdown("**✏️ Editar datos del candidato**")
                                with st.form(key=f"form_edit_cand_{c['id']}"):
                                    st.text_input("Cédula", value=c.get("cedula", ""), disabled=True)
                                    _ef1, _ef2 = st.columns(2)
                                    _e_name  = _ef1.text_input("Apellidos y Nombres", value=c.get("name", ""))
                                    _e_cargo = _ef2.text_input("Cargo", value=c.get("position", ""))
                                    _ef3, _ef4, _ef5 = st.columns(3)
                                    _e_age   = _ef3.number_input("Edad", min_value=0, max_value=100, value=int(c.get("age") or 0))
                                    _e_sex   = _ef4.selectbox("Sexo", ["", "Masculino", "Femenino", "Otro"],
                                                              index=["", "Masculino", "Femenino", "Otro"].index(c.get("sex") or "") if (c.get("sex") or "") in ["", "Masculino", "Femenino", "Otro"] else 0)
                                    _e_edu   = _ef5.text_input("Educación", value=c.get("education", "") or "")
                                    _ef6, _ef7 = st.columns(2)
                                    _e_correo = _ef6.text_input("Correo", value=c.get("correo", "") or "")
                                    _e_jefe   = _ef7.text_input("Jefe Inmediato", value=c.get("jefe_inmediato", "") or "")
                                    _ef8, _ef9 = st.columns(2)
                                    _nc_opts = ["Operativo", "Líder", "Directivo", "Administrativo"]
                                    _nc_cur = c.get("nivel_cargo") or "Operativo"
                                    _e_nivel = _ef8.selectbox("Nivel de Cargo", _nc_opts,
                                                              index=_nc_opts.index(_nc_cur) if _nc_cur in _nc_opts else 0)
                                    _e_regional = _ef9.text_input("Regional", value=c.get("regional", "") or "")
                                    _es1, _es2, _ = st.columns([1, 1, 2])
                                    _edit_ok = _es1.form_submit_button("💾 Guardar", use_container_width=True)
                                    _edit_cancel = _es2.form_submit_button("✖ Cancelar", use_container_width=True)
                                    if _edit_ok:
                                        db.update_empleado(
                                            candidate_id=c["id"],
                                            name=_e_name.strip(),
                                            age=_e_age if _e_age > 0 else None,
                                            sex=_e_sex if _e_sex else None,
                                            education=_e_edu.strip() or None,
                                            position=_e_cargo.strip(),
                                            correo=_e_correo.strip() or None,
                                            jefe_inmediato=_e_jefe.strip() or None,
                                            nivel_cargo=_e_nivel,
                                            regional=_e_regional.strip() or None,
                                        )
                                        st.session_state.pop(_edit_key, None)
                                        st.success(f"✅ Candidato **{_e_name.strip()}** actualizado.")
                                        st.rerun()
                                    if _edit_cancel:
                                        st.session_state.pop(_edit_key, None)
                                        st.rerun()

                            if admin.get("role") == "superadmin":
                                st.markdown("---")
                                if st.button(f"🗑️ Eliminar candidato", key=f"del_cand_{c['id']}"):
                                    st.session_state[f"confirm_del_cand_{c['id']}"] = True
                                if st.session_state.get(f"confirm_del_cand_{c['id']}", False):
                                    _ns = len(_csessions)
                                    st.warning(f"⚠️ ¿Eliminar **{c['name']}** (CC: {c['cedula']})? Se eliminarán {_ns} evaluación(es). Irreversible.")
                                    _cy, _cn, _ = st.columns([1, 1, 2])
                                    with _cy:
                                        if st.button("✅ Sí, eliminar", key=f"confirm_yes_cand_{c['id']}"):
                                            db.delete_candidate(c['id'])
                                            st.session_state.pop(f"confirm_del_cand_{c['id']}", None)
                                            st.success(f"Candidato **{c['name']}** eliminado.")
                                            st.rerun()
                                    with _cn:
                                        if st.button("❌ Cancelar", key=f"confirm_no_cand_{c['id']}"):
                                            st.session_state.pop(f"confirm_del_cand_{c['id']}", None)
                                            st.rerun()
                    st.divider()

    # ----- SECCIÓN: Dashboard Gerencial -----
    if _active_section == "dashboard":
        st.markdown("## 📈 Dashboard Gerencial de Evaluaciones")
        st.caption("Resumen ejecutivo de pruebas realizadas — actualizado en tiempo real.")
        st.markdown("---")

        _all_completed = [s for s in _all_raw if s["status"] in ("completed", "expired")]
        _all_pending   = [s for s in _all_raw if s["status"] in ("pending", "in_progress", "employee_done")]
        _total         = len(_all_raw)
        _pct_done = int(len(_all_completed) / _total * 100) if _total else 0

        # ── KPIs principales ──────────────────────────────────────────────
        _k1, _k2, _k3, _k4 = st.columns(4)
        _k1.metric("📋 Total Evaluaciones", _total)
        _k2.metric("✅ Completadas", len(_all_completed), f"{_pct_done}% del total")
        _k3.metric("⏳ Pendientes / En curso", len(_all_pending))
        _k4.metric("👥 Candidatos activos", len(set(s["candidate_id"] for s in _all_raw)))

        st.markdown("---")

        # ── Filtros del dashboard ─────────────────────────────────────────
        _TEST_LABELS = {
            "disc": "🎯 DISC", "valanti": "🧭 VALANTI", "wpi": "💼 WPI",
            "eri": "🔐 ERI", "talent_map": "🌟 Talent Map",
            "desempeno": "📊 Desempeño Op.", "desempeno_lider": "📊 Desempeño Líd.",
            "periodo_prueba": "📋 Per. Prueba",
        }
        from collections import Counter, defaultdict

        # Fechas disponibles
        _all_dates = []
        for _s in _all_completed:
            _dr = _s.get("completed_at") or _s.get("started_at") or _s.get("created_at") or ""
            try:
                _all_dates.append(datetime.strptime(_dr, "%Y-%m-%d %H:%M:%S").date())
            except:
                pass

        _min_date = min(_all_dates) if _all_dates else datetime.today().date()
        _max_date = max(_all_dates) if _all_dates else datetime.today().date()

        # Cargos disponibles
        _all_cargos = sorted(set(
            (s.get("position") or "Sin cargo").strip() or "Sin cargo"
            for s in _all_completed
        ))

        with st.expander("🔎 Filtros del Dashboard", expanded=False):
            _filt_col1, _filt_col2, _filt_col3 = st.columns(3)
            with _filt_col1:
                _f_desde = st.date_input("Desde", value=_min_date, key="db_f_desde")
                _f_hasta = st.date_input("Hasta", value=_max_date, key="db_f_hasta")
            with _filt_col2:
                _f_tipos = st.multiselect(
                    "Tipo de prueba",
                    options=list(_TEST_LABELS.keys()),
                    default=[],
                    format_func=lambda x: _TEST_LABELS.get(x, x),
                    key="db_f_tipos",
                )
            with _filt_col3:
                _f_cargos = st.multiselect(
                    "Cargo",
                    options=_all_cargos,
                    default=[],
                    key="db_f_cargos",
                )

        # Aplicar filtros a completadas
        def _dash_date(s):
            _dr = s.get("completed_at") or s.get("started_at") or s.get("created_at") or ""
            try:
                return datetime.strptime(_dr, "%Y-%m-%d %H:%M:%S").date()
            except:
                return None

        _comp_filt = _all_completed
        if _f_desde or _f_hasta:
            _comp_filt = [s for s in _comp_filt
                          if _dash_date(s) and _f_desde <= _dash_date(s) <= _f_hasta]
        if _f_tipos:
            _comp_filt = [s for s in _comp_filt if s["test_type"] in _f_tipos]
        if _f_cargos:
            _comp_filt = [s for s in _comp_filt
                          if ((s.get("position") or "Sin cargo").strip() or "Sin cargo") in _f_cargos]

        if len(_comp_filt) < len(_all_completed):
            st.caption(f"🔎 Mostrando **{len(_comp_filt)}** de {len(_all_completed)} evaluaciones completadas según filtros.")

        st.markdown("---")

        # ── Distribución por tipo de prueba (tabla) ───────────────────────
        _counts_total     = Counter(s["test_type"] for s in _all_raw)
        _counts_completed = Counter(s["test_type"] for s in _comp_filt)
        _counts_pending   = Counter(s["test_type"] for s in _all_pending)

        st.markdown("### 🔢 Pruebas por Tipo")
        _tipo_data = []
        for _tt, _lbl in _TEST_LABELS.items():
            _tot = _counts_total.get(_tt, 0)
            if _tot == 0:
                continue
            _done = _counts_completed.get(_tt, 0)
            _pend = _counts_pending.get(_tt, 0)
            _pct  = int(_done / _tot * 100) if _tot else 0
            _tipo_data.append((_lbl, _tot, _done, _pend, _pct))

        if _tipo_data:
            _th1, _th2, _th3, _th4, _th5 = st.columns([2.2, 0.9, 0.9, 0.9, 3])
            for _hdr_txt, _hdr_col in [("Tipo de Prueba", _th1), ("Total", _th2),
                                        ("Completadas", _th3), ("Pendientes", _th4), ("Progreso", _th5)]:
                _hdr_col.markdown(
                    f"<div style='font-size:11px;font-weight:700;color:#aaa;"
                    f"border-bottom:1px solid #444;padding-bottom:3px'>{_hdr_txt}</div>",
                    unsafe_allow_html=True,
                )
            for _lbl, _tot, _done, _pend, _pct in sorted(_tipo_data, key=lambda x: -x[1]):
                _c1, _c2, _c3, _c4, _c5 = st.columns([2.2, 0.9, 0.9, 0.9, 3])
                _c1.markdown(f"<div style='padding-top:6px;font-size:13px'>{_lbl}</div>", unsafe_allow_html=True)
                _c2.markdown(f"<div style='padding-top:6px;text-align:center;font-weight:600'>{_tot}</div>", unsafe_allow_html=True)
                _c3.markdown(f"<div style='padding-top:6px;text-align:center;color:#10B981;font-weight:600'>{_done}</div>", unsafe_allow_html=True)
                _c4.markdown(f"<div style='padding-top:6px;text-align:center;color:#F59E0B;font-weight:600'>{_pend}</div>", unsafe_allow_html=True)
                _bar_fill = "#10B981" if _pct >= 75 else ("#F59E0B" if _pct >= 40 else "#EF4444")
                _c5.markdown(
                    f"<div style='background:rgba(128,128,128,0.15);border-radius:6px;height:18px;margin-top:7px'>"
                    f"<div style='background:{_bar_fill};width:{_pct}%;height:18px;border-radius:6px;"
                    f"display:flex;align-items:center;justify-content:center;"
                    f"font-size:10px;color:white;font-weight:700'>{_pct}%</div></div>",
                    unsafe_allow_html=True,
                )
        else:
            st.info("Aún no hay evaluaciones registradas.")

        st.markdown("---")

        # ── Gráficos: Mensual | Por tipo | Por cargo (Plotly interactivo) ──
        st.markdown("### 📅 Distribución de Evaluaciones")
        import plotly.graph_objects as _pgo
        from collections import defaultdict

        # Detectar tema claro/oscuro
        try:
            _theme_base = st.get_option("theme.base") or "light"
        except Exception:
            _theme_base = "light"
        _ui_mode = st.session_state.get("ui_theme_mode", "Sistema")
        if _ui_mode == "Oscuro":
            _dash_dark = True
        elif _ui_mode == "Claro":
            _dash_dark = False
        else:
            _dash_dark = (_theme_base == "dark")
        _fc = "#FFFFFF" if _dash_dark else "#1E293B"
        _gc = "rgba(255,255,255,0.15)" if _dash_dark else "rgba(0,0,0,0.08)"
        _plotly_tpl = "plotly_dark" if _dash_dark else "plotly_white"

        _FIG_H = 380
        _gcol1, _gcol2, _gcol3 = st.columns(3)

        # --- Gráfico 1: Completadas por mes ---
        _mes_counts = defaultdict(int)
        for s in _comp_filt:
            _date_ref = s.get("completed_at") or s.get("started_at") or s.get("created_at") or ""
            try:
                _mes = datetime.strptime(_date_ref, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m")
                _mes_counts[_mes] += 1
            except:
                pass
        with _gcol1:
            st.markdown("**📅 Completadas por Mes**")
            if _mes_counts:
                _meses = sorted(_mes_counts.keys())
                _vals_m = [_mes_counts[m] for m in _meses]
                _meses_lbl = [datetime.strptime(m, "%Y-%m").strftime("%b %Y") for m in _meses]
                _fig_m = _pgo.Figure(_pgo.Bar(
                    x=_meses_lbl, y=_vals_m,
                    marker_color="#3B82F6",
                    text=_vals_m, textposition="outside",
                    textfont=dict(color=_fc),
                    hovertemplate="<b>%{x}</b><br>Completadas: %{y}<extra></extra>",
                ))
                _fig_m.update_layout(
                    template=_plotly_tpl,
                    height=_FIG_H, margin=dict(l=20, r=20, t=20, b=50),
                    plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                    font=dict(color=_fc),
                    xaxis=dict(tickangle=-30, tickfont=dict(size=10, color=_fc),
                               gridcolor=_gc, linecolor=_gc),
                    yaxis=dict(title="Cantidad", tickfont=dict(color=_fc),
                               gridcolor=_gc, linecolor=_gc),
                    showlegend=False,
                )
                st.plotly_chart(_fig_m, use_container_width=True)
            else:
                st.info("Sin datos para los filtros seleccionados.")

        # --- Gráfico 2: Por tipo de prueba (pie interactivo) ---
        with _gcol2:
            st.markdown("**🔢 Completadas por Tipo de Prueba**")
            _tipo_done_filt = {_TEST_LABELS[t].split(" ", 1)[-1]: _counts_completed.get(t, 0)
                               for t in _TEST_LABELS if _counts_completed.get(t, 0) > 0}
            if _tipo_done_filt:
                _tipo_colors = ["#3B82F6", "#10B981", "#8B5CF6", "#EF4444",
                                 "#F59E0B", "#06B6D4", "#EC4899", "#84CC16"]
                _fig_t = _pgo.Figure(_pgo.Pie(
                    labels=list(_tipo_done_filt.keys()),
                    values=list(_tipo_done_filt.values()),
                    marker=dict(colors=_tipo_colors[:len(_tipo_done_filt)],
                                line=dict(color="white", width=2)),
                    textinfo="label+value",
                    textfont=dict(color=_fc),
                    hovertemplate="<b>%{label}</b><br>Cantidad: %{value}<br>%{percent}<extra></extra>",
                    hole=0.0,
                ))
                _fig_t.update_layout(
                    template=_plotly_tpl,
                    height=_FIG_H, margin=dict(l=10, r=10, t=20, b=10),
                    paper_bgcolor="rgba(0,0,0,0)",
                    font=dict(color=_fc),
                    legend=dict(orientation="v", x=1.02, y=0.5,
                                font=dict(size=10, color=_fc)),
                    showlegend=True,
                )
                st.plotly_chart(_fig_t, use_container_width=True)
            else:
                st.info("Sin evaluaciones completadas con los filtros actuales.")

        # --- Gráfico 3: Por cargo (barras horizontales interactivas) ---
        with _gcol3:
            st.markdown("**💼 Completadas por Cargo**")
            _cargo_counts = defaultdict(int)
            for s in _comp_filt:
                _cargo = (s.get("position") or "Sin cargo").strip() or "Sin cargo"
                _cargo_counts[_cargo] += 1
            if _cargo_counts:
                _top_cargos = sorted(_cargo_counts.items(), key=lambda x: x[1])[-8:]
                _cargo_lbl  = [c[:25] for c, _ in _top_cargos]
                _cargo_vals = [v for _, v in _top_cargos]
                _cargo_colors_c = ["#6366F1" if i % 2 == 0 else "#818CF8" for i in range(len(_cargo_lbl))]
                _fig_c = _pgo.Figure(_pgo.Bar(
                    x=_cargo_vals, y=_cargo_lbl,
                    orientation="h",
                    marker_color=_cargo_colors_c,
                    text=_cargo_vals, textposition="outside",
                    textfont=dict(color=_fc),
                    hovertemplate="<b>%{y}</b><br>Completadas: %{x}<extra></extra>",
                ))
                _fig_c.update_layout(
                    template=_plotly_tpl,
                    height=_FIG_H, margin=dict(l=20, r=40, t=20, b=30),
                    plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                    font=dict(color=_fc),
                    xaxis=dict(title="Cantidad", tickfont=dict(color=_fc),
                               gridcolor=_gc, linecolor=_gc),
                    yaxis=dict(tickfont=dict(size=9, color=_fc)),
                    showlegend=False,
                )
                st.plotly_chart(_fig_c, use_container_width=True)
            else:
                st.info("Sin datos de cargo para los filtros seleccionados.")

        st.markdown("---")

        # ── Top 10 candidatos con más evaluaciones ─────────────────────────
        st.markdown("### 🏆 Top Candidatos con Más Evaluaciones")
        _cand_counts = Counter(s["candidate_name"] for s in _all_raw)
        _top_cands = _cand_counts.most_common(10)
        if _top_cands:
            _tc1, _tc2, _tc3 = st.columns([0.4, 0.8, 0.8])
            for _rk, (_cname, _cnt) in enumerate(_top_cands, 1):
                _done_cand = sum(1 for s in _all_completed if s["candidate_name"] == _cname)
                _medal = "🥇" if _rk == 1 else ("🥈" if _rk == 2 else ("🥉" if _rk == 3 else f"{_rk}."))
                st.markdown(
                    f"<div style='display:flex;align-items:center;gap:12px;padding:6px 8px;"
                    f"background:{'#1E3A5F11' if _rk % 2 == 0 else 'transparent'};"
                    f"border-radius:6px;margin-bottom:3px'>"
                    f"<span style='font-size:16px;min-width:28px'>{_medal}</span>"
                    f"<span style='font-weight:600;flex:1'>{_cname}</span>"
                    f"<span style='color:#3B82F6;font-weight:700;min-width:60px;text-align:right'>"
                    f"{_cnt} total</span>"
                    f"<span style='color:#10B981;min-width:80px;text-align:right'>"
                    f"{_done_cand} completadas</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

    # ----- SECCIÓN: Resultados -----
    elif _active_section == "results":
        st.markdown("### Resultados de Evaluaciones")
        _res3 = [
            s for s in _all_raw
            if (_ft is None or s["test_type"] == _ft) and s["status"] in ("completed", "expired")
        ]
        if _filter_cand != "Todos":
            _res3 = [s for s in _res3 if s["candidate_name"] == _filter_cand]
        _sort_sessions(_res3)
        _render_sessions_list(_res3, "res")

    # ----- SECCIÓN: Pruebas Pendientes -----
    elif _active_section == "pending":
        st.markdown("### Pruebas Pendientes")
        _pend4 = [
            s for s in _all_raw
            if (_ft is None or s["test_type"] == _ft) and s["status"] in ("pending", "in_progress", "employee_done")
        ]
        if _filter_cand != "Todos":
            _pend4 = [s for s in _pend4 if s["candidate_name"] == _filter_cand]
        _sort_sessions(_pend4)
        _render_sessions_list(_pend4, "pend")

    # ----- SECCIÓN: Cargue Masivo -----
    elif _active_section == "bulk":
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        _BULK_THIN = Side(border_style="thin", color="BFBFBF")

        def _bulk_border():
            s = _BULK_THIN
            return Border(left=s, right=s, top=s, bottom=s)

        def _bulk_hdr(ws, r, c, v, w=None):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _bulk_border()
            if w:
                ws.column_dimensions[get_column_letter(c)].width = w

        def _bulk_data(ws, r, c, v="", bg="FFFFFF"):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = _bulk_border()

        def _gen_plantilla_candidatos():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BD empleados"
            ws.sheet_view.showGridLines = False
            ws.freeze_panes = "A2"
            cols = [
                ("CEDULA", 18, True), ("EMPRESA", 12, True), ("REGIONAL", 15, False),
                ("APELLIDOS Y NOMBRES", 30, True), ("CORREO", 28, False),
                ("CARGO", 22, False), ("INVITAR", 10, False),
                ("JEFE INMEDIATO", 25, False), ("NIVEL DE CARGO", 18, False),
            ]
            ws.row_dimensions[1].height = 32
            for ci, (nm, wd, req) in enumerate(cols, 1):
                _bulk_hdr(ws, 1, ci, nm, wd)
                bg = "FFF2CC" if req else "E2EFDA"
                for r in range(2, 502):
                    _bulk_data(ws, r, ci, bg=bg)
            # Fila ejemplo
            ej = ["1020304050","HESEGO","Bogotá","García López Juan Carlos",
                  "juan@hesego.com","Analista","SI","María Rodríguez","Operativo"]
            for ci, v in enumerate(ej, 1):
                c = ws.cell(row=2, column=ci, value=v)
                c.fill = PatternFill("solid", fgColor="DDEEFF")
                c.font = Font(italic=True, size=10, color="2F5496")
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = _bulk_border()
            # Validaciones
            dv1 = DataValidation(type="list", formula1='"HESEGO"', allow_blank=True, showDropDown=False)
            dv1.sqref = "B2:B501"
            ws.add_data_validation(dv1)
            dv2 = DataValidation(type="list", formula1='"SI,NO"', allow_blank=True, showDropDown=False)
            dv2.sqref = "G2:G501"
            ws.add_data_validation(dv2)
            dv3 = DataValidation(type="list", formula1='"Operativo,Líder,Directivo,Administrativo"', allow_blank=True, showDropDown=False)
            dv3.sqref = "I2:I501"
            ws.add_data_validation(dv3)
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf.read()

        def _gen_plantilla_pruebas():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Pruebas"
            ws.sheet_view.showGridLines = False
            ws.freeze_panes = "A2"
            cols = [
                ("CEDULA_CANDIDATO", 18, True), ("TIPO_PRUEBA", 22, True),
                ("TIEMPO_LIMITE_MINUTOS", 22, False),
                ("CEDULA_EVALUADOR", 18, False), ("NOMBRE_EVALUADOR", 30, False),
            ]
            ws.row_dimensions[1].height = 32
            for ci, (nm, wd, req) in enumerate(cols, 1):
                _bulk_hdr(ws, 1, ci, nm, wd)
                bg = "FFF2CC" if req else "E2EFDA"
                for r in range(2, 502):
                    _bulk_data(ws, r, ci, bg=bg)
            ejemplos = [
                ("1020304050","disc","30","",""),
                ("1020304051","wpi","30","",""),
                ("1020304052","desempeno_lider","60","9876543210","María Rodríguez"),
                ("1020304053","periodo_prueba","60","9876543210","María Rodríguez"),
            ]
            for ri, vals in enumerate(ejemplos):
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(row=2+ri, column=ci, value=v)
                    c.fill = PatternFill("solid", fgColor="DDEEFF")
                    c.font = Font(italic=True, size=10, color="2F5496")
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    c.border = _bulk_border()
            tipos = "disc,valanti,wpi,eri,talent_map,desempeno,desempeno_lider,periodo_prueba"
            dv = DataValidation(type="list", formula1=f'"{tipos}"', allow_blank=False, showDropDown=False)
            dv.sqref = "B2:B501"
            ws.add_data_validation(dv)
            dv_t = DataValidation(type="whole", operator="between", formula1="5", formula2="180", allow_blank=True)
            dv_t.sqref = "C2:C501"
            ws.add_data_validation(dv_t)
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf.read()

        st.markdown("### 📤 Cargue Masivo")

        # ── Descarga de plantillas ────────────────────────────────────────────
        st.markdown("#### 📥 Descargar plantillas de ejemplo")
        _dc1, _dc2 = st.columns(2)
        with _dc1:
            st.download_button(
                label="⬇️ Plantilla Candidatos (.xlsx)",
                data=_gen_plantilla_candidatos(),
                file_name="plantilla_cargue_candidatos.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.caption("Columnas: CEDULA, EMPRESA, REGIONAL, APELLIDOS Y NOMBRES, CORREO, CARGO, INVITAR, JEFE INMEDIATO, NIVEL DE CARGO")
        with _dc2:
            st.download_button(
                label="⬇️ Plantilla Pruebas (.xlsx)",
                data=_gen_plantilla_pruebas(),
                file_name="plantilla_cargue_pruebas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.caption("Columnas: CEDULA_CANDIDATO, TIPO_PRUEBA, TIEMPO_LIMITE_MINUTOS, CEDULA_EVALUADOR, NOMBRE_EVALUADOR")

        st.markdown("---")

        # ── Cargue masivo de candidatos ───────────────────────────────────────
        st.markdown("#### 👥 Cargar candidatos desde Excel")
        _uploaded_cands = st.file_uploader(
            "Selecciona el Excel de candidatos (hoja: BD empleados)",
            type=["xlsx"],
            key="bulk_upload_cands",
        )
        if _uploaded_cands:
            try:
                _wb_c = openpyxl.load_workbook(_uploaded_cands, data_only=True)
                if "BD empleados" not in _wb_c.sheetnames:
                    st.error("❌ El archivo no tiene una hoja llamada 'BD empleados'.")
                else:
                    _ws_c = _wb_c["BD empleados"]
                    _ok_c, _dup_c, _err_c = 0, 0, []
                    for _ri in range(2, _ws_c.max_row + 1):
                        _cedula = _ws_c.cell(_ri, 1).value
                        _emp_cod = _ws_c.cell(_ri, 2).value
                        _regional = _ws_c.cell(_ri, 3).value
                        _nombre = _ws_c.cell(_ri, 4).value
                        _correo = _ws_c.cell(_ri, 5).value
                        _cargo = _ws_c.cell(_ri, 6).value
                        _invitar = _ws_c.cell(_ri, 7).value
                        _jefe = _ws_c.cell(_ri, 8).value
                        _nivel = _ws_c.cell(_ri, 9).value
                        if not _cedula or not _nombre:
                            continue
                        try:
                            _cedula = str(int(_cedula)) if isinstance(_cedula, float) else str(_cedula).strip()
                            _res = db.create_empleado(
                                cedula=_cedula,
                                name=str(_nombre).strip(),
                                empresa_codigo=str(_emp_cod).strip() if _emp_cod else "",
                                regional=str(_regional).strip() if _regional else "",
                                correo=str(_correo).strip() if _correo else "",
                                position=str(_cargo).strip() if _cargo else "",
                                jefe_inmediato=str(_jefe).strip() if _jefe and str(_jefe) != "#N/A" else "",
                                nivel_cargo=str(_nivel).strip() if _nivel else "",
                                invitar=str(_invitar).strip().upper() if _invitar else "SI",
                            )
                            if _res:
                                _ok_c += 1
                            else:
                                _dup_c += 1
                        except Exception as _e:
                            _err_c.append(f"Fila {_ri}: {_e}")
                    if _ok_c > 0:
                        st.success(f"✅ {_ok_c} candidato(s) importado(s) correctamente.")
                    if _dup_c > 0:
                        st.warning(f"⚠️ {_dup_c} candidato(s) omitido(s) por cédula duplicada.")
                    for _em in _err_c[:5]:
                        st.error(f"❌ {_em}")
            except Exception as _ex:
                st.error(f"❌ Error al leer el archivo: {_ex}")

        st.markdown("---")

        # ── Cargue masivo de pruebas ──────────────────────────────────────────
        st.markdown("#### 🧪 Cargar pruebas desde Excel")
        _TIEMPOS_DEFAULT = {
            "disc": 30, "valanti": 30, "wpi": 30,
            "eri": 20, "talent_map": 25,
            "desempeno": 60, "desempeno_lider": 60, "periodo_prueba": 60,
        }
        _TIPOS_VALIDOS = set(_TIEMPOS_DEFAULT.keys())
        _uploaded_tests = st.file_uploader(
            "Selecciona el Excel de pruebas (hoja: Pruebas)",
            type=["xlsx"],
            key="bulk_upload_tests",
        )
        if _uploaded_tests:
            try:
                _wb_t = openpyxl.load_workbook(_uploaded_tests, data_only=True)
                if "Pruebas" not in _wb_t.sheetnames:
                    st.error("❌ El archivo no tiene una hoja llamada 'Pruebas'.")
                else:
                    _ws_t = _wb_t["Pruebas"]
                    _ok_t, _err_t, _warn_t = 0, [], []
                    for _ri in range(2, _ws_t.max_row + 1):
                        _ced_t = _ws_t.cell(_ri, 1).value
                        _tipo = _ws_t.cell(_ri, 2).value
                        _tiempo = _ws_t.cell(_ri, 3).value
                        _eval_ced_bulk = _ws_t.cell(_ri, 4).value
                        _eval_nom_bulk = _ws_t.cell(_ri, 5).value
                        # Omitir filas completamente vacías
                        if not _ced_t and not _tipo:
                            continue
                        # Validar campos obligatorios
                        if not _ced_t:
                            _err_t.append(f"Fila {_ri}: falta CEDULA_CANDIDATO.")
                            continue
                        if not _tipo:
                            _err_t.append(f"Fila {_ri}: falta TIPO_PRUEBA.")
                            continue
                        try:
                            _ced_t = str(int(_ced_t)) if isinstance(_ced_t, float) else str(_ced_t).strip()
                            _tipo = str(_tipo).strip().lower()
                            if _tipo not in _TIPOS_VALIDOS:
                                _err_t.append(f"Fila {_ri}: tipo '{_tipo}' no válido. Opciones: {', '.join(_TIPOS_VALIDOS)}")
                                continue
                            # Validar que el candidato exista
                            _cand_t = db.get_candidate_by_cedula(_ced_t)
                            if not _cand_t:
                                _err_t.append(f"Fila {_ri}: candidato con cédula **{_ced_t}** no existe en el sistema.")
                                continue
                            # Tiempo
                            _tl = int(_tiempo) if _tiempo else _TIEMPOS_DEFAULT[_tipo]
                            if _tl < 5 or _tl > 180:
                                _err_t.append(f"Fila {_ri}: tiempo {_tl} fuera de rango (5–180 min).")
                                continue
                            # Evaluador
                            _ec = str(_eval_ced_bulk).strip() if _eval_ced_bulk else None
                            _en = str(_eval_nom_bulk).strip() if _eval_nom_bulk else None
                            if _tipo in ("desempeno", "desempeno_lider", "periodo_prueba") and not _ec:
                                _err_t.append(f"Fila {_ri} ({_ced_t}): '{_tipo}' requiere CEDULA_EVALUADOR.")
                                continue
                            # Lookup automático del nombre del evaluador si no fue especificado
                            if _ec and not _en:
                                _jefe_bulk = db.get_candidate_by_cedula(_ec)
                                if _jefe_bulk:
                                    _en = _jefe_bulk["name"]
                                else:
                                    _warn_t.append(f"Fila {_ri}: evaluador cédula {_ec} no registrado — se guardó sin nombre.")
                            _sid, _serr = db.create_test_session(
                                _cand_t["id"], _tipo, _tl, admin["id"],
                                evaluador_cedula=_ec, evaluador_nombre=_en,
                            )
                            if _serr:
                                _err_t.append(f"Fila {_ri} ({_cand_t['name']}): {_serr}")
                            else:
                                _ok_t += 1
                        except Exception as _e:
                            _err_t.append(f"Fila {_ri}: {_e}")
                    if _ok_t > 0:
                        st.success(f"✅ {_ok_t} prueba(s) creada(s) correctamente.")
                    for _wm in _warn_t:
                        st.warning(f"⚠️ {_wm}")
                    for _em in _err_t[:15]:
                        st.error(f"❌ {_em}")
                    if len(_err_t) > 15:
                        st.error(f"... y {len(_err_t) - 15} error(es) más. Corrige el archivo y vuelve a subirlo.")
            except Exception as _ex:
                st.error(f"❌ Error al leer el archivo: {_ex}")

    # ----- SECCIÓN: Configuración -----
    elif _active_section == "settings":
        st.markdown("### Cambiar Contraseña de Administrador")
        with st.form("change_pw"):
            new_pw = st.text_input("Nueva Contraseña", type="password")
            confirm_pw = st.text_input("Confirmar Contraseña", type="password")
            if st.form_submit_button("Cambiar Contraseña"):
                if new_pw and new_pw == confirm_pw:
                    db.change_admin_password(admin["id"], new_pw)
                    st.success("✅ Contraseña actualizada.")
                else:
                    st.error("Las contraseñas no coinciden o están vacías.")


def show_disc_results_admin(results, candidate, session):
    """Show DISC results in the admin panel."""
    normalized = results.get("normalized", {})
    relative = results.get("relative", {})

    # Análisis de aptitud
    analysis = analyze_disc_aptitude(normalized, relative)

    # Estilos conductuales, temperamento y mega resumen (nuevas funcionalidades THT-inspired)
    behavioral_styles = calculate_behavioral_styles(normalized)
    temperament = get_disc_temperament(normalized)
    mega_summary = generate_disc_mega_summary(normalized)

    # Banner de aptitud
    st.markdown(f"""
    <div style="background: {analysis['aptitude_color']}22; border-left: 5px solid {analysis['aptitude_color']};
                padding: 15px 20px; border-radius: 8px; margin-bottom: 15px;">
        <h3 style="margin: 0; color: {analysis['aptitude_color']};">{analysis['aptitude_emoji']} {analysis['aptitude_level']} — Puntaje: {analysis['aptitude_score']}/100</h3>
        <p style="margin: 5px 0 0 0; color: #374151;">{analysis['aptitude_desc']}</p>
        <p style="margin: 5px 0 0 0; color: #6B7280;"><b>Perfil:</b> {analysis['profile_name']} ({analysis['dominant_name']} + {analysis['secondary_name']})</p>
        <p style="margin: 5px 0 0 0; color: #6B7280;"><b>Temperamento:</b> {temperament['label']}</p>
    </div>
    """, unsafe_allow_html=True)

    cols = st.columns(4)
    disc_colors = {"D": "#EF4444", "I": "#F59E0B", "S": "#10B981", "C": "#3B82F6"}
    disc_names = {"D": "Dominancia", "I": "Influencia", "S": "Estabilidad", "C": "Cumplimiento"}
    for idx, style in enumerate("DISC"):
        with cols[idx]:
            st.metric(f"{style} — {disc_names[style]}", f"{normalized.get(style, 0):.1f}%",
                      f"Rel: {relative.get(style, 0):.1f}%")

    fig = create_disc_plot(normalized)
    st.pyplot(fig)

    # ── MEGA RESUMEN ──────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 📋 Resumen Conductual")
    st.caption("Descripción detallada del perfil conductual en 16 dimensiones, derivada del resultado DISC.")

    cols_summary = st.columns(2)
    items = list(mega_summary.items())
    half = len(items) // 2
    for col, chunk in zip(cols_summary, [items[:half], items[half:]]):
        with col:
            for label, text in chunk:
                st.markdown(f"""
                <div style="background:#F8FAFC; border-left:3px solid #3B82F6; padding:10px 14px;
                            border-radius:6px; margin-bottom:8px;">
                    <b style="color:#1E40AF; font-size:0.85em;">{label}</b><br>
                    <span style="color:#374151; font-size:0.92em;">{text}</span>
                </div>""", unsafe_allow_html=True)

    # ── 9 ESTILOS CONDUCTUALES ────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 🎯 Estilos Conductuales Derivados")
    st.caption("9 estilos con 4 sub-dimensiones cada uno, inferidos matemáticamente del perfil DISC (metodología THT).")

    styles_fig = create_behavioral_styles_chart(behavioral_styles)
    st.pyplot(styles_fig)

    # Detalle expandible de cada estilo
    with st.expander("🔍 Ver descripción detallada de cada sub-dimensión"):
        for style_name, style_data in behavioral_styles.items():
            st.markdown(f"**{style_name}**")
            for sub_name, sub_val in style_data["subs"].items():
                desc = style_data["desc"][sub_name]
                bar_w = int(sub_val)
                color = "#EF4444" if sub_name in ["Franqueza", "Control", "Insistencia", "Por Resultados",
                                                  "Resolución", "Confrontación", "Priorización",
                                                  "Pragmatismo"] else (
                        "#F59E0B" if sub_name in ["Expresividad", "Inspiración", "Optimismo", "Por Inspiración",
                                                  "Positivismo", "Apasionamiento", "Entusiasmo",
                                                  "Extroversión", "Persuasión"] else (
                        "#10B981" if sub_name in ["Autoregulación", "Moderación", "Focalización", "Democrático",
                                                  "Resistencia", "Inalterabilidad", "Pausa", "Autocontrol",
                                                  "Calma"] else "#3B82F6"))
                st.markdown(f"""
                <div style="display:flex; align-items:center; gap:10px; margin:4px 0;">
                    <span style="min-width:140px; font-size:0.85em; color:#374151;">{sub_name}</span>
                    <div style="flex:1; background:#E2E8F0; border-radius:4px; height:14px; position:relative;">
                        <div style="width:{bar_w}%; background:{color}; border-radius:4px; height:14px;"></div>
                    </div>
                    <span style="font-weight:bold; color:{color}; min-width:30px;">{bar_w}</span>
                    <span style="font-size:0.75em; color:#94A3B8; flex:1;">{desc}</span>
                </div>""", unsafe_allow_html=True)
            st.markdown("")

    # ── FORTALEZAS Y ALERTAS ──────────────────────────────────────────────
    st.markdown("---")
    col_f, col_a = st.columns(2)
    with col_f:
        st.markdown("#### 💪 Fortalezas")
        for f in analysis['fortalezas']:
            st.markdown(f"- ✅ {f}")
    with col_a:
        st.markdown("#### ⚠️ Alertas")
        for a in analysis['alertas']:
            st.markdown(f"- 🔸 {a}")

    st.markdown("#### 📋 Recomendaciones para el Candidato")
    for r in analysis['recomendaciones']:
        st.markdown(f"- 💡 {r}")

    if analysis['ideal_para']:
        st.markdown("#### 🎯 Ideal para roles de")
        st.markdown(", ".join([f"**{r}**" for r in analysis['ideal_para']]))

    if analysis['cuidado_en']:
        st.markdown("#### ⛔ Tener cuidado en")
        st.markdown(", ".join([f"*{r}*" for r in analysis['cuidado_en']]))

    session_id = session if isinstance(session, str) else session.get("id")
    completed_at = session.get("completed_at") if isinstance(session, dict) else None
    pdf = generate_disc_pdf(candidate, normalized, relative, fig, session_id, completed_at, analysis,
                            behavioral_styles=behavioral_styles, temperament=temperament,
                            mega_summary=mega_summary, styles_fig=styles_fig)
    c1, c2 = st.columns(2)
    with c1:
        st.download_button("📑 Descargar PDF", data=pdf.getvalue(), file_name=f"disc_{candidate['cedula']}.pdf", mime="application/pdf", key=f"pdf_disc_{session_id}")
    with c2:
        st.download_button("📄 Descargar JSON", data=json.dumps(results, indent=2, ensure_ascii=False), file_name=f"disc_{candidate['cedula']}.json", mime="application/json", key=f"json_disc_{session_id}")


def show_valanti_results_admin(results, candidate, session):
    """Show VALANTI results in the admin panel."""
    direct = results.get("direct", {})
    # Recalcular siempre los puntajes estándar desde los directos usando las normas actuales
    # (evita errores si los valores guardados fueron calculados con normas incorrectas)
    standard = {}
    for trait in VALANTI_TRAITS:
        if trait in direct:
            z = (direct[trait] - VALANTI_AVGS[trait]) / VALANTI_SDS[trait]
            standard[trait] = round(z * 10 + 50)
        else:
            standard[trait] = results.get("standard", {}).get(trait, 50)
    
    # Análisis de aptitud
    analysis = analyze_valanti_aptitude(standard)
    
    # Banner de aptitud
    st.markdown(f"""
    <div style="background: {analysis['aptitude_color']}22; border-left: 5px solid {analysis['aptitude_color']};
                padding: 15px 20px; border-radius: 8px; margin-bottom: 15px;">
        <h3 style="margin: 0; color: {analysis['aptitude_color']};">{analysis['aptitude_emoji']} {analysis['aptitude_level']} — Puntaje: {analysis['aptitude_score']}/100</h3>
        <p style="margin: 5px 0 0 0; color: #374151;">{analysis['aptitude_desc']}</p>
        <p style="margin: 5px 0 0 0; color: #6B7280;"><b>Valor más fuerte:</b> {analysis['strongest_value']} (T={analysis['strongest_score']}) | <b>Valor más bajo:</b> {analysis['weakest_value']} (T={analysis['weakest_score']})</p>
    </div>
    """, unsafe_allow_html=True)

    cols = st.columns(5)
    for idx, trait in enumerate(VALANTI_TRAITS):
        with cols[idx]:
            st.metric(trait, standard.get(trait, 0), f"Dir: {direct.get(trait, 0)}")

    radar_fig = create_valanti_radar(standard)
    st.pyplot(radar_fig)

    bar_fig = create_valanti_bars(direct, standard)
    st.pyplot(bar_fig)

    sorted_scores = sorted(standard.items(), key=lambda x: x[1], reverse=True)
    st.markdown(f"**Valor más prominente:** {sorted_scores[0][0]} ({sorted_scores[0][1]})")
    st.markdown(f"**Valor menos enfatizado:** {sorted_scores[-1][0]} ({sorted_scores[-1][1]})")
    for trait, score in sorted_scores:
        desc = VALANTI_DESCRIPTIONS[trait]
        level = "Alto" if score >= 55 else ("Bajo" if score <= 45 else "Promedio")
        text = desc["high"] if score >= 55 else (desc["low"] if score <= 45 else "Puntaje dentro del rango promedio.")
        st.markdown(f"**{desc['title']}** — {level} ({score}): {text}")
    
    # Fortalezas y alertas
    if analysis['fortalezas']:
        st.markdown("#### 💪 Fortalezas Valorales")
        for f in analysis['fortalezas']:
            st.markdown(f"- ✅ {f}")
    
    if analysis['alertas']:
        st.markdown("#### ⚠️ Alertas")
        for a in analysis['alertas']:
            st.markdown(f"- 🔸 {a}")
    
    if analysis['recomendaciones']:
        st.markdown("#### 📋 Recomendaciones")
        for r in analysis['recomendaciones']:
            st.markdown(f"- {r}")

    session_id = session if isinstance(session, str) else session.get("id")
    completed_at = session.get("completed_at") if isinstance(session, dict) else None
    pdf = generate_valanti_pdf(candidate, direct, standard, radar_fig, session_id, completed_at, analysis)
    c1, c2 = st.columns(2)
    with c1:
        st.download_button("📑 Descargar PDF", data=pdf.getvalue(), file_name=f"valanti_{candidate['cedula']}.pdf", mime="application/pdf", key=f"pdf_val_{session_id}")
    with c2:
        st.download_button("📄 Descargar JSON", data=json.dumps(results, indent=2, ensure_ascii=False), file_name=f"valanti_{candidate['cedula']}.json", mime="application/json", key=f"json_val_{session_id}")


def show_wpi_results_admin(results, candidate, session):
    """
    Muestra los resultados del WPI en el panel de administración.
    
    Args:
        results: Dict con raw, normalized y percentages
        candidate: Dict con información del candidato
        session: Dict con información de la sesión o str con session_id
    """
    raw = results.get("raw", {})
    normalized = results.get("normalized", {})
    percentages = results.get("percentages", {})
    
    # Análisis de aptitud
    analysis = analyze_wpi_aptitude(normalized)
    
    # === BANNER DE APTITUD ===
    st.markdown(f"""
    <div style="background: {analysis['aptitude_color']}22; border-left: 5px solid {analysis['aptitude_color']};
                padding: 15px 20px; border-radius: 8px; margin-bottom: 15px;">
        <h3 style="margin: 0; color: {analysis['aptitude_color']};">
            {analysis['aptitude_emoji']} {analysis['aptitude_level']} — Puntaje: {analysis['aptitude_score']}/100
        </h3>
        <p style="margin: 5px 0 0 0; color: #374151;">{analysis['aptitude_desc']}</p>
        <p style="margin: 5px 0 0 0; color: #6B7280;">
            <b>Dimensión más fuerte:</b> {analysis['strongest_dimension']} ({int(analysis['strongest_score'])}/100) | 
            <b>Dimensión a desarrollar:</b> {analysis['weakest_dimension']} ({int(analysis['weakest_score'])}/100) |
            <b>Promedio:</b> {analysis['average_score']}/100
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # === MÉTRICAS POR DIMENSIÓN ===
    st.markdown("### 📊 Puntajes por Dimensión")
    
    # Crear 6 columnas para las 6 dimensiones
    cols = st.columns(3)
    for idx, dim in enumerate(WPI_DIMENSIONS):
        with cols[idx % 3]:
            score = normalized.get(dim, 0)
            nivel = "🟢 Alto" if score >= 70 else ("🟡 Medio" if score >= 45 else "🔴 Bajo")
            st.metric(
                label=dim,
                value=f"{int(score)}/100",
                delta=nivel,
                delta_color="off"
            )
    
    st.markdown("---")
    
    # === GRÁFICOS ===
    col_radar, col_bars = st.columns(2)
    
    with col_radar:
        st.markdown("#### 🎯 Perfil Radar")
        radar_fig = create_wpi_radar(normalized)
        st.pyplot(radar_fig)
    
    with col_bars:
        st.markdown("#### 📊 Puntajes por Dimensión")
        bar_fig = create_wpi_bars(normalized)
        st.pyplot(bar_fig)
    
    st.markdown("---")
    
    # === ANÁLISIS POR DIMENSIÓN ===
    st.markdown("### 📋 Análisis Detallado por Dimensión")
    
    sorted_scores = sorted(normalized.items(), key=lambda x: x[1], reverse=True)
    
    for dim, score in sorted_scores:
        desc_info = WPI_DESCRIPTIONS[dim]
        
        # Determinar nivel
        if score >= 70:
            level = "🟢 Alto"
            text = desc_info["high"]
            color = "#10B981"
        elif score >= 45:
            level = "🟡 Medio"
            text = desc_info["medium"]
            color = "#F59E0B"
        else:
            level = "🔴 Bajo"
            text = desc_info["low"]
            color = "#EF4444"
        
        st.markdown(f"""
        <div style="background: {color}15; border-left: 3px solid {color}; 
                    padding: 12px; border-radius: 6px; margin-bottom: 10px;">
            <b style="color: {color};">{desc_info['title']}</b> — {level} ({int(score)}/100)
            <br><span style="color: #374151;">{text}</span>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # === FORTALEZAS ===
    if analysis.get('fortalezas'):
        st.markdown("### 💪 Fortalezas Destacadas")
        for f in analysis['fortalezas']:
            # Limpiar markdown
            f_clean = f.replace("**", "")
            st.markdown(f"- ✅ {f_clean}")
        st.markdown("")
    
    # === ALERTAS ===
    if analysis.get('alertas'):
        st.markdown("### ⚠️ Áreas de Atención")
        for a in analysis['alertas']:
            # Limpiar markdown
            a_clean = a.replace("**", "").replace("⚠️ ", "")
            st.markdown(f"- 🔸 {a_clean}")
        st.markdown("")
    
    # === ROLES IDEALES ===
    if analysis.get('ideal_para'):
        st.markdown("### 🎯 Roles Ideales para el Candidato")
        for role in analysis['ideal_para']:
            st.markdown(f"- 🎯 {role}")
        st.markdown("")
    
    # === ROLES A EVITAR ===
    if analysis.get('avoid_roles'):
        st.markdown("### ⛔ Roles No Recomendados")
        for role in analysis['avoid_roles']:
            st.markdown(f"- ⛔ {role}")
        st.markdown("")
    
    # === RECOMENDACIONES ===
    if analysis.get('recomendaciones'):
        st.markdown("### 💡 Recomendaciones Específicas")
        for r in analysis['recomendaciones']:
            # Limpiar markdown
            r_clean = r.replace("**", "")
            st.markdown(f"- {r_clean}")
    
    st.markdown("---")
    
    # === DESCARGA DE REPORTES ===
    st.markdown("### 📥 Descargar Reportes")
    
    session_id = session if isinstance(session, str) else session.get("id")
    completed_at = session.get("completed_at") if isinstance(session, dict) else None
    
    # Generar PDF
    pdf_buffer = generate_wpi_pdf(candidate, raw, normalized, radar_fig, session_id, completed_at, analysis)
    
    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            "📑 Descargar PDF Completo",
            data=pdf_buffer.getvalue(),
            file_name=f"wpi_{candidate['cedula']}_{session_id}.pdf",
            mime="application/pdf",
            key=f"pdf_wpi_{session_id}"
        )
    with col2:
        st.download_button(
            "📄 Descargar JSON",
            data=json.dumps(results, indent=2, ensure_ascii=False),
            file_name=f"wpi_{candidate['cedula']}_{session_id}.json",
            mime="application/json",
            key=f"json_wpi_{session_id}"
        )


def show_eri_results_admin(results, candidate, session):
    """
    Muestra los resultados del ERI en el panel de administración.
    
    Args:
        results: Dict con raw, normalized, percentages, validity_score, validity_flags
        candidate: Dict con información del candidato
        session: Dict con información de la sesión o str con session_id
    """
    raw = results.get("raw", {})
    normalized = results.get("normalized", {})
    percentages = results.get("percentages", {})
    validity_score = results.get("validity_score", ERI_VALIDITY_QUESTIONS_COUNT)
    validity_flags = results.get("validity_flags", [])
    
    # Análisis de riesgo
    analysis = analyze_eri_aptitude(normalized, validity_score, validity_flags)
    
    # === BANNER DE VALIDEZ (si aplica) ===
    if analysis.get('validity_warning'):
        st.markdown(f"""
        <div style="background: #FEF2F2; border-left: 5px solid #DC2626;
                    padding: 15px 20px; border-radius: 8px; margin-bottom: 15px; border: 1px solid #FCA5A5;">
            <h4 style="margin: 0; color: #DC2626;">
                ⚠️ ALERTA DE VALIDEZ DEL TEST
            </h4>
            <p style="margin: 5px 0 0 0; color: #991B1B;">{analysis['validity_warning']}</p>
            <p style="margin: 5px 0 0 0; color: #7F1D1D; font-size: 0.9em;">
                El test detectó {len(validity_flags)} respuestas poco realistas. Considerar entrevista profunda adicional.
            </p>
        </div>
        """, unsafe_allow_html=True)
    
    # === BANNER DE RIESGO ===
    st.markdown(f"""
    <div style="background: {analysis['risk_color']}22; border-left: 5px solid {analysis['risk_color']};
                padding: 15px 20px; border-radius: 8px; margin-bottom: 15px;">
        <h3 style="margin: 0; color: {analysis['risk_color']};">
            {analysis['risk_emoji']} {analysis['risk_level']} — Puntaje: {analysis['risk_score']:.1f}/100
        </h3>
        <p style="margin: 5px 0 0 0; color: #374151;">{analysis['risk_desc']}</p>
        <p style="margin: 5px 0 0 0; color: #6B7280;">
            <b>Dimensión de menor riesgo:</b> {analysis['safest_dimension']} ({int(analysis['safest_score'])}/100) | 
            <b>Dimensión de mayor riesgo:</b> {analysis['riskiest_dimension']} ({int(analysis['riskiest_score'])}/100) |
            <b>Promedio:</b> {analysis['average_score']:.1f}/100
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # === DECISIÓN DE CONTRATACIÓN ===
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #1e293b 0%, #334155 100%); 
                padding: 20px; border-radius: 12px; margin-bottom: 20px; color: white;">
        <h4 style="margin: 0 0 10px 0;">📋 Decisión Recomendada de Contratación</h4>
        <h2 style="margin: 0; color: {analysis['risk_color']};">{analysis['hiring_decision']}</h2>
    </div>
    """, unsafe_allow_html=True)
    
    # === MÉTRICAS POR DIMENSIÓN ===
    st.markdown("### 📊 Puntajes por Dimensión de Riesgo")
    st.caption("⚠️ Recuerda: Puntajes más ALTOS = MENOR riesgo (Verde ✅), puntajes más BAJOS = MAYOR riesgo (Rojo 🚨)")
    
    # Crear 6 columnas para las 6 dimensiones
    cols = st.columns(3)
    for idx, dim in enumerate(ERI_DIMENSIONS):
        with cols[idx % 3]:
            score = normalized.get(dim, 0)
            if score >= ERI_RISK_THRESHOLDS["low_risk"]:
                nivel = "✅ Bajo Riesgo"
                delta_color = "normal"
            elif score >= ERI_RISK_THRESHOLDS["medium_risk"]:
                nivel = "⚠️ Moderado"
                delta_color = "off"
            else:
                nivel = "🚨 Alto Riesgo"
                delta_color = "inverse"
            
            st.metric(
                label=dim,
                value=f"{int(score)}/100",
                delta=nivel,
                delta_color=delta_color
            )
    
    st.markdown("---")
    
    # === GRÁFICOS ===
    col_radar, col_bars = st.columns(2)
    
    with col_radar:
        st.markdown("#### 🎯 Perfil de Riesgo (Radar)")
        radar_fig = create_eri_radar(normalized)
        st.pyplot(radar_fig)
    
    with col_bars:
        st.markdown("#### 📊 Puntajes por Dimensión")
        bar_fig = create_eri_bars(normalized)
        st.pyplot(bar_fig)
    
    st.markdown("---")
    
    # === ANÁLISIS POR DIMENSIÓN ===
    st.markdown("### 📋 Análisis Detallado por Dimensión")
    
    sorted_scores = sorted(normalized.items(), key=lambda x: x[1], reverse=False)  # Menor a mayor (más riesgo primero)
    
    for dim, score in sorted_scores:
        desc_info = ERI_DESCRIPTIONS[dim]
        
        # Determinar nivel de riesgo
        if score >= ERI_RISK_THRESHOLDS["low_risk"]:
            level = "✅ Bajo Riesgo"
            text = desc_info["low_risk"]
            color = "#10B981"
        elif score >= ERI_RISK_THRESHOLDS["medium_risk"]:
            level = "⚠️ Riesgo Moderado"
            text = desc_info["medium_risk"]
            color = "#F59E0B"
        else:
            level = "🚨 Alto Riesgo"
            text = desc_info["high_risk"]
            color = "#EF4444"
        
        st.markdown(f"""
        <div style="background: {color}15; border-left: 3px solid {color}; 
                    padding: 12px; border-radius: 6px; margin-bottom: 10px;">
            <b style="color: {color};">{desc_info['title']}</b> — {level} ({int(score)}/100)
            <br><span style="color: #374151;">{text}</span>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # === ASPECTOS POSITIVOS ===
    if analysis.get('fortalezas'):
        st.markdown("### 💚 Aspectos Positivos (Bajo Riesgo)")
        for f in analysis['fortalezas']:
            # Limpiar markdown
            f_clean = f.replace("**", "")
            st.markdown(f"- ✅ {f_clean}")
        st.markdown("")
    
    # === SEÑALES DE ALERTA ===
    if analysis.get('alertas'):
        st.markdown("### 🚨 Señales de Alerta y Factores de Riesgo")
        for a in analysis['alertas']:
            # Limpiar markdown
            a_clean = a.replace("**", "").replace("⚠️ ", "").replace("🚨 ", "")
            st.markdown(f"- 🔴 {a_clean}")
        st.markdown("")
    
    # === RECOMENDACIONES DE CONTRATACIÓN ===
    if analysis.get('recomendaciones'):
        st.markdown("### 💼 Recomendaciones de Contratación")
        for r in analysis['recomendaciones']:
            # Limpiar markdown (pero mantener bullets internos)
            r_clean = r.replace("**", "")
            st.markdown(f"{r_clean}")
        st.markdown("")
    
    # === DETALLES DE VALIDEZ ===
    if validity_flags and len(validity_flags) > 0:
        with st.expander(f"⚠️ Ver Detalles de Validez del Test ({len(validity_flags)} respuestas sospechosas)"):
            st.markdown(f"""
            Se detectaron **{len(validity_flags)}** respuestas poco realistas en preguntas de validez.
            
            Esto puede indicar:
            - El candidato está tratando de presentarse de forma irrealmente perfecta
            - Falta de sinceridad en las respuestas
            - No comprendió las instrucciones
            
            **Recomendación:** Explorar estos aspectos en entrevista personal.
            """)
            
            st.markdown("**Ejemplos de respuestas sospechosas:**")
            for flag in validity_flags[:10]:  # Mostrar máximo 10
                st.markdown(f"- {flag}")
            if len(validity_flags) > 10:
                st.caption(f"... y {len(validity_flags) - 10} respuestas más.")
    
    st.markdown("---")
    
    # === DESCARGA DE REPORTES ===
    st.markdown("### 📥 Descargar Reportes")
    
    session_id = session if isinstance(session, str) else session.get("id")
    completed_at = session.get("completed_at") if isinstance(session, dict) else None
    
    # Generar PDF
    pdf_buffer = generate_eri_pdf(candidate, raw, normalized, radar_fig, session_id, completed_at, analysis, validity_score, validity_flags)
    
    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            "📑 Descargar PDF Completo",
            data=pdf_buffer.getvalue(),
            file_name=f"eri_{candidate['cedula']}_{session_id}.pdf",
            mime="application/pdf",
            key=f"pdf_eri_{session_id}"
        )
    with col2:
        st.download_button(
            "📄 Descargar JSON",
            data=json.dumps(results, indent=2, ensure_ascii=False),
            file_name=f"eri_{candidate['cedula']}_{session_id}.json",
            mime="application/json",
            key=f"json_eri_{session_id}"
        )


