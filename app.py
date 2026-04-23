import streamlit as st
import streamlit.components.v1 as components
import numpy as np
import matplotlib.pyplot as plt
import random
import json
import os
import math
import base64
import hmac
import hashlib
from io import BytesIO
from datetime import datetime, timedelta, timezone as _tz_mod
from reportlab.lib.pagesizes import letter

_GMT5 = _tz_mod(timedelta(hours=-5))
def _now_gmt5(): return datetime.now(_GMT5)
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

import database as db

# =========================================================================
# IMPORTAR MÓDULOS REFACTORIZADOS
# =========================================================================
from constants import *
from calculations import (
    normalize_disc_scores,
    calculate_disc_results,
    calculate_behavioral_styles,
    get_disc_temperament,
    generate_disc_mega_summary,
    calculate_valanti_results,
    calculate_wpi_results,
    load_eri_questions,
    calculate_eri_results,
    load_talent_map_questions,
    calculate_talent_map_results,
    calculate_desempeno_results,
    calculate_desempeno_lider_results,
    calculate_periodo_prueba_results,
)
from analysis import (
    analyze_disc_aptitude,
    analyze_valanti_aptitude,
    analyze_wpi_aptitude,
    analyze_eri_aptitude,
    analyze_talent_map_match
)
from utils import load_disc_questions, load_disc_descriptions, load_wpi_questions, nav

# =========================================================================
# CONFIGURACIÓN DE PÁGINA
# =========================================================================
st.set_page_config(
    page_title="Evaluaciones Psicométricas RH",
    layout="wide",
    page_icon="🧠",
)

st.markdown("""
<style>
    .stApp { max-width: 100% !important; padding: 0 !important; }
    .block-container { max-width: 100% !important; padding-left: 1rem !important; padding-right: 1rem !important; }
    .stButton>button { font-weight: bold; }
    div[data-testid="stMetric"] {
        background: var(--secondary-background-color, #f8fafc);
        padding: 12px;
        border-radius: 10px;
        border: 1px solid rgba(148, 163, 184, 0.35);
    }
    div[data-testid="stMetricLabel"],
    div[data-testid="stMetricValue"],
    div[data-testid="stMetricDelta"],
    div[data-testid="stMetricLabel"] *,
    div[data-testid="stMetricValue"] *,
    div[data-testid="stMetricDelta"] * {
        color: var(--text-color, #0f172a) !important;
    }
</style>
""", unsafe_allow_html=True)


def _apply_theme_override(theme_mode):
    """Permite forzar apariencia clara u oscura sin depender del modo del sistema."""
    if theme_mode == "Claro":
        st.markdown("""
        <style>
            .stApp, [data-testid="stAppViewContainer"] {
                background-color: #f8fafc !important;
                color: #0f172a !important;
            }
            [data-testid="stHeader"], [data-testid="stToolbar"] {
                background-color: #f8fafc !important;
            }
            [data-testid="stSidebar"] {
                background-color: #eef2f7 !important;
                color: #0f172a !important;
            }
            h1, h2, h3, h4, h5, h6, p, span, label, div {
                color: #0f172a;
            }
            .stTextInput input, .stTextArea textarea, .stSelectbox div[data-baseweb="select"] > div,
            .stNumberInput input {
                background-color: #ffffff !important;
                color: #0f172a !important;
                border-color: #cbd5e1 !important;
            }
            /* Dropdown popup del selectbox/multiselect — portal nivel body */
            [data-baseweb="popover"], [data-baseweb="popover"] > div,
            div[data-baseweb="menu"], ul[data-baseweb="menu"],
            [data-baseweb="select-dropdown"],
            div[class*="Menu"], div[class*="menu"],
            div[data-baseweb="popover"] * {
                background-color: #ffffff !important;
                color: #0f172a !important;
            }
            li[data-baseweb="option"], li[data-baseweb="option"] > div,
            li[data-baseweb="option"] * {
                background-color: #ffffff !important;
                color: #0f172a !important;
            }
            li[data-baseweb="option"]:hover,
            li[data-baseweb="option"]:hover *,
            [data-baseweb="option"][aria-selected="true"],
            [data-baseweb="option"][aria-selected="true"] * {
                background-color: #e2e8f0 !important;
                color: #0f172a !important;
            }
            /* File uploader */
            [data-testid="stFileUploader"] section,
            [data-testid="stFileUploadDropzone"],
            [data-testid="stFileUploader"] > div {
                background-color: #f8fafc !important;
                color: #0f172a !important;
                border-color: #cbd5e1 !important;
            }
            [data-testid="stFileUploader"] small,
            [data-testid="stFileUploader"] span,
            [data-testid="stFileUploader"] p {
                color: #64748b !important;
            }
            /* Expander header y contenido */
            [data-testid="stExpander"],
            [data-testid="stExpander"] > details,
            [data-testid="stExpander"] > details > summary,
            [data-testid="stExpander"] > details > div {
                background-color: #f8fafc !important;
                color: #0f172a !important;
                border-color: #cbd5e1 !important;
            }
            [data-testid="stExpander"] > details > summary svg {
                fill: #0f172a !important;
            }
            .stButton > button,
            [data-testid="stDownloadButton"] > button,
            [data-testid="stFormSubmitButton"] > button,
            button[kind="formSubmit"],
            button[kind="secondaryFormSubmit"] {
                background: #ffffff !important;
                color: #0f172a !important;
                border: 1px solid #cbd5e1 !important;
            }
            [data-testid="stDownloadButton"] > button:hover,
            [data-testid="stFormSubmitButton"] > button:hover,
            .stButton > button:hover {
                background: #f1f5f9 !important;
                border-color: #94a3b8 !important;
            }
            ::selection {
                background-color: #bfdbfe !important;
                color: #1e3a5f !important;
            }
            div[data-testid="stMetric"] {
                background: #ffffff !important;
                border: 1px solid #dbe3ef !important;
            }
            div[data-testid="stMetricLabel"],
            div[data-testid="stMetricValue"],
            div[data-testid="stMetricDelta"],
            div[data-testid="stMetricLabel"] *,
            div[data-testid="stMetricValue"] *,
            div[data-testid="stMetricDelta"] * {
                color: #0f172a !important;
            }
        </style>
        """, unsafe_allow_html=True)
    elif theme_mode == "Oscuro":
        st.markdown("""
        <style>
            .stApp, [data-testid="stAppViewContainer"] {
                background-color: #0b1220 !important;
                color: #e2e8f0 !important;
            }
            [data-testid="stHeader"], [data-testid="stToolbar"] {
                background-color: #0b1220 !important;
            }
            [data-testid="stSidebar"] {
                background-color: #101a2e !important;
                color: #e2e8f0 !important;
            }
            div[data-testid="stMetric"] {
                background: #111827 !important;
                border: 1px solid #334155 !important;
            }
            div[data-testid="stMetricLabel"],
            div[data-testid="stMetricValue"],
            div[data-testid="stMetricDelta"],
            div[data-testid="stMetricLabel"] *,
            div[data-testid="stMetricValue"] *,
            div[data-testid="stMetricDelta"] * {
                color: #e2e8f0 !important;
            }
            /* Dropdown popup oscuro */
            [data-baseweb="popover"], [data-baseweb="popover"] > div,
            div[data-baseweb="menu"], ul[data-baseweb="menu"] {
                background-color: #1e293b !important;
                color: #e2e8f0 !important;
            }
            li[data-baseweb="option"], li[data-baseweb="option"] > div {
                background-color: #1e293b !important;
                color: #e2e8f0 !important;
            }
            li[data-baseweb="option"]:hover,
            [data-baseweb="option"][aria-selected="true"] {
                background-color: #334155 !important;
                color: #e2e8f0 !important;
            }
            .stTextInput input, .stTextArea textarea,
            .stSelectbox div[data-baseweb="select"] > div,
            .stNumberInput input {
                background-color: #1e293b !important;
                color: #e2e8f0 !important;
                border-color: #475569 !important;
            }
        </style>
        """, unsafe_allow_html=True)


def _render_theme_switcher():
    """Control global de apariencia."""
    if "ui_theme_mode" not in st.session_state:
        st.session_state["ui_theme_mode"] = "Sistema"

    with st.sidebar:
        st.markdown("---")
        st.markdown("### 🎨 Apariencia")
        st.radio(
            "Tema visual",
            options=["Sistema", "Claro", "Oscuro"],
            key="ui_theme_mode",
            horizontal=False,
            help="Elige 'Claro' para ver la app con colores normales aunque tu PC esté en modo oscuro.",
        )

    _apply_theme_override(st.session_state.get("ui_theme_mode", "Sistema"))


_render_theme_switcher()

ADMIN_IDLE_TIMEOUT_MINUTES = 60
ADMIN_SESSION_SECRET = os.getenv("ADMIN_SESSION_SECRET", "rh-evaluaciones-secret-key")


def _get_admin_token_from_query():
    """Lee el token admin desde query params (permite restaurar sesión tras refresh)."""
    try:
        return st.query_params.get("admin_token", None)
    except Exception:
        return None


def _set_admin_token_in_query(token):
    """Escribe o elimina el token admin en query params preservando los demás parámetros."""
    try:
        if token:
            st.query_params["admin_token"] = token
        else:
            st.query_params.pop("admin_token", None)
    except Exception:
        pass


def _create_admin_session_token(admin_id, expires_at=None):
    """Crea token firmado con expiración para restaurar sesión tras refresh."""
    if expires_at is None:
        expires_at = datetime.utcnow() + timedelta(minutes=ADMIN_IDLE_TIMEOUT_MINUTES)
    exp_ts = int(expires_at.timestamp())
    payload = f"{admin_id}:{exp_ts}"
    signature = hmac.new(ADMIN_SESSION_SECRET.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
    raw = f"{payload}:{signature}"
    return base64.urlsafe_b64encode(raw.encode("utf-8")).decode("utf-8")


def _parse_admin_session_token(token):
    """Valida token firmado y devuelve admin_id si está vigente; si no, None."""
    try:
        raw = base64.urlsafe_b64decode(token.encode("utf-8")).decode("utf-8")
        admin_id_str, exp_ts_str, signature = raw.split(":", 2)
        payload = f"{admin_id_str}:{exp_ts_str}"
        expected_sig = hmac.new(
            ADMIN_SESSION_SECRET.encode("utf-8"),
            payload.encode("utf-8"),
            hashlib.sha256,
        ).hexdigest()
        if not hmac.compare_digest(signature, expected_sig):
            return None
        if datetime.utcnow().timestamp() > int(exp_ts_str):
            return None
        return int(admin_id_str)
    except Exception:
        return None


def _get_admin_by_id(admin_id):
    """Obtiene admin por ID para restaurar sesión desde token."""
    conn = db.get_connection()
    row = conn.execute("SELECT * FROM admins WHERE id = ?", (admin_id,)).fetchone()
    conn.close()
    return dict(row) if row else None


def _start_admin_session(admin):
    """Inicia sesión admin con token de 10 minutos renovable por actividad."""
    token = _create_admin_session_token(admin["id"])
    st.session_state["admin"] = admin
    st.session_state["admin_session_token"] = token
    st.session_state["admin_last_seen_at"] = datetime.utcnow().isoformat()
    _set_admin_token_in_query(token)


def _touch_admin_session():
    """Renueva ventana de inactividad cuando hay uso de la app."""
    admin = st.session_state.get("admin")
    if not admin:
        return

    token = _create_admin_session_token(admin["id"])
    st.session_state["admin_session_token"] = token
    st.session_state["admin_last_seen_at"] = datetime.utcnow().isoformat()
    _set_admin_token_in_query(token)


def _restore_admin_session():
    """Restaura admin desde token en URL si la sesión aún no expiró."""
    if st.session_state.get("admin"):
        if not st.session_state.get("admin_session_token"):
            _start_admin_session(st.session_state["admin"])
        else:
            _touch_admin_session()
        return

    token = _get_admin_token_from_query()
    if not token:
        return

    admin_id = _parse_admin_session_token(token)
    if not admin_id:
        _set_admin_token_in_query(None)
        return

    admin = _get_admin_by_id(admin_id)
    if not admin:
        _set_admin_token_in_query(None)
        return

    st.session_state["admin"] = admin
    st.session_state["admin_session_token"] = token
    _touch_admin_session()


def _logout_admin():
    """Cierra sesión admin local y elimina token persistido."""
    st.session_state.pop("admin", None)
    st.session_state.pop("admin_session_token", None)
    st.session_state.pop("admin_last_seen_at", None)
    _set_admin_token_in_query(None)

# =========================================================================
# NOTA: Constantes y funciones movidas a módulos separados
# =========================================================================
# - constants.py: Todas las constantes (VALANTI, WPI, ERI, TALENT MAP, DESEMPEÑO, DISC)
# - calculations.py: Todas las funciones de cálculo
# - analysis.py: Todas las funciones de análisis
# - utils.py: Funciones auxiliares
# =========================================================================

# =========================================================================
# FUNCIONES DE GRÁFICOS
# =========================================================================

def create_disc_plot(normalized_score):
    categories = ["D", "I", "S", "C"]
    labels = ["D\nDominancia", "I\nInfluencia", "S\nEstabilidad", "C\nCumplimiento"]
    disc_colors = {"D": "#EF4444", "I": "#F59E0B", "S": "#10B981", "C": "#3B82F6"}
    
    # Gráfico de barras horizontales + radar pequeño
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(9, 3.5), gridspec_kw={'width_ratios': [3, 2]})
    
    # --- Barras horizontales ---
    vals = [normalized_score.get(s, 0) for s in categories]
    colors = [disc_colors[s] for s in categories]
    bars = ax1.barh(labels, vals, color=colors, height=0.6, edgecolor='white', linewidth=1.5)
    
    for bar, val, cat in zip(bars, vals, categories):
        ax1.text(bar.get_width() + 1.5, bar.get_y() + bar.get_height()/2, 
                f"{val:.1f}%", va='center', fontweight='bold', fontsize=12, color=disc_colors[cat])
    
    ax1.set_xlim(0, 110)
    ax1.axvline(x=50, color='#94A3B8', linestyle='--', alpha=0.6, label='Promedio')
    ax1.set_title("Puntajes por Estilo DISC", fontsize=14, fontweight='bold', pad=15)
    ax1.spines['top'].set_visible(False)
    ax1.spines['right'].set_visible(False)
    ax1.spines['bottom'].set_color('#CBD5E1')
    ax1.spines['left'].set_color('#CBD5E1')
    ax1.tick_params(axis='y', labelsize=11)
    ax1.set_facecolor('#FAFBFC')
    ax1.legend(fontsize=9)
    
    # --- Radar ---
    angles = [7 * np.pi / 4, np.pi / 4, 3 * np.pi / 4, 5 * np.pi / 4]
    scaled = {s: v / 100 for s, v in normalized_score.items()}
    
    # Dibujar áreas por estilo
    ax2 = fig.add_subplot(122, projection='polar')
    ax2.set_theta_offset(np.pi / 2)
    ax2.set_theta_direction(-1)
    ax2.set_ylim(0, 1.01)
    
    for i, s in enumerate(categories):
        ax2.bar(angles[i], scaled[s], width=np.pi/2.5, alpha=0.35, color=disc_colors[s], edgecolor=disc_colors[s], linewidth=2)
    
    # Punto central del perfil
    x = sum(scaled[s] * np.cos(angles[i]) for i, s in enumerate(categories))
    y = sum(scaled[s] * np.sin(angles[i]) for i, s in enumerate(categories))
    mag = np.sqrt(x**2 + y**2)
    ang = np.arctan2(y, x)
    ax2.plot(ang, mag, "o", markersize=16, color="#1E293B", zorder=5)
    ax2.plot(ang, mag, "o", markersize=10, color="#FBBF24", zorder=6)
    
    ax2.set_xticks(angles)
    ax2.set_xticklabels(categories, fontsize=13, fontweight="bold")
    tick_colors = ['#EF4444', '#F59E0B', '#10B981', '#3B82F6']
    for label, color in zip(ax2.get_xticklabels(), tick_colors):
        label.set_color(color)
    ax2.set_yticklabels([])
    ax2.grid(True, alpha=0.2)
    ax2.spines["polar"].set_visible(False)
    ax2.set_facecolor('#FAFBFC')
    ax2.set_title("Perfil DISC", fontsize=13, fontweight='bold', pad=20)
    
    fig.patch.set_facecolor('white')
    plt.tight_layout()
    return fig


def create_behavioral_styles_chart(behavioral_styles):
    """
    Crea un gráfico de barras horizontales para los 9 estilos conductuales
    derivados del perfil DISC, con sus 4 sub-dimensiones cada uno.
    Inspirado en el modelo de reporte THT.
    """
    style_names = list(behavioral_styles.keys())
    disc_colors = {"D": "#EF4444", "I": "#F59E0B", "S": "#10B981", "C": "#3B82F6"}
    sub_order = ["D", "I", "S", "C"]  # orden estándar de sub-dimensiones

    # Mapeo de sub-dimensión → estilo DISC para colorear
    sub_to_disc_idx = {0: "D", 1: "I", 2: "S", 3: "C"}

    n_styles = len(style_names)
    fig, axes = plt.subplots(n_styles, 1, figsize=(10, n_styles * 1.3 + 1))
    fig.patch.set_facecolor('white')
    fig.suptitle("Estilos Conductuales Derivados del Perfil DISC", fontsize=14,
                 fontweight='bold', color='#1E293B', y=1.01)

    for ax_idx, (style_name, style_data) in enumerate(behavioral_styles.items()):
        ax = axes[ax_idx]
        subs = style_data["subs"]
        sub_names = list(subs.keys())
        sub_values = list(subs.values())
        colors = [disc_colors[sub_to_disc_idx[i]] for i in range(len(sub_names))]

        bars = ax.barh(sub_names, sub_values, color=colors, height=0.55,
                       edgecolor='white', linewidth=1.2)

        for bar, val, color in zip(bars, sub_values, colors):
            ax.text(min(val + 2, 102), bar.get_y() + bar.get_height() / 2,
                    f"{val}", va='center', fontweight='bold', fontsize=9, color=color)

        ax.set_xlim(0, 110)
        ax.axvline(x=50, color='#CBD5E1', linestyle='--', alpha=0.6, linewidth=0.8)

        # Fondo de la fila con color alternado
        ax.set_facecolor('#F8FAFC' if ax_idx % 2 == 0 else '#FFFFFF')

        # Título del estilo a la izquierda como etiqueta del eje y
        ax.set_title(f"  {ax_idx + 1}. {style_name}", fontsize=10, fontweight='bold',
                     color='#1E293B', loc='left', pad=4)

        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_color('#E2E8F0')
        ax.spines['left'].set_color('#E2E8F0')
        ax.tick_params(axis='y', labelsize=8.5, colors='#475569')
        ax.tick_params(axis='x', labelsize=7, colors='#94A3B8')
        ax.set_xticks([0, 25, 50, 75, 100])
        ax.set_xticklabels(['0', '25', '50', '75', '100'])

    plt.tight_layout(pad=1.2)
    return fig


def create_valanti_radar(standard_scores):
    cats = list(standard_scores.keys())
    vals = list(standard_scores.values()) + [list(standard_scores.values())[0]]
    angles = np.linspace(0, 2 * np.pi, len(cats), endpoint=False).tolist() + [0]
    
    valanti_radar_colors = ["#3B82F6", "#10B981", "#8B5CF6", "#EF4444", "#F59E0B"]
    
    fig, ax = plt.subplots(figsize=(5, 5), subplot_kw=dict(polar=True))
    
    # Línea principal con gradiente
    ax.plot(angles, vals, "o-", linewidth=2.5, color="#6366F1", markersize=10, 
            markerfacecolor="#818CF8", markeredgecolor="white", markeredgewidth=2, zorder=5)
    ax.fill(angles, vals, alpha=0.15, color="#6366F1")
    
    # Colorear cada punto según su valor
    for i, (angle, val) in enumerate(zip(angles[:-1], vals[:-1])):
        color = valanti_radar_colors[i]
        ax.plot(angle, val, "o", markersize=14, color=color, zorder=6, markeredgecolor='white', markeredgewidth=2)
        ax.text(angle, val + 6, str(val), ha='center', va='center', fontsize=10, fontweight='bold', color=color)
    
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(cats, fontsize=12, fontweight="bold",
                       color='#1E293B')
    ax.set_ylim(0, 100)
    ax.set_yticks([20, 40, 50, 60, 80])
    ax.set_yticklabels(['20', '40', '50', '60', '80'], fontsize=8, color='#94A3B8')
    
    # Línea de referencia promedio
    ref = [50] * (len(cats) + 1)
    ax.plot(angles, ref, "--", linewidth=1.5, color="#F59E0B", alpha=0.6, label="Promedio (50)")
    
    # Zonas de color
    theta = np.linspace(0, 2*np.pi, 100)
    ax.fill_between(theta, 0, 40, alpha=0.05, color='#EF4444')  # zona baja
    ax.fill_between(theta, 55, 100, alpha=0.05, color='#10B981')  # zona alta
    
    ax.grid(True, alpha=0.2, color='#CBD5E1')
    ax.spines["polar"].set_visible(False)
    ax.set_facecolor('#FAFBFC')
    fig.patch.set_facecolor('white')
    plt.title("Perfil Valoral - VALANTI", fontsize=15, fontweight="bold", pad=25, color='#1E293B')
    plt.legend(loc="upper right", bbox_to_anchor=(1.35, 1.1), fontsize=10)
    return fig


def create_valanti_bars(direct_scores, standard_scores):
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(9, 3.5))
    fig.patch.set_facecolor('white')
    cats = list(direct_scores.keys())
    bar_colors = [VALANTI_COLORS[c] for c in cats]
    
    # --- Puntajes Directos ---
    dv = list(direct_scores.values())
    bars1 = ax1.bar(cats, dv, color=bar_colors, alpha=0.85, edgecolor='white', linewidth=1.5, width=0.6)
    ax1.set_title("Puntajes Directos", fontsize=13, fontweight="bold", color='#1E293B', pad=15)
    for b, v, c in zip(bars1, dv, bar_colors):
        ax1.text(b.get_x() + b.get_width() / 2, b.get_height() + 0.5, str(v), 
                ha="center", fontweight="bold", fontsize=12, color=c)
    ax1.set_ylim(0, max(dv) * 1.3 if max(dv) > 0 else 15)
    ax1.spines['top'].set_visible(False)
    ax1.spines['right'].set_visible(False)
    ax1.spines['bottom'].set_color('#CBD5E1')
    ax1.spines['left'].set_color('#CBD5E1')
    ax1.set_facecolor('#FAFBFC')
    ax1.tick_params(axis='x', labelsize=10)
    ax1.tick_params(axis='y', colors='#94A3B8')
    
    # --- Puntajes Estándar ---
    sv = list(standard_scores.values())
    bars2 = ax2.bar(cats, sv, color=bar_colors, alpha=0.85, edgecolor='white', linewidth=1.5, width=0.6)
    ax2.axhline(y=50, color="#F59E0B", linestyle="--", alpha=0.7, linewidth=1.5, label="Promedio (50)")
    ax2.axhspan(0, 40, alpha=0.04, color='#EF4444')  # zona baja
    ax2.axhspan(55, max(sv)*1.3 if max(sv) > 0 else 100, alpha=0.04, color='#10B981')  # zona alta
    ax2.set_title("Puntajes Estándar (Escala T)", fontsize=13, fontweight="bold", color='#1E293B', pad=15)
    for b, v, c in zip(bars2, sv, bar_colors):
        ax2.text(b.get_x() + b.get_width() / 2, b.get_height() + 0.5, str(v), 
                ha="center", fontweight="bold", fontsize=12, color=c)
    ax2.set_ylim(0, max(sv) * 1.3 if max(sv) > 0 else 100)
    ax2.spines['top'].set_visible(False)
    ax2.spines['right'].set_visible(False)
    ax2.spines['bottom'].set_color('#CBD5E1')
    ax2.spines['left'].set_color('#CBD5E1')
    ax2.set_facecolor('#FAFBFC')
    ax2.tick_params(axis='x', labelsize=10)
    ax2.tick_params(axis='y', colors='#94A3B8')
    ax2.legend(fontsize=10, loc='upper right')
    plt.tight_layout()
    return fig


def create_wpi_radar(normalized_scores):
    """
    Crea un gráfico de radar para visualizar las 6 dimensiones del WPI.
    
    Args:
        normalized_scores: Dict con puntajes normalizados (0-100) por dimensión
        
    Returns:
        matplotlib.figure.Figure: Gráfico de radar
    """
    # Preparar datos para el radar
    dimensions = WPI_DIMENSIONS
    values = [normalized_scores[dim] for dim in dimensions]
    values_closed = values + [values[0]]  # Cerrar el polígono
    
    # Calcular ángulos para cada dimensión
    angles = np.linspace(0, 2 * np.pi, len(dimensions), endpoint=False).tolist()
    angles_closed = angles + [angles[0]]
    
    # Colores para cada dimensión
    dim_colors = [WPI_COLORS[dim] for dim in dimensions]
    
    # Crear figura
    fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
    
    # Línea principal del perfil
    ax.plot(angles_closed, values_closed, "o-", linewidth=2.5, color="#6366F1", 
            markersize=8, markerfacecolor="#818CF8", markeredgecolor="white", 
            markeredgewidth=2, zorder=5)
    
    # Rellenar área
    ax.fill(angles_closed, values_closed, alpha=0.2, color="#6366F1")
    
    # Puntos coloreados por dimensión con valores
    for i, (angle, val, color) in enumerate(zip(angles, values, dim_colors)):
        # Punto
        ax.plot(angle, val, "o", markersize=16, color=color, zorder=6, 
                markeredgecolor='white', markeredgewidth=2.5)
        # Valor del punto
        ax.text(angle, val + 7, f"{int(val)}", ha='center', va='center', 
                fontsize=11, fontweight='bold', color=color)
    
    # Configurar etiquetas de dimensiones
    ax.set_xticks(angles)
    ax.set_xticklabels(dimensions, fontsize=11, fontweight="bold", color='#1E293B')
    
    # Configurar escala radial
    ax.set_ylim(0, 100)
    ax.set_yticks([20, 40, 60, 80, 100])
    ax.set_yticklabels(['20', '40', '60', '80', '100'], fontsize=9, color='#94A3B8')
    
    # Líneas de referencia
    ref_50 = [50] * (len(dimensions) + 1)
    ref_70 = [70] * (len(dimensions) + 1)
    ax.plot(angles_closed, ref_50, "--", linewidth=1.5, color="#F59E0B", 
            alpha=0.6, label="Promedio (50)")
    ax.plot(angles_closed, ref_70, ":", linewidth=1.5, color="#10B981", 
            alpha=0.6, label="Alto (70)")
    
    # Zonas de color de fondo
    theta = np.linspace(0, 2*np.pi, 100)
    ax.fill_between(theta, 0, 45, alpha=0.04, color='#EF4444')   # zona baja (rojo)
    ax.fill_between(theta, 70, 100, alpha=0.05, color='#10B981') # zona alta (verde)
    
    # Estilo del gráfico
    ax.grid(True, alpha=0.3, color='#CBD5E1', linestyle='-', linewidth=0.8)
    ax.spines["polar"].set_visible(False)
    ax.set_facecolor('#FAFBFC')
    fig.patch.set_facecolor('white')
    
    # Título y leyenda
    plt.title("Perfil de Personalidad Laboral - WPI", fontsize=16, fontweight="bold", 
              pad=30, color='#1E293B')
    plt.legend(loc="upper right", bbox_to_anchor=(1.25, 1.1), fontsize=10)
    
    plt.tight_layout()
    return fig


def create_wpi_bars(normalized_scores):
    """
    Crea un gráfico de barras horizontales para visualizar las dimensiones del WPI.
    
    Args:
        normalized_scores: Dict con puntajes normalize (0-100) por dimensión
        
    Returns:
        matplotlib.figure.Figure: Gráfico de barras horizontales
    """
    fig, ax = plt.subplots(figsize=(8, 5))
    fig.patch.set_facecolor('white')
    
    dimensions = WPI_DIMENSIONS
    values = [normalized_scores[dim] for dim in dimensions]
    colors = [WPI_COLORS[dim] for dim in dimensions]
    
    # Crear barras horizontales (de abajo hacia arriba)
    y_positions = np.arange(len(dimensions))
    bars = ax.barh(y_positions, values, color=colors, alpha=0.85, 
                   edgecolor='white', linewidth=2, height=0.7)
    
    # Agregar valores al final de cada barra
    for i, (bar, val, color) in enumerate(zip(bars, values, colors)):
        ax.text(val + 2, bar.get_y() + bar.get_height()/2, f"{int(val)}/100", 
                va='center', fontweight='bold', fontsize=12, color=color)
    
    # Líneas de referencia verticales
    ax.axvline(x=50, color="#F59E0B", linestyle="--", alpha=0.7, linewidth=2, 
               label="Promedio (50)")
    ax.axvline(x=70, color="#10B981", linestyle=":", alpha=0.7, linewidth=2, 
               label="Alto (70)")
    
    # Zonas de color de fondo
    ax.axvspan(0, 45, alpha=0.05, color='#EF4444')   # zona baja
    ax.axvspan(70, 100, alpha=0.05, color='#10B981') # zona alta
    
    # Configuración de ejes
    ax.set_yticks(y_positions)
    ax.set_yticklabels(dimensions, fontsize=12, fontweight='bold', color='#1E293B')
    ax.set_xlabel('Puntaje (0-100)', fontsize=12, fontweight='bold', color='#475569')
    ax.set_xlim(0, 105)
    ax.set_ylim(-0.5, len(dimensions) - 0.5)
    
    # Título
    ax.set_title("Dimensiones de Personalidad Laboral", fontsize=14, 
                 fontweight="bold", pad=20, color='#1E293B')
    
    # Estilo
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_color('#CBD5E1')
    ax.spines['left'].set_color('#CBD5E1')
    ax.set_facecolor('#FAFBFC')
    ax.tick_params(axis='x', colors='#94A3B8')
    ax.tick_params(axis='y', colors='#475569')
    ax.grid(axis='x', alpha=0.2, color='#CBD5E1', linestyle='-')
    
    # Leyenda
    ax.legend(fontsize=10, loc='lower right', framealpha=0.95)
    
    plt.tight_layout()
    return fig


def create_eri_radar(normalized_scores):
    """
    Crea un gráfico de radar para visualizar las 6 dimensiones del ERI.
    IMPORTANTE: Valores altos = BAJO riesgo (verde), valores bajos = ALTO riesgo (rojo)
    
    Args:
        normalized_scores: Dict con puntajes normalizados (0-100) por dimensión (100 = bajo riesgo)
        
    Returns:
        matplotlib.figure.Figure: Gráfico de radar
    """
    # Preparar datos para el radar
    dimensions = ERI_DIMENSIONS
    values = [normalized_scores[dim] for dim in dimensions]
    values_closed = values + [values[0]]  # Cerrar el polígono
    
    # Calcular ángulos para cada dimensión
    angles = np.linspace(0, 2 * np.pi, len(dimensions), endpoint=False).tolist()
    angles_closed = angles + [angles[0]]
    
    # Colores para cada dimensión
    dim_colors = [ERI_COLORS[dim] for dim in dimensions]
    
    # Crear figura
    fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
    
    # Línea principal del perfil
    ax.plot(angles_closed, values_closed, "o-", linewidth=3, color="#6366F1", 
            markersize=10, markerfacecolor="#818CF8", markeredgecolor="white", 
            markeredgewidth=2.5, zorder=5)
    
    # Rellenar área
    ax.fill(angles_closed, values_closed, alpha=0.2, color="#6366F1")
    
    # Puntos coloreados por dimensión con valores
    for i, (angle, val, color) in enumerate(zip(angles, values, dim_colors)):
        # Determinar riesgo por color del punto
        if val >= ERI_RISK_THRESHOLDS["low_risk"]:
            point_color = "#10B981"  # Verde - Bajo riesgo
        elif val >= ERI_RISK_THRESHOLDS["medium_risk"]:
            point_color = "#F59E0B"  # Amarillo - Riesgo moderado
        else:
            point_color = "#EF4444"  # Rojo - Alto riesgo
        
        # Punto
        ax.plot(angle, val, "o", markersize=18, color=point_color, zorder=6, 
                markeredgecolor='white', markeredgewidth=3)
        # Valor del punto
        ax.text(angle, val + 7, f"{int(val)}", ha='center', va='center', 
                fontsize=12, fontweight='bold', color=point_color)
    
    # Configurar etiquetas de dimensiones con ajuste de tamaño
    ax.set_xticks(angles)
    labels = []
    for dim in dimensions:
        # Dividir nombres largos en dos líneas
        if len(dim) > 15:
            words = dim.split()
            if len(words) >= 2:
                mid = len(words) // 2
                line1 = " ".join(words[:mid])
                line2 = " ".join(words[mid:])
                labels.append(f"{line1}\n{line2}")
            else:
                labels.append(dim)
        else:
            labels.append(dim)
    ax.set_xticklabels(labels, fontsize=10, fontweight="bold", color='#1E293B')
    
    # Configurar escala radial
    ax.set_ylim(0, 100)
    ax.set_yticks([20, 40, 66, 80, 100])
    ax.set_yticklabels(['20', '40', '66\n(Umbral)', '80', '100'], fontsize=9, color='#94A3B8')
    
    # Líneas de referencia (umbrales de riesgo)
    ref_low_risk = [ERI_RISK_THRESHOLDS["low_risk"]] * (len(dimensions) + 1)
    ref_medium_risk = [ERI_RISK_THRESHOLDS["medium_risk"]] * (len(dimensions) + 1)
    
    ax.plot(angles_closed, ref_low_risk, "-", linewidth=2, color="#10B981", 
            alpha=0.7, label="Bajo Riesgo (≥66)")
    ax.plot(angles_closed, ref_medium_risk, "--", linewidth=2, color="#F59E0B", 
            alpha=0.7, label="Riesgo Moderado (≥41)")
    
    # Zonas de color de fondo (invertidas: alto score = bajo riesgo)
    theta = np.linspace(0, 2*np.pi, 100)
    ax.fill_between(theta, 0, ERI_RISK_THRESHOLDS["medium_risk"], 
                     alpha=0.08, color='#EF4444')  # zona alto riesgo (rojo)
    ax.fill_between(theta, ERI_RISK_THRESHOLDS["medium_risk"], ERI_RISK_THRESHOLDS["low_risk"], 
                     alpha=0.06, color='#F59E0B')  # zona riesgo moderado (amarillo)
    ax.fill_between(theta, ERI_RISK_THRESHOLDS["low_risk"], 100, 
                     alpha=0.08, color='#10B981')  # zona bajo riesgo (verde)
    
    # Estilo del gráfico
    ax.grid(True, alpha=0.3, color='#CBD5E1', linestyle='-', linewidth=0.8)
    ax.spines["polar"].set_visible(False)
    ax.set_facecolor('#FAFBFC')
    fig.patch.set_facecolor('white')
    
    # Título y leyenda
    plt.title("Perfil de Riesgo e Integridad - ERI\n(Puntajes altos = BAJO riesgo)", 
              fontsize=16, fontweight="bold", pad=35, color='#1E293B')
    plt.legend(loc="upper right", bbox_to_anchor=(1.3, 1.1), fontsize=10)
    
    plt.tight_layout()
    return fig


def create_eri_bars(normalized_scores):
    """
    Crea un gráfico de barras horizontales para visualizar las dimensiones del ERI con zonas de riesgo.
    IMPORTANTE: Valores altos = BAJO riesgo (verde), valores bajos = ALTO riesgo (rojo)
    
    Args:
        normalized_scores: Dict con puntajes normalizados (0-100) por dimensión (100 = bajo riesgo)
        
    Returns:
        matplotlib.figure.Figure: Gráfico de barras horizontales
    """
    fig, ax = plt.subplots(figsize=(12, 8))
    fig.patch.set_facecolor('white')
    
    dimensions = ERI_DIMENSIONS
    values = [normalized_scores[dim] for dim in dimensions]
    
    # Colores de barras según nivel de riesgo
    colors = []
    for val in values:
        if val >= ERI_RISK_THRESHOLDS["low_risk"]:
            colors.append("#10B981")  # Verde - Bajo riesgo
        elif val >= ERI_RISK_THRESHOLDS["medium_risk"]:
            colors.append("#F59E0B")  # Amarillo - Riesgo moderado
        else:
            colors.append("#EF4444")  # Rojo - Alto riesgo
    
    # Crear barras horizontales (de abajo hacia arriba)
    y_positions = np.arange(len(dimensions))
    bars = ax.barh(y_positions, values, color=colors, alpha=0.85, 
                   edgecolor='white', linewidth=2.5, height=0.7)
    
    # Agregar valores al final de cada barra con etiqueta de riesgo
    for i, (bar, val, color) in enumerate(zip(bars, values, colors)):
        if val >= ERI_RISK_THRESHOLDS["low_risk"]:
            risk_label = "✅ Bajo Riesgo"
        elif val >= ERI_RISK_THRESHOLDS["medium_risk"]:
            risk_label = "⚠️ Moderado"
        else:
            risk_label = "🚨 Alto Riesgo"
        
        ax.text(val + 2, bar.get_y() + bar.get_height()/2, f"{int(val)}  {risk_label}", 
                va='center', fontweight='bold', fontsize=11, color=color)
    
    # Líneas de referencia verticales (umbrales)
    ax.axvline(x=ERI_RISK_THRESHOLDS["low_risk"], color="#10B981", linestyle="-", 
               alpha=0.8, linewidth=2.5, label="Bajo Riesgo (≥66)")
    ax.axvline(x=ERI_RISK_THRESHOLDS["medium_risk"], color="#F59E0B", linestyle="--", 
               alpha=0.8, linewidth=2.5, label="Riesgo Moderado (≥41)")
    
    # Zonas de color de fondo
    ax.axvspan(0, ERI_RISK_THRESHOLDS["medium_risk"], alpha=0.08, color='#EF4444')  # Alto riesgo
    ax.axvspan(ERI_RISK_THRESHOLDS["medium_risk"], ERI_RISK_THRESHOLDS["low_risk"], 
               alpha=0.06, color='#F59E0B')  # Riesgo moderado
    ax.axvspan(ERI_RISK_THRESHOLDS["low_risk"], 100, alpha=0.08, color='#10B981')  # Bajo riesgo
    
    # Configuración de ejes
    ax.set_yticks(y_positions)
    ax.set_yticklabels(dimensions, fontsize=11, fontweight='bold', color='#1E293B')
    ax.set_xlabel('Puntaje (0-100) - Mayor puntaje = MENOR riesgo', fontsize=12, 
                  fontweight='bold', color='#475569')
    ax.set_xlim(0, 110)
    ax.set_ylim(-0.5, len(dimensions) - 0.5)
    
    # Título
    ax.set_title("Evaluación de Riesgo e Integridad por Dimensión", fontsize=15, 
                 fontweight="bold", pad=20, color='#1E293B')
    
    # Estilo
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_color('#CBD5E1')
    ax.spines['left'].set_color('#CBD5E1')
    ax.set_facecolor('#FAFBFC')
    ax.tick_params(axis='x', colors='#94A3B8')
    ax.tick_params(axis='y', colors='#475569')
    ax.grid(axis='x', alpha=0.2, color='#CBD5E1', linestyle='-')
    
    # Leyenda
    ax.legend(fontsize=11, loc='lower right', framealpha=0.95)
    
    plt.tight_layout()
    return fig


def create_talent_map_radar(normalized_scores, job_profile_scores=None):
    """
    Crea un gráfico de radar para visualizar las 8 competencias del Talent Map.
    Opcionalmente muestra overlay con perfil de puesto para comparación.
    
    Args:
        normalized_scores: Dict con puntajes del candidato (0-100) por competencia
        job_profile_scores: Dict opcional con puntajes del perfil de puesto para comparar
        
    Returns:
        matplotlib.figure.Figure: Gráfico de radar
    """
    # Preparar datos para el radar
    competencies = TALENT_MAP_COMPETENCIES
    values = [normalized_scores[comp] for comp in competencies]
    values_closed = values + [values[0]]  # Cerrar el polígono
    
    # Calcular ángulos para cada competencia
    angles = np.linspace(0, 2 * np.pi, len(competencies), endpoint=False).tolist()
    angles_closed = angles + [angles[0]]
    
    # Crear figura
    fig, ax = plt.subplots(figsize=(7, 7), subplot_kw=dict(polar=True))
    
    # Línea principal del perfil del candidato
    ax.plot(angles_closed, values_closed, "o-", linewidth=3.5, color="#6366F1", 
            markersize=12, markerfacecolor="#818CF8", markeredgecolor="white", 
            markeredgewidth=3, zorder=5, label="Candidato")
    
    # Rellenar área del candidato
    ax.fill(angles_closed, values_closed, alpha=0.2, color="#6366F1")
    
    # Si hay perfil de puesto, agregarlo como comparación
    if job_profile_scores:
        profile_values = [job_profile_scores[comp] for comp in competencies]
        profile_values_closed = profile_values + [profile_values[0]]
        
        ax.plot(angles_closed, profile_values_closed, "s--", linewidth=2.5, color="#EF4444", 
                markersize=8, markerfacecolor="#FCA5A5", markeredgecolor="white", 
                markeredgewidth=2, zorder=4, label="Perfil Requerido", alpha=0.8)
        ax.fill(angles_closed, profile_values_closed, alpha=0.15, color="#EF4444")
    
    # Puntos coloreados por competencia con valores
    for i, (angle, val) in enumerate(zip(angles, values)):
        comp = competencies[i]
        point_color = TALENT_MAP_COLORS[comp]
        
        # Punto
        ax.plot(angle, val, "o", markersize=16, color=point_color, zorder=6, 
                markeredgecolor='white', markeredgewidth=2.5)
        # Valor del punto
        ax.text(angle, val + 6, f"{int(val)}", ha='center', va='center', 
                fontsize=11, fontweight='bold', color=point_color)
    
    # Configurar etiquetas de competencias con ajuste de tamaño
    ax.set_xticks(angles)
    labels = []
    for comp in competencies:
        # Dividir nombres largos en dos líneas
        if len(comp) > 15:
            words = comp.split()
            if len(words) >= 2:
                mid = len(words) // 2
                line1 = " ".join(words[:mid])
                line2 = " ".join(words[mid:])
                labels.append(f"{line1}\n{line2}")
            else:
                labels.append(comp)
        else:
            labels.append(comp)
    ax.set_xticklabels(labels, fontsize=10, fontweight="bold", color='#1E293B')
    
    # Configurar escala radial
    ax.set_ylim(0, 100)
    ax.set_yticks([25, 50, 75, 100])
    ax.set_yticklabels(['25', '50\n(Promedio)', '75', '100'], fontsize=9, color='#94A3B8')
    
    # Líneas de referencia
    ref_levels = [[50] * (len(competencies) + 1), [75] * (len(competencies) + 1)]
    ax.plot(angles_closed, ref_levels[0], ":", linewidth=1.5, color="#94A3B8", 
            alpha=0.6, label="Nivel Promedio (50)")
    ax.plot(angles_closed, ref_levels[1], "--", linewidth=1.5, color="#10B981", 
            alpha=0.6, label="Nivel Alto (75)")
    
    # Zonas de color de fondo
    theta = np.linspace(0, 2*np.pi, 100)
    ax.fill_between(theta, 0, 50, alpha=0.05, color='#EF4444')  # zona baja
    ax.fill_between(theta, 50, 75, alpha=0.05, color='#F59E0B')  # zona media
    ax.fill_between(theta, 75, 100, alpha=0.08, color='#10B981')  # zona alta
    
    # Estilo del gráfico
    ax.grid(True, alpha=0.3, color='#CBD5E1', linestyle='-', linewidth=0.8)
    ax.spines["polar"].set_visible(False)
    ax.set_facecolor('#FAFBFC')
    fig.patch.set_facecolor('white')
    
    # Título y leyenda
    title = "Mapeo de Competencias y Talentos"
    if job_profile_scores:
        title += "\n(Candidato vs. Perfil Requerido)"
    plt.title(title, fontsize=16, fontweight="bold", pad=40, color='#1E293B')
    plt.legend(loc="upper right", bbox_to_anchor=(1.35, 1.1), fontsize=11)
    
    plt.tight_layout()
    return fig


def create_talent_map_bars(normalized_scores, job_profile_scores=None):
    """
    Crea un gráfico de barras horizontales para visualizar las competencias del Talent Map.
    Opcionalmente incluye barras del perfil de puesto para comparación.
    
    Args:
        normalized_scores: Dict con puntajes del candidato (0-100) por competencia
        job_profile_scores: Dict opcional con puntajes del perfil de puesto
        
    Returns:
        matplotlib.figure.Figure: Gráfico de barras horizontales
    """
    fig, ax = plt.subplots(figsize=(9, 6))
    fig.patch.set_facecolor('white')
    
    competencies = TALENT_MAP_COMPETENCIES
    values = [normalized_scores[comp] for comp in competencies]
    
    # Si hay perfil de puesto, crear barras agrupadas
    y_positions = np.arange(len(competencies))
    bar_height = 0.35 if job_profile_scores else 0.7
    
    # Colores de barras según nivel
    colors = []
    for val in values:
        if val >= 75:
            colors.append("#10B981")  # Verde - Alto
        elif val >= 50:
            colors.append("#F59E0B")  # Amarillo - Medio
        else:
            colors.append("#EF4444")  # Rojo - Bajo
    
    # Crear barras del candidato
    if job_profile_scores:
        bars1 = ax.barh(y_positions - bar_height/2, values, bar_height, 
                       color=colors, alpha=0.85, edgecolor='white', 
                       linewidth=2, label="Candidato")
        
        # Barras del perfil requerido
        profile_values = [job_profile_scores[comp] for comp in competencies]
        bars2 = ax.barh(y_positions + bar_height/2, profile_values, bar_height, 
                       color="#94A3B8", alpha=0.7, edgecolor='white', 
                       linewidth=2, label="Perfil Requerido")
        
        # Agregar valores en las barras
        for bar, val in zip(bars1, values):
            ax.text(val + 2, bar.get_y() + bar.get_height()/2, f"{int(val)}", 
                    va='center', fontweight='bold', fontsize=10, color='#1E293B')
        
        for bar, val in zip(bars2, profile_values):
            ax.text(val + 2, bar.get_y() + bar.get_height()/2, f"{int(val)}", 
                    va='center', fontweight='bold', fontsize=10, color='#64748B')
    else:
        bars = ax.barh(y_positions, values, bar_height, color=colors, 
                      alpha=0.85, edgecolor='white', linewidth=2.5)
        
        # Agregar valores y nivel al final de cada barra
        for i, (bar, val, color) in enumerate(zip(bars, values, colors)):
            if val >= 75:
                level_label = "🌟 Alto"
            elif val >= 50:
                level_label = "👍 Medio"
            else:
                level_label = "📈 En Desarrollo"
            
            ax.text(val + 2, bar.get_y() + bar.get_height()/2, 
                    f"{int(val)}  {level_label}", 
                    va='center', fontweight='bold', fontsize=11, color=color)
    
    # Líneas de referencia verticales
    ax.axvline(x=50, color="#94A3B8", linestyle=":", alpha=0.6, linewidth=2, 
               label="Nivel Promedio (50)")
    ax.axvline(x=75, color="#10B981", linestyle="--", alpha=0.7, linewidth=2, 
               label="Nivel Alto (75)")
    
    # Zonas de color de fondo
    ax.axvspan(0, 50, alpha=0.05, color='#EF4444')  # Bajo
    ax.axvspan(50, 75, alpha=0.05, color='#F59E0B')  # Medio
    ax.axvspan(75, 100, alpha=0.08, color='#10B981')  # Alto
    
    # Configuración de ejes
    ax.set_yticks(y_positions)
    ax.set_yticklabels(competencies, fontsize=11, fontweight='bold', color='#1E293B')
    ax.set_xlabel('Puntuación (0-100)', fontsize=12, fontweight='bold', color='#475569')
    ax.set_xlim(0, 110)
    ax.set_ylim(-0.5, len(competencies) - 0.5)
    
    # Título
    title = "Evaluación de Competencias por Dimensión"
    if job_profile_scores:
        title += "\n(Candidato vs. Perfil Requerido)"
    ax.set_title(title, fontsize=15, fontweight="bold", pad=20, color='#1E293B')
    
    # Estilo
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_color('#CBD5E1')
    ax.spines['left'].set_color('#CBD5E1')
    ax.set_facecolor('#FAFBFC')
    ax.tick_params(axis='x', colors='#94A3B8')
    ax.tick_params(axis='y', colors='#475569')
    ax.grid(axis='x', alpha=0.2, color='#CBD5E1', linestyle='-')
    
    # Leyenda
    ax.legend(fontsize=11, loc='lower right', framealpha=0.95)
    
    plt.tight_layout()
    return fig


def create_talent_map_comparison(normalized_scores, job_profile_name, job_profile_scores):
    """
    Crea un gráfico de comparación detallada mostrando gaps y strengths vs perfil de puesto.
    
    Args:
        normalized_scores: Dict con puntajes del candidato
        job_profile_name: Nombre del perfil de puesto
        job_profile_scores: Dict con puntajes del perfil
        
    Returns:
        matplotlib.figure.Figure: Gráfico de comparación de gaps
    """
    fig, ax = plt.subplots(figsize=(9, 6))
    fig.patch.set_facecolor('white')
    
    competencies = TALENT_MAP_COMPETENCIES
    gaps = []
    gap_colors = []
    
    # Calcular gaps (positivo = excede, negativo = deficit)
    for comp in competencies:
        candidate = normalized_scores[comp]
        required = job_profile_scores[comp]
        gap = candidate - required
        gaps.append(gap)
        
        # Color según gap
        if gap >= 0:
            gap_colors.append("#10B981")  # Verde - Excede o cumple
        elif gap >= -15:
            gap_colors.append("#F59E0B")  # Amarillo - Gap moderado
        else:
            gap_colors.append("#EF4444")  # Rojo - Gap significativo
    
    # Crear barras de gap
    y_positions = np.arange(len(competencies))
    bars = ax.barh(y_positions, gaps, color=gap_colors, alpha=0.85, 
                   edgecolor='white', linewidth=2.5, height=0.7)
    
    # Agregar valores y etiquetas
    for i, (bar, gap) in enumerate(zip(bars, gaps)):
        comp = competencies[i]
        candidate_score = normalized_scores[comp]
        required_score = job_profile_scores[comp]
        
        # Texto del gap
        gap_text = f"{gap:+.0f}"
        if gap >= 0:
            label = f"{gap_text}  ✅ Excede"
            x_pos = gap + 2
        elif gap >= -15:
            label = f"{gap_text}  ⚠️ Gap moderado"
            x_pos = gap - 2
        else:
            label = f"{gap_text}  🚨 Gap crítico"
            x_pos = gap - 2
        
        ha = 'left' if gap >= 0 else 'right'
        ax.text(x_pos, bar.get_y() + bar.get_height()/2, label, 
                va='center', ha=ha, fontweight='bold', fontsize=10, 
                color=gap_colors[i])
        
        # Texto de puntajes (candidato vs requerido)
        score_text = f"Candidato: {candidate_score:.0f}  |  Requerido: {required_score:.0f}"
        ax.text(-42, bar.get_y() + bar.get_height()/2, score_text, 
                va='center', ha='left', fontsize=9, color='#64748B', style='italic')
    
    # Línea de referencia (gap = 0)
    ax.axvline(x=0, color='#1E293B', linestyle='-', linewidth=2.5, alpha=0.8)
    
    # Configuración de ejes
    ax.set_yticks(y_positions)
    ax.set_yticklabels(competencies, fontsize=11, fontweight='bold', color='#1E293B')
    ax.set_xlabel('Gap de Competencia (Candidato - Requerido)', fontsize=12, 
                  fontweight='bold', color='#475569')
    
    # Ajustar límites del eje X
    max_abs_gap = max(abs(min(gaps)), abs(max(gaps)))
    ax.set_xlim(-max_abs_gap - 20, max_abs_gap + 20)
    ax.set_ylim(-0.5, len(competencies) - 0.5)
    
    # Título
    profile_info = TALENT_MAP_JOB_PROFILES[job_profile_name]
    title = f"Análisis de Brechas vs. {profile_info['emoji']} {job_profile_name}\n{profile_info['descripcion']}"
    ax.set_title(title, fontsize=14, fontweight="bold", pad=20, color='#1E293B')
    
    # Estilo
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_color('#CBD5E1')
    ax.spines['left'].set_color('#CBD5E1')
    ax.set_facecolor('#FAFBFC')
    ax.tick_params(axis='both', colors='#475569')
    ax.grid(axis='x', alpha=0.2, color='#CBD5E1', linestyle='-')
    
    plt.tight_layout()
    return fig


def create_desempeno_radar(potencial_scores):
    """Crea radar chart para las 5 dimensiones de potencial."""
    fig = plt.figure(figsize=(10, 10))
    ax = fig.add_subplot(111, projection='polar')
    
    # Datos
    dimensiones = [dim["nombre"] for dim in DESEMPENO_DIMENSIONES]
    valores = [potencial_scores.get(i+1, 0) for i in range(5)]
    
    # Cerrar el polígono
    valores_plot = valores + [valores[0]]
    
    # Ángulos
    angulos = [n / 5 * 2 * np.pi for n in range(5)]
    angulos_plot = angulos + [angulos[0]]
    
    # Dibujar área
    ax.plot(angulos_plot, valores_plot, 'o-', linewidth=2.5, color='#3B82F6', markersize=8)
    ax.fill(angulos_plot, valores_plot, alpha=0.25, color='#3B82F6')
    
    # Zonas de fondo (0-1: rojo, 1-2: amarillo, 2-3: verde)
    for level, color, alpha in [(1, '#FEE2E2', 0.3), (2, '#FEF3C7', 0.3), (3, '#D1FAE5', 0.3)]:
        circle_angles = np.linspace(0, 2 * np.pi, 100)
        circle_values = [level] * 100
        ax.fill(circle_angles, circle_values, color=color, alpha=alpha)
    
    # Configuración
    ax.set_ylim(0, 3)
    ax.set_xticks(angulos)
    ax.set_xticklabels(dimensiones, size=11, fontweight='bold', color='#1E293B')
    ax.set_yticks([0.5, 1, 1.5, 2, 2.5, 3])
    ax.set_yticklabels(['', '1', '', '2', '', '3'], size=10, color='#64748B')
    ax.set_theta_offset(np.pi / 2)
    ax.set_theta_direction(-1)
    ax.grid(True, color='#CBD5E1', linestyle='-', linewidth=0.5, alpha=0.7)
    ax.set_facecolor('#FFFFFF')
    
    # Título
    ax.set_title('Evaluación de Potencial\n5 Dimensiones', size=14, fontweight='bold', 
                 pad=30, color='#1E293B')
    
    plt.tight_layout()
    return fig


def create_desempeno_bars(rendimiento_scores):
    """Crea gráfico de barras para los 6 objetivos de rendimiento."""
    fig, ax = plt.subplots(figsize=(12, 8))
    
    # Datos
    objetivos = [obj["titulo"] for obj in DESEMPENO_OBJETIVOS]
    valores = [rendimiento_scores.get(i+1, 0) for i in range(6)]
    
    # Colores según nivel
    colores = []
    for valor in valores:
        if valor >= 4.5:
            colores.append('#10B981')  # Verde
        elif valor >= 3.5:
            colores.append('#3B82F6')  # Azul
        elif valor >= 2.5:
            colores.append('#F59E0B')  # Amarillo
        elif valor >= 1.5:
            colores.append('#EF4444')  # Rojo
        else:
            colores.append('#991B1B')  # Rojo oscuro
    
    # Crear barras horizontales
    y_positions = range(len(objetivos))
    bars = ax.barh(y_positions, valores, color=colores, alpha=0.8, height=0.6, 
                   edgecolor='#1E293B', linewidth=1.5)
    
    # Agregar valores en las barras
    for i, (bar, valor) in enumerate(zip(bars, valores)):
        label = DESEMPENO_ESCALA_RENDIMIENTO[int(valor)]["label"]
        ax.text(valor + 0.15, bar.get_y() + bar.get_height()/2, 
                f'{valor:.1f} - {label}', 
                va='center', ha='left', fontsize=10, fontweight='bold', color='#1E293B')
    
    # Zonas de fondo
    ax.axvspan(0, 1.5, alpha=0.1, color='#EF4444', label='Insatisfactorio')
    ax.axvspan(1.5, 2.5, alpha=0.1, color='#F59E0B', label='Debajo')
    ax.axvspan(2.5, 3.5, alpha=0.1, color='#3B82F6', label='Cumple')
    ax.axvspan(3.5, 5, alpha=0.1, color='#10B981', label='Supera/Sobresaliente')
    
    # Configuración
    ax.set_yticks(y_positions)
    ax.set_yticklabels(objetivos, fontsize=11, fontweight='bold', color='#1E293B')
    ax.set_xlabel('Calificación (1-5)', fontsize=12, fontweight='bold', color='#475569')
    ax.set_xlim(0, 5.5)
    ax.set_title('Evaluación de Rendimiento\n6 Objetivos de Desempeño', 
                 fontsize=14, fontweight='bold', pad=20, color='#1E293B')
    
    # Estilo
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_color('#CBD5E1')
    ax.spines['left'].set_color('#CBD5E1')
    ax.set_facecolor('#FAFBFC')
    ax.tick_params(axis='both', colors='#475569')
    ax.grid(axis='x', alpha=0.3, color='#CBD5E1', linestyle='--')
    ax.legend(loc='lower right', fontsize=9, framealpha=0.9)
    
    plt.tight_layout()
    return fig


# =========================================================================
# TIMER (JavaScript countdown)
# =========================================================================

def render_timer(deadline_ts, session_id):
    """Render a real-time JavaScript countdown timer."""
    html = f"""
    <div id="timer-box" style="
        display: flex; align-items: center; justify-content: center; gap: 12px;
        background: linear-gradient(135deg, #1e40af, #3b82f6); color: white;
        padding: 14px 24px; border-radius: 12px; font-family: 'Segoe UI', sans-serif;
        box-shadow: 0 4px 12px rgba(30,64,175,0.3); margin-bottom: 8px;">
        <span style="font-size: 16px;">⏱️ Tiempo restante:</span>
        <span id="countdown" style="font-size: 28px; font-weight: bold; font-family: monospace; letter-spacing: 2px;">--:--</span>
        <span style="font-size: 12px; opacity: 0.8;">ID: {session_id}</span>
    </div>
    <script>
    var deadline = new Date({deadline_ts * 1000});
    function updateTimer() {{
        var now = new Date();
        var remaining = deadline - now;
        var box = document.getElementById("timer-box");
        var cd = document.getElementById("countdown");
        if (remaining <= 0) {{
            cd.textContent = "⏰ TIEMPO AGOTADO";
            box.style.background = "linear-gradient(135deg, #dc2626, #ef4444)";
        }} else {{
            var hrs = Math.floor(remaining / 3600000);
            var mins = Math.floor((remaining % 3600000) / 60000);
            var secs = Math.floor((remaining % 60000) / 1000);
            var display = "";
            if (hrs > 0) display = String(hrs).padStart(2,"0") + ":";
            display += String(mins).padStart(2,"0") + ":" + String(secs).padStart(2,"0");
            cd.textContent = display;
            if (remaining < 300000) {{
                box.style.background = "linear-gradient(135deg, #dc2626, #f59e0b)";
            }} else if (remaining < 600000) {{
                box.style.background = "linear-gradient(135deg, #f59e0b, #eab308)";
            }}
        }}
    }}
    updateTimer();
    setInterval(updateTimer, 1000);
    </script>
    """
    components.html(html, height=65)


# =========================================================================
# PDF GENERATION
# =========================================================================

def generate_disc_pdf(candidate, normalized, relative, fig, session_id, completed_at=None, analysis=None,
                      behavioral_styles=None, temperament=None, mega_summary=None, styles_fig=None):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Justify", alignment=4, leading=14))
    styles.add(ParagraphStyle(name="SmallBold", parent=styles["Normal"], fontSize=9, leading=12, fontName="Helvetica-Bold"))
    styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], fontSize=9, leading=12))
    story = []
    story.append(Paragraph("Evaluación de Personalidad DISC - Reporte", styles["Title"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"<b>ID Evaluación:</b> {session_id}", styles["Normal"]))
    story.append(Paragraph(f"<b>Candidato:</b> {candidate['name']} (Cédula: {candidate['cedula']})", styles["Normal"]))
    
    # Formatear la fecha de presentación
    if completed_at:
        try:
            fecha_obj = datetime.strptime(completed_at, "%Y-%m-%d %H:%M:%S")
            fecha_str = fecha_obj.strftime('%d/%m/%Y %H:%M')
        except:
            fecha_str = completed_at
    else:
        fecha_str = _now_gmt5().strftime('%d/%m/%Y %H:%M')
    
    story.append(Paragraph(f"<b>Cargo:</b> {candidate.get('position','N/A')} | <b>Fecha de Presentación:</b> {fecha_str}", styles["Normal"]))
    
    # Generar análisis si no se proporcionó
    if analysis is None:
        analysis = analyze_disc_aptitude(normalized, relative)
    
    # Sección de aptitud
    story.append(Spacer(1, 12))
    apt_color = analysis['aptitude_color']
    story.append(Paragraph(f"<b>RESULTADO DE APTITUD: {analysis['aptitude_level']} ({analysis['aptitude_score']}/100)</b>", styles["Heading2"]))
    story.append(Paragraph(f"{analysis['aptitude_desc']}", styles["Normal"]))
    story.append(Paragraph(f"<b>Perfil:</b> {analysis['profile_name']} ({analysis['dominant_name']} + {analysis['secondary_name']})", styles["Normal"]))
    story.append(Spacer(1, 12))
    
    # Tabla de puntajes
    data = [["Estilo", "Puntaje Normalizado", "Porcentaje Relativo"]]
    for s in "DISC":
        data.append([s, f"{normalized[s]:.1f}%", f"{relative[s]:.1f}%"])
    t = Table(data, colWidths=[100, 150, 150])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(t)
    story.append(Spacer(1, 16))
    if fig:
        img_buf = BytesIO()
        fig.savefig(img_buf, format="png", dpi=150, bbox_inches="tight")
        img_buf.seek(0)
        story.append(Image(img_buf, width=280, height=280))
    
    # Página de recomendaciones
    story.append(PageBreak())
    story.append(Paragraph("Análisis y Recomendaciones", styles["Heading1"]))
    story.append(Spacer(1, 10))
    
    # Fortalezas
    story.append(Paragraph("FORTALEZAS DEL CANDIDATO", styles["Heading2"]))
    for f in analysis.get('fortalezas', []):
        story.append(Paragraph(f"• {f}", styles["Small"]))
    story.append(Spacer(1, 10))
    
    # Alertas
    story.append(Paragraph("ALERTAS Y ÁREAS DE ATENCIÓN", styles["Heading2"]))
    for a in analysis.get('alertas', []):
        story.append(Paragraph(f"• {a}", styles["Small"]))
    story.append(Spacer(1, 10))
    
    # Recomendaciones
    story.append(Paragraph("RECOMENDACIONES", styles["Heading2"]))
    for r in analysis.get('recomendaciones', []):
        story.append(Paragraph(f"• {r}", styles["Small"]))
    story.append(Spacer(1, 10))
    
    # Roles ideales
    if analysis.get('ideal_para'):
        story.append(Paragraph("ROLES IDEALES", styles["Heading2"]))
        for r in analysis['ideal_para']:
            story.append(Paragraph(f"• {r}", styles["Small"]))
        story.append(Spacer(1, 10))
    
    if analysis.get('cuidado_en'):
        story.append(Paragraph("PRECAUCIÓN EN ROLES DE", styles["Heading2"]))
        for r in analysis['cuidado_en']:
            story.append(Paragraph(f"• {r}", styles["Small"]))

    # ── MEGA RESUMEN CONDUCTUAL ──────────────────────────────────────────
    if mega_summary:
        story.append(PageBreak())
        story.append(Paragraph("Resumen Conductual Detallado", styles["Heading1"]))
        story.append(Spacer(1, 8))
        if temperament:
            story.append(Paragraph(
                f"<b>Temperamento:</b> {temperament['label'].capitalize()} — {temperament['description']}",
                styles["Normal"]
            ))
            story.append(Spacer(1, 8))
        data_rows = [["Dimensión", "Descripción Conductual"]]
        for label, text in mega_summary.items():
            data_rows.append([label, text])
        t_sum = Table(data_rows, colWidths=[130, 360])
        t_sum.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F8FAFC"), colors.white]),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, 1), (0, -1), colors.HexColor("#1e40af")),
        ]))
        story.append(t_sum)

    # ── ESTILOS CONDUCTUALES ─────────────────────────────────────────────
    if behavioral_styles:
        story.append(PageBreak())
        story.append(Paragraph("9 Estilos Conductuales Derivados", styles["Heading1"]))
        story.append(Spacer(1, 6))
        story.append(Paragraph(
            "Puntajes derivados matemáticamente del perfil DISC. Cada estilo presenta 4 sub-dimensiones "
            "mapeadas a Dominancia (D), Influencia (I), Estabilidad (S) y Cumplimiento (C).",
            styles["Small"]
        ))
        story.append(Spacer(1, 10))

        if styles_fig:
            try:
                styles_buf = BytesIO()
                styles_fig.savefig(styles_buf, format="png", dpi=130, bbox_inches="tight")
                styles_buf.seek(0)
                story.append(Image(styles_buf, width=480, height=len(behavioral_styles) * 55 + 30))
            except Exception:
                pass
        
        story.append(Spacer(1, 12))
        for style_name, style_data in behavioral_styles.items():
            story.append(Paragraph(f"<b>{style_name}</b>", styles["Heading3"]))
            sub_data = [["Sub-dimensión", "Puntaje", "Descripción"]]
            for sub_name, sub_val in style_data["subs"].items():
                sub_data.append([sub_name, str(sub_val), style_data["desc"][sub_name]])
            t_sub = Table(sub_data, colWidths=[110, 45, 335])
            t_sub.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (1, 0), (1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F9FAFB"), colors.white]),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(t_sub)
            story.append(Spacer(1, 8))

    story.append(Spacer(1, 20))
    story.append(Paragraph("<i>Este reporte es generado automáticamente como herramienta de apoyo para Recursos Humanos. Los resultados deben complementarse con entrevistas y otras evaluaciones.</i>", styles["Small"]))
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_valanti_pdf(candidate, direct, standard, radar_fig, session_id, completed_at=None, analysis=None):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Justify", alignment=4, leading=14))
    styles.add(ParagraphStyle(name="SmallBold", parent=styles["Normal"], fontSize=9, leading=12, fontName="Helvetica-Bold"))
    styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], fontSize=9, leading=12))
    story = []
    story.append(Paragraph("Cuestionario VALANTI - Reporte de Resultados", styles["Title"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"<b>ID Evaluación:</b> {session_id}", styles["Normal"]))
    story.append(Paragraph(f"<b>Candidato:</b> {candidate['name']} (Cédula: {candidate['cedula']})", styles["Normal"]))
    
    # Formatear la fecha de presentación
    if completed_at:
        try:
            fecha_obj = datetime.strptime(completed_at, "%Y-%m-%d %H:%M:%S")
            fecha_str = fecha_obj.strftime('%d/%m/%Y %H:%M')
        except:
            fecha_str = completed_at
    else:
        fecha_str = _now_gmt5().strftime('%d/%m/%Y %H:%M')
    
    story.append(Paragraph(f"<b>Cargo:</b> {candidate.get('position','N/A')} | <b>Fecha de Presentación:</b> {fecha_str}", styles["Normal"]))
    
    # Generar análisis si no se proporcionó
    if analysis is None:
        analysis = analyze_valanti_aptitude(standard)
    
    # Sección de aptitud
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"<b>RESULTADO DE APTITUD: {analysis['aptitude_level']} ({analysis['aptitude_score']}/100)</b>", styles["Heading2"]))
    story.append(Paragraph(f"{analysis['aptitude_desc']}", styles["Normal"]))
    story.append(Paragraph(f"<b>Valor más fuerte:</b> {analysis['strongest_value']} (T={analysis['strongest_score']}) | <b>Valor más bajo:</b> {analysis['weakest_value']} (T={analysis['weakest_score']})", styles["Normal"]))
    story.append(Spacer(1, 12))
    
    # Tabla de puntajes
    data = [["Valor", "Puntaje Directo", "Puntaje Estándar (T)"]]
    for trait in VALANTI_TRAITS:
        data.append([trait, str(direct[trait]), str(standard[trait])])
    t = Table(data, colWidths=[120, 120, 150])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(t)
    story.append(Spacer(1, 16))
    if radar_fig:
        img_buf = BytesIO()
        radar_fig.savefig(img_buf, format="png", dpi=150, bbox_inches="tight")
        img_buf.seek(0)
        story.append(Image(img_buf, width=300, height=300))
    
    # Página de recomendaciones
    story.append(PageBreak())
    story.append(Paragraph("Análisis y Recomendaciones", styles["Heading1"]))
    story.append(Spacer(1, 10))
    
    # Fortalezas
    if analysis.get('fortalezas'):
        story.append(Paragraph("FORTALEZAS VALORALES", styles["Heading2"]))
        for f in analysis['fortalezas']:
            story.append(Paragraph(f"• {f}", styles["Small"]))
        story.append(Spacer(1, 10))
    
    # Alertas
    if analysis.get('alertas'):
        story.append(Paragraph("ALERTAS Y ÁREAS DE ATENCIÓN", styles["Heading2"]))
        for a in analysis['alertas']:
            story.append(Paragraph(f"• {a}", styles["Small"]))
        story.append(Spacer(1, 10))
    
    # Recomendaciones
    if analysis.get('recomendaciones'):
        story.append(Paragraph("RECOMENDACIONES", styles["Heading2"]))
        for r in analysis['recomendaciones']:
            # Limpiar markdown para PDF
            r_clean = r.replace("**", "")
            story.append(Paragraph(f"• {r_clean}", styles["Small"]))
        story.append(Spacer(1, 10))
    
    story.append(Spacer(1, 20))
    story.append(Paragraph("<i>Este reporte es generado automáticamente como herramienta de apoyo para Recursos Humanos. Los resultados deben complementarse con entrevistas y otras evaluaciones.</i>", styles["Small"]))
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_wpi_pdf(candidate, raw_scores, normalized, radar_fig, session_id, completed_at=None, analysis=None):
    """
    Genera un PDF con los resultados del WPI (Work Personality Index).
    
    Args:
        candidate: Dict con información del candidato
        raw_scores: Puntajes directos por dimensión
        normalized: Puntajes normalizados (0-100) por dimensión
        radar_fig: Figura matplotlib del radar
        session_id: ID de la sesión
        completed_at: Fecha de completación (opcional)
        analysis: Dict con análisis de aptitud (opcional, se genera si no se proporciona)
        
    Returns:
        BytesIO: Buffer con el PDF generado
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=40, bottomMargin=40)
    
    # Estilos
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Justify", alignment=4, leading=14))
    styles.add(ParagraphStyle(name="SmallBold", parent=styles["Normal"], 
                             fontSize=9, leading=12, fontName="Helvetica-Bold"))
    styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], 
                             fontSize=9, leading=12))
    
    story = []
    
    # === PÁGINA 1: PORTADA Y RESULTADOS ===
    story.append(Paragraph("WPI - Work Personality Index", styles["Title"]))
    story.append(Paragraph("Evaluación de Personalidad Laboral", styles["Heading2"]))
    story.append(Spacer(1, 12))
    
    # Información del candidato
    story.append(Paragraph(f"<b>ID Evaluación:</b> {session_id}", styles["Normal"]))
    story.append(Paragraph(f"<b>Candidato:</b> {candidate['name']}", styles["Normal"]))
    story.append(Paragraph(f"<b>Cédula:</b> {candidate['cedula']}", styles["Normal"]))
    story.append(Paragraph(f"<b>Cargo:</b> {candidate.get('position', 'N/A')}", styles["Normal"]))
    
    # Formatear fecha
    if completed_at:
        try:
            fecha_obj = datetime.strptime(completed_at, "%Y-%m-%d %H:%M:%S")
            fecha_str = fecha_obj.strftime('%d/%m/%Y %H:%M')
        except:
            fecha_str = completed_at
    else:
        fecha_str = _now_gmt5().strftime('%d/%m/%Y %H:%M')
    
    story.append(Paragraph(f"<b>Fecha de Presentación:</b> {fecha_str}", styles["Normal"]))
    story.append(Spacer(1, 16))
    
    # Generar análisis si no se proporcionó
    if analysis is None:
        analysis = analyze_wpi_aptitude(normalized)
    
    # === RESULTADO DE APTITUD ===
    story.append(Paragraph(
        f"<b>RESULTADO: {analysis['aptitude_level']} ({analysis['aptitude_score']}/100)</b>", 
        styles["Heading2"]
    ))
    story.append(Paragraph(f"{analysis['aptitude_desc']}", styles["Normal"]))
    story.append(Spacer(1, 8))
    story.append(Paragraph(
        f"<b>Dimensión más fuerte:</b> {analysis['strongest_dimension']} "
        f"({int(analysis['strongest_score'])}/100) | "
        f"<b>Dimensión a desarrollar:</b> {analysis['weakest_dimension']} "
        f"({int(analysis['weakest_score'])}/100)",
        styles["Normal"]
    ))
    story.append(Paragraph(
        f"<b>Promedio general:</b> {analysis['average_score']}/100",
        styles["Normal"]
    ))
    story.append(Spacer(1, 16))
    
    # === TABLA DE PUNTAJES ===
    story.append(Paragraph("Puntajes por Dimensión", styles["Heading2"]))
    story.append(Spacer(1, 8))
    
    data = [["Dimensión", "Puntaje Directo", "Puntaje Normalizado (0-100)", "Nivel"]]
    for dim in WPI_DIMENSIONS:
        nivel = "Alto" if normalized[dim] >= 70 else ("Medio" if normalized[dim] >= 45 else "Bajo")
        data.append([
            dim,
            str(int(raw_scores[dim])),
            f"{int(normalized[dim])}/100",
            nivel
        ])
    
    t = Table(data, colWidths=[140, 80, 130, 60])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 16))
    
    # === GRÁFICO RADAR ===
    if radar_fig:
        img_buf = BytesIO()
        radar_fig.savefig(img_buf, format="png", dpi=150, bbox_inches="tight")
        img_buf.seek(0)
        story.append(Image(img_buf, width=320, height=320))
    
    # === PÁGINA 2: ANÁLISIS DETALLADO ===
    story.append(PageBreak())
    story.append(Paragraph("Análisis Detallado y Recomendaciones", styles["Heading1"]))
    story.append(Spacer(1, 12))
    
    # === FORTALEZAS ===
    if analysis.get('fortalezas'):
        story.append(Paragraph("✅ FORTALEZAS DESTACADAS", styles["Heading2"]))
        story.append(Spacer(1, 6))
        for f in analysis['fortalezas']:
            # Limpiar markdown para PDF
            f_clean = f.replace("**", "")
            story.append(Paragraph(f"• {f_clean}", styles["Small"]))
        story.append(Spacer(1, 12))
    
    # === ALERTAS ===
    if analysis.get('alertas'):
        story.append(Paragraph("⚠️ ÁREAS DE ATENCIÓN", styles["Heading2"]))
        story.append(Spacer(1, 6))
        for a in analysis['alertas']:
            # Limpiar markdown para PDF
            a_clean = a.replace("**", "").replace("⚠️ ", "")
            story.append(Paragraph(f"• {a_clean}", styles["Small"]))
        story.append(Spacer(1, 12))
    
    # === ROLES IDEALES ===
    if analysis.get('ideal_para'):
        story.append(Paragraph("🎯 ROLES IDEALES PARA EL CANDIDATO", styles["Heading2"]))
        story.append(Spacer(1, 6))
        for role in analysis['ideal_para']:
            story.append(Paragraph(f"• {role}", styles["Small"]))
        story.append(Spacer(1, 12))
    
    # === ROLES A EVITAR ===
    if analysis.get('avoid_roles'):
        story.append(Paragraph("⛔ ROLES NO RECOMENDADOS", styles["Heading2"]))
        story.append(Spacer(1, 6))
        for role in analysis['avoid_roles']:
            story.append(Paragraph(f"• {role}", styles["Small"]))
        story.append(Spacer(1, 12))
    
    # === RECOMENDACIONES ===
    if analysis.get('recomendaciones'):
        story.append(Paragraph("💡 RECOMENDACIONES ESPECÍFICAS", styles["Heading2"]))
        story.append(Spacer(1, 6))
        for r in analysis['recomendaciones']:
            # Limpiar markdown para PDF
            r_clean = r.replace("**", "")
            story.append(Paragraph(f"• {r_clean}", styles["Small"]))
        story.append(Spacer(1, 12))
    
    # === DESCRIPCIÓN DE DIMENSIONES ===
    story.append(PageBreak())
    story.append(Paragraph("Descripción de las Dimensiones del WPI", styles["Heading1"]))
    story.append(Spacer(1, 12))
    
    for dim in WPI_DIMENSIONS:
        score = normalized[dim]
        desc_info = WPI_DESCRIPTIONS[dim]
        
        # Determinar nivel y descripción
        if score >= 70:
            level_text = "ALTO"
            desc_text = desc_info["high"]
        elif score >= 45:
            level_text = "MEDIO"
            desc_text = desc_info["medium"]
        else:
            level_text = "BAJO"
            desc_text = desc_info["low"]
        
        story.append(Paragraph(
            f"<b>{desc_info['title']}</b> - Nivel: {level_text} ({int(score)}/100)",
            styles["Heading3"]
        ))
        story.append(Paragraph(desc_text, styles["Small"]))
        story.append(Spacer(1, 8))
    
    # === FOOTER ===
    story.append(Spacer(1, 20))
    story.append(Paragraph(
        "<i>Este reporte es generado automáticamente como herramienta de apoyo para "
        "Recursos Humanos. Los resultados deben complementarse con entrevistas, "
        "referencias laborales y otras evaluaciones pertinentes.</i>",
        styles["Small"]
    ))
    
    # Construir PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_eri_pdf(candidate, raw_scores, normalized, radar_fig, session_id, completed_at=None, analysis=None, validity_score=None, validity_flags=None):
    """
    Genera un PDF con los resultados del ERI (Evaluación de Riesgo e Integridad).
    
    Args:
        candidate: Dict con información del candidato
        raw_scores: Puntajes directos por dimensión
        normalized: Puntajes normalizados (0-100) por dimensión (100 = bajo riesgo)
        radar_fig: Figura matplotlib del radar
        session_id: ID de la sesión
        completed_at: Fecha de completación (opcional)
        analysis: Dict con análisis de riesgo (opcional, se genera si no se proporciona)
        validity_score: Puntaje de validez del test (0-12)
        validity_flags: Lista de alertas de validez
        
    Returns:
        BytesIO: Buffer con el PDF generado
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=40, bottomMargin=40)
    
    # Estilos
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Justify", alignment=4, leading=14))
    styles.add(ParagraphStyle(name="SmallBold", parent=styles["Normal"], 
                             fontSize=9, leading=12, fontName="Helvetica-Bold"))
    styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], 
                             fontSize=9, leading=12))
    styles.add(ParagraphStyle(name="AlertBold", parent=styles["Normal"],
                             fontSize=11, leading=14, fontName="Helvetica-Bold",
                             textColor=colors.HexColor("#DC2626")))
    
    story = []
    
    # === PÁGINA 1: PORTADA Y RESULTADOS ===
    story.append(Paragraph("ERI - Evaluación de Riesgo e Integridad", styles["Title"]))
    story.append(Paragraph("Screening de Confiabilidad y Comportamiento Laboral", styles["Heading2"]))
    story.append(Spacer(1, 12))
    
    # Información del candidato
    story.append(Paragraph(f"<b>ID Evaluación:</b> {session_id}", styles["Normal"]))
    story.append(Paragraph(f"<b>Candidato:</b> {candidate['name']}", styles["Normal"]))
    story.append(Paragraph(f"<b>Cédula:</b> {candidate['cedula']}", styles["Normal"]))
    story.append(Paragraph(f"<b>Cargo:</b> {candidate.get('position', 'N/A')}", styles["Normal"]))
    
    # Formatear fecha
    if completed_at:
        try:
            fecha_obj = datetime.strptime(completed_at, "%Y-%m-%d %H:%M:%S")
            fecha_str = fecha_obj.strftime('%d/%m/%Y %H:%M')
        except:
            fecha_str = completed_at
    else:
        fecha_str = _now_gmt5().strftime('%d/%m/%Y %H:%M')
    
    story.append(Paragraph(f"<b>Fecha de Presentación:</b> {fecha_str}", styles["Normal"]))
    story.append(Spacer(1, 16))
    
    # Generar análisis si no se proporcionó
    if analysis is None:
        if validity_score is None:
            validity_score = ERI_VALIDITY_QUESTIONS_COUNT
        if validity_flags is None:
            validity_flags = []
        analysis = analyze_eri_aptitude(normalized, validity_score, validity_flags)
    
    # === BANNER DE VALIDEZ (si aplica) ===
    if analysis.get('validity_warning'):
        story.append(Paragraph("⚠️ ALERTA DE VALIDEZ DEL TEST", styles["AlertBold"]))
        story.append(Paragraph(analysis['validity_warning'], styles["Small"]))
        story.append(Spacer(1, 12))
    
    # === RESULTADO DE RIESGO ===
    risk_color_map = {
        "#10B981": "✅ BAJO RIESGO",
        "#F59E0B": "⚠️ RIESGO MODERADO",
        "#EF4444": "🚫 ALTO RIESGO"
    }
    risk_banner = risk_color_map.get(analysis['risk_color'], analysis['risk_level'])
    
    story.append(Paragraph(
        f"<b>RESULTADO: {risk_banner} ({analysis['risk_score']:.1f}/100)</b>", 
        styles["Heading2"]
    ))
    story.append(Paragraph(f"{analysis['risk_desc']}", styles["Normal"]))
    story.append(Spacer(1, 8))
    story.append(Paragraph(
        f"<b>Dimensión de menor riesgo:</b> {analysis['safest_dimension']} "
        f"({int(analysis['safest_score'])}/100) | "
        f"<b>Dimensión de mayor riesgo:</b> {analysis['riskiest_dimension']} "
        f"({int(analysis['riskiest_score'])}/100)",
        styles["Normal"]
    ))
    story.append(Paragraph(
        f"<b>Promedio de riesgo:</b> {analysis['average_score']:.1f}/100 "
        f"(Puntajes altos = BAJO riesgo)",
        styles["Normal"]
    ))
    story.append(Spacer(1, 8))
    story.append(Paragraph(
        f"<b>Decisión Recomendada:</b> {analysis['hiring_decision']}",
        styles["Heading3"]
    ))
    story.append(Spacer(1, 16))
    
    # === TABLA DE PUNTAJES ===
    story.append(Paragraph("Puntajes por Dimensión de Riesgo", styles["Heading2"]))
    story.append(Spacer(1, 8))
    
    data = [["Dimensión", "Puntaje", "Nivel de Riesgo", "Estado"]]
    for dim in ERI_DIMENSIONS:
        score = normalized[dim]
        if score >= ERI_RISK_THRESHOLDS["low_risk"]:
            nivel = "Bajo Riesgo"
            estado = "✅"
        elif score >= ERI_RISK_THRESHOLDS["medium_risk"]:
            nivel = "Riesgo Moderado"
            estado = "⚠️"
        else:
            nivel = "Alto Riesgo"
            estado = "🚨"
        
        data.append([
            dim,
            f"{int(score)}/100",
            nivel,
            estado
        ])
    
    t = Table(data, colWidths=[140, 70, 110, 50])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DC2626")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 16))
    
    # === GRÁFICO RADAR ===
    if radar_fig:
        img_buf = BytesIO()
        radar_fig.savefig(img_buf, format="png", dpi=150, bbox_inches="tight")
        img_buf.seek(0)
        story.append(Image(img_buf, width=350, height=350))
    
    # === PÁGINA 2: ANÁLISIS DETALLADO ===
    story.append(PageBreak())
    story.append(Paragraph("Análisis Detallado y Recomendaciones de Contratación", styles["Heading1"]))
    story.append(Spacer(1, 12))
    
    # === FORTALEZAS ===
    if analysis.get('fortalezas'):
        story.append(Paragraph("✅ ASPECTOS POSITIVOS (Bajo Riesgo)", styles["Heading2"]))
        story.append(Spacer(1, 6))
        for f in analysis['fortalezas']:
            # Limpiar markdown para PDF
            f_clean = f.replace("**", "")
            story.append(Paragraph(f"• {f_clean}", styles["Small"]))
        story.append(Spacer(1, 12))
    
    # === ALERTAS ===
    if analysis.get('alertas'):
        story.append(Paragraph("🚨 SEÑALES DE ALERTA Y FACTORES DE RIESGO", styles["Heading2"]))
        story.append(Spacer(1, 6))
        for a in analysis['alertas']:
            # Limpiar markdown para PDF
            a_clean = a.replace("**", "").replace("⚠️ ", "").replace("🚨 ", "")
            story.append(Paragraph(f"• {a_clean}", styles["Small"]))
        story.append(Spacer(1, 12))
    
    # === RECOMENDACIONES ===
    if analysis.get('recomendaciones'):
        story.append(Paragraph("💼 RECOMENDACIONES DE CONTRATACIÓN", styles["Heading2"]))
        story.append(Spacer(1, 6))
        for r in analysis['recomendaciones']:
            # Limpiar markdown para PDF
            r_clean = r.replace("**", "")
            story.append(Paragraph(f"{r_clean}", styles["Small"]))
        story.append(Spacer(1, 8))
    
    # === FLAGS DE VALIDEZ ===
    if validity_flags and len(validity_flags) > 0:
        story.append(PageBreak())
        story.append(Paragraph("⚠️ DETALLES DE VALIDEZ DEL TEST", styles["Heading2"]))
        story.append(Spacer(1, 8))
        story.append(Paragraph(
            f"Se detectaron {len(validity_flags)} respuestas poco realistas en preguntas de validez. "
            "Esto puede indicar que el candidato está tratando de presentarse de forma irrealmente perfecta.",
            styles["Small"]
        ))
        story.append(Spacer(1, 8))
        story.append(Paragraph("Ejemplos de respuestas sospechosas:", styles["SmallBold"]))
        story.append(Spacer(1, 4))
        for flag in validity_flags[:5]:  # Mostrar máximo 5 ejemplos
            story.append(Paragraph(f"• {flag}", styles["Small"]))
        if len(validity_flags) > 5:
            story.append(Paragraph(f"... y {len(validity_flags) - 5} más.", styles["Small"]))
        story.append(Spacer(1, 12))
    
    # === DESCRIPCIÓN DE DIMENSIONES ===
    story.append(PageBreak())
    story.append(Paragraph("Descripción de las Dimensiones del ERI", styles["Heading1"]))
    story.append(Spacer(1, 12))
    
    for dim in ERI_DIMENSIONS:
        score = normalized[dim]
        desc_info = ERI_DESCRIPTIONS[dim]
        
        # Determinar nivel y descripción (invertido: alto score = bajo riesgo)
        if score >= ERI_RISK_THRESHOLDS["low_risk"]:
            level_text = "BAJO RIESGO ✅"
            desc_text = desc_info["low_risk"]
        elif score >= ERI_RISK_THRESHOLDS["medium_risk"]:
            level_text = "RIESGO MODERADO ⚠️"
            desc_text = desc_info["medium_risk"]
        else:
            level_text = "ALTO RIESGO 🚨"
            desc_text = desc_info["high_risk"]
        
        story.append(Paragraph(
            f"<b>{desc_info['title']}</b> - {level_text} ({int(score)}/100)",
            styles["Heading3"]
        ))
        story.append(Paragraph(desc_text, styles["Small"]))
        story.append(Spacer(1, 8))
    
    # === INTERPRETACIÓN DE UMBRALES ===
    story.append(PageBreak())
    story.append(Paragraph("Interpretación de Umbrales de Riesgo", styles["Heading1"]))
    story.append(Spacer(1, 12))
    
    story.append(Paragraph("<b>✅ BAJO RIESGO (66-100 puntos):</b>", styles["Heading3"]))
    story.append(Paragraph(
        "Sin indicadores significativos de riesgo. El candidato muestra actitudes y comportamientos "
        "compatibles con un desempeño confiable y ético en el entorno laboral.",
        styles["Small"]
    ))
    story.append(Spacer(1, 8))
    
    story.append(Paragraph("<b>⚠️ RIESGO MODERADO (41-65 puntos):</b>", styles["Heading3"]))
    story.append(Paragraph(
        "Señales de alerta moderadas. Se recomienda profundizar con entrevistas enfocadas, "
        "referencias laborales exhaustivas y período de prueba con supervisión cercana.",
        styles["Small"]
    ))
    story.append(Spacer(1, 8))
    
    story.append(Paragraph("<b>🚨 ALTO RIESGO (0-40 puntos):</b>", styles["Heading3"]))
    story.append(Paragraph(
        "Múltiples indicadores de riesgo significativo. La contratación representa riesgo elevado "
        "para la organización en términos de pérdidas, conflictos, accidentes o incumplimiento normativo. "
        "Se recomienda NO CONTRATAR o requerir evaluación psicológica profesional adicional.",
        styles["Small"]
    ))
    story.append(Spacer(1, 12))
    
    # === LIMITACIONES Y DISCLAIMERS ===
    story.append(Paragraph("Limitaciones y Consideraciones Importantes", styles["Heading2"]))
    story.append(Spacer(1, 8))
    story.append(Paragraph(
        "• Este test es una herramienta de SCREENING, no un diagnóstico psicológico definitivo.",
        styles["Small"]
    ))
    story.append(Paragraph(
        "• Los resultados deben complementarse con: entrevistas conductuales (STAR), "
        "referencias laborales verificables, verificación de antecedentes penales y laborales.",
        styles["Small"]
    ))
    story.append(Paragraph(
        "• Ningún test psicométrico predice el comportamiento futuro con 100% de certeza.",
        styles["Small"]
    ))
    story.append(Paragraph(
        "• En casos de alto riesgo en dimensiones críticas (violencia, sustancias, deshonestidad), "
        "se recomienda evaluación por psicólogo organizacional certificado.",
        styles["Small"]
    ))
    story.append(Paragraph(
        "• Este reporte es CONFIDENCIAL y debe manejarse según políticas de protección de datos.",
        styles["Small"]
    ))
    story.append(Spacer(1, 20))
    
    # === FOOTER ===
    story.append(Paragraph(
        "<i>Este reporte es generado automáticamente como herramienta de apoyo para "
        "Recursos Humanos en procesos de selección. Los resultados deben ser interpretados "
        "por personal capacitado y complementados con otras fuentes de información.</i>",
        styles["Small"]
    ))
    
    # Construir PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_talent_map_pdf(candidate, raw_scores, normalized, radar_fig, session_id, completed_at=None, analysis=None, job_profile_name=None, comparison_fig=None):
    """
    Genera un PDF con los resultados del Talent Map (Mapeo de Competencias).
    
    Args:
        candidate: Dict con información del candidato
        raw_scores: Puntajes directos por competencia
        normalized: Puntajes normalizados (0-100) por competencia
        radar_fig: Figura matplotlib del radar
        session_id: ID de la sesión
        completed_at: Fecha de completación (opcional)
        analysis: Dict con análisis de competencias (opcional)
        job_profile_name: Nombre del perfil de puesto para match (opcional)
        comparison_fig: Figura matplotlib de comparación (opcional)
        
    Returns:
        BytesIO: Buffer con el PDF generado
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=40, bottomMargin=40)
    
    # Estilos
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Justify", alignment=4, leading=14))
    styles.add(ParagraphStyle(name="SmallBold", parent=styles["Normal"], 
                             fontSize=9, leading=12, fontName="Helvetica-Bold"))
    styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], 
                             fontSize=9, leading=12))
    styles.add(ParagraphStyle(name="MatchHighlight", parent=styles["Normal"],
                             fontSize=13, leading=16, fontName="Helvetica-Bold",
                             textColor=colors.HexColor("#1E40AF")))
    
    story = []
    
    # === PÁGINA 1: PORTADA Y RESULTADOS ===
    story.append(Paragraph("Talent Map - Mapeo de Competencias y Talentos", styles["Title"]))
    story.append(Paragraph("Evaluación de 8 Competencias Universales", styles["Heading2"]))
    story.append(Spacer(1, 12))
    
    # Información del candidato
    story.append(Paragraph(f"<b>ID Evaluación:</b> {session_id}", styles["Normal"]))
    story.append(Paragraph(f"<b>Candidato:</b> {candidate['name']}", styles["Normal"]))
    story.append(Paragraph(f"<b>Cédula:</b> {candidate['cedula']}", styles["Normal"]))
    story.append(Paragraph(f"<b>Cargo Evaluado:</b> {candidate.get('position', 'N/A')}", styles["Normal"]))
    
    # Formatear fecha
    if completed_at:
        try:
            fecha_obj = datetime.strptime(completed_at, "%Y-%m-%d %H:%M:%S")
            fecha_str = fecha_obj.strftime('%d/%m/%Y %H:%M')
        except:
            fecha_str = completed_at
    else:
        fecha_str = _now_gmt5().strftime('%d/%m/%Y %H:%M')
    
    story.append(Paragraph(f"<b>Fecha de Presentación:</b> {fecha_str}", styles["Normal"]))
    story.append(Spacer(1, 16))
    
    # Generar análisis si no se proporcionó
    if analysis is None:
        profile_scores = TALENT_MAP_JOB_PROFILES[job_profile_name]["competencias"] if job_profile_name else None
        analysis = analyze_talent_map_match(normalized, job_profile_name)
    
    # === RESULTADO GENERAL ===
    story.append(Paragraph(
        f"<b>PERFIL DE COMPETENCIAS: Promedio {analysis['average_score']:.1f}/100</b>", 
        styles["Heading2"]
    ))
    story.append(Paragraph(
        f"<b>Competencia más fuerte:</b> {analysis['strongest_competency']} "
        f"({int(analysis['strongest_score'])}/100) | "
        f"<b>Área de mayor desarrollo:</b> {analysis['weakest_competency']} "
        f"({int(analysis['weakest_score'])}/100)",
        styles["Normal"]
    ))
    story.append(Spacer(1, 12))
    
    # === ANÁLISIS DE MATCH (si aplica) ===
    if analysis.get('match_analysis'):
        match = analysis['match_analysis']
        story.append(Paragraph(
            f"{match['match_label']}: {match['match_percentage']:.1f}%",
            styles["MatchHighlight"]
        ))
        story.append(Paragraph(
            f"<b>Perfil de Puesto:</b> {match['job_emoji']} {match['job_profile']} - {match['job_description']}",
            styles["Normal"]
        ))
        story.append(Paragraph(
            f"<b>Evaluación:</b> {match['match_desc']}",
            styles["Normal"]
        ))
        story.append(Spacer(1, 16))
    else:
        story.append(Spacer(1, 12))
    
    # === TABLA DE COMPETENCIAS ===
    story.append(Paragraph("Puntajes por Competencia", styles["Heading2"]))
    story.append(Spacer(1, 8))
    
    data = [["Competencia", "Puntaje", "Nivel", "Estado"]]
    for comp in TALENT_MAP_COMPETENCIES:
        score = normalized[comp]
        if score >= 75:
            nivel = "Alto"
            estado = "🌟"
        elif score >= 50:
            nivel = "Medio"
            estado = "👍"
        else:
            nivel = "En Desarrollo"
            estado = "📈"
        
        data.append([
            comp,
            f"{int(score)}/100",
            nivel,
            estado
        ])
    
    t = Table(data, colWidths=[140, 70, 90, 50])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 16))
    
    # === GRÁFICO RADAR ===
    if radar_fig:
        img_buf = BytesIO()
        radar_fig.savefig(img_buf, format="png", dpi=150, bbox_inches="tight")
        img_buf.seek(0)
        story.append(Image(img_buf, width=350, height=350))
    
    # === PÁGINA 2: ANÁLISIS DETALLADO ===
    story.append(PageBreak())
    story.append(Paragraph("Análisis Detallado de Competencias", styles["Heading1"]))
    story.append(Spacer(1, 12))
    
    # === FORTALEZAS ===
    if analysis.get('fortalezas'):
        story.append(Paragraph("🌟 FORTALEZAS CLAVE", styles["Heading2"]))
        story.append(Spacer(1, 6))
        for f in analysis['fortalezas']:
            # Limpiar markdown para PDF
            f_clean = f.replace("**", "")
            story.append(Paragraph(f"• {f_clean}", styles["Small"]))
        story.append(Spacer(1, 12))
    
    # === ÁREAS DE DESARROLLO ===
    if analysis.get('areas_desarrollo'):
        story.append(Paragraph("📈 ÁREAS DE DESARROLLO", styles["Heading2"]))
        story.append(Spacer(1, 6))
        for a in analysis['areas_desarrollo']:
            # Limpiar markdown para PDF
            a_clean = a.replace("**", "")
            story.append(Paragraph(f"• {a_clean}", styles["Small"]))
        story.append(Spacer(1, 12))
    
    # === ANÁLISIS DE MATCH CON PERFIL (si aplica) ===
    if analysis.get('match_analysis'):
        match = analysis['match_analysis']
        
        story.append(Paragraph(f"🎯 ANÁLISIS DE MATCH CON {match['job_profile'].upper()}", styles["Heading2"]))
        story.append(Spacer(1, 8))
        
        # Fortalezas del match
        if match.get('match_strengths'):
            story.append(Paragraph("<b>✅ Competencias que EXCEDEN el perfil:</b>", styles["SmallBold"]))
            story.append(Spacer(1, 4))
            for s in match['match_strengths']:
                s_clean = s.replace("**", "")
                story.append(Paragraph(f"• {s_clean}", styles["Small"]))
            story.append(Spacer(1, 8))
        
        # Gaps del match
        if match.get('match_gaps'):
            story.append(Paragraph("<b>⚠️ Brechas a cerrar:</b>", styles["SmallBold"]))
            story.append(Spacer(1, 4))
            for g in match['match_gaps']:
                g_clean = g.replace("**", "")
                story.append(Paragraph(f"• {g_clean}", styles["Small"]))
            story.append(Spacer(1, 12))
    
    # === GRÁFICO DE COMPARACIÓN (si aplica) ===
    if comparison_fig:
        story.append(PageBreak())
        story.append(Paragraph("Análisis de Brechas de Competencia", styles["Heading1"]))
        story.append(Spacer(1, 12))
        img_buf = BytesIO()
        comparison_fig.savefig(img_buf, format="png", dpi=150, bbox_inches="tight")
        img_buf.seek(0)
        story.append(Image(img_buf, width=500, height=420))
    
    # === PÁGINA 3: RECOMENDACIONES ===
    story.append(PageBreak())
    story.append(Paragraph("💼 Recomendaciones y Plan de Desarrollo", styles["Heading1"]))
    story.append(Spacer(1, 12))
    
    if analysis.get('recomendaciones'):
        for r in analysis['recomendaciones']:
            # Limpiar markdown para PDF
            r_clean = r.replace("**", "")
            story.append(Paragraph(f"{r_clean}", styles["Small"]))
            story.append(Spacer(1, 6))
    
    # === DESCRIPCIÓN DE LAS 8 COMPETENCIAS ===
    story.append(PageBreak())
    story.append(Paragraph("Descripción de las 8 Competencias Evaluadas", styles["Heading1"]))
    story.append(Spacer(1, 12))
    
    for comp in TALENT_MAP_COMPETENCIES:
        score = normalized[comp]
        desc_info = TALENT_MAP_DESCRIPTIONS[comp]
        
        # Determinar nivel y descripción
        if score >= 75:
            level_text = "ALTO 🌟"
            desc_text = desc_info["high"]
        elif score >= 50:
            level_text = "MEDIO 👍"
            desc_text = desc_info["medium"]
        else:
            level_text = "EN DESARROLLO 📈"
            desc_text = desc_info["low"]
        
        story.append(Paragraph(
            f"<b>{desc_info['title']}</b> - {level_text} ({int(score)}/100)",
            styles["Heading3"]
        ))
        story.append(Paragraph(desc_text, styles["Small"]))
        story.append(Spacer(1, 8))
    
    # === PERFILES DE PUESTOS DISPONIBLES ===
    story.append(PageBreak())
    story.append(Paragraph("Perfiles de Puestos de Referencia", styles["Heading1"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph(
        "El sistema incluye perfiles de referencia (benchmarks) para los siguientes puestos:",
        styles["Normal"]
    ))
    story.append(Spacer(1, 8))
    
    for job_name, job_info in TALENT_MAP_JOB_PROFILES.items():
        story.append(Paragraph(
            f"<b>{job_info['emoji']} {job_name}:</b> {job_info['descripcion']}",
            styles["Small"]
        ))
        story.append(Spacer(1, 4))
    
    story.append(Spacer(1, 12))
    story.append(Paragraph(
        "Estos perfiles sirven como referencia para evaluar el ajuste (fit) entre "
        "las competencias del candidato y los requisitos del puesto.",
        styles["Small"]
    ))
    
    # === DISCLAIMER ===
    story.append(Spacer(1, 20))
    story.append(Paragraph(
        "<i>Este reporte es generado automáticamente como herramienta de apoyo para "
        "Recursos Humanos en procesos de selección y desarrollo. Los resultados deben ser "
        "interpretados por personal capacitado y complementados con entrevistas, evaluaciones "
        "de desempeño y otras fuentes de información. Las competencias son desarrollables mediante "
        "capacitación, coaching y experiencia práctica.</i>",
        styles["Small"]
    ))
    
    # Construir PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_desempeno_pdf(candidate, rendimiento_scores, potencial_scores, radar_fig, bars_fig, 
                           session_id, completed_at=None, analysis=None, evaluador_nombre=None, iniciativas=None):
    """Genera PDF de Evaluación de Desempeño."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50)
    
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Title', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor("#1E40AF"), alignment=1, spaceAfter=14))
    styles.add(ParagraphStyle(name='SubTitle', parent=styles['Heading2'], fontSize=13, textColor=colors.HexColor("#374151"), spaceAfter=10))
    styles.add(ParagraphStyle(name='Small', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor("#6B7280")))
    styles.add(ParagraphStyle(name='ListItem', parent=styles['Normal'], fontSize=10, leftIndent=20, spaceAfter=6))
    
    story = []
    
    # Página 1: Portada
    story.append(Spacer(1, 72))
    story.append(Paragraph("📊 EVALUACIÓN DE DESEMPEÑO", styles['Title']))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"<b>Colaborador Evaluado:</b> {candidate['name']}", styles['Normal']))
    story.append(Paragraph(f"<b>Cédula:</b> {candidate['cedula']}", styles['Normal']))
    story.append(Paragraph(f"<b>Cargo:</b> {candidate.get('position', 'N/A')}", styles['Normal']))
    if evaluador_nombre:
        story.append(Paragraph(f"<b>Evaluador:</b> {evaluador_nombre}", styles['Normal']))
    story.append(Paragraph(f"<b>Fecha de Evaluación:</b> {completed_at or 'N/A'}", styles['Normal']))
    story.append(Paragraph(f"<b>ID de Sesión:</b> {session_id}", styles['Small']))
    story.append(Spacer(1, 24))
    
    # Banner de clasificación
    if analysis and analysis.get("clasificacion"):
        clasif = analysis["clasificacion"]
        banner_color = colors.HexColor(clasif["color"])
        banner_table = Table([[clasif["label"]]], colWidths=[450])
        banner_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), banner_color),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 16),
            ('TOPPADDING', (0, 0), (-1, -1), 16),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 16),
        ]))
        story.append(banner_table)
        story.append(Spacer(1, 6))
        story.append(Paragraph(f"<i>{clasif['descripcion']}</i>", styles['Small']))
    
    story.append(Spacer(1, 24))
    
    # Tabla de puntajes
    if analysis:
        puntajes_data = [
            ["Componente", "Promedio", "Máximo"],
            ["Evaluación de Rendimiento (6 objetivos)", f"{analysis['promedio_rendimiento']:.2f}", "5.00"],
            ["Evaluación de Potencial (5 dimensiones)", f"{analysis['promedio_potencial']:.2f}", "3.00"],
            ["Puntaje Global Ponderado", f"<b>{analysis['puntaje_global']:.2f}</b>", "5.00"]
        ]
        
        puntajes_table = Table(puntajes_data, colWidths=[250, 100, 100])
        puntajes_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#3B82F6")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor("#F3F4F6")),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#DBEAFE")),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(puntajes_table)
    
    story.append(PageBreak())
    
    # Página 2: Gráfico de Rendimiento
    story.append(Paragraph("EVALUACIÓN DE RENDIMIENTO", styles['SubTitle']))
    story.append(Spacer(1, 12))
    
    if bars_fig:
        img_buffer = BytesIO()
        bars_fig.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
        img_buffer.seek(0)
        img = Image(img_buffer, width=480, height=320)
        story.append(img)
        plt.close(bars_fig)
    
    story.append(Spacer(1, 12))
    
    # Detalles de cada objetivo
    story.append(Paragraph("<b>Detalle por Objetivo:</b>", styles['Normal']))
    story.append(Spacer(1, 6))
    
    for obj_id, score in rendimiento_scores.items():
        objetivo = DESEMPENO_OBJETIVOS[obj_id - 1]
        nivel = DESEMPENO_ESCALA_RENDIMIENTO[score]
        story.append(Paragraph(
            f"<b>{objetivo['titulo']}</b> - {score:.1f}/5.0 ({nivel['label']})",
            styles['ListItem']
        ))
    
    story.append(PageBreak())
    
    # Página 3: Gráfico de Potencial
    story.append(Paragraph("EVALUACIÓN DE POTENCIAL", styles['SubTitle']))
    story.append(Spacer(1, 12))
    
    if radar_fig:
        img_buffer = BytesIO()
        radar_fig.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
        img_buffer.seek(0)
        img = Image(img_buffer, width=400, height=400)
        story.append(img)
        plt.close(radar_fig)
    
    story.append(Spacer(1, 12))
    
    # Detalles de cada dimensión
    story.append(Paragraph("<b>Detalle por Dimensión:</b>", styles['Normal']))
    story.append(Spacer(1, 6))
    
    for dim_id, score in potencial_scores.items():
        dimension = DESEMPENO_DIMENSIONES[dim_id - 1]
        story.append(Paragraph(
            f"<b>{dimension['nombre']}</b> - Nivel {score}/3",
            styles['ListItem']
        ))
        story.append(Paragraph(
            f"<i>{dimension['niveles'][score]}</i>",
            ParagraphStyle(name='DimDesc', parent=styles['Small'], leftIndent=30, spaceAfter=8)
        ))
    
    story.append(PageBreak())
    
    # Página 4: Fortalezas y Áreas de Mejora
    story.append(Paragraph("ANÁLISIS DE FORTALEZAS Y ÁREAS DE MEJORA", styles['SubTitle']))
    story.append(Spacer(1, 12))
    
    if analysis:
        # Fortalezas de Rendimiento
        if analysis.get("fortalezas_rendimiento"):
            story.append(Paragraph("<b>✅ Fortalezas de Rendimiento:</b>", styles['Normal']))
            story.append(Spacer(1, 6))
            for item in analysis["fortalezas_rendimiento"]:
                story.append(Paragraph(
                    f"• {item['titulo']} ({item['score']:.1f}/5.0 - {item['label']})",
                    styles['ListItem']
                ))
            story.append(Spacer(1, 12))
        
        # Fortalezas de Potencial
        if analysis.get("fortalezas_potencial"):
            story.append(Paragraph("<b>⭐ Fortalezas de Potencial:</b>", styles['Normal']))
            story.append(Spacer(1, 6))
            for item in analysis["fortalezas_potencial"]:
                story.append(Paragraph(
                    f"• {item['nombre']} ({item['nivel']})",
                    styles['ListItem']
                ))
            story.append(Spacer(1, 12))
        
        # Áreas de Mejora de Rendimiento
        if analysis.get("areas_mejora_rendimiento"):
            story.append(Paragraph("<b>⚠️ Áreas de Mejora en Rendimiento:</b>", styles['Normal']))
            story.append(Spacer(1, 6))
            for item in analysis["areas_mejora_rendimiento"]:
                story.append(Paragraph(
                    f"• {item['titulo']} ({item['score']:.1f}/5.0 - {item['label']})",
                    styles['ListItem']
                ))
            story.append(Spacer(1, 12))
        
        # Áreas de Desarrollo de Potencial
        if analysis.get("areas_desarrollo_potencial"):
            story.append(Paragraph("<b>📈 Áreas de Desarrollo en Potencial:</b>", styles['Normal']))
            story.append(Spacer(1, 6))
            for item in analysis["areas_desarrollo_potencial"]:
                story.append(Paragraph(
                    f"• {item['nombre']} ({item['nivel']})",
                    styles['ListItem']
                ))
            story.append(Spacer(1, 12))
    
    story.append(PageBreak())
    
    # Página 5: Recomendaciones e Iniciativas
    story.append(Paragraph("RECOMENDACIONES Y PLAN DE ACCIÓN", styles['SubTitle']))
    story.append(Spacer(1, 12))
    
    if analysis and analysis.get("recomendaciones"):
        story.append(Paragraph("<b>💡 Recomendaciones Generales:</b>", styles['Normal']))
        story.append(Spacer(1, 6))
        for recom in analysis["recomendaciones"]:
            story.append(Paragraph(f"• {recom}", styles['ListItem']))
        story.append(Spacer(1, 18))
    
    # Iniciativas de Mejora
    if iniciativas and len(iniciativas) > 0:
        story.append(Paragraph("<b>🎯 Iniciativas de Mejora Definidas:</b>", styles['Normal']))
        story.append(Spacer(1, 6))
        for i, iniciativa in enumerate(iniciativas, 1):
            if iniciativa and iniciativa.strip():
                story.append(Paragraph(f"<b>Iniciativa {i}:</b>", styles['Normal']))
                story.append(Paragraph(iniciativa, styles['ListItem']))
                story.append(Spacer(1, 8))
    elif analysis and analysis.get("requiere_iniciativas"):
        story.append(Paragraph(
            "<b>⚠️ NOTA:</b> El promedio de evaluación requiere establecer iniciativas de mejora específicas.",
            ParagraphStyle(name='Alert', parent=styles['Normal'], textColor=colors.HexColor("#EF4444"))
        ))
    
    story.append(Spacer(1, 24))
    
    # Footer
    story.append(Paragraph(
        f"<i>Documento generado automáticamente el {completed_at or _now_gmt5().strftime('%Y-%m-%d %H:%M:%S')}</i>",
        styles['Small']
    ))
    
    # Construir PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


# =========================================================================
# HELPER: Load DISC questions
# =========================================================================

def load_disc_questions():
    qfile = os.path.join(os.path.dirname(__file__), "questions_es.json")
    with open(qfile, "r", encoding="utf-8") as f:
        return json.load(f)


def load_disc_descriptions():
    dfile = os.path.join(os.path.dirname(__file__), "disc_descriptions_es.json")
    with open(dfile, "r", encoding="utf-8") as f:
        return json.load(f)


def load_wpi_questions():
    """Carga las preguntas del WPI desde el archivo JSON."""
    qfile = os.path.join(os.path.dirname(__file__), "questions_wpi.json")
    with open(qfile, "r", encoding="utf-8") as f:
        return json.load(f)


# =========================================================================
# NAVIGATION HELPERS
# =========================================================================

def nav(page):
    st.session_state.page = page


# =========================================================================
# PÁGINAS
# =========================================================================

def page_home():
    # Logo HESEGO centrado con HTML
    _logo_path = os.path.join(os.path.dirname(__file__), "logo.png")
    if os.path.exists(_logo_path):
        import base64
        with open(_logo_path, "rb") as _lf:
            _logo_b64 = base64.b64encode(_lf.read()).decode()
        st.markdown(
            f"<div style='display:flex;justify-content:center;margin-bottom:8px'>"
            f"<img src='data:image/png;base64,{_logo_b64}' style='width:200px;height:auto'/>"
            f"</div>",
            unsafe_allow_html=True,
        )
    st.markdown("<h1 style='text-align:center; color:#1e3a5f;'>Plataforma de Evaluaciones Psicométricas</h1>", unsafe_allow_html=True)
    st.markdown("<p style='text-align:center; color:#555;'>Sistema de evaluación para Recursos Humanos</p>", unsafe_allow_html=True)
    st.markdown("---")

    text_col1, text_col2, text_col3 = st.columns(3, gap="large")
    with text_col1:
        st.markdown(
            """
            ### 👤 Soy Candidato / Empleado
            Ingresa con tu número de cédula para realizar la evaluación asignada por Recursos Humanos.
            """
        )

    with text_col2:
        st.markdown(
            """
            ### 👔 Soy Jefe / Evaluador
            Ingresa para completar las evaluaciones de tus colaboradores una vez que ellos hayan realizado su auto-evaluación.
            """
        )

    with text_col3:
        st.markdown(
            """
            ### 🔒 Soy Administrador RH
            Accede al panel de administración para gestionar evaluaciones y ver resultados.
            """
        )

    st.markdown("<div style='height: 8px;'></div>", unsafe_allow_html=True)

    btn_col1, btn_col2, btn_col3 = st.columns(3, gap="large")
    with btn_col1:
        if st.button("🔑 Ingresar como Candidato / Empleado", use_container_width=True, key="btn_candidate"):
            nav("candidate_login")
            st.rerun()

    with btn_col2:
        if st.button("✏️ Ingresar como Evaluador", use_container_width=True, key="btn_evaluador"):
            nav("evaluador_login")
            st.rerun()

    with btn_col3:
        if st.button("🛡️ Ingresar como Administrador", use_container_width=True, key="btn_admin"):
            nav("admin_login")
            st.rerun()


# -------------------------------------------------------------------------
# ADMIN: LOGIN
# -------------------------------------------------------------------------
def page_admin_login():
    st.markdown("## 🔒 Acceso Administrador RH")
    if st.button("⬅️ Volver al inicio"):
        nav("home")
        st.rerun()

    with st.form("admin_login_form"):
        username = st.text_input("Usuario", key="admin_user")
        password = st.text_input("Contraseña", type="password", key="admin_pass")
        submitted = st.form_submit_button("Iniciar Sesión")
        
        if submitted:
            if not username or not password:
                st.error("❌ Por favor completa todos los campos.")
            else:
                username = username.strip()
                password = password.strip()
                
                admin = db.verify_admin(username, password)
                if admin:
                    _start_admin_session(admin)
                    nav("admin_dashboard")
                    st.rerun()
                else:
                    st.error("❌ Credenciales incorrectas. Verifica usuario y contraseña.")


# -------------------------------------------------------------------------
# ADMIN: DASHBOARD
# -------------------------------------------------------------------------
def page_admin_dashboard():
    _restore_admin_session()
    admin = st.session_state.get("admin")
    if not admin:
        nav("admin_login")
        st.rerun()
        return

    st.markdown(f"## 🛡️ Panel de Administración")
    st.caption(f"Bienvenido, {admin['name']}")

    st.markdown(
        """
        <style>
        div[class*="st-key-"][class*="_pending_edit_btn_"] button {
            border: 1px solid #cbd5e1;
        }
        div[class*="st-key-"][class*="_pending_delete_btn_"] button {
            border: 1px solid #dc2626;
            color: #b91c1c;
            background: #fff;
        }
        div[class*="st-key-"][class*="_pending_delete_btn_"] button:hover {
            border-color: #b91c1c;
            background: #fee2e2;
            color: #7f1d1d;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    # ----- Filtros compartidos (Resultados / Pruebas Pendientes) -----
    _all_raw = db.get_all_sessions()
    _all_candidates = db.get_all_candidates()
    _candidate_name_by_cedula = {c["cedula"]: c.get("name", "N/A") for c in _all_candidates}
    _sessions_by_cedula = {}
    for _s in _all_raw:
        _sessions_by_cedula.setdefault(_s["cedula"], []).append(_s)
    _cand_names = sorted(set(s["candidate_name"] for s in _all_raw)) if _all_raw else []
    _FILTER_LABELS = {
        "Todos": "📋 Todos", "disc": "🎯 DISC", "valanti": "🧭 VALANTI",
        "wpi": "💼 WPI", "eri": "🔐 ERI", "talent_map": "🌟 Talent Map",
        "desempeno": "📊 Desempeño Operativo",
        "desempeno_lider": "📊 Desempeño Líderes",
        "periodo_prueba": "📋 Período de Prueba",
    }
    _SECTION_OPTIONS = {
        "dashboard": "📈 Dashboard Gerencial",
        "create": "📋 Crear Evaluación",
        "new_candidate": "➕ Nuevo Candidato",
        "candidates": "👥 Candidatos",
        "results": "📊 Resultados",
        "pending": "⏳ Pruebas Pendientes",
        "bulk": "📤 Cargue Masivo",
        "settings": "⚙️ Configuración",
    }

    with st.sidebar:
        st.markdown("### 🛡️ Panel Admin")
        st.caption(f"Sesión: {admin['name']}")
        if st.button("🚪 Cerrar Sesión", use_container_width=True):
            _logout_admin()
            nav("home")
            st.rerun()

        st.markdown("---")
        _active_section = st.radio(
            "Ir a",
            options=list(_SECTION_OPTIONS.keys()),
            key="admin_sidebar_section",
            format_func=lambda value: _SECTION_OPTIONS[value],
        )

        _filter_type = "Todos"
        _filter_cand = "Todos"
        _sort_opt = "Fecha (reciente)"
        if _active_section in ("results", "pending"):
            st.markdown("---")
            st.markdown("### 🔎 Filtros")
            _filter_type = st.selectbox(
                "Filtrar por tipo",
                list(_FILTER_LABELS.keys()),
                key="filter_type",
                format_func=lambda x: _FILTER_LABELS.get(x, x),
            )
            _filter_cand = st.selectbox(
                "Filtrar por candidato",
                ["Todos"] + _cand_names,
                key="filter_candidate",
            )
            _sort_opt = st.selectbox(
                "Ordenar por",
                ["Fecha (reciente)", "Fecha (antigua)", "Candidato A-Z", "Candidato Z-A", "Tipo prueba"],
                key="sort_option",
            )

    _ft = _filter_type if _filter_type != "Todos" else None
    _EDITABLE_TEST_TYPES = [k for k in _FILTER_LABELS.keys() if k != "Todos"]

    def _get_sort_date(s):
        date_str = s.get("completed_at") or s.get("started_at") or s.get("created_at") or ""
        try:
            return datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
        except:
            return datetime.min

    def _sort_sessions(lst):
        if _sort_opt == "Fecha (reciente)":
            lst.sort(key=_get_sort_date, reverse=True)
        elif _sort_opt == "Fecha (antigua)":
            lst.sort(key=_get_sort_date)
        elif _sort_opt == "Candidato A-Z":
            lst.sort(key=lambda s: s["candidate_name"].lower())
        elif _sort_opt == "Candidato Z-A":
            lst.sort(key=lambda s: s["candidate_name"].lower(), reverse=True)
        elif _sort_opt == "Tipo prueba":
            lst.sort(key=lambda s: s["test_type"])

    def _render_sessions_list(sessions, tab_key):
        if not sessions:
            st.info("No hay evaluaciones que coincidan con los filtros.")
            return
        _TEST_NAME_MAP = {
            "disc": "DISC", "valanti": "VALANTI", "wpi": "WPI", "eri": "ERI",
            "talent_map": "TALENT MAP", "desempeno": "DESEMPEÑO OP.",
            "desempeno_lider": "DESEMPEÑO LÍD.", "periodo_prueba": "PER. PRUEBA",
        }
        _TEST_EMOJI_MAP = {"disc": "🎯", "valanti": "🧭", "wpi": "💼", "eri": "🔐"}
        _STATUS_EMOJI = {"pending": "⏳", "in_progress": "▶️", "completed": "✅", "expired": "⏰", "employee_done": "📝"}
        # Estado de ordenamiento por tabla (independiente por tab_key)
        _sc_key = f"_rsort_col_{tab_key}"
        _sd_key = f"_rsort_dir_{tab_key}"
        if _sc_key not in st.session_state:
            st.session_state[_sc_key] = "FECHA"
            st.session_state[_sd_key] = "desc"
        _curr_col = st.session_state[_sc_key]
        _curr_dir = st.session_state[_sd_key]

        def _col_val_sort(s, col):
            if col == "CANDIDATO": return s["candidate_name"].lower()
            if col == "PRUEBA": return s["test_type"]
            if col == "CÉDULA": return str(s["cedula"])
            if col == "FECHA": return _get_sort_date(s)
            return ""

        sessions = sorted(sessions, key=lambda s: _col_val_sort(s, _curr_col), reverse=(_curr_dir == "desc"))

        _W = [0.35, 2.4, 1.6, 1.5, 1.05, 1.65, 0.45]
        st.caption(f"📊 {len(sessions)} evaluación(es) encontrada(s)")
        # Cabecera con ordenamiento por columna
        _hdr = st.columns(_W)
        _hdr[0].markdown("<div style='height:30px'></div>", unsafe_allow_html=True)
        _hdr[4].markdown(
            "<div style='font-size:11px;font-weight:700;color:#aaa;padding-top:6px;border-bottom:1px solid #444'>ID</div>",
            unsafe_allow_html=True,
        )
        _hdr[6].markdown("<div style='height:30px'></div>", unsafe_allow_html=True)
        for _hcol_idx, _hcol_name in [(1, "CANDIDATO"), (2, "PRUEBA"), (3, "CÉDULA"), (5, "FECHA")]:
            _arrow = " ↓" if (_curr_col == _hcol_name and _curr_dir == "desc") else (
                     " ↑" if _curr_col == _hcol_name else " ↕")
            if _hdr[_hcol_idx].button(
                f"{_hcol_name}{_arrow}",
                key=f"hdr_{tab_key}_{_hcol_name}",
                use_container_width=True,
            ):
                if _curr_col == _hcol_name:
                    st.session_state[_sd_key] = "asc" if _curr_dir == "desc" else "desc"
                else:
                    st.session_state[_sc_key] = _hcol_name
                    st.session_state[_sd_key] = "desc"
                st.rerun()
        for sess in sessions:
            status_emoji = _STATUS_EMOJI.get(sess["status"], "❓")
            test_emoji = _TEST_EMOJI_MAP.get(sess["test_type"], "📝")
            test_name = _TEST_NAME_MAP.get(sess["test_type"], sess["test_type"].upper())
            evaluador_ced = sess.get("evaluador_cedula")
            evaluador_nombre = sess.get("evaluador_nombre")
            _date_ref = sess.get("completed_at") or sess.get("started_at") or sess.get("created_at")
            fecha_str = "—"
            if _date_ref:
                try:
                    fecha_str = datetime.strptime(_date_ref, "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y %H:%M")
                except:
                    fecha_str = str(_date_ref)
            toggle_key = f"show_detail_{tab_key}_{sess['id']}"
            is_open = st.session_state.get(toggle_key, False)
            # Fila de la tabla
            rc = st.columns(_W)
            rc[0].markdown(f"<div style='font-size:17px;padding-top:3px'>{status_emoji}</div>", unsafe_allow_html=True)
            rc[1].markdown(f"<div style='font-weight:600;padding-top:5px;font-size:13px'>{sess['candidate_name']}</div>", unsafe_allow_html=True)
            rc[2].markdown(f"<div style='font-size:12px;padding-top:6px'>{test_emoji} {test_name}</div>", unsafe_allow_html=True)
            rc[3].markdown(f"<div style='font-size:12px;font-family:monospace;padding-top:6px'>{sess['cedula']}</div>", unsafe_allow_html=True)
            rc[4].markdown(f"<div style='font-size:12px;font-family:monospace;padding-top:6px'>{sess['id']}</div>", unsafe_allow_html=True)
            rc[5].markdown(f"<div style='font-size:12px;padding-top:6px'>{fecha_str}</div>", unsafe_allow_html=True)
            if rc[6].button("▲" if is_open else "▼", key=f"tog_{tab_key}_{sess['id']}", use_container_width=True):
                st.session_state[toggle_key] = not is_open
                st.rerun()
            if is_open:
                with st.container(border=True):
                    ec1, ec2, ec3, ec4 = st.columns(4)
                    ec1.metric("Estado", sess["status"].upper())
                    ec2.metric("Tiempo Límite", f"{sess['time_limit_minutes']} min")
                    _s_at = sess.get("started_at")
                    _c_at = sess.get("completed_at")
                    try:
                        started_str = datetime.strptime(_s_at, "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y %H:%M") if _s_at else "N/A"
                    except:
                        started_str = _s_at or "N/A"
                    try:
                        completed_str = datetime.strptime(_c_at, "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y %H:%M") if _c_at else "N/A"
                    except:
                        completed_str = _c_at or "N/A"
                    ec3.metric("Iniciado", started_str)
                    ec4.metric("Completado", completed_str)
                    if evaluador_ced:
                        jefe_nombre = evaluador_nombre or _candidate_name_by_cedula.get(evaluador_ced)
                        if jefe_nombre:
                            st.info(f"👔 **Jefe asignado:** {jefe_nombre} (Cédula `{evaluador_ced}`)")
                        else:
                            st.info(f"👔 **Jefe asignado:** Cédula `{evaluador_ced}`")
                    if sess["status"] == "completed":
                        details_key = f"{tab_key}_show_details_{sess['id']}"
                        if not st.session_state.get(details_key, False):
                            st.caption("Resultados detallados ocultos para acelerar carga.")
                            _dc1, _dc2 = st.columns(2)
                            with _dc1:
                                if st.button("📊 Ver resultados", key=f"{details_key}_btn", use_container_width=True):
                                    st.session_state[details_key] = True
                                    st.session_state.pop(f"{details_key}_focus_pdf", None)
                                    st.rerun()
                            with _dc2:
                                if st.button("📥 Descargar PDF", key=f"{details_key}_pdf_btn", use_container_width=True):
                                    st.session_state[details_key] = True
                                    st.session_state[f"{details_key}_focus_pdf"] = True
                                    st.rerun()
                        else:
                            if st.session_state.get(f"{details_key}_focus_pdf", False):
                                st.info("📥 El PDF de descarga está disponible al final de esta sección ↓")
                            results = db.get_results(sess["id"])
                            candidate = db.get_candidate_by_cedula(sess["cedula"])
                            if results:
                                if sess["test_type"] == "disc":
                                    show_disc_results_admin(results, candidate, sess)
                                elif sess["test_type"] == "valanti":
                                    show_valanti_results_admin(results, candidate, sess)
                                elif sess["test_type"] == "wpi":
                                    show_wpi_results_admin(results, candidate, sess)
                                elif sess["test_type"] == "eri":
                                    show_eri_results_admin(results, candidate, sess)
                                elif sess["test_type"] == "talent_map":
                                    show_talent_map_results_admin(results, candidate, sess)
                                elif sess["test_type"] == "desempeno":
                                    show_desempeno_results_admin(results, candidate, sess)
                                elif sess["test_type"] == "desempeno_lider":
                                    show_desempeno_lider_results_admin(results, candidate, sess)
                                elif sess["test_type"] == "periodo_prueba":
                                    show_periodo_prueba_results_admin(results, candidate, sess)
                            else:
                                st.warning("Resultados no disponibles.")
                    elif sess["status"] == "pending" and sess["test_type"] == "desempeno":
                        st.info("⏳ Esta evaluación está pendiente de ser completada por un evaluador.")
                        if st.button("✏️ Evaluar Ahora", key=f"{tab_key}_eval_desemp_{sess['id']}"):
                            st.session_state["desempeno_session_id"] = sess["id"]
                            nav("desempeno_eval")
                            st.rerun()
                    elif sess["status"] == "pending" and sess["test_type"] == "desempeno_lider":
                        st.info("⏳ Pendiente. El empleado debe completar su auto-evaluación primero.")
                    elif sess["status"] == "employee_done" and sess["test_type"] == "desempeno_lider":
                        st.success("📝 El empleado completó su parte. El jefe asignado debe completar la evaluación.")
                        if st.button("✏️ Completar como Admin", key=f"{tab_key}_eval_lider_admin_{sess['id']}"):
                            st.session_state["desempeno_lider_session_id"] = sess["id"]
                            nav("desempeno_lider_jefe_eval")
                            st.rerun()
                    elif sess["status"] == "pending" and sess["test_type"] == "periodo_prueba":
                        st.info("⏳ Pendiente. El empleado debe completar su auto-evaluación primero.")
                    elif sess["status"] == "employee_done" and sess["test_type"] == "periodo_prueba":
                        st.success("📝 El empleado completó su parte. El jefe asignado debe completar la evaluación.")
                        if st.button("✏️ Completar como Admin", key=f"{tab_key}_eval_pp_admin_{sess['id']}"):
                            st.session_state["periodo_prueba_session_id"] = sess["id"]
                            nav("periodo_prueba_jefe_eval")
                            st.rerun()
                    if sess["status"] == "pending":
                        st.markdown("---")
                        pending_actions_col, pending_delete_col = st.columns([3, 1])
                        edit_toggle_key = f"show_edit_pending_{tab_key}_{sess['id']}"
                        with pending_actions_col:
                            if st.button(
                                "✏️ Editar prueba pendiente",
                                key=f"{tab_key}_pending_edit_btn_{sess['id']}",
                                use_container_width=True,
                            ):
                                st.session_state[edit_toggle_key] = not st.session_state.get(edit_toggle_key, False)
                        with pending_delete_col:
                            if st.button(
                                "🗑️ Eliminar",
                                key=f"{tab_key}_pending_delete_btn_{sess['id']}",
                                use_container_width=True,
                            ):
                                st.session_state[f"confirm_del_{sess['id']}"] = True
                        if st.session_state.get(edit_toggle_key, False):
                            with st.container(border=True):
                                with st.form(f"edit_pending_{tab_key}_{sess['id']}"):
                                    cedit1, cedit2 = st.columns(2)
                                    with cedit1:
                                        edit_test_type = st.selectbox(
                                            "Tipo de prueba",
                                            options=_EDITABLE_TEST_TYPES,
                                            index=_EDITABLE_TEST_TYPES.index(sess["test_type"]) if sess["test_type"] in _EDITABLE_TEST_TYPES else 0,
                                            format_func=lambda x: _FILTER_LABELS.get(x, x),
                                            key=f"edit_type_{tab_key}_{sess['id']}",
                                        )
                                    with cedit2:
                                        _tl_opts = [15, 20, 30, 45, 60, 90]
                                        _curr_tl = sess.get("time_limit_minutes", 45)
                                        edit_time_limit = st.selectbox(
                                            "Tiempo límite",
                                            options=_tl_opts,
                                            index=_tl_opts.index(_curr_tl) if _curr_tl in _tl_opts else 3,
                                            format_func=lambda x: f"{x} minutos",
                                            key=f"edit_tl_{tab_key}_{sess['id']}",
                                        )
                                    edit_jefe_ced = st.text_input(
                                        "Cédula del jefe/evaluador",
                                        value=(sess.get("evaluador_cedula") or ""),
                                        key=f"edit_jefe_ced_{tab_key}_{sess['id']}",
                                    )
                                    edit_jefe_nom = st.text_input(
                                        "Nombre del jefe/evaluador",
                                        value=(sess.get("evaluador_nombre") or ""),
                                        key=f"edit_jefe_nom_{tab_key}_{sess['id']}",
                                    )
                                    save_edit = st.form_submit_button("💾 Guardar cambios")
                                    if save_edit:
                                        if edit_test_type in ("desempeno", "desempeno_lider", "periodo_prueba") and not edit_jefe_ced.strip():
                                            st.error("❌ La cédula del jefe/evaluador es obligatoria para este tipo de prueba.")
                                        else:
                                            ok, err = db.update_pending_session(
                                                sess["id"],
                                            edit_test_type,
                                            edit_time_limit,
                                            evaluador_cedula=edit_jefe_ced.strip() or None,
                                            evaluador_nombre=edit_jefe_nom.strip() or None,
                                        )
                                        if not ok:
                                            st.error(f"❌ {err}")
                                        else:
                                            st.success("✅ Prueba actualizada correctamente.")
                                            st.session_state[edit_toggle_key] = False
                                            st.rerun()

                    if st.session_state.get(f"confirm_del_{sess['id']}", False):
                        st.warning(f"⚠️ ¿Eliminar prueba **{sess['id']}** de **{sess['candidate_name']}**? Esta acción no se puede deshacer.")
                        col_yes, col_no, _ = st.columns([1, 1, 2])
                        with col_yes:
                            if st.button("✅ Sí, eliminar", key=f"{tab_key}_confirm_yes_{sess['id']}"):
                                db.delete_test_session(sess['id'])
                                st.session_state.pop(f"confirm_del_{sess['id']}", None)
                                st.success("✅ Prueba pendiente eliminada.")
                                st.rerun()
                        with col_no:
                            if st.button("❌ Cancelar", key=f"{tab_key}_confirm_no_{sess['id']}"):
                                st.session_state.pop(f"confirm_del_{sess['id']}", None)
                                st.rerun()

    # ----- SECCIÓN: Crear Evaluación (solo candidatos existentes) -----
    if _active_section == "create":
        st.markdown("### 📋 Crear Evaluación")

        _TEST_LABELS_SHARED = {
            "disc": "🎯 DISC", "valanti": "🧭 VALANTI", "wpi": "💼 WPI",
            "eri": "🔐 ERI", "talent_map": "🌟 Talent Map",
            "desempeno": "📊 Desempeño Operativo",
            "desempeno_lider": "📊 Desempeño Líderes",
            "periodo_prueba": "📋 Período de Prueba",
        }
        _TIPOS_CON_EVALUADOR = ("desempeno", "desempeno_lider", "periodo_prueba")

        _candidates_all = db.get_all_candidates()
        if not _candidates_all:
            st.warning("No hay candidatos registrados. Ve a **➕ Nuevo Candidato** en el menú para agregar uno.")
        else:
            _cand_opts = {f"{c['cedula']} — {c['name']}": c for c in _candidates_all}
            _PLACEHOLDER_CAND = "— Escribe o selecciona un candidato —"
            _cand_sel = st.selectbox(
                "Candidato *",
                [_PLACEHOLDER_CAND] + list(_cand_opts.keys()),
                key="create_cand_sel",
                placeholder="Busca por cédula o nombre...",
            )
            if _cand_sel == _PLACEHOLDER_CAND:
                st.info("Selecciona un candidato para continuar.")
                _cand = None
            else:
                _cand = _cand_opts[_cand_sel]
            if _cand:
                _empresa_str = _cand.get("empresa_codigo") or ""
                _regional_str = _cand.get("regional") or ""
                _cargo_str = _cand.get("position") or "N/A"
                st.caption(f"Cargo: {_cargo_str}  |  Empresa: {_empresa_str}  |  Regional: {_regional_str}")
            # Mostrar pruebas activas del candidato seleccionado
            _cand_active = []
            _pending_types = set()
            if _cand:
                _cand_active = [
                    s for s in _all_raw
                    if s["cedula"] == _cand["cedula"] and s["status"] in ("pending", "in_progress", "employee_done")
                ]
                _pending_types = {s["test_type"] for s in _cand_active}
            if _cand_active:
                _pa_lines = "  \n".join(
                    f"• ⏳ **{_TEST_LABELS_SHARED.get(s['test_type'], s['test_type'].upper())}** — ID: `{s['id']}` — Estado: *{s['status']}*"
                    for s in _cand_active
                )
                st.warning(
                    f"⚠️ Este candidato ya tiene {len(_cand_active)} prueba(s) activa(s). "
                    f"No podrás asignar el mismo tipo nuevamente hasta que finalice:\n\n{_pa_lines}"
                )

            st.markdown("---")
            st.markdown("**👤 Jefe / Evaluador** _(requerido para Desempeño y Período de Prueba)_")

            # Lookup automático por cédula — fuera del formulario para reactividad inmediata
            def _on_change_eval_ced_create():
                _ced_v = st.session_state.get("create_eval_ced", "").strip()
                _found_v = db.get_candidate_by_cedula(_ced_v) if _ced_v else None
                if _found_v:
                    st.session_state["create_eval_nom"] = _found_v["name"]
                else:
                    st.session_state["create_eval_nom"] = ""

            _eval_ced = st.text_input(
                "Cédula del Jefe / Evaluador",
                key="create_eval_ced",
                placeholder="Ingresa la cédula para buscar automáticamente",
                on_change=_on_change_eval_ced_create,
            )
            _jefe_found = db.get_candidate_by_cedula(_eval_ced.strip()) if _eval_ced.strip() else None
            _jefe_ok = True
            if _jefe_found:
                st.success(f"✅ {_jefe_found['name']} | Cargo: {_jefe_found.get('position', 'N/A')}")
                _eval_nom = _jefe_found["name"]
            elif _eval_ced.strip():
                st.error("❌ Cédula no registrada. Regístralo primero en la sección **👥 Candidatos** antes de asignarlo como evaluador.")
                _jefe_ok = False
                _eval_nom = ""
            else:
                _eval_nom = ""

            st.markdown("---")
            with st.form("create_eval_form"):
                _fcol1, _fcol2 = st.columns(2)
                with _fcol1:
                    _test_type = st.selectbox(
                        "Tipo de Evaluación",
                        list(_TEST_LABELS_SHARED.keys()),
                        format_func=lambda x: _TEST_LABELS_SHARED.get(x, x),
                        key="create_test_type",
                    )
                with _fcol2:
                    _time_limit = st.selectbox(
                        "Tiempo Límite",
                        [15, 20, 30, 45, 60, 90],
                        index=3,
                        format_func=lambda x: f"{x} minutos",
                        key="create_time_limit",
                    )
                _create_submitted = st.form_submit_button("✅ Crear Evaluación", type="primary")
                if _create_submitted:
                    _ec_val = st.session_state.get("create_eval_ced", "").strip()
                    if _test_type in _TIPOS_CON_EVALUADOR and not _ec_val:
                        st.error("❌ La cédula del Jefe/Evaluador es obligatoria para este tipo de evaluación.")
                    elif _test_type in _TIPOS_CON_EVALUADOR and not _jefe_ok:
                        st.error("❌ Debes registrar al jefe como candidato antes de continuar.")
                    elif _test_type in _pending_types:
                        st.error(
                            f"❌ El candidato ya tiene una prueba **{_TEST_LABELS_SHARED.get(_test_type, _test_type.upper())}** pendiente o en curso. "
                            f"Completa o elimina la evaluación activa antes de crear una nueva."
                        )
                    else:
                        _ec_save = _ec_val if _test_type in _TIPOS_CON_EVALUADOR else None
                        _en_save = _eval_nom if _eval_nom else None
                        _sid, _serr = db.create_test_session(
                            _cand["id"], _test_type, _time_limit, admin["id"],
                            evaluador_cedula=_ec_save, evaluador_nombre=_en_save,
                        )
                        if _serr:
                            st.warning(f"⚠️ {_serr}")
                        else:
                            _extra = f" | **Evaluador:** {_ec_save}" if _ec_save else ""
                            st.success(f"✅ Evaluación creada!\n\n**ID:** `{_sid}` | **Tipo:** {_test_type.upper()}{_extra}")

    # ----- SECCIÓN: Nuevo Candidato -----
    elif _active_section == "new_candidate":
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        def _nc_border():
            s = Side(border_style="thin", color="BFBFBF")
            return Border(left=s, right=s, top=s, bottom=s)

        def _nc_hdr(ws, r, c, v, w=None):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _nc_border()
            if w:
                ws.column_dimensions[get_column_letter(c)].width = w

        _CIUDADES_CO = [
            "Bogotá", "Medellín", "Cali", "Barranquilla", "Cartagena", "Cúcuta",
            "Bucaramanga", "Pereira", "Santa Marta", "Ibagué", "Pasto", "Manizales",
            "Neiva", "Villavicencio", "Armenia", "Valledupar", "Montería", "Sincelejo",
            "Riohacha", "Tunja", "Popayán", "Florencia", "Quibdó", "Mocoa",
            "Yopal", "Arauca", "Palmira", "Buenaventura", "Bello", "Soledad",
            "Itagüí", "Soacha", "Barrancabermeja", "Dos Quebradas", "Envigado",
            "Floridablanca", "Girón", "Piedecuesta", "Turbo", "Magangué",
            "Maicao", "Pitalito", "Sogamoso", "Duitama", "Zipaquirá", "Chía",
            "Fusagasugá", "Facatativá", "Girardot", "Espinal", "Honda",
            "Cartago", "Tuluá", "Buga", "Ipiales", "Tumaco", "Ocaña", "Pamplona",
            "San José del Guaviare", "Mitú", "Puerto Carreño", "Leticia", "Inírida",
            "Apartadó", "Caucasia", "Rionegro", "Bello", "Sabaneta", "La Estrella",
            "Caldas", "Copacabana", "Girardota", "Barbosa", "Jamundí", "Yumbo",
            "Mosquera", "Madrid", "Funza", "Cajicá", "Sopó", "La Calera",
            "Otra",
        ]
        _EMPRESAS_LIST = ["HESEGO"]

        def _nc_gen_plantilla():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BD empleados"
            ws.sheet_view.showGridLines = False
            ws.freeze_panes = "A2"
            # Todas las columnas son obligatorias
            cols = [
                ("CEDULA", 18), ("EMPRESA", 12), ("REGIONAL", 20),
                ("APELLIDOS Y NOMBRES", 30), ("CORREO", 28),
                ("CARGO", 22), ("INVITAR", 10),
                ("JEFE INMEDIATO", 25), ("NIVEL DE CARGO", 18),
            ]
            ws.row_dimensions[1].height = 32
            for ci, (nm, wd) in enumerate(cols, 1):
                _nc_hdr(ws, 1, ci, nm, wd)
                for r in range(2, 502):
                    cell = ws.cell(row=r, column=ci)
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")  # todo obligatorio
                    cell.border = _nc_border()
            ej = ["1020304050", "HESEGO", "Bogotá", "García López Juan Carlos",
                  "juan@hesego.com", "Analista de Calidad", "SI", "María Rodríguez", "Operativo"]
            for ci, v in enumerate(ej, 1):
                c = ws.cell(row=2, column=ci, value=v)
                c.fill = PatternFill("solid", fgColor="DDEEFF")
                c.font = Font(italic=True, size=10, color="2F5496")
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = _nc_border()
            # Dropdowns
            dv1 = DataValidation(type="list", formula1='"HESEGO"', allow_blank=False, showDropDown=False)
            dv1.sqref = "B2:B501"
            ws.add_data_validation(dv1)
            dv2 = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False, showDropDown=False)
            dv2.sqref = "G2:G501"
            ws.add_data_validation(dv2)
            dv3 = DataValidation(type="list", formula1='"Operativo,Líder,Directivo,Administrativo"', allow_blank=False, showDropDown=False)
            dv3.sqref = "I2:I501"
            ws.add_data_validation(dv3)
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf.read()

        st.markdown("### ➕ Nuevo Candidato")

        # ── Registro uno a uno ────────────────────────────────────────────
        _nc_cedula = st.text_input("Cédula *", placeholder="Número de identificación", key="nc_cedula_lookup")
        _nc_existing = db.get_candidate_by_cedula(_nc_cedula.strip()) if _nc_cedula.strip() else None
        if _nc_existing:
            st.warning(f"⚠️ La cédula **{_nc_cedula.strip()}** ya está registrada como **{_nc_existing['name']}**. Usa 'Crear Evaluación' para asignarle una prueba.")
        else:
            with st.form("new_candidate_form"):
                _ncf1, _ncf2 = st.columns(2)
                with _ncf1:
                    _nc_name = st.text_input("Apellidos y Nombres *", placeholder="García López Juan Carlos")
                    _nc_empresa = st.selectbox("Empresa *", _EMPRESAS_LIST)
                    _nc_regional = st.selectbox("Regional *", _CIUDADES_CO)
                    _nc_cargo = st.text_input("Cargo *", placeholder="Analista de Calidad")
                    _nc_nivel = st.selectbox("Nivel de Cargo *", ["Operativo", "Líder", "Directivo", "Administrativo"])
                with _ncf2:
                    _nc_correo = st.text_input("Correo *", placeholder="juan@empresa.com")
                    _nc_jefe = st.text_input("Jefe Inmediato *", placeholder="Nombre completo del jefe")
                    _nc_invitar = st.selectbox("Invitar a evaluación *", ["SI", "NO"])
                    _nc_sex = st.selectbox("Sexo", ["", "Masculino", "Femenino", "Otro"],
                                          format_func=lambda x: x if x else "— Opcional —")
                    _nc_edu = st.text_input("Nivel Educativo", placeholder="Universitario")
                    _nc_age = st.number_input("Edad", min_value=0, max_value=100, value=0)
                _nc_submit = st.form_submit_button("💾 Guardar Candidato", type="primary")
                if _nc_submit:
                    _nc_errs = []
                    if not _nc_cedula.strip():
                        _nc_errs.append("Cédula")
                    if not _nc_name.strip():
                        _nc_errs.append("Apellidos y Nombres")
                    if not _nc_cargo.strip():
                        _nc_errs.append("Cargo")
                    if not _nc_correo.strip():
                        _nc_errs.append("Correo")
                    if not _nc_jefe.strip():
                        _nc_errs.append("Jefe Inmediato")
                    if _nc_errs:
                        st.error(f"❌ Campos obligatorios faltantes: {', '.join(_nc_errs)}.")
                    else:
                        _nc_result = db.create_empleado(
                            cedula=_nc_cedula.strip(),
                            name=_nc_name.strip(),
                            empresa_codigo=_nc_empresa,
                            regional=_nc_regional,
                            correo=_nc_correo.strip(),
                            position=_nc_cargo.strip(),
                            jefe_inmediato=_nc_jefe.strip(),
                            nivel_cargo=_nc_nivel,
                            invitar=_nc_invitar,
                            age=_nc_age if _nc_age > 0 else None,
                            sex=_nc_sex if _nc_sex else None,
                            education=_nc_edu.strip() if _nc_edu.strip() else None,
                        )
                        if _nc_result:
                            st.success(f"✅ Candidato **{_nc_name.strip()}** (CC: {_nc_cedula.strip()}) registrado correctamente.")
                        else:
                            st.error("❌ Error al guardar. La cédula puede estar duplicada.")

    # ----- SECCIÓN: Candidatos -----
    elif _active_section == "candidates":
        st.markdown("### Candidatos Registrados")
        _cands2 = db.get_all_candidates()
        if not _cands2:
            st.info("No hay candidatos registrados.")
        else:
            _sc1, _sc2 = st.columns([3, 1])
            with _sc1:
                _cand_search = st.text_input(
                    "🔍 Buscar por nombre o cédula",
                    placeholder="Escribe para filtrar...",
                    key="cand_search_filter",
                )
            with _sc2:
                st.markdown("<div style='margin-top:28px'></div>", unsafe_allow_html=True)
                if st.button("✖ Limpiar", key="cand_search_clear", use_container_width=True):
                    st.session_state["cand_search_filter"] = ""
                    st.rerun()
            if _cand_search.strip():
                _q = _cand_search.strip().lower()
                _cands2 = [
                    c for c in _cands2
                    if _q in c.get("name", "").lower() or _q in str(c.get("cedula", "")).lower()
                ]
            st.caption(f"📋 {len(_cands2)} candidato(s) encontrado(s)")
            if not _cands2:
                st.info("No se encontraron candidatos con ese criterio.")
            else:
                # ── Ordenamiento de tabla ──────────────────────────────────
                _csort_col_key = "_csort_col"
                _csort_dir_key = "_csort_dir"
                if _csort_col_key not in st.session_state:
                    st.session_state[_csort_col_key] = "NOMBRE"
                    st.session_state[_csort_dir_key] = "asc"
                _csort_col = st.session_state[_csort_col_key]
                _csort_dir = st.session_state[_csort_dir_key]

                def _cand_sort_val(c, col):
                    if col == "NOMBRE":    return c.get("name", "").lower()
                    if col == "CÉDULA":    return str(c.get("cedula", ""))
                    if col == "CARGO":     return c.get("position", "").lower()
                    if col == "NIVEL":     return c.get("nivel_cargo", "").lower()
                    if col == "REGIONAL":  return c.get("regional", "").lower()
                    if col == "EMPRESA":   return c.get("empresa_codigo", "").lower()
                    return ""

                _cands2 = sorted(_cands2, key=lambda c: _cand_sort_val(c, _csort_col), reverse=(_csort_dir == "desc"))

                # ── Cabecera de tabla ──────────────────────────────────────
                # Proporciones: nombre, cédula, cargo, nivel, regional, empresa, acciones
                _CW = [2.5, 1.5, 2.0, 1.4, 1.5, 1.2, 0.45]
                _ch = st.columns(_CW)
                for _cidx, _cname in [(0, "NOMBRE"), (1, "CÉDULA"), (2, "CARGO"), (3, "NIVEL"), (4, "REGIONAL"), (5, "EMPRESA")]:
                    _arrow = " ↓" if (_csort_col == _cname and _csort_dir == "desc") else (
                             " ↑" if _csort_col == _cname else " ↕")
                    if _ch[_cidx].button(f"{_cname}{_arrow}", key=f"chdr_{_cname}", use_container_width=True):
                        if _csort_col == _cname:
                            st.session_state[_csort_dir_key] = "asc" if _csort_dir == "desc" else "desc"
                        else:
                            st.session_state[_csort_col_key] = _cname
                            st.session_state[_csort_dir_key] = "asc"
                        st.rerun()
                _ch[6].markdown("<div style='height:30px'></div>", unsafe_allow_html=True)

                st.markdown("---")

                # ── Filas de tabla ─────────────────────────────────────────
                for c in _cands2:
                    _toggle_key = f"cand_open_{c['id']}"
                    _is_open = st.session_state.get(_toggle_key, False)
                    _csessions = _sessions_by_cedula.get(c["cedula"], [])
                    _n_evals = len(_csessions)

                    rc = st.columns(_CW)
                    rc[0].markdown(
                        f"<div style='font-weight:600;padding-top:5px;font-size:13px'>👤 {c['name']}</div>",
                        unsafe_allow_html=True,
                    )
                    rc[1].markdown(
                        f"<div style='font-size:12px;font-family:monospace;padding-top:6px'>{c.get('cedula','—')}</div>",
                        unsafe_allow_html=True,
                    )
                    rc[2].markdown(
                        f"<div style='font-size:12px;padding-top:6px'>{c.get('position','—')}</div>",
                        unsafe_allow_html=True,
                    )
                    rc[3].markdown(
                        f"<div style='font-size:12px;padding-top:6px'>{c.get('nivel_cargo','—')}</div>",
                        unsafe_allow_html=True,
                    )
                    rc[4].markdown(
                        f"<div style='font-size:12px;padding-top:6px'>{c.get('regional','—')}</div>",
                        unsafe_allow_html=True,
                    )
                    rc[5].markdown(
                        f"<div style='font-size:12px;padding-top:6px'>{c.get('empresa_codigo','—')}</div>",
                        unsafe_allow_html=True,
                    )
                    if rc[6].button("▲" if _is_open else "▼", key=f"ctog_{c['id']}", use_container_width=True):
                        st.session_state[_toggle_key] = not _is_open
                        st.rerun()

                    if _is_open:
                        with st.container(border=True):
                            _di1, _di2, _di3, _di4 = st.columns(4)
                            _di1.metric("Edad", c.get("age") or "—")
                            _di2.metric("Sexo", c.get("sex") or "—")
                            _di3.metric("Educación", c.get("education") or "—")
                            _di4.metric("Evaluaciones", _n_evals)
                            if c.get("correo"):    st.caption(f"📧 {c['correo']}")
                            if c.get("jefe_inmediato"): st.caption(f"👔 Jefe: {c['jefe_inmediato']}")
                            if _csessions:
                                st.markdown("**Evaluaciones:**")
                                for s in _csessions:
                                    _se = {"pending": "⏳", "in_progress": "▶️", "completed": "✅", "expired": "⏰", "employee_done": "📝"}.get(s["status"], "❓")
                                    _fp = ""
                                    _dr = s.get("completed_at") or s.get("started_at")
                                    if _dr:
                                        try: _fp = f" — {datetime.strptime(_dr, '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y %H:%M')}"
                                        except: _fp = f" — {_dr}"
                                    st.markdown(f"  - {_se} {s['test_type'].upper()} (ID: `{s['id']}`) — {s['status']}{_fp}")

                            # Botón editar
                            _edit_key = f"edit_cand_{c['id']}"
                            if st.button("✏️ Editar datos", key=f"edit_btn_{c['id']}", use_container_width=False):
                                st.session_state[_edit_key] = not st.session_state.get(_edit_key, False)
                                st.rerun()

                            if st.session_state.get(_edit_key, False):
                                st.markdown("---")
                                st.markdown("**✏️ Editar datos del candidato**")
                                with st.form(key=f"form_edit_cand_{c['id']}"):
                                    st.text_input("Cédula", value=c.get("cedula", ""), disabled=True)
                                    _ef1, _ef2 = st.columns(2)
                                    _e_name  = _ef1.text_input("Apellidos y Nombres", value=c.get("name", ""))
                                    _e_cargo = _ef2.text_input("Cargo", value=c.get("position", ""))
                                    _ef3, _ef4, _ef5 = st.columns(3)
                                    _e_age   = _ef3.number_input("Edad", min_value=0, max_value=100, value=int(c.get("age") or 0))
                                    _e_sex   = _ef4.selectbox("Sexo", ["", "Masculino", "Femenino", "Otro"],
                                                              index=["", "Masculino", "Femenino", "Otro"].index(c.get("sex") or "") if (c.get("sex") or "") in ["", "Masculino", "Femenino", "Otro"] else 0)
                                    _e_edu   = _ef5.text_input("Educación", value=c.get("education", "") or "")
                                    _ef6, _ef7 = st.columns(2)
                                    _e_correo = _ef6.text_input("Correo", value=c.get("correo", "") or "")
                                    _e_jefe   = _ef7.text_input("Jefe Inmediato", value=c.get("jefe_inmediato", "") or "")
                                    _ef8, _ef9 = st.columns(2)
                                    _nc_opts = ["Operativo", "Líder", "Directivo", "Administrativo"]
                                    _nc_cur = c.get("nivel_cargo") or "Operativo"
                                    _e_nivel = _ef8.selectbox("Nivel de Cargo", _nc_opts,
                                                              index=_nc_opts.index(_nc_cur) if _nc_cur in _nc_opts else 0)
                                    _e_regional = _ef9.text_input("Regional", value=c.get("regional", "") or "")
                                    _es1, _es2, _ = st.columns([1, 1, 2])
                                    _edit_ok = _es1.form_submit_button("💾 Guardar", use_container_width=True)
                                    _edit_cancel = _es2.form_submit_button("✖ Cancelar", use_container_width=True)
                                    if _edit_ok:
                                        db.update_empleado(
                                            candidate_id=c["id"],
                                            name=_e_name.strip(),
                                            age=_e_age if _e_age > 0 else None,
                                            sex=_e_sex if _e_sex else None,
                                            education=_e_edu.strip() or None,
                                            position=_e_cargo.strip(),
                                            correo=_e_correo.strip() or None,
                                            jefe_inmediato=_e_jefe.strip() or None,
                                            nivel_cargo=_e_nivel,
                                            regional=_e_regional.strip() or None,
                                        )
                                        st.session_state.pop(_edit_key, None)
                                        st.success(f"✅ Candidato **{_e_name.strip()}** actualizado.")
                                        st.rerun()
                                    if _edit_cancel:
                                        st.session_state.pop(_edit_key, None)
                                        st.rerun()

                            if admin.get("role") == "superadmin":
                                st.markdown("---")
                                if st.button(f"🗑️ Eliminar candidato", key=f"del_cand_{c['id']}"):
                                    st.session_state[f"confirm_del_cand_{c['id']}"] = True
                                if st.session_state.get(f"confirm_del_cand_{c['id']}", False):
                                    _ns = len(_csessions)
                                    st.warning(f"⚠️ ¿Eliminar **{c['name']}** (CC: {c['cedula']})? Se eliminarán {_ns} evaluación(es). Irreversible.")
                                    _cy, _cn, _ = st.columns([1, 1, 2])
                                    with _cy:
                                        if st.button("✅ Sí, eliminar", key=f"confirm_yes_cand_{c['id']}"):
                                            db.delete_candidate(c['id'])
                                            st.session_state.pop(f"confirm_del_cand_{c['id']}", None)
                                            st.success(f"Candidato **{c['name']}** eliminado.")
                                            st.rerun()
                                    with _cn:
                                        if st.button("❌ Cancelar", key=f"confirm_no_cand_{c['id']}"):
                                            st.session_state.pop(f"confirm_del_cand_{c['id']}", None)
                                            st.rerun()
                    st.divider()

    # ----- SECCIÓN: Dashboard Gerencial -----
    if _active_section == "dashboard":
        st.markdown("## 📈 Dashboard Gerencial de Evaluaciones")
        st.caption("Resumen ejecutivo de pruebas realizadas — actualizado en tiempo real.")
        st.markdown("---")

        _all_completed = [s for s in _all_raw if s["status"] in ("completed", "expired")]
        _all_pending   = [s for s in _all_raw if s["status"] in ("pending", "in_progress", "employee_done")]
        _total         = len(_all_raw)
        _pct_done = int(len(_all_completed) / _total * 100) if _total else 0

        # ── KPIs principales ──────────────────────────────────────────────
        _k1, _k2, _k3, _k4 = st.columns(4)
        _k1.metric("📋 Total Evaluaciones", _total)
        _k2.metric("✅ Completadas", len(_all_completed), f"{_pct_done}% del total")
        _k3.metric("⏳ Pendientes / En curso", len(_all_pending))
        _k4.metric("👥 Candidatos activos", len(set(s["candidate_id"] for s in _all_raw)))

        st.markdown("---")

        # ── Filtros del dashboard ─────────────────────────────────────────
        _TEST_LABELS = {
            "disc": "🎯 DISC", "valanti": "🧭 VALANTI", "wpi": "💼 WPI",
            "eri": "🔐 ERI", "talent_map": "🌟 Talent Map",
            "desempeno": "📊 Desempeño Op.", "desempeno_lider": "📊 Desempeño Líd.",
            "periodo_prueba": "📋 Per. Prueba",
        }
        from collections import Counter, defaultdict

        # Fechas disponibles
        _all_dates = []
        for _s in _all_completed:
            _dr = _s.get("completed_at") or _s.get("started_at") or _s.get("created_at") or ""
            try:
                _all_dates.append(datetime.strptime(_dr, "%Y-%m-%d %H:%M:%S").date())
            except:
                pass

        _min_date = min(_all_dates) if _all_dates else datetime.today().date()
        _max_date = max(_all_dates) if _all_dates else datetime.today().date()

        # Cargos disponibles
        _all_cargos = sorted(set(
            (s.get("position") or "Sin cargo").strip() or "Sin cargo"
            for s in _all_completed
        ))

        with st.expander("🔎 Filtros del Dashboard", expanded=False):
            _filt_col1, _filt_col2, _filt_col3 = st.columns(3)
            with _filt_col1:
                _f_desde = st.date_input("Desde", value=_min_date, key="db_f_desde")
                _f_hasta = st.date_input("Hasta", value=_max_date, key="db_f_hasta")
            with _filt_col2:
                _f_tipos = st.multiselect(
                    "Tipo de prueba",
                    options=list(_TEST_LABELS.keys()),
                    default=[],
                    format_func=lambda x: _TEST_LABELS.get(x, x),
                    key="db_f_tipos",
                )
            with _filt_col3:
                _f_cargos = st.multiselect(
                    "Cargo",
                    options=_all_cargos,
                    default=[],
                    key="db_f_cargos",
                )

        # Aplicar filtros a completadas
        def _dash_date(s):
            _dr = s.get("completed_at") or s.get("started_at") or s.get("created_at") or ""
            try:
                return datetime.strptime(_dr, "%Y-%m-%d %H:%M:%S").date()
            except:
                return None

        _comp_filt = _all_completed
        if _f_desde or _f_hasta:
            _comp_filt = [s for s in _comp_filt
                          if _dash_date(s) and _f_desde <= _dash_date(s) <= _f_hasta]
        if _f_tipos:
            _comp_filt = [s for s in _comp_filt if s["test_type"] in _f_tipos]
        if _f_cargos:
            _comp_filt = [s for s in _comp_filt
                          if ((s.get("position") or "Sin cargo").strip() or "Sin cargo") in _f_cargos]

        if len(_comp_filt) < len(_all_completed):
            st.caption(f"🔎 Mostrando **{len(_comp_filt)}** de {len(_all_completed)} evaluaciones completadas según filtros.")

        st.markdown("---")

        # ── Distribución por tipo de prueba (tabla) ───────────────────────
        _counts_total     = Counter(s["test_type"] for s in _all_raw)
        _counts_completed = Counter(s["test_type"] for s in _comp_filt)
        _counts_pending   = Counter(s["test_type"] for s in _all_pending)

        st.markdown("### 🔢 Pruebas por Tipo")
        _tipo_data = []
        for _tt, _lbl in _TEST_LABELS.items():
            _tot = _counts_total.get(_tt, 0)
            if _tot == 0:
                continue
            _done = _counts_completed.get(_tt, 0)
            _pend = _counts_pending.get(_tt, 0)
            _pct  = int(_done / _tot * 100) if _tot else 0
            _tipo_data.append((_lbl, _tot, _done, _pend, _pct))

        if _tipo_data:
            _th1, _th2, _th3, _th4, _th5 = st.columns([2.2, 0.9, 0.9, 0.9, 3])
            for _hdr_txt, _hdr_col in [("Tipo de Prueba", _th1), ("Total", _th2),
                                        ("Completadas", _th3), ("Pendientes", _th4), ("Progreso", _th5)]:
                _hdr_col.markdown(
                    f"<div style='font-size:11px;font-weight:700;color:#aaa;"
                    f"border-bottom:1px solid #444;padding-bottom:3px'>{_hdr_txt}</div>",
                    unsafe_allow_html=True,
                )
            for _lbl, _tot, _done, _pend, _pct in sorted(_tipo_data, key=lambda x: -x[1]):
                _c1, _c2, _c3, _c4, _c5 = st.columns([2.2, 0.9, 0.9, 0.9, 3])
                _c1.markdown(f"<div style='padding-top:6px;font-size:13px'>{_lbl}</div>", unsafe_allow_html=True)
                _c2.markdown(f"<div style='padding-top:6px;text-align:center;font-weight:600'>{_tot}</div>", unsafe_allow_html=True)
                _c3.markdown(f"<div style='padding-top:6px;text-align:center;color:#10B981;font-weight:600'>{_done}</div>", unsafe_allow_html=True)
                _c4.markdown(f"<div style='padding-top:6px;text-align:center;color:#F59E0B;font-weight:600'>{_pend}</div>", unsafe_allow_html=True)
                _bar_fill = "#10B981" if _pct >= 75 else ("#F59E0B" if _pct >= 40 else "#EF4444")
                _c5.markdown(
                    f"<div style='background:rgba(128,128,128,0.15);border-radius:6px;height:18px;margin-top:7px'>"
                    f"<div style='background:{_bar_fill};width:{_pct}%;height:18px;border-radius:6px;"
                    f"display:flex;align-items:center;justify-content:center;"
                    f"font-size:10px;color:white;font-weight:700'>{_pct}%</div></div>",
                    unsafe_allow_html=True,
                )
        else:
            st.info("Aún no hay evaluaciones registradas.")

        st.markdown("---")

        # ── Gráficos: Mensual | Por tipo | Por cargo (Plotly interactivo) ──
        st.markdown("### 📅 Distribución de Evaluaciones")
        import plotly.graph_objects as _pgo
        from collections import defaultdict

        # Detectar tema claro/oscuro
        try:
            _theme_base = st.get_option("theme.base") or "light"
        except Exception:
            _theme_base = "light"
        _ui_mode = st.session_state.get("ui_theme_mode", "Sistema")
        if _ui_mode == "Oscuro":
            _dash_dark = True
        elif _ui_mode == "Claro":
            _dash_dark = False
        else:
            _dash_dark = (_theme_base == "dark")
        _fc = "#FFFFFF" if _dash_dark else "#1E293B"
        _gc = "rgba(255,255,255,0.15)" if _dash_dark else "rgba(0,0,0,0.08)"
        _plotly_tpl = "plotly_dark" if _dash_dark else "plotly_white"

        _FIG_H = 380
        _gcol1, _gcol2, _gcol3 = st.columns(3)

        # --- Gráfico 1: Completadas por mes ---
        _mes_counts = defaultdict(int)
        for s in _comp_filt:
            _date_ref = s.get("completed_at") or s.get("started_at") or s.get("created_at") or ""
            try:
                _mes = datetime.strptime(_date_ref, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m")
                _mes_counts[_mes] += 1
            except:
                pass
        with _gcol1:
            st.markdown("**📅 Completadas por Mes**")
            if _mes_counts:
                _meses = sorted(_mes_counts.keys())
                _vals_m = [_mes_counts[m] for m in _meses]
                _meses_lbl = [datetime.strptime(m, "%Y-%m").strftime("%b %Y") for m in _meses]
                _fig_m = _pgo.Figure(_pgo.Bar(
                    x=_meses_lbl, y=_vals_m,
                    marker_color="#3B82F6",
                    text=_vals_m, textposition="outside",
                    textfont=dict(color=_fc),
                    hovertemplate="<b>%{x}</b><br>Completadas: %{y}<extra></extra>",
                ))
                _fig_m.update_layout(
                    template=_plotly_tpl,
                    height=_FIG_H, margin=dict(l=20, r=20, t=20, b=50),
                    plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                    font=dict(color=_fc),
                    xaxis=dict(tickangle=-30, tickfont=dict(size=10, color=_fc),
                               gridcolor=_gc, linecolor=_gc),
                    yaxis=dict(title="Cantidad", tickfont=dict(color=_fc),
                               gridcolor=_gc, linecolor=_gc),
                    showlegend=False,
                )
                st.plotly_chart(_fig_m, use_container_width=True)
            else:
                st.info("Sin datos para los filtros seleccionados.")

        # --- Gráfico 2: Por tipo de prueba (pie interactivo) ---
        with _gcol2:
            st.markdown("**🔢 Completadas por Tipo de Prueba**")
            _tipo_done_filt = {_TEST_LABELS[t].split(" ", 1)[-1]: _counts_completed.get(t, 0)
                               for t in _TEST_LABELS if _counts_completed.get(t, 0) > 0}
            if _tipo_done_filt:
                _tipo_colors = ["#3B82F6", "#10B981", "#8B5CF6", "#EF4444",
                                 "#F59E0B", "#06B6D4", "#EC4899", "#84CC16"]
                _fig_t = _pgo.Figure(_pgo.Pie(
                    labels=list(_tipo_done_filt.keys()),
                    values=list(_tipo_done_filt.values()),
                    marker=dict(colors=_tipo_colors[:len(_tipo_done_filt)],
                                line=dict(color="white", width=2)),
                    textinfo="label+value",
                    textfont=dict(color=_fc),
                    hovertemplate="<b>%{label}</b><br>Cantidad: %{value}<br>%{percent}<extra></extra>",
                    hole=0.0,
                ))
                _fig_t.update_layout(
                    template=_plotly_tpl,
                    height=_FIG_H, margin=dict(l=10, r=10, t=20, b=10),
                    paper_bgcolor="rgba(0,0,0,0)",
                    font=dict(color=_fc),
                    legend=dict(orientation="v", x=1.02, y=0.5,
                                font=dict(size=10, color=_fc)),
                    showlegend=True,
                )
                st.plotly_chart(_fig_t, use_container_width=True)
            else:
                st.info("Sin evaluaciones completadas con los filtros actuales.")

        # --- Gráfico 3: Por cargo (barras horizontales interactivas) ---
        with _gcol3:
            st.markdown("**💼 Completadas por Cargo**")
            _cargo_counts = defaultdict(int)
            for s in _comp_filt:
                _cargo = (s.get("position") or "Sin cargo").strip() or "Sin cargo"
                _cargo_counts[_cargo] += 1
            if _cargo_counts:
                _top_cargos = sorted(_cargo_counts.items(), key=lambda x: x[1])[-8:]
                _cargo_lbl  = [c[:25] for c, _ in _top_cargos]
                _cargo_vals = [v for _, v in _top_cargos]
                _cargo_colors_c = ["#6366F1" if i % 2 == 0 else "#818CF8" for i in range(len(_cargo_lbl))]
                _fig_c = _pgo.Figure(_pgo.Bar(
                    x=_cargo_vals, y=_cargo_lbl,
                    orientation="h",
                    marker_color=_cargo_colors_c,
                    text=_cargo_vals, textposition="outside",
                    textfont=dict(color=_fc),
                    hovertemplate="<b>%{y}</b><br>Completadas: %{x}<extra></extra>",
                ))
                _fig_c.update_layout(
                    template=_plotly_tpl,
                    height=_FIG_H, margin=dict(l=20, r=40, t=20, b=30),
                    plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                    font=dict(color=_fc),
                    xaxis=dict(title="Cantidad", tickfont=dict(color=_fc),
                               gridcolor=_gc, linecolor=_gc),
                    yaxis=dict(tickfont=dict(size=9, color=_fc)),
                    showlegend=False,
                )
                st.plotly_chart(_fig_c, use_container_width=True)
            else:
                st.info("Sin datos de cargo para los filtros seleccionados.")

        st.markdown("---")

        # ── Top 10 candidatos con más evaluaciones ─────────────────────────
        st.markdown("### 🏆 Top Candidatos con Más Evaluaciones")
        _cand_counts = Counter(s["candidate_name"] for s in _all_raw)
        _top_cands = _cand_counts.most_common(10)
        if _top_cands:
            _tc1, _tc2, _tc3 = st.columns([0.4, 0.8, 0.8])
            for _rk, (_cname, _cnt) in enumerate(_top_cands, 1):
                _done_cand = sum(1 for s in _all_completed if s["candidate_name"] == _cname)
                _medal = "🥇" if _rk == 1 else ("🥈" if _rk == 2 else ("🥉" if _rk == 3 else f"{_rk}."))
                st.markdown(
                    f"<div style='display:flex;align-items:center;gap:12px;padding:6px 8px;"
                    f"background:{'#1E3A5F11' if _rk % 2 == 0 else 'transparent'};"
                    f"border-radius:6px;margin-bottom:3px'>"
                    f"<span style='font-size:16px;min-width:28px'>{_medal}</span>"
                    f"<span style='font-weight:600;flex:1'>{_cname}</span>"
                    f"<span style='color:#3B82F6;font-weight:700;min-width:60px;text-align:right'>"
                    f"{_cnt} total</span>"
                    f"<span style='color:#10B981;min-width:80px;text-align:right'>"
                    f"{_done_cand} completadas</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

    # ----- SECCIÓN: Resultados -----
    elif _active_section == "results":
        st.markdown("### Resultados de Evaluaciones")
        _res3 = [
            s for s in _all_raw
            if (_ft is None or s["test_type"] == _ft) and s["status"] in ("completed", "expired")
        ]
        if _filter_cand != "Todos":
            _res3 = [s for s in _res3 if s["candidate_name"] == _filter_cand]
        _sort_sessions(_res3)
        _render_sessions_list(_res3, "res")

    # ----- SECCIÓN: Pruebas Pendientes -----
    elif _active_section == "pending":
        st.markdown("### Pruebas Pendientes")
        _pend4 = [
            s for s in _all_raw
            if (_ft is None or s["test_type"] == _ft) and s["status"] in ("pending", "in_progress", "employee_done")
        ]
        if _filter_cand != "Todos":
            _pend4 = [s for s in _pend4 if s["candidate_name"] == _filter_cand]
        _sort_sessions(_pend4)
        _render_sessions_list(_pend4, "pend")

    # ----- SECCIÓN: Cargue Masivo -----
    elif _active_section == "bulk":
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        _BULK_THIN = Side(border_style="thin", color="BFBFBF")

        def _bulk_border():
            s = _BULK_THIN
            return Border(left=s, right=s, top=s, bottom=s)

        def _bulk_hdr(ws, r, c, v, w=None):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _bulk_border()
            if w:
                ws.column_dimensions[get_column_letter(c)].width = w

        def _bulk_data(ws, r, c, v="", bg="FFFFFF"):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = _bulk_border()

        def _gen_plantilla_candidatos():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BD empleados"
            ws.sheet_view.showGridLines = False
            ws.freeze_panes = "A2"
            cols = [
                ("CEDULA", 18, True), ("EMPRESA", 12, True), ("REGIONAL", 15, False),
                ("APELLIDOS Y NOMBRES", 30, True), ("CORREO", 28, False),
                ("CARGO", 22, False), ("INVITAR", 10, False),
                ("JEFE INMEDIATO", 25, False), ("NIVEL DE CARGO", 18, False),
            ]
            ws.row_dimensions[1].height = 32
            for ci, (nm, wd, req) in enumerate(cols, 1):
                _bulk_hdr(ws, 1, ci, nm, wd)
                bg = "FFF2CC" if req else "E2EFDA"
                for r in range(2, 502):
                    _bulk_data(ws, r, ci, bg=bg)
            # Fila ejemplo
            ej = ["1020304050","HESEGO","Bogotá","García López Juan Carlos",
                  "juan@hesego.com","Analista","SI","María Rodríguez","Operativo"]
            for ci, v in enumerate(ej, 1):
                c = ws.cell(row=2, column=ci, value=v)
                c.fill = PatternFill("solid", fgColor="DDEEFF")
                c.font = Font(italic=True, size=10, color="2F5496")
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = _bulk_border()
            # Validaciones
            dv1 = DataValidation(type="list", formula1='"HESEGO"', allow_blank=True, showDropDown=False)
            dv1.sqref = "B2:B501"
            ws.add_data_validation(dv1)
            dv2 = DataValidation(type="list", formula1='"SI,NO"', allow_blank=True, showDropDown=False)
            dv2.sqref = "G2:G501"
            ws.add_data_validation(dv2)
            dv3 = DataValidation(type="list", formula1='"Operativo,Líder,Directivo,Administrativo"', allow_blank=True, showDropDown=False)
            dv3.sqref = "I2:I501"
            ws.add_data_validation(dv3)
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf.read()

        def _gen_plantilla_pruebas():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Pruebas"
            ws.sheet_view.showGridLines = False
            ws.freeze_panes = "A2"
            cols = [
                ("CEDULA_CANDIDATO", 18, True), ("TIPO_PRUEBA", 22, True),
                ("TIEMPO_LIMITE_MINUTOS", 22, False),
                ("CEDULA_EVALUADOR", 18, False), ("NOMBRE_EVALUADOR", 30, False),
            ]
            ws.row_dimensions[1].height = 32
            for ci, (nm, wd, req) in enumerate(cols, 1):
                _bulk_hdr(ws, 1, ci, nm, wd)
                bg = "FFF2CC" if req else "E2EFDA"
                for r in range(2, 502):
                    _bulk_data(ws, r, ci, bg=bg)
            ejemplos = [
                ("1020304050","disc","30","",""),
                ("1020304051","wpi","30","",""),
                ("1020304052","desempeno_lider","60","9876543210","María Rodríguez"),
                ("1020304053","periodo_prueba","60","9876543210","María Rodríguez"),
            ]
            for ri, vals in enumerate(ejemplos):
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(row=2+ri, column=ci, value=v)
                    c.fill = PatternFill("solid", fgColor="DDEEFF")
                    c.font = Font(italic=True, size=10, color="2F5496")
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    c.border = _bulk_border()
            tipos = "disc,valanti,wpi,eri,talent_map,desempeno,desempeno_lider,periodo_prueba"
            dv = DataValidation(type="list", formula1=f'"{tipos}"', allow_blank=False, showDropDown=False)
            dv.sqref = "B2:B501"
            ws.add_data_validation(dv)
            dv_t = DataValidation(type="whole", operator="between", formula1="5", formula2="180", allow_blank=True)
            dv_t.sqref = "C2:C501"
            ws.add_data_validation(dv_t)
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf.read()

        st.markdown("### 📤 Cargue Masivo")

        # ── Descarga de plantillas ────────────────────────────────────────────
        st.markdown("#### 📥 Descargar plantillas de ejemplo")
        _dc1, _dc2 = st.columns(2)
        with _dc1:
            st.download_button(
                label="⬇️ Plantilla Candidatos (.xlsx)",
                data=_gen_plantilla_candidatos(),
                file_name="plantilla_cargue_candidatos.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.caption("Columnas: CEDULA, EMPRESA, REGIONAL, APELLIDOS Y NOMBRES, CORREO, CARGO, INVITAR, JEFE INMEDIATO, NIVEL DE CARGO")
        with _dc2:
            st.download_button(
                label="⬇️ Plantilla Pruebas (.xlsx)",
                data=_gen_plantilla_pruebas(),
                file_name="plantilla_cargue_pruebas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.caption("Columnas: CEDULA_CANDIDATO, TIPO_PRUEBA, TIEMPO_LIMITE_MINUTOS, CEDULA_EVALUADOR, NOMBRE_EVALUADOR")

        st.markdown("---")

        # ── Cargue masivo de candidatos ───────────────────────────────────────
        st.markdown("#### 👥 Cargar candidatos desde Excel")
        _uploaded_cands = st.file_uploader(
            "Selecciona el Excel de candidatos (hoja: BD empleados)",
            type=["xlsx"],
            key="bulk_upload_cands",
        )
        if _uploaded_cands:
            try:
                _wb_c = openpyxl.load_workbook(_uploaded_cands, data_only=True)
                if "BD empleados" not in _wb_c.sheetnames:
                    st.error("❌ El archivo no tiene una hoja llamada 'BD empleados'.")
                else:
                    _ws_c = _wb_c["BD empleados"]
                    _ok_c, _dup_c, _err_c = 0, 0, []
                    for _ri in range(2, _ws_c.max_row + 1):
                        _cedula = _ws_c.cell(_ri, 1).value
                        _emp_cod = _ws_c.cell(_ri, 2).value
                        _regional = _ws_c.cell(_ri, 3).value
                        _nombre = _ws_c.cell(_ri, 4).value
                        _correo = _ws_c.cell(_ri, 5).value
                        _cargo = _ws_c.cell(_ri, 6).value
                        _invitar = _ws_c.cell(_ri, 7).value
                        _jefe = _ws_c.cell(_ri, 8).value
                        _nivel = _ws_c.cell(_ri, 9).value
                        if not _cedula or not _nombre:
                            continue
                        try:
                            _cedula = str(int(_cedula)) if isinstance(_cedula, float) else str(_cedula).strip()
                            _res = db.create_empleado(
                                cedula=_cedula,
                                name=str(_nombre).strip(),
                                empresa_codigo=str(_emp_cod).strip() if _emp_cod else "",
                                regional=str(_regional).strip() if _regional else "",
                                correo=str(_correo).strip() if _correo else "",
                                position=str(_cargo).strip() if _cargo else "",
                                jefe_inmediato=str(_jefe).strip() if _jefe and str(_jefe) != "#N/A" else "",
                                nivel_cargo=str(_nivel).strip() if _nivel else "",
                                invitar=str(_invitar).strip().upper() if _invitar else "SI",
                            )
                            if _res:
                                _ok_c += 1
                            else:
                                _dup_c += 1
                        except Exception as _e:
                            _err_c.append(f"Fila {_ri}: {_e}")
                    if _ok_c > 0:
                        st.success(f"✅ {_ok_c} candidato(s) importado(s) correctamente.")
                    if _dup_c > 0:
                        st.warning(f"⚠️ {_dup_c} candidato(s) omitido(s) por cédula duplicada.")
                    for _em in _err_c[:5]:
                        st.error(f"❌ {_em}")
            except Exception as _ex:
                st.error(f"❌ Error al leer el archivo: {_ex}")

        st.markdown("---")

        # ── Cargue masivo de pruebas ──────────────────────────────────────────
        st.markdown("#### 🧪 Cargar pruebas desde Excel")
        _TIEMPOS_DEFAULT = {
            "disc": 30, "valanti": 30, "wpi": 30,
            "eri": 20, "talent_map": 25,
            "desempeno": 60, "desempeno_lider": 60, "periodo_prueba": 60,
        }
        _TIPOS_VALIDOS = set(_TIEMPOS_DEFAULT.keys())
        _uploaded_tests = st.file_uploader(
            "Selecciona el Excel de pruebas (hoja: Pruebas)",
            type=["xlsx"],
            key="bulk_upload_tests",
        )
        if _uploaded_tests:
            try:
                _wb_t = openpyxl.load_workbook(_uploaded_tests, data_only=True)
                if "Pruebas" not in _wb_t.sheetnames:
                    st.error("❌ El archivo no tiene una hoja llamada 'Pruebas'.")
                else:
                    _ws_t = _wb_t["Pruebas"]
                    _ok_t, _err_t, _warn_t = 0, [], []
                    for _ri in range(2, _ws_t.max_row + 1):
                        _ced_t = _ws_t.cell(_ri, 1).value
                        _tipo = _ws_t.cell(_ri, 2).value
                        _tiempo = _ws_t.cell(_ri, 3).value
                        _eval_ced_bulk = _ws_t.cell(_ri, 4).value
                        _eval_nom_bulk = _ws_t.cell(_ri, 5).value
                        # Omitir filas completamente vacías
                        if not _ced_t and not _tipo:
                            continue
                        # Validar campos obligatorios
                        if not _ced_t:
                            _err_t.append(f"Fila {_ri}: falta CEDULA_CANDIDATO.")
                            continue
                        if not _tipo:
                            _err_t.append(f"Fila {_ri}: falta TIPO_PRUEBA.")
                            continue
                        try:
                            _ced_t = str(int(_ced_t)) if isinstance(_ced_t, float) else str(_ced_t).strip()
                            _tipo = str(_tipo).strip().lower()
                            if _tipo not in _TIPOS_VALIDOS:
                                _err_t.append(f"Fila {_ri}: tipo '{_tipo}' no válido. Opciones: {', '.join(_TIPOS_VALIDOS)}")
                                continue
                            # Validar que el candidato exista
                            _cand_t = db.get_candidate_by_cedula(_ced_t)
                            if not _cand_t:
                                _err_t.append(f"Fila {_ri}: candidato con cédula **{_ced_t}** no existe en el sistema.")
                                continue
                            # Tiempo
                            _tl = int(_tiempo) if _tiempo else _TIEMPOS_DEFAULT[_tipo]
                            if _tl < 5 or _tl > 180:
                                _err_t.append(f"Fila {_ri}: tiempo {_tl} fuera de rango (5–180 min).")
                                continue
                            # Evaluador
                            _ec = str(_eval_ced_bulk).strip() if _eval_ced_bulk else None
                            _en = str(_eval_nom_bulk).strip() if _eval_nom_bulk else None
                            if _tipo in ("desempeno", "desempeno_lider", "periodo_prueba") and not _ec:
                                _err_t.append(f"Fila {_ri} ({_ced_t}): '{_tipo}' requiere CEDULA_EVALUADOR.")
                                continue
                            # Lookup automático del nombre del evaluador si no fue especificado
                            if _ec and not _en:
                                _jefe_bulk = db.get_candidate_by_cedula(_ec)
                                if _jefe_bulk:
                                    _en = _jefe_bulk["name"]
                                else:
                                    _warn_t.append(f"Fila {_ri}: evaluador cédula {_ec} no registrado — se guardó sin nombre.")
                            _sid, _serr = db.create_test_session(
                                _cand_t["id"], _tipo, _tl, admin["id"],
                                evaluador_cedula=_ec, evaluador_nombre=_en,
                            )
                            if _serr:
                                _err_t.append(f"Fila {_ri} ({_cand_t['name']}): {_serr}")
                            else:
                                _ok_t += 1
                        except Exception as _e:
                            _err_t.append(f"Fila {_ri}: {_e}")
                    if _ok_t > 0:
                        st.success(f"✅ {_ok_t} prueba(s) creada(s) correctamente.")
                    for _wm in _warn_t:
                        st.warning(f"⚠️ {_wm}")
                    for _em in _err_t[:15]:
                        st.error(f"❌ {_em}")
                    if len(_err_t) > 15:
                        st.error(f"... y {len(_err_t) - 15} error(es) más. Corrige el archivo y vuelve a subirlo.")
            except Exception as _ex:
                st.error(f"❌ Error al leer el archivo: {_ex}")

    # ----- SECCIÓN: Configuración -----
    elif _active_section == "settings":
        st.markdown("### Cambiar Contraseña de Administrador")
        with st.form("change_pw"):
            new_pw = st.text_input("Nueva Contraseña", type="password")
            confirm_pw = st.text_input("Confirmar Contraseña", type="password")
            if st.form_submit_button("Cambiar Contraseña"):
                if new_pw and new_pw == confirm_pw:
                    db.change_admin_password(admin["id"], new_pw)
                    st.success("✅ Contraseña actualizada.")
                else:
                    st.error("Las contraseñas no coinciden o están vacías.")


def show_disc_results_admin(results, candidate, session):
    """Show DISC results in the admin panel."""
    normalized = results.get("normalized", {})
    relative = results.get("relative", {})

    # Análisis de aptitud
    analysis = analyze_disc_aptitude(normalized, relative)

    # Estilos conductuales, temperamento y mega resumen (nuevas funcionalidades THT-inspired)
    behavioral_styles = calculate_behavioral_styles(normalized)
    temperament = get_disc_temperament(normalized)
    mega_summary = generate_disc_mega_summary(normalized)

    # Banner de aptitud
    st.markdown(f"""
    <div style="background: {analysis['aptitude_color']}22; border-left: 5px solid {analysis['aptitude_color']};
                padding: 15px 20px; border-radius: 8px; margin-bottom: 15px;">
        <h3 style="margin: 0; color: {analysis['aptitude_color']};">{analysis['aptitude_emoji']} {analysis['aptitude_level']} — Puntaje: {analysis['aptitude_score']}/100</h3>
        <p style="margin: 5px 0 0 0; color: #374151;">{analysis['aptitude_desc']}</p>
        <p style="margin: 5px 0 0 0; color: #6B7280;"><b>Perfil:</b> {analysis['profile_name']} ({analysis['dominant_name']} + {analysis['secondary_name']})</p>
        <p style="margin: 5px 0 0 0; color: #6B7280;"><b>Temperamento:</b> {temperament['label']}</p>
    </div>
    """, unsafe_allow_html=True)

    cols = st.columns(4)
    disc_colors = {"D": "#EF4444", "I": "#F59E0B", "S": "#10B981", "C": "#3B82F6"}
    disc_names = {"D": "Dominancia", "I": "Influencia", "S": "Estabilidad", "C": "Cumplimiento"}
    for idx, style in enumerate("DISC"):
        with cols[idx]:
            st.metric(f"{style} — {disc_names[style]}", f"{normalized.get(style, 0):.1f}%",
                      f"Rel: {relative.get(style, 0):.1f}%")

    fig = create_disc_plot(normalized)
    st.pyplot(fig)

    # ── MEGA RESUMEN ──────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 📋 Resumen Conductual")
    st.caption("Descripción detallada del perfil conductual en 16 dimensiones, derivada del resultado DISC.")

    cols_summary = st.columns(2)
    items = list(mega_summary.items())
    half = len(items) // 2
    for col, chunk in zip(cols_summary, [items[:half], items[half:]]):
        with col:
            for label, text in chunk:
                st.markdown(f"""
                <div style="background:#F8FAFC; border-left:3px solid #3B82F6; padding:10px 14px;
                            border-radius:6px; margin-bottom:8px;">
                    <b style="color:#1E40AF; font-size:0.85em;">{label}</b><br>
                    <span style="color:#374151; font-size:0.92em;">{text}</span>
                </div>""", unsafe_allow_html=True)

    # ── 9 ESTILOS CONDUCTUALES ────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 🎯 Estilos Conductuales Derivados")
    st.caption("9 estilos con 4 sub-dimensiones cada uno, inferidos matemáticamente del perfil DISC (metodología THT).")

    styles_fig = create_behavioral_styles_chart(behavioral_styles)
    st.pyplot(styles_fig)

    # Detalle expandible de cada estilo
    with st.expander("🔍 Ver descripción detallada de cada sub-dimensión"):
        for style_name, style_data in behavioral_styles.items():
            st.markdown(f"**{style_name}**")
            for sub_name, sub_val in style_data["subs"].items():
                desc = style_data["desc"][sub_name]
                bar_w = int(sub_val)
                color = "#EF4444" if sub_name in ["Franqueza", "Control", "Insistencia", "Por Resultados",
                                                  "Resolución", "Confrontación", "Priorización",
                                                  "Pragmatismo"] else (
                        "#F59E0B" if sub_name in ["Expresividad", "Inspiración", "Optimismo", "Por Inspiración",
                                                  "Positivismo", "Apasionamiento", "Entusiasmo",
                                                  "Extroversión", "Persuasión"] else (
                        "#10B981" if sub_name in ["Autoregulación", "Moderación", "Focalización", "Democrático",
                                                  "Resistencia", "Inalterabilidad", "Pausa", "Autocontrol",
                                                  "Calma"] else "#3B82F6"))
                st.markdown(f"""
                <div style="display:flex; align-items:center; gap:10px; margin:4px 0;">
                    <span style="min-width:140px; font-size:0.85em; color:#374151;">{sub_name}</span>
                    <div style="flex:1; background:#E2E8F0; border-radius:4px; height:14px; position:relative;">
                        <div style="width:{bar_w}%; background:{color}; border-radius:4px; height:14px;"></div>
                    </div>
                    <span style="font-weight:bold; color:{color}; min-width:30px;">{bar_w}</span>
                    <span style="font-size:0.75em; color:#94A3B8; flex:1;">{desc}</span>
                </div>""", unsafe_allow_html=True)
            st.markdown("")

    # ── FORTALEZAS Y ALERTAS ──────────────────────────────────────────────
    st.markdown("---")
    col_f, col_a = st.columns(2)
    with col_f:
        st.markdown("#### 💪 Fortalezas")
        for f in analysis['fortalezas']:
            st.markdown(f"- ✅ {f}")
    with col_a:
        st.markdown("#### ⚠️ Alertas")
        for a in analysis['alertas']:
            st.markdown(f"- 🔸 {a}")

    st.markdown("#### 📋 Recomendaciones para el Candidato")
    for r in analysis['recomendaciones']:
        st.markdown(f"- 💡 {r}")

    if analysis['ideal_para']:
        st.markdown("#### 🎯 Ideal para roles de")
        st.markdown(", ".join([f"**{r}**" for r in analysis['ideal_para']]))

    if analysis['cuidado_en']:
        st.markdown("#### ⛔ Tener cuidado en")
        st.markdown(", ".join([f"*{r}*" for r in analysis['cuidado_en']]))

    session_id = session if isinstance(session, str) else session.get("id")
    completed_at = session.get("completed_at") if isinstance(session, dict) else None
    pdf = generate_disc_pdf(candidate, normalized, relative, fig, session_id, completed_at, analysis,
                            behavioral_styles=behavioral_styles, temperament=temperament,
                            mega_summary=mega_summary, styles_fig=styles_fig)
    c1, c2 = st.columns(2)
    with c1:
        st.download_button("📑 Descargar PDF", data=pdf.getvalue(), file_name=f"disc_{candidate['cedula']}.pdf", mime="application/pdf", key=f"pdf_disc_{session_id}")
    with c2:
        st.download_button("📄 Descargar JSON", data=json.dumps(results, indent=2, ensure_ascii=False), file_name=f"disc_{candidate['cedula']}.json", mime="application/json", key=f"json_disc_{session_id}")


def show_valanti_results_admin(results, candidate, session):
    """Show VALANTI results in the admin panel."""
    direct = results.get("direct", {})
    standard = results.get("standard", {})
    
    # Análisis de aptitud
    analysis = analyze_valanti_aptitude(standard)
    
    # Banner de aptitud
    st.markdown(f"""
    <div style="background: {analysis['aptitude_color']}22; border-left: 5px solid {analysis['aptitude_color']};
                padding: 15px 20px; border-radius: 8px; margin-bottom: 15px;">
        <h3 style="margin: 0; color: {analysis['aptitude_color']};">{analysis['aptitude_emoji']} {analysis['aptitude_level']} — Puntaje: {analysis['aptitude_score']}/100</h3>
        <p style="margin: 5px 0 0 0; color: #374151;">{analysis['aptitude_desc']}</p>
        <p style="margin: 5px 0 0 0; color: #6B7280;"><b>Valor más fuerte:</b> {analysis['strongest_value']} (T={analysis['strongest_score']}) | <b>Valor más bajo:</b> {analysis['weakest_value']} (T={analysis['weakest_score']})</p>
    </div>
    """, unsafe_allow_html=True)

    cols = st.columns(5)
    for idx, trait in enumerate(VALANTI_TRAITS):
        with cols[idx]:
            st.metric(trait, standard.get(trait, 0), f"Dir: {direct.get(trait, 0)}")

    radar_fig = create_valanti_radar(standard)
    st.pyplot(radar_fig)

    bar_fig = create_valanti_bars(direct, standard)
    st.pyplot(bar_fig)

    sorted_scores = sorted(standard.items(), key=lambda x: x[1], reverse=True)
    st.markdown(f"**Valor más prominente:** {sorted_scores[0][0]} ({sorted_scores[0][1]})")
    st.markdown(f"**Valor menos enfatizado:** {sorted_scores[-1][0]} ({sorted_scores[-1][1]})")
    for trait, score in sorted_scores:
        desc = VALANTI_DESCRIPTIONS[trait]
        level = "Alto" if score >= 55 else ("Bajo" if score <= 45 else "Promedio")
        text = desc["high"] if score >= 55 else (desc["low"] if score <= 45 else "Puntaje dentro del rango promedio.")
        st.markdown(f"**{desc['title']}** — {level} ({score}): {text}")
    
    # Fortalezas y alertas
    if analysis['fortalezas']:
        st.markdown("#### 💪 Fortalezas Valorales")
        for f in analysis['fortalezas']:
            st.markdown(f"- ✅ {f}")
    
    if analysis['alertas']:
        st.markdown("#### ⚠️ Alertas")
        for a in analysis['alertas']:
            st.markdown(f"- 🔸 {a}")
    
    if analysis['recomendaciones']:
        st.markdown("#### 📋 Recomendaciones")
        for r in analysis['recomendaciones']:
            st.markdown(f"- {r}")

    session_id = session if isinstance(session, str) else session.get("id")
    completed_at = session.get("completed_at") if isinstance(session, dict) else None
    pdf = generate_valanti_pdf(candidate, direct, standard, radar_fig, session_id, completed_at, analysis)
    c1, c2 = st.columns(2)
    with c1:
        st.download_button("📑 Descargar PDF", data=pdf.getvalue(), file_name=f"valanti_{candidate['cedula']}.pdf", mime="application/pdf", key=f"pdf_val_{session_id}")
    with c2:
        st.download_button("📄 Descargar JSON", data=json.dumps(results, indent=2, ensure_ascii=False), file_name=f"valanti_{candidate['cedula']}.json", mime="application/json", key=f"json_val_{session_id}")


def show_wpi_results_admin(results, candidate, session):
    """
    Muestra los resultados del WPI en el panel de administración.
    
    Args:
        results: Dict con raw, normalized y percentages
        candidate: Dict con información del candidato
        session: Dict con información de la sesión o str con session_id
    """
    raw = results.get("raw", {})
    normalized = results.get("normalized", {})
    percentages = results.get("percentages", {})
    
    # Análisis de aptitud
    analysis = analyze_wpi_aptitude(normalized)
    
    # === BANNER DE APTITUD ===
    st.markdown(f"""
    <div style="background: {analysis['aptitude_color']}22; border-left: 5px solid {analysis['aptitude_color']};
                padding: 15px 20px; border-radius: 8px; margin-bottom: 15px;">
        <h3 style="margin: 0; color: {analysis['aptitude_color']};">
            {analysis['aptitude_emoji']} {analysis['aptitude_level']} — Puntaje: {analysis['aptitude_score']}/100
        </h3>
        <p style="margin: 5px 0 0 0; color: #374151;">{analysis['aptitude_desc']}</p>
        <p style="margin: 5px 0 0 0; color: #6B7280;">
            <b>Dimensión más fuerte:</b> {analysis['strongest_dimension']} ({int(analysis['strongest_score'])}/100) | 
            <b>Dimensión a desarrollar:</b> {analysis['weakest_dimension']} ({int(analysis['weakest_score'])}/100) |
            <b>Promedio:</b> {analysis['average_score']}/100
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # === MÉTRICAS POR DIMENSIÓN ===
    st.markdown("### 📊 Puntajes por Dimensión")
    
    # Crear 6 columnas para las 6 dimensiones
    cols = st.columns(3)
    for idx, dim in enumerate(WPI_DIMENSIONS):
        with cols[idx % 3]:
            score = normalized.get(dim, 0)
            nivel = "🟢 Alto" if score >= 70 else ("🟡 Medio" if score >= 45 else "🔴 Bajo")
            st.metric(
                label=dim,
                value=f"{int(score)}/100",
                delta=nivel,
                delta_color="off"
            )
    
    st.markdown("---")
    
    # === GRÁFICOS ===
    col_radar, col_bars = st.columns(2)
    
    with col_radar:
        st.markdown("#### 🎯 Perfil Radar")
        radar_fig = create_wpi_radar(normalized)
        st.pyplot(radar_fig)
    
    with col_bars:
        st.markdown("#### 📊 Puntajes por Dimensión")
        bar_fig = create_wpi_bars(normalized)
        st.pyplot(bar_fig)
    
    st.markdown("---")
    
    # === ANÁLISIS POR DIMENSIÓN ===
    st.markdown("### 📋 Análisis Detallado por Dimensión")
    
    sorted_scores = sorted(normalized.items(), key=lambda x: x[1], reverse=True)
    
    for dim, score in sorted_scores:
        desc_info = WPI_DESCRIPTIONS[dim]
        
        # Determinar nivel
        if score >= 70:
            level = "🟢 Alto"
            text = desc_info["high"]
            color = "#10B981"
        elif score >= 45:
            level = "🟡 Medio"
            text = desc_info["medium"]
            color = "#F59E0B"
        else:
            level = "🔴 Bajo"
            text = desc_info["low"]
            color = "#EF4444"
        
        st.markdown(f"""
        <div style="background: {color}15; border-left: 3px solid {color}; 
                    padding: 12px; border-radius: 6px; margin-bottom: 10px;">
            <b style="color: {color};">{desc_info['title']}</b> — {level} ({int(score)}/100)
            <br><span style="color: #374151;">{text}</span>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # === FORTALEZAS ===
    if analysis.get('fortalezas'):
        st.markdown("### 💪 Fortalezas Destacadas")
        for f in analysis['fortalezas']:
            # Limpiar markdown
            f_clean = f.replace("**", "")
            st.markdown(f"- ✅ {f_clean}")
        st.markdown("")
    
    # === ALERTAS ===
    if analysis.get('alertas'):
        st.markdown("### ⚠️ Áreas de Atención")
        for a in analysis['alertas']:
            # Limpiar markdown
            a_clean = a.replace("**", "").replace("⚠️ ", "")
            st.markdown(f"- 🔸 {a_clean}")
        st.markdown("")
    
    # === ROLES IDEALES ===
    if analysis.get('ideal_para'):
        st.markdown("### 🎯 Roles Ideales para el Candidato")
        for role in analysis['ideal_para']:
            st.markdown(f"- 🎯 {role}")
        st.markdown("")
    
    # === ROLES A EVITAR ===
    if analysis.get('avoid_roles'):
        st.markdown("### ⛔ Roles No Recomendados")
        for role in analysis['avoid_roles']:
            st.markdown(f"- ⛔ {role}")
        st.markdown("")
    
    # === RECOMENDACIONES ===
    if analysis.get('recomendaciones'):
        st.markdown("### 💡 Recomendaciones Específicas")
        for r in analysis['recomendaciones']:
            # Limpiar markdown
            r_clean = r.replace("**", "")
            st.markdown(f"- {r_clean}")
    
    st.markdown("---")
    
    # === DESCARGA DE REPORTES ===
    st.markdown("### 📥 Descargar Reportes")
    
    session_id = session if isinstance(session, str) else session.get("id")
    completed_at = session.get("completed_at") if isinstance(session, dict) else None
    
    # Generar PDF
    pdf_buffer = generate_wpi_pdf(candidate, raw, normalized, radar_fig, session_id, completed_at, analysis)
    
    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            "📑 Descargar PDF Completo",
            data=pdf_buffer.getvalue(),
            file_name=f"wpi_{candidate['cedula']}_{session_id}.pdf",
            mime="application/pdf",
            key=f"pdf_wpi_{session_id}"
        )
    with col2:
        st.download_button(
            "📄 Descargar JSON",
            data=json.dumps(results, indent=2, ensure_ascii=False),
            file_name=f"wpi_{candidate['cedula']}_{session_id}.json",
            mime="application/json",
            key=f"json_wpi_{session_id}"
        )


def show_eri_results_admin(results, candidate, session):
    """
    Muestra los resultados del ERI en el panel de administración.
    
    Args:
        results: Dict con raw, normalized, percentages, validity_score, validity_flags
        candidate: Dict con información del candidato
        session: Dict con información de la sesión o str con session_id
    """
    raw = results.get("raw", {})
    normalized = results.get("normalized", {})
    percentages = results.get("percentages", {})
    validity_score = results.get("validity_score", ERI_VALIDITY_QUESTIONS_COUNT)
    validity_flags = results.get("validity_flags", [])
    
    # Análisis de riesgo
    analysis = analyze_eri_aptitude(normalized, validity_score, validity_flags)
    
    # === BANNER DE VALIDEZ (si aplica) ===
    if analysis.get('validity_warning'):
        st.markdown(f"""
        <div style="background: #FEF2F2; border-left: 5px solid #DC2626;
                    padding: 15px 20px; border-radius: 8px; margin-bottom: 15px; border: 1px solid #FCA5A5;">
            <h4 style="margin: 0; color: #DC2626;">
                ⚠️ ALERTA DE VALIDEZ DEL TEST
            </h4>
            <p style="margin: 5px 0 0 0; color: #991B1B;">{analysis['validity_warning']}</p>
            <p style="margin: 5px 0 0 0; color: #7F1D1D; font-size: 0.9em;">
                El test detectó {len(validity_flags)} respuestas poco realistas. Considerar entrevista profunda adicional.
            </p>
        </div>
        """, unsafe_allow_html=True)
    
    # === BANNER DE RIESGO ===
    st.markdown(f"""
    <div style="background: {analysis['risk_color']}22; border-left: 5px solid {analysis['risk_color']};
                padding: 15px 20px; border-radius: 8px; margin-bottom: 15px;">
        <h3 style="margin: 0; color: {analysis['risk_color']};">
            {analysis['risk_emoji']} {analysis['risk_level']} — Puntaje: {analysis['risk_score']:.1f}/100
        </h3>
        <p style="margin: 5px 0 0 0; color: #374151;">{analysis['risk_desc']}</p>
        <p style="margin: 5px 0 0 0; color: #6B7280;">
            <b>Dimensión de menor riesgo:</b> {analysis['safest_dimension']} ({int(analysis['safest_score'])}/100) | 
            <b>Dimensión de mayor riesgo:</b> {analysis['riskiest_dimension']} ({int(analysis['riskiest_score'])}/100) |
            <b>Promedio:</b> {analysis['average_score']:.1f}/100
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # === DECISIÓN DE CONTRATACIÓN ===
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #1e293b 0%, #334155 100%); 
                padding: 20px; border-radius: 12px; margin-bottom: 20px; color: white;">
        <h4 style="margin: 0 0 10px 0;">📋 Decisión Recomendada de Contratación</h4>
        <h2 style="margin: 0; color: {analysis['risk_color']};">{analysis['hiring_decision']}</h2>
    </div>
    """, unsafe_allow_html=True)
    
    # === MÉTRICAS POR DIMENSIÓN ===
    st.markdown("### 📊 Puntajes por Dimensión de Riesgo")
    st.caption("⚠️ Recuerda: Puntajes más ALTOS = MENOR riesgo (Verde ✅), puntajes más BAJOS = MAYOR riesgo (Rojo 🚨)")
    
    # Crear 6 columnas para las 6 dimensiones
    cols = st.columns(3)
    for idx, dim in enumerate(ERI_DIMENSIONS):
        with cols[idx % 3]:
            score = normalized.get(dim, 0)
            if score >= ERI_RISK_THRESHOLDS["low_risk"]:
                nivel = "✅ Bajo Riesgo"
                delta_color = "normal"
            elif score >= ERI_RISK_THRESHOLDS["medium_risk"]:
                nivel = "⚠️ Moderado"
                delta_color = "off"
            else:
                nivel = "🚨 Alto Riesgo"
                delta_color = "inverse"
            
            st.metric(
                label=dim,
                value=f"{int(score)}/100",
                delta=nivel,
                delta_color=delta_color
            )
    
    st.markdown("---")
    
    # === GRÁFICOS ===
    col_radar, col_bars = st.columns(2)
    
    with col_radar:
        st.markdown("#### 🎯 Perfil de Riesgo (Radar)")
        radar_fig = create_eri_radar(normalized)
        st.pyplot(radar_fig)
    
    with col_bars:
        st.markdown("#### 📊 Puntajes por Dimensión")
        bar_fig = create_eri_bars(normalized)
        st.pyplot(bar_fig)
    
    st.markdown("---")
    
    # === ANÁLISIS POR DIMENSIÓN ===
    st.markdown("### 📋 Análisis Detallado por Dimensión")
    
    sorted_scores = sorted(normalized.items(), key=lambda x: x[1], reverse=False)  # Menor a mayor (más riesgo primero)
    
    for dim, score in sorted_scores:
        desc_info = ERI_DESCRIPTIONS[dim]
        
        # Determinar nivel de riesgo
        if score >= ERI_RISK_THRESHOLDS["low_risk"]:
            level = "✅ Bajo Riesgo"
            text = desc_info["low_risk"]
            color = "#10B981"
        elif score >= ERI_RISK_THRESHOLDS["medium_risk"]:
            level = "⚠️ Riesgo Moderado"
            text = desc_info["medium_risk"]
            color = "#F59E0B"
        else:
            level = "🚨 Alto Riesgo"
            text = desc_info["high_risk"]
            color = "#EF4444"
        
        st.markdown(f"""
        <div style="background: {color}15; border-left: 3px solid {color}; 
                    padding: 12px; border-radius: 6px; margin-bottom: 10px;">
            <b style="color: {color};">{desc_info['title']}</b> — {level} ({int(score)}/100)
            <br><span style="color: #374151;">{text}</span>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # === ASPECTOS POSITIVOS ===
    if analysis.get('fortalezas'):
        st.markdown("### 💚 Aspectos Positivos (Bajo Riesgo)")
        for f in analysis['fortalezas']:
            # Limpiar markdown
            f_clean = f.replace("**", "")
            st.markdown(f"- ✅ {f_clean}")
        st.markdown("")
    
    # === SEÑALES DE ALERTA ===
    if analysis.get('alertas'):
        st.markdown("### 🚨 Señales de Alerta y Factores de Riesgo")
        for a in analysis['alertas']:
            # Limpiar markdown
            a_clean = a.replace("**", "").replace("⚠️ ", "").replace("🚨 ", "")
            st.markdown(f"- 🔴 {a_clean}")
        st.markdown("")
    
    # === RECOMENDACIONES DE CONTRATACIÓN ===
    if analysis.get('recomendaciones'):
        st.markdown("### 💼 Recomendaciones de Contratación")
        for r in analysis['recomendaciones']:
            # Limpiar markdown (pero mantener bullets internos)
            r_clean = r.replace("**", "")
            st.markdown(f"{r_clean}")
        st.markdown("")
    
    # === DETALLES DE VALIDEZ ===
    if validity_flags and len(validity_flags) > 0:
        with st.expander(f"⚠️ Ver Detalles de Validez del Test ({len(validity_flags)} respuestas sospechosas)"):
            st.markdown(f"""
            Se detectaron **{len(validity_flags)}** respuestas poco realistas en preguntas de validez.
            
            Esto puede indicar:
            - El candidato está tratando de presentarse de forma irrealmente perfecta
            - Falta de sinceridad en las respuestas
            - No comprendió las instrucciones
            
            **Recomendación:** Explorar estos aspectos en entrevista personal.
            """)
            
            st.markdown("**Ejemplos de respuestas sospechosas:**")
            for flag in validity_flags[:10]:  # Mostrar máximo 10
                st.markdown(f"- {flag}")
            if len(validity_flags) > 10:
                st.caption(f"... y {len(validity_flags) - 10} respuestas más.")
    
    st.markdown("---")
    
    # === DESCARGA DE REPORTES ===
    st.markdown("### 📥 Descargar Reportes")
    
    session_id = session if isinstance(session, str) else session.get("id")
    completed_at = session.get("completed_at") if isinstance(session, dict) else None
    
    # Generar PDF
    pdf_buffer = generate_eri_pdf(candidate, raw, normalized, radar_fig, session_id, completed_at, analysis, validity_score, validity_flags)
    
    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            "📑 Descargar PDF Completo",
            data=pdf_buffer.getvalue(),
            file_name=f"eri_{candidate['cedula']}_{session_id}.pdf",
            mime="application/pdf",
            key=f"pdf_eri_{session_id}"
        )
    with col2:
        st.download_button(
            "📄 Descargar JSON",
            data=json.dumps(results, indent=2, ensure_ascii=False),
            file_name=f"eri_{candidate['cedula']}_{session_id}.json",
            mime="application/json",
            key=f"json_eri_{session_id}"
        )


def page_talent_map_test():
    """
    Página del test Talent Map (Mapeo de Competencias) - 80 preguntas con escala Likert 1-5.
    """
    session = st.session_state.get("test_session")
    candidate = st.session_state.get("candidate")
    
    if not session or not candidate:
        nav("candidate_login")
        st.rerun()
        return

    session = db.get_session_by_id(session["id"])
    if not session or session["status"] not in ("in_progress",):
        if session and session["status"] == "expired":
            st.error("⏰ El tiempo de esta evaluación ha expirado.")
            if st.button("Volver"):
                nav("candidate_select_test")
                st.rerun()
            return
        nav("candidate_select_test")
        st.rerun()
        return

    # Verificar tiempo restante
    remaining = db.check_session_time(session)
    if remaining == -1:
        st.error("⏰ El tiempo de esta evaluación ha expirado.")
        if st.button("Volver"):
            nav("candidate_select_test")
            st.rerun()
        return

    # Mostrar timer
    deadline_ts = db.get_session_deadline_timestamp(session)
    if deadline_ts:
        render_timer(deadline_ts, session["id"])

    st.markdown(f"### 🎯 Talent Map - Mapeo de Competencias y Talentos")
    st.caption(f"Candidato: {candidate['name']} | ID: {session['id']}")
    
    # Cargar preguntas si no están en session_state
    if "tm_questions" not in st.session_state:
        all_questions = load_talent_map_questions()
        # Mezclar preguntas de manera consistente por sesión
        rng = random.Random(session["id"])
        rng.shuffle(all_questions)
        st.session_state.tm_questions = all_questions
        db.update_session_questions(session["id"], all_questions)

    # Inicializar respuestas
    if "tm_responses" not in st.session_state:
        st.session_state.tm_responses = [None] * len(st.session_state.tm_questions)

    # Inicializar página
    if "tm_page" not in st.session_state:
        st.session_state.tm_page = 0

    questions = st.session_state.tm_questions
    total = len(questions)
    questions_per_page = 10  # 10 preguntas por página
    page = st.session_state.tm_page
    q_start = page * questions_per_page
    q_end = min(q_start + questions_per_page, total)

    # Barra de progreso
    progress = q_end / total
    st.progress(progress)
    st.markdown(f"**Preguntas {q_start + 1} - {q_end} de {total}**")

    # Instrucciones
    st.info("""
    **Instrucciones:** Responde con HONESTIDAD sobre cómo te comportas habitualmente en situaciones laborales.
    
    Escala:
    - **5** = Totalmente de acuerdo (Siempre me describe)
    - **4** = De acuerdo (Frecuentemente me describe)
    - **3** = Neutral / A veces (Depende de la situación)
    - **2** = En desacuerdo (Raramente me describe)
    - **1** = Totalmente en desacuerdo (Nunca me describe)
    
    💡 No hay respuestas correctas o incorrectas. Este test evalúa tu perfil de competencias.
    """)

    # Mostrar preguntas de la página actual
    all_answered = True
    
    for i in range(q_start, q_end):
        q = questions[i]
        q_text = q["question"]
        comp = q["competency"]
        
        # Crear tarjeta visual para cada pregunta con colores de Talent Map
        st.markdown(
            f"""
            <div style="background: linear-gradient(135deg, #1e40af 0%, #3b82f6 100%);
                        border-radius: 12px; padding: 20px; margin: 15px 0;
                        border-left: 4px solid {TALENT_MAP_COLORS.get(comp, '#3b82f6')};">
                <div style="margin-bottom: 8px;">
                    <span style="background: {TALENT_MAP_COLORS.get(comp, '#3b82f6')}; color: white; 
                                padding: 4px 12px; border-radius: 20px; 
                                font-size: 0.85em; font-weight: bold;">
                        Pregunta {i + 1} - {comp}
                    </span>
                </div>
                <p style="color: #e2e8f0; font-size: 1.1em; margin: 12px 0;">
                    {q_text}
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        
        # Radio buttons para la respuesta
        response_key = f"tm_q_{i}"
        
        # Inicializar desde respuestas guardadas
        if response_key not in st.session_state and st.session_state.tm_responses[i] is not None:
            st.session_state[response_key] = st.session_state.tm_responses[i]
        
        col1, col2 = st.columns([4, 1])
        with col1:
            response = st.radio(
                f"Respuesta {i + 1}",
                options=[1, 2, 3, 4, 5],
                format_func=lambda x: {
                    1: "1 - Totalmente en desacuerdo",
                    2: "2 - En desacuerdo",
                    3: "3 - Neutral",
                    4: "4 - De acuerdo",
                    5: "5 - Totalmente de acuerdo"
                }[x],
                key=response_key,
                horizontal=False,
                index=None if response_key not in st.session_state or st.session_state[response_key] is None else st.session_state[response_key] - 1
            )
        
        with col2:
            st.markdown("<br>" * 2, unsafe_allow_html=True)
            if response is not None:
                st.success("✅")
                st.session_state.tm_responses[i] = response
            else:
                st.warning("⚠️")
                all_answered = False

    # Navegación
    st.markdown("---")
    col_prev, col_space, col_next = st.columns([1, 4, 1])

    with col_prev:
        if page > 0:
            if st.button("⬅️ Anterior", key="tm_prev"):
                st.session_state.tm_page -= 1
                st.rerun()

    with col_next:
        is_last = q_end >= total
        btn_label = "✅ Finalizar Evaluación" if is_last else "Siguiente ➡️"
        if st.button(btn_label, key="tm_next", disabled=not all_answered):
            # Verificar tiempo nuevamente
            remaining = db.check_session_time(db.get_session_by_id(session["id"]))
            if remaining == -1:
                st.error("⏰ El tiempo ha expirado.")
                return

            if is_last:
                # Verificar que todas las preguntas estén respondidas
                if None in st.session_state.tm_responses:
                    st.warning("⚠️ Hay preguntas sin responder. Revisa las páginas anteriores.")
                else:
                    # Calcular resultados
                    responses = st.session_state.tm_responses
                    raw, normalized, percentages = calculate_talent_map_results(responses, questions)

                    # Guardar respuestas
                    answer_records = []
                    for i in range(total):
                        answer_records.append({
                            "question_index": i,
                            "question_text": questions[i]["question"],
                            "answer_value": responses[i],
                            "answer_b_value": None,  # No aplica para Talent Map
                        })
                    db.save_answers(session["id"], answer_records)

                    # Guardar resultados
                    results = {
                        "raw": raw,
                        "normalized": normalized,
                        "percentages": percentages
                    }
                    db.save_results(session["id"], results)
                    db.complete_test_session(session["id"])

                    # Limpiar session state
                    for key in ["tm_questions", "tm_responses", "tm_page", "test_session"]:
                        st.session_state.pop(key, None)

                    nav("candidate_done")
                    st.rerun()
            else:
                st.session_state.tm_page += 1
                st.rerun()


def show_talent_map_results_admin(results, candidate, session):
    """
    Muestra los resultados del Talent Map en el panel de administración.
    
    Args:
        results: Dict con raw, normalized, percentages
        candidate: Dict con información del candidato
        session: Dict con información de la sesión o str con session_id
    """
    raw = results.get("raw", {})
    normalized = results.get("normalized", {})
    percentages = results.get("percentages", {})
    
    # Selector de perfil de puesto para comparación
    st.markdown("### 🎯 Comparación con Perfil de Puesto")
    
    job_profile_name = st.selectbox(
        "Selecciona un perfil de puesto para comparar competencias:",
        options=["(Sin comparación)"] + list(TALENT_MAP_JOB_PROFILES.keys()),
        key="tm_job_profile_selector"
    )
    
    # Análisis de competencias con o sin match
    if job_profile_name and job_profile_name != "(Sin comparación)":
        analysis = analyze_talent_map_match(normalized, job_profile_name)
    else:
        analysis = analyze_talent_map_match(normalized, None)
    
    # === BANNER DE RESULTADO GENERAL ===
    avg_color = "#10B981" if analysis['average_score'] >= 75 else ("#F59E0B" if analysis['average_score'] >= 50 else "#EF4444")
    
    st.markdown(f"""
    <div style="background: {avg_color}22; border-left: 5px solid {avg_color};
                padding: 15px 20px; border-radius: 8px; margin-bottom: 15px;">
        <h3 style="margin: 0; color: {avg_color};">
            🎯 Perfil de Competencias — Promedio: {analysis['average_score']:.1f}/100
        </h3>
        <p style="margin: 5px 0 0 0; color: #374151;">
            <b>Competencia más fuerte:</b> {analysis['strongest_competency']} ({int(analysis['strongest_score'])}/100) | 
            <b>Área de mayor desarrollo:</b> {analysis['weakest_competency']} ({int(analysis['weakest_score'])}/100)
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # === ANÁLISIS DE MATCH (si aplica) ===
    if analysis.get('match_analysis'):
        match = analysis['match_analysis']
        match_color = match['match_color']
        
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #1e40af 0%, #3b82f6 100%); 
                    padding: 20px; border-radius: 12px; margin-bottom: 20px; color: white;">
            <h4 style="margin: 0 0 5px 0;">📊 Match con {match['job_emoji']} {match['job_profile']}</h4>
            <h2 style="margin: 0; color: {match_color};">{match['match_label']}: {match['match_percentage']:.1f}%</h2>
            <p style="margin: 8px 0 0 0; opacity: 0.9;">{match['match_desc']}</p>
            <p style="margin: 8px 0 0 0; font-size: 0.9em; opacity: 0.85;">{match['job_description']}</p>
        </div>
        """, unsafe_allow_html=True)
    
    # === MÉTRICAS POR COMPETENCIA ===
    st.markdown("### 📊 Puntajes por Competencia")
    
    # Crear columnas para las 8 competencias
    cols = st.columns(4)
    for idx, comp in enumerate(TALENT_MAP_COMPETENCIES):
        with cols[idx % 4]:
            score = normalized.get(comp, 0)
            if score >= 75:
                nivel = "🌟 Alto"
                delta_color = "normal"
            elif score >= 50:
                nivel = "👍 Medio"
                delta_color = "off"
            else:
                nivel = "📈 Desarrollo"
                delta_color = "inverse"
            
            st.metric(
                label=comp,
                value=f"{int(score)}/100",
                delta=nivel,
                delta_color=delta_color
            )
    
    st.markdown("---")
    
    # === GRÁFICOS ===
    # Si hay perfil de puesto seleccionado, crear gráficos con comparación
    if job_profile_name and job_profile_name != "(Sin comparación)":
        job_profile_scores = TALENT_MAP_JOB_PROFILES[job_profile_name]["competencias"]
        
        col_radar = st.container()
        with col_radar:
            st.markdown("#### 🎯 Perfil de Competencias (Candidato vs. Perfil Requerido)")
            radar_fig = create_talent_map_radar(normalized, job_profile_scores)
            st.pyplot(radar_fig)
        
        st.markdown("---")
        
        col_bars = st.container()
        with col_bars:
            st.markdown("#### 📊 Comparación de Competencias")
            bar_fig = create_talent_map_bars(normalized, job_profile_scores)
            st.pyplot(bar_fig)
        
        st.markdown("---")
        
        col_comparison = st.container()
        with col_comparison:
            st.markdown("#### 📈 Análisis de Brechas de Competencia")
            comparison_fig = create_talent_map_comparison(normalized, job_profile_name, job_profile_scores)
            st.pyplot(comparison_fig)
    else:
        col_radar, col_bars = st.columns(2)
        
        with col_radar:
            st.markdown("#### 🎯 Perfil de Competencias (Radar)")
            radar_fig = create_talent_map_radar(normalized)
            st.pyplot(radar_fig)
        
        with col_bars:
            st.markdown("#### 📊 Puntajes por Competencia")
            bar_fig = create_talent_map_bars(normalized)
            st.pyplot(bar_fig)
        
        comparison_fig = None
    
    st.markdown("---")
    
    # === ANÁLISIS POR COMPETENCIA ===
    st.markdown("### 📋 Análisis Detallado por Competencia")
    
    sorted_scores = sorted(normalized.items(), key=lambda x: x[1], reverse=True)
    
    for comp, score in sorted_scores:
        desc_info = TALENT_MAP_DESCRIPTIONS[comp]
        
        # Determinar nivel
        if score >= 75:
            level = "🌟 Alto"
            text = desc_info["high"]
            color = "#10B981"
        elif score >= 50:
            level = "👍 Medio"
            text = desc_info["medium"]
            color = "#F59E0B"
        else:
            level = "📈 En Desarrollo"
            text = desc_info["low"]
            color = "#EF4444"
        
        st.markdown(f"""
        <div style="background: {color}15; border-left: 3px solid {color}; 
                    padding: 12px; border-radius: 6px; margin-bottom: 10px;">
            <b style="color: {color};">{desc_info['title']}</b> — {level} ({int(score)}/100)
            <br><span style="color: #374151;">{text}</span>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # === FORTALEZAS ===
    if analysis.get('fortalezas'):
        st.markdown("### 💚 Fortalezas Clave")
        for f in analysis['fortalezas']:
            # Limpiar markdown
            f_clean = f.replace("**", "")
            st.markdown(f"- ✅ {f_clean}")
        st.markdown("")
    
    # === ÁREAS DE DESARROLLO ===
    if analysis.get('areas_desarrollo'):
        st.markdown("### 📈 Áreas de Desarrollo")
        for a in analysis['areas_desarrollo']:
            # Limpiar markdown
            a_clean = a.replace("**", "")
            st.markdown(f"- 🔵 {a_clean}")
        st.markdown("")
    
    # === GAPS Y STRENGTHS DE MATCH (si aplica) ===
    if analysis.get('match_analysis'):
        match = analysis['match_analysis']
        
        # Strengths del match
        if match.get('match_strengths'):
            st.markdown("### ✅ Competencias que Exceden el Perfil")
            for s in match['match_strengths']:
                s_clean = s.replace("**", "")
                st.markdown(f"- ✨ {s_clean}")
            st.markdown("")
        
        # Gaps del match
        if match.get('match_gaps'):
            st.markdown("### ⚠️ Brechas a Cerrar")
            for g in match['match_gaps']:
                g_clean = g.replace("**", "")
                st.markdown(f"- 📊 {g_clean}")
            st.markdown("")
    
    # === RECOMENDACIONES ===
    if analysis.get('recomendaciones'):
        with st.expander("💼 Ver Recomendaciones y Plan de Desarrollo"):
            for r in analysis['recomendaciones']:
                # Limpiar markdown (pero mantener bullets internos)
                r_clean = r.replace("**", "")
                st.markdown(f"{r_clean}")
    
    st.markdown("---")
    
    # === DESCARGA DE REPORTES ===
    st.markdown("### 📥 Descargar Reportes")
    
    session_id = session if isinstance(session, str) else session.get("id")
    completed_at = session.get("completed_at") if isinstance(session, dict) else None
    
    # Generar PDF (con o sin comparación de perfil)
    if job_profile_name and job_profile_name != "(Sin comparación)":
        pdf_buffer = generate_talent_map_pdf(
            candidate, raw, normalized, radar_fig, session_id, 
            completed_at, analysis, job_profile_name, comparison_fig
        )
    else:
        pdf_buffer = generate_talent_map_pdf(
            candidate, raw, normalized, radar_fig, session_id, 
            completed_at, analysis
        )
    
    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            label="📄 Descargar Reporte PDF Completo",
            data=pdf_buffer,
            file_name=f"talent_map_{candidate['cedula']}_{session_id}.pdf",
            mime="application/pdf",
            key=f"pdf_tm_{session_id}"
        )
    
    with col2:
        st.download_button(
            label="📊 Descargar Datos JSON",
            data=json.dumps(results, indent=2, ensure_ascii=False),
            file_name=f"talent_map_{candidate['cedula']}_{session_id}.json",
            mime="application/json",
            key=f"json_tm_{session_id}"
        )


def page_desempeno_eval():
    """Página de evaluación de desempeño (completada por el administra

dor)."""
    session_id = st.session_state.get("desempeno_session_id")
    
    if not session_id:
        st.error("No se encontró una sesión de evaluación activa.")
        if st.button("Volver al Dashboard"):
            nav("admin_dashboard")
            st.rerun()
        return
    
    session = db.get_session_by_id(session_id)
    if not session:
        st.error("Sesión no válida.")
        return
    
    candidate = db.get_candidate_by_cedula(
        db.get_connection().execute("SELECT cedula FROM candidates WHERE id = ?", (session["candidate_id"],)).fetchone()["cedula"]
    )
    
    admin = st.session_state.get("admin")
    evaluador = st.session_state.get("evaluador")
    if evaluador:
        evaluador_nombre = evaluador.get("name", "Evaluador")
    elif admin:
        evaluador_nombre = admin.get("name", "Administrador")
    else:
        evaluador_nombre = session.get("evaluador_nombre") or "Evaluador"
    
    st.markdown(f"## 📊 Evaluación de Desempeño")
    st.markdown(f"**Colaborador:** {candidate['name']} (Cédula: {candidate['cedula']})")
    st.markdown(f"**Cargo:** {candidate.get('position', 'N/A')}")
    st.markdown(f"**Evaluador:** {evaluador_nombre}")
    st.markdown("---")
    
    # Formulario de evaluación
    with st.form("evaluacion_desempeno_form"):
        st.markdown("### 📝 SECCIÓN 1: Evaluación de Rendimiento")
        st.markdown("Califique los siguientes 6 objetivos con una escala del 1 al 5:")
        st.markdown("**5** = Sobresaliente | **4** = Supera | **3** = Cumple | **2** = Debajo | **1** = Insatisfactorio")
        
        rendimiento_scores = {}
        
        for obj in DESEMPENO_OBJETIVOS:
            st.markdown(f"**{obj['titulo']}**")
            st.caption(obj['descripcion'])
            
            rendimiento_scores[obj['id']] = st.select_slider(
                f"Calificación Objetivo {obj['id']}",
                options=[1, 2, 3, 4, 5],
                value=3,
                format_func=lambda x: DESEMPENO_ESCALA_RENDIMIENTO[x]['label'],
                key=f"rend_{obj['id']}",
                label_visibility="collapsed"
            )
            st.markdown("---")
        
        st.markdown("### 🎯 SECCIÓN 2: Evaluación de Potencial")
        st.markdown("Seleccione el nivel que mejor describe al colaborador en cada dimensión (0-3):")
        
        potencial_scores = {}
        
        for dim in DESEMPENO_DIMENSIONES:
            st.markdown(f"**{dim['nombre']}**")
            st.caption(dim['descripcion'])
            
            opciones_texto = [f"Nivel {nivel}: {descripcion[:80]}..." for nivel, descripcion in dim['niveles'].items()]
            nivel_seleccionado = st.radio(
                f"Nivel para {dim['nombre']}",
                options=[3, 2, 1, 0],
                format_func=lambda x: f"Nivel {x}",
                key=f"pot_{dim['id']}",
                horizontal=True,
                label_visibility="collapsed"
            )
            
            potencial_scores[dim['id']] = nivel_seleccionado
            
            # Mostrar descripción del nivel seleccionado
            with st.expander("📄 Ver descripción completa del nivel seleccionado"):
                st.info(dim['niveles'][nivel_seleccionado])
            
            st.markdown("---")
        
        st.markdown("### 💡 SECCIÓN 3: Iniciativas de Mejora (Opcional)")
        st.markdown("Si el desempeño lo requiere, defina hasta 3 iniciativas de mejora:")
        
        iniciativa_1 = st.text_area("Iniciativa 1", placeholder="Descripción de la primera iniciativa...", key="init_1")
        iniciativa_2 = st.text_area("Iniciativa 2", placeholder="Descripción de la segunda iniciativa...", key="init_2")
        iniciativa_3 = st.text_area("Iniciativa 3", placeholder="Descripción de la tercera iniciativa...", key="init_3")
        
        iniciativas = [ini for ini in [iniciativa_1, iniciativa_2, iniciativa_3] if ini and ini.strip()]
        
        submitted = st.form_submit_button("✅ Completar Evaluación y Calcular Resultados", type="primary")
        
        if submitted:
            # Calcular resultados
            analysis = calculate_desempeno_results(rendimiento_scores, potencial_scores, iniciativas)
            
            # Guardar resultados en BD
            results_data = {
                "rendimiento_scores": rendimiento_scores,
                "potencial_scores": potencial_scores,
                "iniciativas": iniciativas,
                "analysis": analysis,
                "evaluador": evaluador_nombre
            }
            
            db.save_results(session_id, results_data)
            db.complete_test_session(session_id)
            
            st.success("✅ Evaluación completada y guardada exitosamente.")
            st.balloons()
            
            # Limpiar session_id y mostrar resultados
            del st.session_state["desempeno_session_id"]
            if evaluador:
                nav("evaluador_dashboard")
            else:
                nav("admin_dashboard")
            st.rerun()
    
    if st.button("❌ Cancelar Evaluación"):
        if "desempeno_session_id" in st.session_state:
            del st.session_state["desempeno_session_id"]
        if evaluador:
            nav("evaluador_dashboard")
        else:
            nav("admin_dashboard")
        st.rerun()


def show_desempeno_results_admin(results, candidate, session):
    """Muestra resultados de Evaluación de Desempeño en el panel de administración."""
    
    rendimiento_scores = results.get("rendimiento_scores", {})
    potencial_scores = results.get("potencial_scores", {})
    iniciativas = results.get("iniciativas", [])
    analysis = results.get("analysis", {})
    evaluador = results.get("evaluador", "N/A")
    
    # Convertir session a session_id si es necesario
    session_id = session["id"] if isinstance(session, dict) else session
    
    # Banner de clasificación
    if analysis.get("clasificacion"):
        clasif = analysis["clasificacion"]
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, {clasif['color']}22 0%, {clasif['color']}44 100%);
                    border-left: 6px solid {clasif['color']}; padding: 20px; border-radius: 12px; margin-bottom: 20px;">
            <h2 style="margin: 0; color: {clasif['color']};">{clasif['label']}</h2>
            <p style="margin: 8px 0 0 0; font-size: 15px; color: #374151;">{clasif['descripcion']}</p>
            <p style="margin: 12px 0 0 0; font-size: 14px; color: #6B7280;">
                <b>Evaluador:</b> {evaluador} | <b>Puntaje Global:</b> {analysis.get('puntaje_global', 0):.2f}/5.00
            </p>
        </div>
        """, unsafe_allow_html=True)
    
    # Métricas principales
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric(
            "🎯 Promedio Rendimiento",
            f"{analysis.get('promedio_rendimiento', 0):.2f}/5.00",
            help="Promedio de los 6 objetivos de rendimiento"
        )
    
    with col2:
        st.metric(
            "⭐ Promedio Potencial",
            f"{analysis.get('promedio_potencial', 0):.2f}/3.00",
            help="Promedio de las 5 dimensiones de potencial"
        )
    
    with col3:
        st.metric(
            "📊 Puntaje Global",
            f"{analysis.get('puntaje_global', 0):.2f}/5.00",
            help="Puntaje ponderado: 60% Rendimiento + 40% Potencial"
        )
    
    st.markdown("---")
    
    # Gráficos
    st.markdown("### 📈 Visualización de Resultados")
    
    col_left, col_right = st.columns(2)
    
    with col_left:
        st.markdown("#### Evaluación de Rendimiento")
        bars_fig = create_desempeno_bars(rendimiento_scores)
        st.pyplot(bars_fig)
        plt.close(bars_fig)
    
    with col_right:
        st.markdown("#### Evaluación de Potencial")
        radar_fig = create_desempeno_radar(potencial_scores)
        st.pyplot(radar_fig)
        plt.close(radar_fig)
    
    st.markdown("---")
    
    # Detalles por secciones
    tab1, tab2, tab3, tab4 = st.tabs(["📝 Rendimiento", "🎯 Potencial", "💡 Análisis", "🎯 Iniciativas"])
    
    with tab1:
        st.markdown("#### Desglose por Objetivo de Rendimiento")
        for obj_id, score in rendimiento_scores.items():
            objetivo = DESEMPENO_OBJETIVOS[obj_id - 1]
            nivel = DESEMPENO_ESCALA_RENDIMIENTO[score]
            
            col_obj1, col_obj2 = st.columns([3, 1])
            with col_obj1:
                st.markdown(f"**{objetivo['titulo']}**")
                st.caption(objetivo['descripcion'])
            with col_obj2:
                st.markdown(f"<div style='background:{nivel['color']}22; padding:12px; border-radius:8px; text-align:center;'>"
                           f"<b>{score:.1f}/5.0</b><br><span style='font-size:12px;'>{nivel['label']}</span></div>",
                           unsafe_allow_html=True)
            st.markdown("---")
    
    with tab2:
        st.markdown("#### Desglose por Dimensión de Potencial")
        for dim_id, score in potencial_scores.items():
            dimension = DESEMPENO_DIMENSIONES[dim_id - 1]
            
            col_dim1, col_dim2 = st.columns([3, 1])
            with col_dim1:
                st.markdown(f"**{dimension['nombre']}**")
                st.caption(dimension['descripcion'])
                with st.expander("📄 Ver descripción del nivel asignado"):
                    st.info(dimension['niveles'][score])
            with col_dim2:
                color = DESEMPENO_COLORES_DIMENSIONES.get(dimension['nombre'], "#6B7280")
                st.markdown(f"<div style='background:{color}22; padding:12px; border-radius:8px; text-align:center;'>"
                           f"<b>Nivel {score}/3</b></div>",
                           unsafe_allow_html=True)
            st.markdown("---")
    
    with tab3:
        col_for, col_mej = st.columns(2)
        
        with col_for:
            st.markdown("#### ✅ Fortalezas")
            
            if analysis.get("fortalezas_rendimiento"):
                st.markdown("**Rendimiento:**")
                for item in analysis["fortalezas_rendimiento"]:
                    st.success(f"**{item['titulo']}** - {item['score']:.1f}/5.0 ({item['label']})")
            
            if analysis.get("fortalezas_potencial"):
                st.markdown("**Potencial:**")
                for item in analysis["fortalezas_potencial"]:
                    st.success(f"**{item['nombre']}** - {item['nivel']}")
        
        with col_mej:
            st.markdown("#### ⚠️ Áreas de Mejora")
            
            if analysis.get("areas_mejora_rendimiento"):
                st.markdown("**Rendimiento:**")
                for item in analysis["areas_mejora_rendimiento"]:
                    st.warning(f"**{item['titulo']}** - {item['score']:.1f}/5.0 ({item['label']})")
            
            if analysis.get("areas_desarrollo_potencial"):
                st.markdown("**Potencial:**")
                for item in analysis["areas_desarrollo_potencial"]:
                    st.warning(f"**{item['nombre']}** - {item['nivel']}")
        
        st.markdown("---")
        st.markdown("#### 💡 Recomendaciones")
        if analysis.get("recomendaciones"):
            for recom in analysis["recomendaciones"]:
                st.info(f"• {recom}")
    
    with tab4:
        st.markdown("#### 🎯 Iniciativas de Mejora Definidas")
        
        if iniciativas and len(iniciativas) > 0:
            for i, iniciativa in enumerate(iniciativas, 1):
                st.markdown(f"**Iniciativa {i}:**")
                st.info(iniciativa)
        else:
            if analysis.get("requiere_iniciativas"):
                st.warning("⚠️ Esta evaluación requiere establecer iniciativas de mejora, pero no se definieron.")
            else:
                st.success("✅ El desempeño es satisfactorio. No se requieren iniciativas de mejora.")
    
    st.markdown("---")
    
    # Descargar PDF y JSON
    st.markdown("### 📥 Descargar Resultados")
    
    col1, col2 = st.columns(2)
    
    # Regenerar gráficos para PDF
    radar_fig_pdf = create_desempeno_radar(potencial_scores)
    bars_fig_pdf = create_desempeno_bars(rendimiento_scores)
    
    pdf_buffer = generate_desempeno_pdf(
        candidate=candidate,
        rendimiento_scores=rendimiento_scores,
        potencial_scores=potencial_scores,
        radar_fig=radar_fig_pdf,
        bars_fig=bars_fig_pdf,
        session_id=session_id,
        completed_at=session.get("completed_at") if isinstance(session, dict) else None,
        analysis=analysis,
        evaluador_nombre=evaluador,
        iniciativas=iniciativas
    )
    
    with col1:
        st.download_button(
            "📄 Descargar PDF",
            data=pdf_buffer,
            file_name=f"evaluacion_desempeno_{candidate['cedula']}_{session_id}.pdf",
            mime="application/pdf",
            key=f"pdf_desempeno_{session_id}"
        )
    
    with col2:
        st.download_button(
            "📄 Descargar JSON",
            data=json.dumps(results, indent=2, ensure_ascii=False),
            file_name=f"evaluacion_desempeno_{candidate['cedula']}_{session_id}.json",
            mime="application/json",
            key=f"json_desempeno_{session_id}"
        )


# -------------------------------------------------------------------------
# ADMIN: EVALUACIÓN DE DESEMPEÑO — LÍDERES (FO-GH-41)
# -------------------------------------------------------------------------
def page_desempeno_lider_eval():
    """Página de evaluación de desempeño para líderes (completada por el administrador)."""
    session_id = st.session_state.get("desempeno_lider_session_id")
    if not session_id:
        st.error("No se encontró una sesión de evaluación activa.")
        if st.button("Volver al Dashboard"):
            nav("admin_dashboard")
            st.rerun()
        return

    session = db.get_session_by_id(session_id)
    if not session:
        st.error("Sesión no válida.")
        return

    candidate = db.get_candidate_by_cedula(
        db.get_connection().execute("SELECT cedula FROM candidates WHERE id = ?", (session["candidate_id"],)).fetchone()["cedula"]
    )
    admin = st.session_state.get("admin")
    evaluador_nombre = admin.get("name", "Administrador") if admin else "Administrador"
    nivel_cargo = candidate.get("nivel_cargo", "ANALISTA") or "ANALISTA"

    st.markdown("## 📊 Evaluación de Desempeño — Líderes")
    st.markdown(f"**Colaborador:** {candidate['name']} (Cédula: {candidate['cedula']})")
    st.markdown(f"**Cargo:** {candidate.get('position', 'N/A')} | **Nivel:** {nivel_cargo}")
    st.markdown(f"**Evaluador:** {evaluador_nombre}")
    st.markdown("---")

    with st.form("evaluacion_desempeno_lider_form"):
        # ---- SECCIÓN 1: COMPETENCIAS ----
        st.markdown("### 🏆 SECCIÓN 1: Evaluación de Competencias Organizacionales")
        st.markdown("Seleccione el nivel alcanzado por el colaborador en cada competencia (1-6):")

        nivel_req_info = COMPETENCIAS_NIVEL_REQUERIDO.get(nivel_cargo.upper(), None)
        competencias_scores = {}

        for comp in COMPETENCIAS_ORGANIZACIONALES:
            req = nivel_req_info["niveles"][comp["id"] - 1] if nivel_req_info else None
            req_text = f" _(Requerido: Nivel {req})_" if req else ""
            st.markdown(f"**{comp['nombre']}**{req_text}")
            st.caption(comp["descripcion"])
            opciones = {n: f"Nivel {n}: {desc[:90]}..." for n, desc in comp["niveles"].items()}
            nivel_sel = st.radio(
                f"Nivel {comp['nombre']}",
                options=[1, 2, 3, 4, 5, 6],
                format_func=lambda x, c=comp: f"Nivel {x} — {c['niveles'][x][:80]}...",
                horizontal=True,
                key=f"comp_{comp['id']}",
                label_visibility="collapsed",
            )
            competencias_scores[comp["id"]] = nivel_sel
            st.markdown("---")

        # ---- SECCIÓN 2: RENDIMIENTO ----
        st.markdown("### 📝 SECCIÓN 2: Evaluación de Rendimiento")
        st.markdown("**5** = Sobresaliente | **4** = Supera | **3** = Cumple | **2** = Debajo | **1** = Insatisfactorio")

        rendimiento_scores = {}
        for obj in DESEMPENO_OBJETIVOS:
            st.markdown(f"**{obj['titulo']}**")
            st.caption(obj["descripcion"])
            rendimiento_scores[obj["id"]] = st.select_slider(
                f"Calificación Objetivo {obj['id']}",
                options=[1, 2, 3, 4, 5],
                value=3,
                format_func=lambda x: DESEMPENO_ESCALA_RENDIMIENTO[x]["label"],
                key=f"rend_lider_{obj['id']}",
                label_visibility="collapsed",
            )
            st.markdown("---")

        # ---- SECCIÓN 3: POTENCIAL ----
        st.markdown("### 🎯 SECCIÓN 3: Evaluación de Potencial (0-3)")
        potencial_scores = {}
        for dim in DESEMPENO_DIMENSIONES:
            st.markdown(f"**{dim['nombre']}**")
            st.caption(dim["descripcion"])
            nivel_sel = st.radio(
                f"Nivel {dim['nombre']}",
                options=[3, 2, 1, 0],
                format_func=lambda x, d=dim: f"Nivel {x}: {d['niveles'][x][:80]}...",
                key=f"pot_lider_{dim['id']}",
                label_visibility="collapsed",
            )
            potencial_scores[dim["id"]] = nivel_sel
            st.markdown("---")

        # ---- INICIATIVAS ----
        st.markdown("### 🚀 Iniciativas de Mejora")
        n_iniciativas = st.selectbox("Número de iniciativas", [0, 1, 2, 3], index=1, key="n_init_lider")
        iniciativas = []
        for i in range(n_iniciativas):
            ini = st.text_area(f"Iniciativa {i+1}", key=f"ini_lider_{i}", height=80)
            if ini.strip():
                iniciativas.append(ini.strip())

        submitted = st.form_submit_button("✅ Guardar Evaluación", use_container_width=True, type="primary")

        if submitted:
            results_calc = calculate_desempeno_lider_results(
                competencias_scores=competencias_scores,
                rendimiento_scores=rendimiento_scores,
                potencial_scores=potencial_scores,
                nivel_cargo=nivel_cargo,
                iniciativas=iniciativas,
            )
            results_to_save = {
                "test_type": "desempeno_lider",
                "evaluador": evaluador_nombre,
                "nivel_cargo": nivel_cargo,
                "competencias_scores": {str(k): v for k, v in competencias_scores.items()},
                "rendimiento_scores": {str(k): v for k, v in rendimiento_scores.items()},
                "potencial_scores": {str(k): v for k, v in potencial_scores.items()},
                "iniciativas": iniciativas,
                "analysis": results_calc,
            }
            db.save_results(session_id, results_to_save)
            db.complete_test_session(session_id)
            st.success("✅ Evaluación de desempeño (líderes) guardada exitosamente.")
            st.session_state.pop("desempeno_lider_session_id", None)
            nav("admin_dashboard")
            st.rerun()

    if st.button("❌ Cancelar"):
        st.session_state.pop("desempeno_lider_session_id", None)
        nav("admin_dashboard")
        st.rerun()


def generate_desempeno_lider_pdf(candidate, competencias_scores, rendimiento_scores, potencial_scores,
                                  session_id, completed_at=None, analysis=None, evaluador_nombre=None,
                                  nivel_cargo=None, iniciativas=None):
    """Genera PDF de Evaluación de Desempeño para Líderes."""
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50)
    styles = getSampleStyleSheet()
    try:
        styles.add(ParagraphStyle(name='DLTitle', parent=styles['Heading1'], fontSize=16,
                                  textColor=colors.HexColor("#1E40AF"), alignment=1, spaceAfter=12))
        styles.add(ParagraphStyle(name='DLSub', parent=styles['Heading2'], fontSize=12,
                                  textColor=colors.HexColor("#374151"), spaceAfter=8))
        styles.add(ParagraphStyle(name='DLSmall', parent=styles['Normal'], fontSize=9,
                                  textColor=colors.HexColor("#6B7280")))
        styles.add(ParagraphStyle(name='DLItem', parent=styles['Normal'], fontSize=10,
                                  leftIndent=16, spaceAfter=4))
    except Exception:
        pass
    DLTitle = styles.get('DLTitle', styles['Title'])
    DLSub = styles.get('DLSub', styles['Heading2'])
    DLSmall = styles.get('DLSmall', styles['Normal'])
    DLItem = styles.get('DLItem', styles['Normal'])

    story = []
    story.append(Spacer(1, 40))
    story.append(Paragraph("EVALUACIÓN DE DESEMPEÑO — LÍDERES", DLTitle))
    story.append(Spacer(1, 10))

    info_rows = [
        ["Colaborador:", candidate['name']],
        ["Cédula:", str(candidate['cedula'])],
        ["Cargo:", candidate.get('position', 'N/A')],
        ["Nivel de Cargo:", nivel_cargo or 'N/A'],
        ["Evaluador:", evaluador_nombre or 'N/A'],
        ["Fecha:", completed_at or 'N/A'],
        ["ID Sesión:", str(session_id)],
    ]
    it = Table(info_rows, colWidths=[130, 360])
    it.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor("#E5E7EB")),
        ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(it)
    story.append(Spacer(1, 14))

    if analysis and analysis.get("clasificacion"):
        clasif = analysis["clasificacion"]
        bn = Table([[clasif.get("label", "")]], colWidths=[450])
        bn.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(clasif.get("color", "#374151"))),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 13),
            ('TOPPADDING', (0, 0), (-1, -1), 10), ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        story.append(bn)
        story.append(Spacer(1, 6))
        if clasif.get("descripcion"):
            story.append(Paragraph(clasif["descripcion"], DLSmall))

    if analysis:
        story.append(Spacer(1, 10))
        pt_data = [
            ["Componente", "Puntaje", "Máximo"],
            ["Competencias Organizacionales", f"{analysis.get('promedio_competencias', 0):.2f}", "6.00"],
            ["Rendimiento (6 objetivos)", f"{analysis.get('promedio_rendimiento', 0):.2f}", "5.00"],
            ["Potencial (5 dimensiones)", f"{analysis.get('promedio_potencial', 0):.2f}", "3.00"],
            ["Puntaje Global Ponderado", f"{analysis.get('puntaje_global', 0):.2f}", "5.00"],
        ]
        pt = Table(pt_data, colWidths=[260, 100, 90])
        pt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#DBEAFE")),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(pt)

    story.append(PageBreak())

    # Competencias
    story.append(Paragraph("COMPETENCIAS ORGANIZACIONALES", DLSub))
    story.append(Spacer(1, 6))
    nivel_req_info = COMPETENCIAS_NIVEL_REQUERIDO.get((nivel_cargo or "").upper(), None)
    comp_data = [["Competencia", "Nivel", "Requerido", "Brecha"]]
    for comp in COMPETENCIAS_ORGANIZACIONALES:
        cid = comp["id"]
        score = competencias_scores.get(cid, 0)
        req = nivel_req_info["niveles"][cid - 1] if nivel_req_info else "-"
        brecha = (score - req) if isinstance(req, int) else "-"
        brecha_str = f"+{brecha}" if isinstance(brecha, int) and brecha > 0 else str(brecha)
        comp_data.append([comp["nombre"][:50], str(score), str(req), brecha_str])
    ct = Table(comp_data, colWidths=[240, 70, 80, 60])
    ct.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#374151")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'), ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
        ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(ct)
    story.append(Spacer(1, 12))

    # Rendimiento
    story.append(Paragraph("EVALUACIÓN DE RENDIMIENTO", DLSub))
    rend_data = [["Objetivo", "Puntaje", "Nivel"]]
    for obj_id, score in rendimiento_scores.items():
        objetivo = DESEMPENO_OBJETIVOS[obj_id - 1]
        nivel = DESEMPENO_ESCALA_RENDIMIENTO.get(score, {})
        rend_data.append([objetivo["titulo"][:60], f"{score}/5", nivel.get("label", "")])
    rt = Table(rend_data, colWidths=[290, 60, 100])
    rt.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
        ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(rt)
    story.append(Spacer(1, 12))

    # Potencial
    story.append(Paragraph("EVALUACIÓN DE POTENCIAL", DLSub))
    pot_data = [["Dimensión", "Nivel"]]
    for dim_id, score in potencial_scores.items():
        dimension = DESEMPENO_DIMENSIONES[dim_id - 1]
        pot_data.append([dimension["nombre"], f"Nivel {score}/3"])
    pot_t = Table(pot_data, colWidths=[370, 80])
    pot_t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
        ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(pot_t)

    if iniciativas:
        story.append(Spacer(1, 10))
        story.append(Paragraph("INICIATIVAS DE MEJORA", DLSub))
        for i, ini in enumerate(iniciativas, 1):
            story.append(Paragraph(f"{i}. {ini}", DLItem))

    if analysis and analysis.get("recomendaciones"):
        story.append(Spacer(1, 10))
        story.append(Paragraph("RECOMENDACIONES", DLSub))
        for recom in analysis["recomendaciones"]:
            story.append(Paragraph(f"• {recom}", DLItem))

    doc.build(story)
    buf.seek(0)
    return buf


def generate_periodo_prueba_pdf(candidate, actuaciones_scores, calificaciones_scores,
                                 session_id, completed_at=None, analysis=None, evaluador_nombre=None,
                                 aprobo=False, llamados_atencion=False, conocimiento_adecuado=True,
                                 observaciones=None):
    """Genera PDF de Evaluación de Período de Prueba."""
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50)
    styles = getSampleStyleSheet()
    try:
        styles.add(ParagraphStyle(name='PPTitle', parent=styles['Heading1'], fontSize=16,
                                  textColor=colors.HexColor("#1E40AF"), alignment=1, spaceAfter=12))
        styles.add(ParagraphStyle(name='PPSub', parent=styles['Heading2'], fontSize=12,
                                  textColor=colors.HexColor("#374151"), spaceAfter=8))
        styles.add(ParagraphStyle(name='PPItem', parent=styles['Normal'], fontSize=10,
                                  leftIndent=16, spaceAfter=4))
    except Exception:
        pass
    PPTitle = styles.get('PPTitle', styles['Title'])
    PPSub = styles.get('PPSub', styles['Heading2'])
    PPItem = styles.get('PPItem', styles['Normal'])

    story = []
    story.append(Spacer(1, 40))
    story.append(Paragraph("EVALUACIÓN PERÍODO DE PRUEBA", PPTitle))
    story.append(Spacer(1, 10))

    info_rows = [
        ["Colaborador:", candidate['name']],
        ["Cédula:", str(candidate['cedula'])],
        ["Cargo:", candidate.get('position', 'N/A')],
        ["Evaluador:", evaluador_nombre or 'N/A'],
        ["Fecha:", completed_at or 'N/A'],
        ["ID Sesión:", str(session_id)],
    ]
    it = Table(info_rows, colWidths=[130, 360])
    it.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor("#E5E7EB")),
        ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(it)
    story.append(Spacer(1, 14))

    aprobacion_color = "#10B981" if aprobo else "#EF4444"
    aprobacion_text = "APROBÓ EL PERÍODO DE PRUEBA" if aprobo else "NO APROBÓ EL PERÍODO DE PRUEBA"
    bn = Table([[aprobacion_text]], colWidths=[450])
    bn.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(aprobacion_color)),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 13),
        ('TOPPADDING', (0, 0), (-1, -1), 10), ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    story.append(bn)

    if analysis:
        story.append(Spacer(1, 10))
        clasif = analysis.get("clasificacion") or {}
        mt_data = [
            ["Promedio Actuaciones", f"{analysis.get('promedio_actuaciones', 0):.2f}/4.00",
             "Promedio Calificaciones", f"{analysis.get('promedio_calificaciones', 0):.2f}/5.00"],
            ["Promedio General", f"{analysis.get('promedio_general', 0):.2f}/4.00",
             "Clasificación", clasif.get("label", "N/A")],
            ["Llamados de atención", "SÍ" if llamados_atencion else "NO",
             "Conocimiento adecuado", "SÍ" if conocimiento_adecuado else "NO"],
        ]
        mt = Table(mt_data, colWidths=[140, 90, 140, 80])
        mt.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#F9FAFB")),
            ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(mt)

    story.append(PageBreak())

    # Actuaciones
    story.append(Paragraph("ACTUACIONES Y COMPORTAMIENTOS", PPSub))
    act_data = [["N°", "Actuación", "Calificación"]]
    for idx, actuacion in enumerate(PERIODO_PRUEBA_ACTUACIONES):
        score = actuaciones_scores.get(idx, 0)
        if score == 0:
            continue
        escala = PERIODO_PRUEBA_ESCALA_ACTUACIONES.get(score, {})
        act_data.append([str(idx + 1), actuacion[:75], escala.get("label", str(score))])
    at = Table(act_data, colWidths=[25, 330, 95])
    at.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#374151")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
        ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(at)
    story.append(Spacer(1, 12))

    # Calificaciones
    story.append(Paragraph("CALIFICACIONES ESPECÍFICAS", PPSub))
    cal_data = [["Criterio", "Calificación"]]
    for idx, cal in enumerate(PERIODO_PRUEBA_CALIFICACIONES):
        score = calificaciones_scores.get(idx, 0)
        if score == 0:
            continue
        escala = PERIODO_PRUEBA_ESCALA_CALIFICACIONES.get(score, {})
        cal_data.append([cal, escala.get("label", str(score))])
    calt = Table(cal_data, colWidths=[360, 90])
    calt.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#374151")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
        ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(calt)

    if observaciones:
        story.append(Spacer(1, 10))
        story.append(Paragraph("OBSERVACIONES", PPSub))
        story.append(Paragraph(str(observaciones), PPItem))

    if analysis and analysis.get("recomendaciones"):
        story.append(Spacer(1, 10))
        story.append(Paragraph("RECOMENDACIONES", PPSub))
        for recom in analysis["recomendaciones"]:
            story.append(Paragraph(f"• {recom}", PPItem))

    doc.build(story)
    buf.seek(0)
    return buf


def show_desempeno_lider_results_admin(results, candidate, session):
    """Muestra resultados de la evaluación de desempeño para líderes."""
    analysis = results.get("analysis", {})
    competencias_scores = {int(k): v for k, v in results.get("competencias_scores", {}).items()}
    rendimiento_scores = {int(k): v for k, v in results.get("rendimiento_scores", {}).items()}
    potencial_scores = {int(k): v for k, v in results.get("potencial_scores", {}).items()}
    iniciativas = results.get("iniciativas", [])
    evaluador = results.get("evaluador", "N/A")
    nivel_cargo = results.get("nivel_cargo", "N/A")
    session_id = session["id"] if isinstance(session, dict) else session

    clasif = analysis.get("clasificacion") or {}
    comp_clasif = analysis.get("clasificacion_comp") or {}

    if clasif:
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, {clasif.get('color','#6B7280')}22, {clasif.get('color','#6B7280')}44);
                    border-left: 6px solid {clasif.get('color','#6B7280')}; padding: 20px; border-radius: 12px; margin-bottom: 20px;">
            <h2 style="margin:0; color:{clasif.get('color','#111')};">{clasif.get('label','')}</h2>
            <p style="margin:8px 0 0 0; font-size:15px; color:#374151;">{clasif.get('descripcion','')}</p>
            <p style="margin:12px 0 0 0; font-size:14px; color:#6B7280;">
                <b>Evaluador:</b> {evaluador} | <b>Nivel Cargo:</b> {nivel_cargo} | <b>Puntaje Global:</b> {analysis.get('puntaje_global', 0):.2f}/5.00
            </p>
        </div>
        """, unsafe_allow_html=True)

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("🏆 Competencias", f"{analysis.get('promedio_competencias', 0):.2f}/6.00")
    col2.metric("🎯 Rendimiento", f"{analysis.get('promedio_rendimiento', 0):.2f}/5.00")
    col3.metric("⭐ Potencial", f"{analysis.get('promedio_potencial', 0):.2f}/3.00")
    col4.metric("📊 Global", f"{analysis.get('puntaje_global', 0):.2f}/5.00")

    st.markdown("---")

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "🏆 Competencias", "📝 Rendimiento", "🎯 Potencial", "💡 Análisis", "🚀 Iniciativas"
    ])

    with tab1:
        st.markdown(f"#### Promedio de Competencias: **{analysis.get('promedio_competencias', 0):.2f}/6.00**")
        if comp_clasif:
            st.markdown(f"**Clasificación:** {comp_clasif.get('label','')}")
        nivel_req_info = COMPETENCIAS_NIVEL_REQUERIDO.get(nivel_cargo.upper(), None)
        for comp in COMPETENCIAS_ORGANIZACIONALES:
            cid = comp["id"]
            score = competencias_scores.get(cid, 0)
            req = nivel_req_info["niveles"][cid - 1] if nivel_req_info else None
            brecha = score - req if req is not None else None
            color = "#10B981" if (brecha is None or brecha >= 0) else "#EF4444"
            col_a, col_b = st.columns([3, 1])
            with col_a:
                st.markdown(f"**{comp['nombre']}**")
                if req:
                    st.caption(f"Requerido: Nivel {req} | Asignado: Nivel {score}")
            with col_b:
                brecha_txt = f"(+{brecha})" if brecha and brecha > 0 else (f"({brecha})" if brecha else "")
                st.markdown(f"<div style='background:{color}22; padding:10px; border-radius:8px; text-align:center;'>"
                           f"<b>Nivel {score}</b><br><span style='color:{color}; font-size:12px;'>{brecha_txt}</span></div>",
                           unsafe_allow_html=True)
            with st.expander("Ver descripción del nivel asignado"):
                st.info(comp["niveles"].get(score, "N/A"))
            st.markdown("---")

    with tab2:
        for obj_id, score in rendimiento_scores.items():
            objetivo = DESEMPENO_OBJETIVOS[obj_id - 1]
            nivel = DESEMPENO_ESCALA_RENDIMIENTO[score]
            c1, c2 = st.columns([3, 1])
            with c1:
                st.markdown(f"**{objetivo['titulo']}**")
                st.caption(objetivo["descripcion"])
            with c2:
                st.markdown(f"<div style='background:{nivel['color']}22; padding:10px; border-radius:8px; text-align:center;'>"
                           f"<b>{score}/5</b><br><span style='font-size:12px;'>{nivel['label']}</span></div>",
                           unsafe_allow_html=True)
            st.markdown("---")

    with tab3:
        for dim_id, score in potencial_scores.items():
            dimension = DESEMPENO_DIMENSIONES[dim_id - 1]
            c1, c2 = st.columns([3, 1])
            with c1:
                st.markdown(f"**{dimension['nombre']}**")
                st.caption(dimension["descripcion"])
                with st.expander("Ver descripción"):
                    st.info(dimension["niveles"][score])
            with c2:
                color = DESEMPENO_COLORES_DIMENSIONES.get(dimension["nombre"], "#6B7280")
                st.markdown(f"<div style='background:{color}22; padding:10px; border-radius:8px; text-align:center;'>"
                           f"<b>Nivel {score}/3</b></div>", unsafe_allow_html=True)
            st.markdown("---")

    with tab4:
        col_for, col_mej = st.columns(2)
        with col_for:
            st.markdown("#### ✅ Fortalezas")
            for item in analysis.get("fortalezas_competencias", []):
                st.success(f"🏆 **{item['nombre']}** — Nivel {item['score']}")
            for item in analysis.get("fortalezas_rendimiento", []):
                st.success(f"🎯 **{item['titulo']}** — {item['score']}/5 ({item['label']})")
            for item in analysis.get("fortalezas_potencial", []):
                st.success(f"⭐ **{item['nombre']}** — {item['nivel']}")
        with col_mej:
            st.markdown("#### ⚠️ Áreas de Mejora")
            for item in analysis.get("brechas_competencias", []):
                st.warning(f"🏆 **{item['nombre']}** — Nivel {item['score']} (req. {item['requerido']}, brecha {item['brecha']})")
            for item in analysis.get("areas_mejora_rendimiento", []):
                st.warning(f"🎯 **{item['titulo']}** — {item['score']}/5 ({item['label']})")
            for item in analysis.get("areas_desarrollo_potencial", []):
                st.warning(f"⭐ **{item['nombre']}** — {item['nivel']}")
        st.markdown("---")
        st.markdown("#### 💡 Recomendaciones")
        for recom in analysis.get("recomendaciones", []):
            st.info(f"• {recom}")

    with tab5:
        if iniciativas:
            for i, ini in enumerate(iniciativas, 1):
                st.markdown(f"**Iniciativa {i}:** {ini}")
        else:
            st.info("No se definieron iniciativas.")

    st.markdown("---")
    col_dl1, col_dl2 = st.columns(2)
    with col_dl1:
        st.download_button(
            "📄 Descargar JSON",
            data=json.dumps(results, indent=2, ensure_ascii=False),
            file_name=f"desempeno_lider_{candidate['cedula']}_{session_id}.json",
            mime="application/json",
            key=f"json_dl_{session_id}",
        )
    with col_dl2:
        try:
            _pdf_dl = generate_desempeno_lider_pdf(
                candidate=candidate,
                competencias_scores=competencias_scores,
                rendimiento_scores=rendimiento_scores,
                potencial_scores=potencial_scores,
                session_id=session_id,
                completed_at=session.get("completed_at") if isinstance(session, dict) else None,
                analysis=analysis,
                evaluador_nombre=evaluador,
                nivel_cargo=nivel_cargo,
                iniciativas=iniciativas,
            )
            st.download_button(
                "📑 Descargar PDF",
                data=_pdf_dl,
                file_name=f"desempeno_lider_{candidate['cedula']}_{session_id}.pdf",
                mime="application/pdf",
                key=f"pdf_dl_{session_id}",
            )
        except Exception as _pdf_err:
            st.warning(f"No se pudo generar el PDF: {_pdf_err}")


# -------------------------------------------------------------------------
# ADMIN: EVALUACIÓN PERÍODO DE PRUEBA (FO-GH-46)
# -------------------------------------------------------------------------
def page_periodo_prueba_eval():
    """Página de evaluación de período de prueba (completada por el administrador/evaluador)."""
    session_id = st.session_state.get("periodo_prueba_session_id")
    if not session_id:
        st.error("No se encontró una sesión de evaluación activa.")
        if st.button("Volver al Dashboard"):
            nav("admin_dashboard")
            st.rerun()
        return

    session = db.get_session_by_id(session_id)
    if not session:
        st.error("Sesión no válida.")
        return

    candidate = db.get_candidate_by_cedula(
        db.get_connection().execute("SELECT cedula FROM candidates WHERE id = ?", (session["candidate_id"],)).fetchone()["cedula"]
    )
    admin = st.session_state.get("admin")
    evaluador_nombre = admin.get("name", "Administrador") if admin else "Administrador"

    st.markdown("## 📋 Evaluación Período de Prueba")
    st.markdown(f"**Trabajador:** {candidate['name']} (Cédula: {candidate['cedula']})")
    st.markdown(f"**Cargo:** {candidate.get('position', 'N/A')} | **Área:** {candidate.get('regional', 'N/A')}")
    st.markdown(f"**Evaluador:** {evaluador_nombre}")
    st.info("Marque la frecuencia con la que observa cada comportamiento durante el desempeño laboral.")
    st.markdown("---")

    with st.form("evaluacion_periodo_prueba_form"):
        # ---- SECCIÓN 1: ACTUACIONES ----
        st.markdown("### 📝 Sección 1: Actuaciones y Comportamientos")
        st.markdown("**Siempre=4 | Casi Siempre=3 | Algunas Veces=2 | Nunca=1**")

        actuaciones_scores = {}
        for idx, actuacion in enumerate(PERIODO_PRUEBA_ACTUACIONES):
            col_act, col_score = st.columns([4, 1])
            with col_act:
                st.markdown(f"**{idx + 1}.** {actuacion}")
            with col_score:
                actuaciones_scores[idx] = st.selectbox(
                    f"Actuación {idx+1}",
                    options=[4, 3, 2, 1],
                    format_func=lambda x: PERIODO_PRUEBA_ESCALA_ACTUACIONES[x]["label"],
                    key=f"act_{idx}",
                    label_visibility="collapsed",
                )
            st.markdown("---")

        # ---- SECCIÓN 2: CALIFICACIONES ----
        st.markdown("### ⭐ Sección 2: Calificaciones Específicas")
        st.markdown("**Excelente=5 | Bueno=4 | Regular=3 | Deficiente=2 | Insuficiente=1**")

        calificaciones_scores = {}
        for idx, calificacion in enumerate(PERIODO_PRUEBA_CALIFICACIONES):
            col_cal, col_cscore = st.columns([4, 1])
            with col_cal:
                st.markdown(f"**{calificacion}**")
            with col_cscore:
                calificaciones_scores[idx] = st.selectbox(
                    f"Cal {idx+1}",
                    options=[5, 4, 3, 2, 1],
                    format_func=lambda x: PERIODO_PRUEBA_ESCALA_CALIFICACIONES[x]["label"],
                    key=f"cal_{idx}",
                    label_visibility="collapsed",
                )
            st.markdown("---")

        # ---- SECCIÓN 3: PREGUNTAS ADICIONALES ----
        st.markdown("### 📌 Sección 3: Información Adicional")
        col_lam, col_con = st.columns(2)
        with col_lam:
            llamados = st.radio("¿Tuvo llamados de atención?", options=[False, True],
                                format_func=lambda x: "SÍ" if x else "NO",
                                key="llamados_atencion", horizontal=True)
        with col_con:
            conocimiento = st.radio("¿Su conocimiento se adecua al perfil del cargo?",
                                    options=[True, False],
                                    format_func=lambda x: "SÍ" if x else "NO",
                                    key="conocimiento_adecuado", horizontal=True)

        observaciones = st.text_area("Observaciones adicionales", height=120, key="obs_pp",
                                     placeholder="Comentarios generales sobre el desempeño durante el período...")

        aprobo = st.radio("¿El evaluado aprobó el período de prueba?",
                          options=[True, False],
                          format_func=lambda x: "✅ SÍ, APROBÓ" if x else "❌ NO APROBÓ",
                          key="aprobo_pp", horizontal=True)

        submitted = st.form_submit_button("✅ Guardar Evaluación", use_container_width=True, type="primary")

        if submitted:
            results_calc = calculate_periodo_prueba_results(
                actuaciones_scores=actuaciones_scores,
                calificaciones_scores=calificaciones_scores,
                aprobo=aprobo,
                llamados_atencion=llamados,
                conocimiento_adecuado=conocimiento,
                observaciones=observaciones,
            )
            results_to_save = {
                "test_type": "periodo_prueba",
                "evaluador": evaluador_nombre,
                "actuaciones_scores": {str(k): v for k, v in actuaciones_scores.items()},
                "calificaciones_scores": {str(k): v for k, v in calificaciones_scores.items()},
                "aprobo": aprobo,
                "llamados_atencion": llamados,
                "conocimiento_adecuado": conocimiento,
                "observaciones": observaciones,
                "analysis": results_calc,
            }
            db.save_results(session_id, results_to_save)
            db.complete_test_session(session_id)
            st.success("✅ Evaluación de período de prueba guardada correctamente.")
            st.session_state.pop("periodo_prueba_session_id", None)
            nav("admin_dashboard")
            st.rerun()

    if st.button("❌ Cancelar"):
        st.session_state.pop("periodo_prueba_session_id", None)
        nav("admin_dashboard")
        st.rerun()


def show_periodo_prueba_results_admin(results, candidate, session):
    """Muestra resultados de la evaluación de período de prueba en el panel de administración."""
    analysis = results.get("analysis", {})
    evaluador = results.get("evaluador", "N/A")
    aprobo = results.get("aprobo", False)
    session_id = session["id"] if isinstance(session, dict) else session

    # Banner de resultado
    clasif = analysis.get("clasificacion") or {}
    aprobacion_color = "#10B981" if aprobo else "#EF4444"
    aprobacion_text = "✅ APROBÓ EL PERÍODO DE PRUEBA" if aprobo else "❌ NO APROBÓ EL PERÍODO DE PRUEBA"

    st.markdown(f"""
    <div style="background: linear-gradient(135deg, {aprobacion_color}22, {aprobacion_color}44);
                border-left: 6px solid {aprobacion_color}; padding: 20px; border-radius: 12px; margin-bottom: 20px;">
        <h2 style="margin:0; color:{aprobacion_color};">{aprobacion_text}</h2>
        <p style="margin:8px 0 0 0; font-size:15px; color:#374151;">{clasif.get('descripcion','')}</p>
        <p style="margin:12px 0 0 0; font-size:14px; color:#6B7280;">
            <b>Evaluador:</b> {evaluador} | <b>Clasificación:</b> {clasif.get('label','')} |
            <b>Promedio General:</b> {analysis.get('promedio_general', 0):.2f}/4.00
        </p>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)
    col1.metric("📝 Actuaciones", f"{analysis.get('promedio_actuaciones', 0):.2f}/4.00")
    col2.metric("⭐ Calificaciones", f"{analysis.get('promedio_calificaciones', 0):.2f}/5.00")
    col3.metric("⚠️ Llamados de atención", "Sí" if results.get("llamados_atencion") else "No")

    st.markdown("---")

    tab1, tab2, tab3 = st.tabs(["📝 Actuaciones", "⭐ Calificaciones", "💡 Análisis"])

    with tab1:
        st.markdown(f"**Promedio de actuaciones:** {analysis.get('promedio_actuaciones', 0):.2f}/4.00")
        actuaciones_scores = {int(k): v for k, v in results.get("actuaciones_scores", {}).items()}
        for idx, actuacion in enumerate(PERIODO_PRUEBA_ACTUACIONES):
            score = actuaciones_scores.get(idx, 0)
            if score == 0:
                continue
            escala = PERIODO_PRUEBA_ESCALA_ACTUACIONES.get(score, {})
            c1, c2 = st.columns([4, 1])
            with c1:
                st.markdown(f"**{idx+1}.** {actuacion}")
            with c2:
                color = escala.get("color", "#6B7280")
                st.markdown(f"<div style='background:{color}22; padding:8px; border-radius:6px; text-align:center;'>"
                           f"<span style='font-size:12px; color:{color};'><b>{escala.get('label','')}</b></span></div>",
                           unsafe_allow_html=True)

    with tab2:
        st.markdown(f"**Promedio de calificaciones:** {analysis.get('promedio_calificaciones', 0):.2f}/5.00")
        calificaciones_scores = {int(k): v for k, v in results.get("calificaciones_scores", {}).items()}
        for idx, cal in enumerate(PERIODO_PRUEBA_CALIFICACIONES):
            score = calificaciones_scores.get(idx, 0)
            if score == 0:
                continue
            escala = PERIODO_PRUEBA_ESCALA_CALIFICACIONES.get(score, {})
            c1, c2 = st.columns([3, 1])
            with c1:
                st.markdown(f"**{cal}**")
            with c2:
                color = escala.get("color", "#6B7280")
                st.markdown(f"<div style='background:{color}22; padding:8px; border-radius:6px; text-align:center;'>"
                           f"<span style='font-size:12px; color:{color};'><b>{escala.get('label','')}</b></span></div>",
                           unsafe_allow_html=True)

        st.markdown("---")
        col_a, col_b = st.columns(2)
        with col_a:
            kon = results.get("conocimiento_adecuado", False)
            st.markdown(f"**¿Conocimiento adecua al perfil?** {'✅ Sí' if kon else '❌ No'}")
        with col_b:
            lam = results.get("llamados_atencion", False)
            st.markdown(f"**¿Llamados de atención?** {'⚠️ Sí' if lam else '✅ No'}")

        if results.get("observaciones"):
            st.markdown("---")
            st.markdown("**Observaciones adicionales:**")
            st.info(results["observaciones"])

    with tab3:
        st.markdown("#### 💡 Recomendaciones")
        for recom in analysis.get("recomendaciones", []):
            if aprobo:
                st.success(f"• {recom}")
            else:
                st.warning(f"• {recom}")

        if analysis.get("actuaciones_destacadas"):
            st.markdown("#### ✅ Comportamientos Destacados")
            for item in analysis["actuaciones_destacadas"]:
                st.success(f"• {item['nombre']}")

        if analysis.get("actuaciones_observacion"):
            st.markdown("#### ⚠️ Comportamientos a Reforzar")
            for item in analysis["actuaciones_observacion"]:
                st.warning(f"• {item['nombre']}")

    st.markdown("---")
    _col_pp1, _col_pp2 = st.columns(2)
    with _col_pp1:
        st.download_button(
            "📄 Descargar JSON",
            data=json.dumps(results, indent=2, ensure_ascii=False),
            file_name=f"periodo_prueba_{candidate['cedula']}_{session_id}.json",
            mime="application/json",
            key=f"json_pp_{session_id}",
        )
    with _col_pp2:
        try:
            _pdf_pp = generate_periodo_prueba_pdf(
                candidate=candidate,
                actuaciones_scores={int(k): v for k, v in results.get("actuaciones_scores", {}).items()},
                calificaciones_scores={int(k): v for k, v in results.get("calificaciones_scores", {}).items()},
                session_id=session_id,
                completed_at=session.get("completed_at") if isinstance(session, dict) else None,
                analysis=analysis,
                evaluador_nombre=evaluador,
                aprobo=aprobo,
                llamados_atencion=results.get("llamados_atencion", False),
                conocimiento_adecuado=results.get("conocimiento_adecuado", True),
                observaciones=results.get("observaciones"),
            )
            st.download_button(
                "📑 Descargar PDF",
                data=_pdf_pp,
                file_name=f"periodo_prueba_{candidate['cedula']}_{session_id}.pdf",
                mime="application/pdf",
                key=f"pdf_pp_{session_id}",
            )
        except Exception as _pdf_err:
            st.warning(f"No se pudo generar el PDF: {_pdf_err}")


# -------------------------------------------------------------------------
# EVALUADOR/JEFE: LOGIN
# -------------------------------------------------------------------------
def page_evaluador_login():
    st.markdown("## 👔 Acceso Evaluador / Jefe")
    st.info("Ingresa tu cédula para ver y completar las evaluaciones de tus colaboradores.")
    if st.button("⬅️ Volver al inicio"):
        nav("home")
        st.rerun()

    with st.form("evaluador_login_form"):
        cedula = st.text_input("Tu Cédula", placeholder="Número de cédula del evaluador/jefe")
        submitted = st.form_submit_button("🔑 Ingresar")
        if submitted:
            cedula = cedula.strip()
            if not cedula:
                st.error("Ingresa tu cédula.")
            else:
                sessions = db.get_sessions_for_evaluador(cedula)
                candidate_info = db.get_candidate_by_cedula(cedula)
                assigned_name = None
                if sessions:
                    assigned_name = next((s.get("evaluador_nombre") for s in sessions if s.get("evaluador_nombre")), None)
                name = candidate_info["name"] if candidate_info else (assigned_name or cedula)
                if not sessions:
                    st.warning("No tienes evaluaciones pendientes de tu parte en este momento.")
                    st.caption("Las evaluaciones aparecerán aquí una vez que el empleado complete su auto-evaluación.")
                st.session_state["evaluador"] = {"cedula": cedula, "name": name}
                nav("evaluador_dashboard")
                st.rerun()


# -------------------------------------------------------------------------
# EVALUADOR/JEFE: DASHBOARD
# -------------------------------------------------------------------------
def page_evaluador_dashboard():
    evaluador = st.session_state.get("evaluador")
    if not evaluador:
        nav("evaluador_login")
        st.rerun()
        return

    st.markdown(f"## 👔 Panel del Evaluador / Jefe")
    st.caption(f"Bienvenido, **{evaluador['name']}** | Cédula: {evaluador['cedula']}")

    if st.button("🚪 Cerrar Sesión"):
        st.session_state.pop("evaluador", None)
        nav("home")
        st.rerun()

    sessions = db.get_sessions_for_evaluador(evaluador["cedula"])

    if not sessions:
        st.info("📋 No tienes evaluaciones pendientes de tu parte en este momento.")
        st.caption("Las evaluaciones aparecerán aquí una vez que el empleado complete su auto-evaluación. Vuelve más tarde.")
        return

    st.success("✅ Tienes evaluaciones asignadas para completar.")
    st.markdown(f"### Evaluaciones Pendientes de tu Parte ({len(sessions)})")
    st.markdown("---")

    for sess in sessions:
        test_label = {
            "desempeno": "📈 Evaluación de Desempeño — Operativo",
            "desempeno_lider": "📊 Evaluación de Desempeño — Líderes",
            "periodo_prueba": "📋 Evaluación Período de Prueba",
        }.get(sess["test_type"], sess["test_type"])

        with st.container():
            c1, c2, c3 = st.columns([4, 1, 1])
            with c1:
                st.markdown(f"### {test_label}")
                st.markdown(f"**Empleado:** {sess['candidate_name']} | **Cédula:** {sess['cedula']}")
                st.caption(f"ID: {sess['id']} | Creado: {sess.get('created_at', 'N/A')[:10]}")
            with c2:
                st.markdown("<br>", unsafe_allow_html=True)
                if sess["test_type"] == "desempeno":
                    st.info("Pendiente ⏳")
                else:
                    st.success("Listo ✅")
            with c3:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("✏️ Completar mi Evaluación", key=f"ev_{sess['id']}", use_container_width=True):
                    st.session_state["evaluador_session_id"] = sess["id"]
                    if sess["test_type"] == "desempeno":
                        st.session_state["desempeno_session_id"] = sess["id"]
                        nav("desempeno_eval")
                    elif sess["test_type"] == "desempeno_lider":
                        nav("desempeno_lider_jefe_eval")
                    elif sess["test_type"] == "periodo_prueba":
                        nav("periodo_prueba_jefe_eval")
                    st.rerun()
            st.markdown("---")


# -------------------------------------------------------------------------
# CANDIDATO: AUTO-EVALUACIÓN DESEMPEÑO LÍDERES
# -------------------------------------------------------------------------
def page_desempeno_lider_employee_eval():
    """Auto-evaluación del empleado para Desempeño Líderes (7 competencias)."""
    session = st.session_state.get("test_session")
    candidate = st.session_state.get("candidate")

    if not session or not candidate:
        nav("candidate_login")
        st.rerun()
        return

    session_id = session["id"]
    nivel_cargo = candidate.get("nivel_cargo", "ANALISTA") or "ANALISTA"

    st.markdown("## 📊 Auto-Evaluación de Competencias")
    st.markdown(f"**Candidato:** {candidate['name']}")
    st.info("Evalúa con honestidad el nivel que consideras que has alcanzado en cada competencia organizacional.")
    st.markdown("---")

    with st.form("employee_competencias_form"):
        st.markdown("### Competencias Organizacionales — Autoevaluación")
        st.markdown("Selecciona el nivel que mejor describe tu desempeño actual:")

        nivel_req_info = COMPETENCIAS_NIVEL_REQUERIDO.get(nivel_cargo.upper(), None)
        competencias_scores = {}

        for comp in COMPETENCIAS_ORGANIZACIONALES:
            req = nivel_req_info["niveles"][comp["id"] - 1] if nivel_req_info else None
            req_text = f" _(Nivel requerido para tu cargo: {req})_" if req else ""
            st.markdown(f"**{comp['nombre']}**{req_text}")
            st.caption(comp["descripcion"])
            nivel_sel = st.radio(
                f"Nivel {comp['nombre']}",
                options=[1, 2, 3, 4, 5, 6],
                format_func=lambda x, c=comp: f"Nivel {x} — {c['niveles'][x][:80]}...",
                horizontal=True,
                key=f"emp_comp_{comp['id']}",
                label_visibility="collapsed",
                index=2,
            )
            competencias_scores[comp["id"]] = nivel_sel
            with st.expander("Ver descripción completa de este nivel"):
                st.info(comp["niveles"][nivel_sel])
            st.markdown("---")

        submitted = st.form_submit_button("✅ Enviar Auto-Evaluación", use_container_width=True, type="primary")

        if submitted:
            partial_results = {
                "employee_self": {
                    "competencias_scores": {str(k): v for k, v in competencias_scores.items()},
                    "nivel_cargo": nivel_cargo,
                }
            }
            db.save_results(session_id, partial_results)
            db.set_employee_done_status(session_id)

            for key in ["test_session"]:
                st.session_state.pop(key, None)

            nav("candidate_done")
            st.rerun()


# -------------------------------------------------------------------------
# CANDIDATO: AUTO-EVALUACIÓN PERÍODO DE PRUEBA
# -------------------------------------------------------------------------
def page_periodo_prueba_employee_eval():
    """Auto-evaluación del empleado para Período de Prueba (18 actuaciones)."""
    session = st.session_state.get("test_session")
    candidate = st.session_state.get("candidate")

    if not session or not candidate:
        nav("candidate_login")
        st.rerun()
        return

    session_id = session["id"]

    st.markdown("## 📋 Auto-Evaluación — Período de Prueba")
    st.markdown(f"**Candidato:** {candidate['name']}")
    st.info("Evalúa con honestidad con qué frecuencia realizas cada comportamiento en tu trabajo.")
    st.markdown("---")

    with st.form("employee_periodo_prueba_form"):
        st.markdown("### Actuaciones y Comportamientos — Autoevaluación")
        st.markdown("**Siempre=4 | Casi Siempre=3 | Algunas Veces=2 | Nunca=1**")

        actuaciones_scores = {}
        for idx, actuacion in enumerate(PERIODO_PRUEBA_ACTUACIONES):
            col_act, col_score = st.columns([4, 1])
            with col_act:
                st.markdown(f"**{idx + 1}.** {actuacion}")
            with col_score:
                actuaciones_scores[idx] = st.selectbox(
                    f"Actuación {idx+1}",
                    options=[4, 3, 2, 1],
                    format_func=lambda x: PERIODO_PRUEBA_ESCALA_ACTUACIONES[x]["label"],
                    key=f"emp_act_{idx}",
                    label_visibility="collapsed",
                )
            st.markdown("---")

        submitted = st.form_submit_button("✅ Enviar Auto-Evaluación", use_container_width=True, type="primary")

        if submitted:
            partial_results = {
                "employee_self": {
                    "actuaciones_scores": {str(k): v for k, v in actuaciones_scores.items()},
                }
            }
            db.save_results(session_id, partial_results)
            db.set_employee_done_status(session_id)

            for key in ["test_session"]:
                st.session_state.pop(key, None)

            nav("candidate_done")
            st.rerun()


# -------------------------------------------------------------------------
# JEFE: EVALUACIÓN DESEMPEÑO LÍDERES (con referencia auto-eval empleado)
# -------------------------------------------------------------------------
def page_desempeno_lider_jefe_eval():
    """Evaluación del jefe para Desempeño Líderes — muestra auto-evaluación del empleado como referencia."""
    session_id = st.session_state.get("evaluador_session_id") or st.session_state.get("desempeno_lider_session_id")
    evaluador = st.session_state.get("evaluador")
    admin = st.session_state.get("admin")

    if not session_id:
        st.error("No se encontró una sesión de evaluación activa.")
        return

    session = db.get_session_by_id(session_id)
    if not session:
        st.error("Sesión no válida.")
        return

    candidate = db.get_candidate_by_cedula(
        db.get_connection().execute("SELECT cedula FROM candidates WHERE id = ?", (session["candidate_id"],)).fetchone()["cedula"]
    )
    evaluador_nombre = (evaluador.get("name") if evaluador else None) or (admin.get("name") if admin else "Evaluador")
    nivel_cargo = candidate.get("nivel_cargo", "ANALISTA") or "ANALISTA"

    # Cargar auto-evaluación del empleado
    existing_results = db.get_results(session_id) or {}
    employee_self = existing_results.get("employee_self", {})
    emp_comp_scores = {int(k): v for k, v in employee_self.get("competencias_scores", {}).items()}

    # Botón de regreso
    if evaluador:
        if st.button("⬅️ Volver al Dashboard del Evaluador"):
            st.session_state.pop("evaluador_session_id", None)
            nav("evaluador_dashboard")
            st.rerun()
    elif admin:
        if st.button("⬅️ Volver al Dashboard Admin"):
            st.session_state.pop("desempeno_lider_session_id", None)
            nav("admin_dashboard")
            st.rerun()

    st.markdown("## 📊 Evaluación de Desempeño — Líderes (Evaluación del Jefe)")
    st.markdown(f"**Colaborador:** {candidate['name']} (Cédula: {candidate['cedula']})")
    st.markdown(f"**Cargo:** {candidate.get('position', 'N/A')} | **Nivel:** {nivel_cargo}")
    st.markdown(f"**Evaluador:** {evaluador_nombre}")

    if emp_comp_scores:
        with st.expander("📋 Ver Auto-Evaluación del Empleado (Referencia)", expanded=False):
            st.markdown("**El empleado se evaluó así en cada competencia:**")
            cols_ref = st.columns(2)
            for i, comp in enumerate(COMPETENCIAS_ORGANIZACIONALES):
                cid = comp["id"]
                emp_score = emp_comp_scores.get(cid, 0)
                with cols_ref[i % 2]:
                    if emp_score:
                        st.markdown(f"- **{comp['nombre']}**: Nivel {emp_score}")
    st.markdown("---")

    with st.form("evaluacion_desempeno_lider_jefe_form"):
        # ---- SECCIÓN 1: COMPETENCIAS (evaluación del jefe) ----
        st.markdown("### 🏆 SECCIÓN 1: Evaluación de Competencias (Tu evaluación como jefe)")
        st.markdown("Selecciona el nivel alcanzado por el colaborador según tu observación:")

        nivel_req_info = COMPETENCIAS_NIVEL_REQUERIDO.get(nivel_cargo.upper(), None)
        competencias_scores = {}

        for comp in COMPETENCIAS_ORGANIZACIONALES:
            req = nivel_req_info["niveles"][comp["id"] - 1] if nivel_req_info else None
            req_text = f" _(Requerido: Nivel {req})_" if req else ""
            emp_score = emp_comp_scores.get(comp["id"])
            emp_ref = f" | _Auto-eval empleado: Nivel {emp_score}_" if emp_score else ""
            st.markdown(f"**{comp['nombre']}**{req_text}{emp_ref}")
            st.caption(comp["descripcion"])
            nivel_sel = st.radio(
                f"Nivel {comp['nombre']}",
                options=[1, 2, 3, 4, 5, 6],
                format_func=lambda x, c=comp: f"Nivel {x} — {c['niveles'][x][:80]}...",
                horizontal=True,
                key=f"jefe_comp_{comp['id']}",
                label_visibility="collapsed",
                index=2,
            )
            competencias_scores[comp["id"]] = nivel_sel
            st.markdown("---")

        # ---- SECCIÓN 2: RENDIMIENTO ----
        st.markdown("### 📝 SECCIÓN 2: Evaluación de Rendimiento")
        st.markdown("**5** = Sobresaliente | **4** = Supera | **3** = Cumple | **2** = Debajo | **1** = Insatisfactorio")

        rendimiento_scores = {}
        for obj in DESEMPENO_OBJETIVOS:
            st.markdown(f"**{obj['titulo']}**")
            st.caption(obj["descripcion"])
            rendimiento_scores[obj["id"]] = st.select_slider(
                f"Calificación Objetivo {obj['id']}",
                options=[1, 2, 3, 4, 5],
                value=3,
                format_func=lambda x: DESEMPENO_ESCALA_RENDIMIENTO[x]["label"],
                key=f"jefe_rend_{obj['id']}",
                label_visibility="collapsed",
            )
            st.markdown("---")

        # ---- SECCIÓN 3: POTENCIAL ----
        st.markdown("### 🎯 SECCIÓN 3: Evaluación de Potencial (0-3)")
        potencial_scores = {}
        for dim in DESEMPENO_DIMENSIONES:
            st.markdown(f"**{dim['nombre']}**")
            st.caption(dim["descripcion"])
            nivel_sel = st.radio(
                f"Nivel {dim['nombre']}",
                options=[3, 2, 1, 0],
                format_func=lambda x, d=dim: f"Nivel {x}: {d['niveles'][x][:80]}...",
                key=f"jefe_pot_{dim['id']}",
                label_visibility="collapsed",
            )
            potencial_scores[dim["id"]] = nivel_sel
            st.markdown("---")

        # ---- INICIATIVAS ----
        st.markdown("### 🚀 Iniciativas de Mejora")
        n_iniciativas = st.selectbox("Número de iniciativas", [0, 1, 2, 3], index=1, key="n_init_jefe_lider")
        iniciativas = []
        for i in range(n_iniciativas):
            ini = st.text_area(f"Iniciativa {i+1}", key=f"ini_jefe_lider_{i}", height=80)
            if ini.strip():
                iniciativas.append(ini.strip())

        submitted = st.form_submit_button("✅ Guardar Evaluación Completa", use_container_width=True, type="primary")

        if submitted:
            results_calc = calculate_desempeno_lider_results(
                competencias_scores=competencias_scores,
                rendimiento_scores=rendimiento_scores,
                potencial_scores=potencial_scores,
                nivel_cargo=nivel_cargo,
                iniciativas=iniciativas,
            )
            results_to_save = {
                "test_type": "desempeno_lider",
                "evaluador": evaluador_nombre,
                "nivel_cargo": nivel_cargo,
                "competencias_scores": {str(k): v for k, v in competencias_scores.items()},
                "rendimiento_scores": {str(k): v for k, v in rendimiento_scores.items()},
                "potencial_scores": {str(k): v for k, v in potencial_scores.items()},
                "iniciativas": iniciativas,
                "analysis": results_calc,
                "employee_self": employee_self,
            }
            db.save_results(session_id, results_to_save)
            db.complete_test_session(session_id)
            st.success("✅ Evaluación de desempeño (líderes) guardada exitosamente.")
            st.balloons()
            st.session_state.pop("evaluador_session_id", None)
            st.session_state.pop("desempeno_lider_session_id", None)
            if evaluador:
                nav("evaluador_dashboard")
            else:
                nav("admin_dashboard")
            st.rerun()


# -------------------------------------------------------------------------
# JEFE: EVALUACIÓN PERÍODO DE PRUEBA (con referencia auto-eval empleado)
# -------------------------------------------------------------------------
def page_periodo_prueba_jefe_eval():
    """Evaluación del jefe para Período de Prueba — muestra auto-evaluación del empleado como referencia."""
    session_id = st.session_state.get("evaluador_session_id") or st.session_state.get("periodo_prueba_session_id")
    evaluador = st.session_state.get("evaluador")
    admin = st.session_state.get("admin")

    if not session_id:
        st.error("No se encontró una sesión de evaluación activa.")
        return

    session = db.get_session_by_id(session_id)
    if not session:
        st.error("Sesión no válida.")
        return

    candidate = db.get_candidate_by_cedula(
        db.get_connection().execute("SELECT cedula FROM candidates WHERE id = ?", (session["candidate_id"],)).fetchone()["cedula"]
    )
    evaluador_nombre = (evaluador.get("name") if evaluador else None) or (admin.get("name") if admin else "Evaluador")

    # Cargar auto-evaluación del empleado
    existing_results = db.get_results(session_id) or {}
    employee_self = existing_results.get("employee_self", {})
    emp_act_scores = {int(k): v for k, v in employee_self.get("actuaciones_scores", {}).items()}

    # Botón de regreso
    if evaluador:
        if st.button("⬅️ Volver al Dashboard del Evaluador"):
            st.session_state.pop("evaluador_session_id", None)
            nav("evaluador_dashboard")
            st.rerun()
    elif admin:
        if st.button("⬅️ Volver al Dashboard Admin"):
            st.session_state.pop("periodo_prueba_session_id", None)
            nav("admin_dashboard")
            st.rerun()

    st.markdown("## 📋 Evaluación Período de Prueba (Evaluación del Jefe)")
    st.markdown(f"**Trabajador:** {candidate['name']} (Cédula: {candidate['cedula']})")
    st.markdown(f"**Cargo:** {candidate.get('position', 'N/A')} | **Área:** {candidate.get('regional', 'N/A')}")
    st.markdown(f"**Evaluador:** {evaluador_nombre}")
    st.info("Marque la frecuencia con la que observa cada comportamiento durante el desempeño laboral.")

    if emp_act_scores:
        with st.expander("📋 Ver Auto-Evaluación del Empleado (Referencia)", expanded=False):
            st.markdown("**El empleado se evaluó así:**")
            for idx, actuacion in enumerate(PERIODO_PRUEBA_ACTUACIONES):
                emp_score = emp_act_scores.get(idx)
                if emp_score:
                    escala = PERIODO_PRUEBA_ESCALA_ACTUACIONES.get(emp_score, {})
                    st.markdown(f"- **{idx+1}. {actuacion[:60]}...**: {escala.get('label', emp_score)}")
    st.markdown("---")

    with st.form("jefe_periodo_prueba_form"):
        st.markdown("### 📝 Sección 1: Actuaciones y Comportamientos (Tu evaluación como jefe)")
        st.markdown("**Siempre=4 | Casi Siempre=3 | Algunas Veces=2 | Nunca=1**")

        actuaciones_scores = {}
        for idx, actuacion in enumerate(PERIODO_PRUEBA_ACTUACIONES):
            emp_score = emp_act_scores.get(idx)
            emp_ref = f" *(Auto-eval: {PERIODO_PRUEBA_ESCALA_ACTUACIONES.get(emp_score, {}).get('label', emp_score)})*" if emp_score else ""
            col_act, col_score = st.columns([4, 1])
            with col_act:
                st.markdown(f"**{idx + 1}.** {actuacion}{emp_ref}")
            with col_score:
                actuaciones_scores[idx] = st.selectbox(
                    f"Actuación {idx+1}",
                    options=[4, 3, 2, 1],
                    format_func=lambda x: PERIODO_PRUEBA_ESCALA_ACTUACIONES[x]["label"],
                    key=f"jefe_act_{idx}",
                    label_visibility="collapsed",
                )
            st.markdown("---")

        st.markdown("### ⭐ Sección 2: Calificaciones Específicas")
        st.markdown("**Excelente=5 | Bueno=4 | Regular=3 | Deficiente=2 | Insuficiente=1**")

        calificaciones_scores = {}
        for idx, calificacion in enumerate(PERIODO_PRUEBA_CALIFICACIONES):
            col_cal, col_cscore = st.columns([4, 1])
            with col_cal:
                st.markdown(f"**{calificacion}**")
            with col_cscore:
                calificaciones_scores[idx] = st.selectbox(
                    f"Cal {idx+1}",
                    options=[5, 4, 3, 2, 1],
                    format_func=lambda x: PERIODO_PRUEBA_ESCALA_CALIFICACIONES[x]["label"],
                    key=f"jefe_cal_{idx}",
                    label_visibility="collapsed",
                )
            st.markdown("---")

        st.markdown("### 📌 Sección 3: Información Adicional")
        col_lam, col_con = st.columns(2)
        with col_lam:
            llamados = st.radio("¿Tuvo llamados de atención?", options=[False, True],
                                format_func=lambda x: "SÍ" if x else "NO",
                                key="jefe_llamados", horizontal=True)
        with col_con:
            conocimiento = st.radio("¿Su conocimiento se adecua al perfil del cargo?",
                                    options=[True, False],
                                    format_func=lambda x: "SÍ" if x else "NO",
                                    key="jefe_conocimiento", horizontal=True)

        observaciones = st.text_area("Observaciones adicionales", height=120, key="jefe_obs_pp",
                                     placeholder="Comentarios generales sobre el desempeño durante el período...")

        aprobo = st.radio("¿El evaluado aprobó el período de prueba?",
                          options=[True, False],
                          format_func=lambda x: "✅ SÍ, APROBÓ" if x else "❌ NO APROBÓ",
                          key="jefe_aprobo_pp", horizontal=True)

        submitted = st.form_submit_button("✅ Guardar Evaluación Completa", use_container_width=True, type="primary")

        if submitted:
            results_calc = calculate_periodo_prueba_results(
                actuaciones_scores=actuaciones_scores,
                calificaciones_scores=calificaciones_scores,
                aprobo=aprobo,
                llamados_atencion=llamados,
                conocimiento_adecuado=conocimiento,
                observaciones=observaciones,
            )
            results_to_save = {
                "test_type": "periodo_prueba",
                "evaluador": evaluador_nombre,
                "actuaciones_scores": {str(k): v for k, v in actuaciones_scores.items()},
                "calificaciones_scores": {str(k): v for k, v in calificaciones_scores.items()},
                "aprobo": aprobo,
                "llamados_atencion": llamados,
                "conocimiento_adecuado": conocimiento,
                "observaciones": observaciones,
                "analysis": results_calc,
                "employee_self": employee_self,
            }
            db.save_results(session_id, results_to_save)
            db.complete_test_session(session_id)
            st.success("✅ Evaluación de período de prueba guardada correctamente.")
            st.balloons()
            st.session_state.pop("evaluador_session_id", None)
            st.session_state.pop("periodo_prueba_session_id", None)
            if evaluador:
                nav("evaluador_dashboard")
            else:
                nav("admin_dashboard")
            st.rerun()


# -------------------------------------------------------------------------
# CANDIDATE: LOGIN
# -------------------------------------------------------------------------
def page_candidate_login():
    st.markdown("## 🔑 Acceso Candidato")
    if st.button("⬅️ Volver al inicio"):
        nav("home")
        st.rerun()

    st.markdown("Ingresa tu número de cédula para acceder a las evaluaciones asignadas.")

    with st.form("candidate_login_form"):
        cedula = st.text_input("Número de Cédula", placeholder="Ingresa tu cédula")
        submitted = st.form_submit_button("Ingresar")

        if submitted:
            if not cedula.strip():
                st.error("Por favor ingresa tu cédula.")
            else:
                candidate = db.get_candidate_by_cedula(cedula.strip())
                if not candidate:
                    st.error("❌ No se encontró un candidato con esa cédula. Contacta a Recursos Humanos.")
                else:
                    pending = db.get_pending_sessions_for_candidate(candidate["id"])
                    if not pending:
                        st.warning("⚠️ No tienes evaluaciones pendientes asignadas. Contacta a Recursos Humanos.")
                    else:
                        st.session_state.candidate = candidate
                        st.session_state.pending_sessions = pending
                        nav("candidate_select_test")
                        st.rerun()


# -------------------------------------------------------------------------
# CANDIDATE: SELECT TEST
# -------------------------------------------------------------------------
def page_candidate_select_test():
    candidate = st.session_state.get("candidate")
    if not candidate:
        nav("candidate_login")
        st.rerun()
        return

    pending_all = db.get_pending_sessions_for_candidate(candidate["id"])
    # Estas evaluaciones existen, pero no son respondidas por el candidato.
    _ADMIN_ONLY_TEST_TYPES = {"desempeno"}
    pending = [s for s in pending_all if s["test_type"] not in _ADMIN_ONLY_TEST_TYPES]
    pending_info_only = [s for s in pending_all if s["test_type"] in _ADMIN_ONLY_TEST_TYPES]
    st.session_state.pending_sessions = pending

    st.markdown(f"## Bienvenido/a, {candidate['name']}")
    st.markdown("Tienes las siguientes evaluaciones asignadas:")

    if not pending and not pending_info_only:
        st.info("✅ No tienes evaluaciones pendientes. ¡Gracias!")
        if st.button("🔑 Cerrar Sesión"):
            for key in ["candidate", "pending_sessions", "test_session", "disc_questions", "disc_page", "disc_answers", "valanti_responses", "valanti_page"]:
                st.session_state.pop(key, None)
            nav("home")
            st.rerun()
        return

    if pending_info_only and not pending:
        st.info("ℹ️ Tienes evaluaciones pendientes asignadas, pero no requieren acción de tu parte por ahora.")

    for sess in pending:
        # Determinar emoji y nombre según tipo de test
        if sess["test_type"] == "disc":
            test_emoji = "🎯"
            test_name = "Evaluación DISC"
        elif sess["test_type"] == "valanti":
            test_emoji = "🧭"
            test_name = "Cuestionario VALANTI"
        elif sess["test_type"] == "wpi":
            test_emoji = "💼"
            test_name = "WPI - Work Personality Index"
        elif sess["test_type"] == "eri":
            test_emoji = "🔐"
            test_name = "ERI - Evaluación de Riesgo e Integridad"
        elif sess["test_type"] == "talent_map":
            test_emoji = "🌟"
            test_name = "Talent Map - Mapeo de Competencias"
        elif sess["test_type"] == "desempeno_lider":
            test_emoji = "📊"
            test_name = "Auto-Evaluación de Competencias (Desempeño Líderes)"
        elif sess["test_type"] == "periodo_prueba":
            test_emoji = "📋"
            test_name = "Auto-Evaluación Período de Prueba"
        else:
            test_emoji = "📝"
            test_name = "Evaluación"
        
        status_text = "En progreso ▶️" if sess["status"] == "in_progress" else "Pendiente ⏳"

        with st.container():
            c1, c2, c3 = st.columns([3, 1, 1])
            with c1:
                st.markdown(f"### {test_emoji} {test_name}")
                st.caption(f"ID: {sess['id']} | Tiempo: {sess['time_limit_minutes']} min | Estado: {status_text}")
                if sess.get("evaluador_nombre") or sess.get("evaluador_cedula"):
                    jefe_info = sess.get("evaluador_nombre") or "N/A"
                    jefe_ced = sess.get("evaluador_cedula") or "N/A"
                    st.caption(f"👔 Jefe asignado: {jefe_info} (CC: {jefe_ced})")
            with c2:
                st.metric("Tiempo", f"{sess['time_limit_minutes']} min")
            with c3:
                button_text = "▶️ Continuar" if sess["status"] == "in_progress" else "🚀 Iniciar"
                if st.button(button_text, key=f"start_{sess['id']}", use_container_width=True):
                    if sess["status"] == "in_progress":
                        remaining = db.check_session_time(sess)
                        if remaining == -1:
                            st.error("⏰ El tiempo de esta evaluación ha expirado.")
                            st.rerun()
                            return

                    if sess["status"] == "pending":
                        db.start_test_session(sess["id"])
                    st.session_state.test_session = db.get_session_by_id(sess["id"])

                    if sess["test_type"] == "disc":
                        nav("disc_test")
                    elif sess["test_type"] == "valanti":
                        nav("valanti_test")
                    elif sess["test_type"] == "wpi":
                        nav("wpi_test")
                    elif sess["test_type"] == "eri":
                        nav("eri_test")
                    elif sess["test_type"] == "talent_map":
                        nav("talent_map_test")
                    elif sess["test_type"] == "desempeno_lider":
                        nav("desempeno_lider_employee_eval")
                    elif sess["test_type"] == "periodo_prueba":
                        nav("periodo_prueba_employee_eval")
                    st.rerun()

    if pending_info_only:
        st.markdown("---")
        st.markdown("### 📌 Pendientes informativas")
        st.caption("Estas evaluaciones están pendientes en el sistema, pero son gestionadas por Evaluador/RH.")
        for sess in pending_info_only:
            with st.container():
                c1, c2 = st.columns([4, 1])
                with c1:
                    st.markdown("### 📊 Evaluación de Desempeño (Gestionada por evaluador)")
                    st.caption(f"ID: {sess['id']} | Tiempo: {sess['time_limit_minutes']} min | Estado: Pendiente ⏳")
                    if sess.get("evaluador_nombre") or sess.get("evaluador_cedula"):
                        jefe_info = sess.get("evaluador_nombre") or "N/A"
                        jefe_ced = sess.get("evaluador_cedula") or "N/A"
                        st.caption(f"👔 Jefe asignado: {jefe_info} (CC: {jefe_ced})")
                with c2:
                    st.metric("Acción", "N/A")

    st.markdown("---")
    if st.button("🔑 Cerrar Sesión"):
        for key in ["candidate", "pending_sessions", "test_session", 
                    "disc_questions", "disc_page", "disc_answers", 
                    "valanti_responses", "valanti_page",
                    "wpi_questions", "wpi_responses", "wpi_page",
                    "eri_questions", "eri_responses", "eri_page",
                    "tm_questions", "tm_responses", "tm_page",
                    "desempeno_session_id"]:
            st.session_state.pop(key, None)
        nav("home")
        st.rerun()


# -------------------------------------------------------------------------
# CANDIDATE: DISC TEST
# -------------------------------------------------------------------------
def page_disc_test():
    session = st.session_state.get("test_session")
    candidate = st.session_state.get("candidate")
    if not session or not candidate:
        nav("candidate_login")
        st.rerun()
        return

    session = db.get_session_by_id(session["id"])
    if not session or session["status"] not in ("in_progress",):
        if session and session["status"] == "expired":
            st.error("⏰ El tiempo de esta evaluación ha expirado.")
            if st.button("Volver"):
                nav("candidate_select_test")
                st.rerun()
            return
        nav("candidate_select_test")
        st.rerun()
        return

    remaining = db.check_session_time(session)
    if remaining == -1:
        st.error("⏰ El tiempo de esta evaluación ha expirado.")
        if st.button("Volver"):
            nav("candidate_select_test")
            st.rerun()
        return

    deadline_ts = db.get_session_deadline_timestamp(session)
    if deadline_ts:
        render_timer(deadline_ts, session["id"])

    st.markdown(f"### 🎯 Evaluación DISC")
    st.caption(f"Candidato: {candidate['name']} | ID: {session['id']}")

    if "disc_questions" not in st.session_state:
        all_questions = load_disc_questions()
        rng = random.Random(session["id"])
        rng.shuffle(all_questions)
        st.session_state.disc_questions = all_questions[:30]
        db.update_session_questions(session["id"], st.session_state.disc_questions)

    if "disc_page" not in st.session_state:
        st.session_state.disc_page = 0

    if "disc_answers" not in st.session_state:
        st.session_state.disc_answers = {}

    questions = st.session_state.disc_questions
    total = len(questions)
    page = st.session_state.disc_page

    progress = page / total
    st.progress(progress)
    st.markdown(f"**Pregunta {page + 1} de {total}**")

    options_map = {
        "Selecciona una opción": None,
        "1 - Totalmente en desacuerdo": 1,
        "2 - Algo en desacuerdo": 2,
        "3 - Neutral": 3,
        "4 - Algo de acuerdo": 4,
        "5 - Totalmente de acuerdo": 5,
    }

    if page < total:
        q = questions[page]
        with st.form(key=f"disc_form_{page}"):
            st.markdown(f"#### {page + 1}) {q['question']}")
            selected = st.radio("Tu respuesta:", list(options_map.keys()), index=0, horizontal=True, key=f"disc_radio_{page}")

            col_prev, col_space, col_next = st.columns([1, 4, 1])
            with col_prev:
                if page > 0:
                    if st.form_submit_button("⬅️ Anterior"):
                        st.session_state.disc_page -= 1
                        st.rerun()
            with col_next:
                if page < total - 1:
                    btn = st.form_submit_button("Siguiente ➡️")
                else:
                    btn = st.form_submit_button("✅ Finalizar")

        if btn:
            remaining = db.check_session_time(db.get_session_by_id(session["id"]))
            if remaining == -1:
                st.error("⏰ El tiempo ha expirado.")
                return

            if options_map[selected] is None:
                st.warning("⚠️ Por favor selecciona una respuesta.")
            else:
                st.session_state.disc_answers[page] = options_map[selected]
                if page < total - 1:
                    st.session_state.disc_page += 1
                    st.rerun()
                else:
                    answers_list = [st.session_state.disc_answers[i] for i in range(total)]
                    raw, normalized, relative = calculate_disc_results(answers_list, questions)

                    answer_records = []
                    for i in range(total):
                        answer_records.append({
                            "question_index": i,
                            "question_text": questions[i]["question"],
                            "answer_value": answers_list[i],
                        })
                    db.save_answers(session["id"], answer_records)

                    results = {"raw": raw, "normalized": normalized, "relative": relative}
                    db.save_results(session["id"], results)
                    db.complete_test_session(session["id"])

                    for key in ["disc_questions", "disc_page", "disc_answers", "test_session"]:
                        st.session_state.pop(key, None)

                    nav("candidate_done")
                    st.rerun()


# -------------------------------------------------------------------------
# CANDIDATE: VALANTI TEST
# -------------------------------------------------------------------------
def page_valanti_test():
    session = st.session_state.get("test_session")
    candidate = st.session_state.get("candidate")
    if not session or not candidate:
        nav("candidate_login")
        st.rerun()
        return

    session = db.get_session_by_id(session["id"])
    if not session or session["status"] not in ("in_progress",):
        if session and session["status"] == "expired":
            st.error("⏰ El tiempo de esta evaluación ha expirado.")
            if st.button("Volver"):
                nav("candidate_select_test")
                st.rerun()
            return
        nav("candidate_select_test")
        st.rerun()
        return

    remaining = db.check_session_time(session)
    if remaining == -1:
        st.error("⏰ El tiempo de esta evaluación ha expirado.")
        if st.button("Volver"):
            nav("candidate_select_test")
            st.rerun()
        return

    deadline_ts = db.get_session_deadline_timestamp(session)
    if deadline_ts:
        render_timer(deadline_ts, session["id"])

    st.markdown(f"### 🧭 Cuestionario VALANTI")
    st.caption(f"Candidato: {candidate['name']} | ID: {session['id']}")

    if "valanti_responses" not in st.session_state:
        st.session_state.valanti_responses = [None] * len(VALANTI_PREGUNTAS)

    if "valanti_page" not in st.session_state:
        st.session_state.valanti_page = 0

    total = len(VALANTI_PREGUNTAS)
    questions_per_page = 5
    page = st.session_state.valanti_page
    q_start = page * questions_per_page
    q_end = min(q_start + questions_per_page, total)

    progress = q_end / total
    st.progress(progress)
    st.markdown(f"**Preguntas {q_start + 1} - {q_end} de {total}**")

    if q_start < 9:
        st.info("**Primera Parte:** Distribuye 3 puntos entre las dos frases. El puntaje más alto para la frase más importante para ti.")
    else:
        st.warning("**Segunda Parte:** Distribuye 3 puntos entre las dos frases. El puntaje más alto para lo que consideres **peor**.")

    # Callbacks de auto-completado
    def make_cb_a(idx):
        def _cb():
            val = st.session_state.get(f"vq_{idx}_a", "--")
            if val != "--":
                st.session_state[f"vq_{idx}_b"] = 3 - int(val)
        return _cb

    def make_cb_b(idx):
        def _cb():
            val = st.session_state.get(f"vq_{idx}_b", "--")
            if val != "--":
                st.session_state[f"vq_{idx}_a"] = 3 - int(val)
        return _cb

    all_answered = True

    for i in range(q_start, q_end):
        par = VALANTI_PREGUNTAS[i]
        a_key = f"vq_{i}_a"
        b_key = f"vq_{i}_b"

        # Inicializar desde respuestas guardadas
        if a_key not in st.session_state:
            if st.session_state.valanti_responses[i] is not None:
                st.session_state[a_key] = st.session_state.valanti_responses[i]
                st.session_state[b_key] = 3 - st.session_state.valanti_responses[i]

        # Tarjeta visual
        st.markdown(
            f"""
            <div style="background: linear-gradient(135deg, #1e293b 0%, #334155 100%);
                        border-radius: 12px; padding: 20px; margin: 15px 0;
                        border-left: 4px solid #3b82f6;">
                <div style="margin-bottom: 8px;">
                    <span style="background: #3b82f6; color: white; padding: 4px 12px;
                                border-radius: 20px; font-size: 0.85em; font-weight: bold;">
                        Pregunta {i + 1}
                    </span>
                </div>
                <div style="display: flex; gap: 20px; margin-top: 10px;">
                    <div style="flex: 1; background: rgba(59,130,246,0.1); border-radius: 8px; padding: 12px;">
                        <span style="color: #60a5fa; font-weight: bold; font-size: 1.1em;">A)</span>
                        <span style="color: #e2e8f0; font-size: 1.05em;"> {par[0]}</span>
                    </div>
                    <div style="flex: 1; background: rgba(245,158,11,0.1); border-radius: 8px; padding: 12px;">
                        <span style="color: #fbbf24; font-weight: bold; font-size: 1.1em;">B)</span>
                        <span style="color: #e2e8f0; font-size: 1.05em;"> {par[1]}</span>
                    </div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        col_sa, col_sb, col_icon = st.columns([3, 3, 1])
        with col_sa:
            st.selectbox(
                f"Puntos para A (P{i+1})",
                options=["--", 0, 1, 2, 3],
                key=a_key,
                on_change=make_cb_a(i),
            )
        with col_sb:
            st.selectbox(
                f"Puntos para B (P{i+1})",
                options=["--", 0, 1, 2, 3],
                key=b_key,
                on_change=make_cb_b(i),
            )

        a_val = st.session_state.get(a_key, "--")
        b_val = st.session_state.get(b_key, "--")

        with col_icon:
            st.markdown("<br>", unsafe_allow_html=True)
            if a_val != "--" and b_val != "--" and int(a_val) + int(b_val) == 3:
                st.success("✅")
            else:
                st.warning("⚠️")
                all_answered = False

    # Navegación
    st.markdown("---")
    col_prev, col_space, col_next = st.columns([1, 4, 1])

    with col_prev:
        if page > 0:
            if st.button("⬅️ Anterior", key="valanti_prev"):
                for j in range(q_start, q_end):
                    a = st.session_state.get(f"vq_{j}_a", "--")
                    if a != "--":
                        st.session_state.valanti_responses[j] = int(a)
                st.session_state.valanti_page -= 1
                st.rerun()

    with col_next:
        is_last = q_end >= total
        btn_label = "✅ Finalizar Evaluación" if is_last else "Siguiente ➡️"
        if st.button(btn_label, key="valanti_next", disabled=not all_answered):
            remaining = db.check_session_time(db.get_session_by_id(session["id"]))
            if remaining == -1:
                st.error("⏰ El tiempo ha expirado.")
            else:
                for j in range(q_start, q_end):
                    a = st.session_state.get(f"vq_{j}_a", "--")
                    if a != "--":
                        st.session_state.valanti_responses[j] = int(a)

                if is_last:
                    if None in st.session_state.valanti_responses:
                        st.warning("⚠️ Hay preguntas sin responder. Revisa las páginas anteriores.")
                    else:
                        responses = st.session_state.valanti_responses
                        direct, standard = calculate_valanti_results(responses)

                        answer_records = []
                        for i in range(total):
                            answer_records.append({
                                "question_index": i,
                                "question_text": f"A: {VALANTI_PREGUNTAS[i][0]} / B: {VALANTI_PREGUNTAS[i][1]}",
                                "answer_value": responses[i],
                                "answer_b_value": 3 - responses[i],
                            })
                        db.save_answers(session["id"], answer_records)

                        results = {"direct": direct, "standard": standard}
                        db.save_results(session["id"], results)
                        db.complete_test_session(session["id"])

                        for key in ["valanti_responses", "valanti_page", "test_session"]:
                            st.session_state.pop(key, None)

                        nav("candidate_done")
                        st.rerun()
                else:
                    st.session_state.valanti_page += 1
                    st.rerun()


def page_wpi_test():
    """
    Página del test WPI (Work Personality Index) - 50 preguntas con escala Likert 1-5.
    """
    session = st.session_state.get("test_session")
    candidate = st.session_state.get("candidate")
    
    if not session or not candidate:
        nav("candidate_login")
        st.rerun()
        return

    session = db.get_session_by_id(session["id"])
    if not session or session["status"] not in ("in_progress",):
        if session and session["status"] == "expired":
            st.error("⏰ El tiempo de esta evaluación ha expirado.")
            if st.button("Volver"):
                nav("candidate_select_test")
                st.rerun()
            return
        nav("candidate_select_test")
        st.rerun()
        return

    # Verificar tiempo restante
    remaining = db.check_session_time(session)
    if remaining == -1:
        st.error("⏰ El tiempo de esta evaluación ha expirado.")
        if st.button("Volver"):
            nav("candidate_select_test")
            st.rerun()
        return

    # Mostrar timer
    deadline_ts = db.get_session_deadline_timestamp(session)
    if deadline_ts:
        render_timer(deadline_ts, session["id"])

    st.markdown(f"### 💼 WPI - Work Personality Index")
    st.caption(f"Candidato: {candidate['name']} | ID: {session['id']}")
    
    # Cargar preguntas si no están en session_state
    if "wpi_questions" not in st.session_state:
        all_questions = load_wpi_questions()
        # Mezclar preguntas de manera consistente por sesión
        rng = random.Random(session["id"])
        rng.shuffle(all_questions)
        st.session_state.wpi_questions = all_questions
        db.update_session_questions(session["id"], all_questions)

    # Inicializar respuestas
    if "wpi_responses" not in st.session_state:
        st.session_state.wpi_responses = [None] * len(st.session_state.wpi_questions)

    # Inicializar página
    if "wpi_page" not in st.session_state:
        st.session_state.wpi_page = 0

    questions = st.session_state.wpi_questions
    total = len(questions)
    questions_per_page = 10  # 10 preguntas por página
    page = st.session_state.wpi_page
    q_start = page * questions_per_page
    q_end = min(q_start + questions_per_page, total)

    # Barra de progreso
    progress = q_end / total
    st.progress(progress)
    st.markdown(f"**Preguntas {q_start + 1} - {q_end} de {total}**")

    # Instrucciones
    st.info("""
    **Instrucciones:** Responde con sinceridad a cada afirmación según la siguiente escala:
    - **1** = Totalmente en desacuerdo
    - **2** = En desacuerdo
    - **3** = Neutral
    - **4** = De acuerdo
    - **5** = Totalmente de acuerdo
    """)

    # Mostrar preguntas de la página actual
    all_answered = True
    
    for i in range(q_start, q_end):
        q = questions[i]
        q_text = q["question"]
        dim = q["dimension"]
        
        # Crear tarjeta visual para cada pregunta
        st.markdown(
            f"""
            <div style="background: linear-gradient(135deg, #1e293b 0%, #334155 100%);
                        border-radius: 12px; padding: 20px; margin: 15px 0;
                        border-left: 4px solid {WPI_COLORS.get(dim, '#3b82f6')};">
                <div style="margin-bottom: 8px;">
                    <span style="background: {WPI_COLORS.get(dim, '#3b82f6')}; color: white; 
                                padding: 4px 12px; border-radius: 20px; 
                                font-size: 0.85em; font-weight: bold;">
                        Pregunta {i + 1} - {dim}
                    </span>
                </div>
                <p style="color: #e2e8f0; font-size: 1.1em; margin: 12px 0;">
                    {q_text}
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        
        # Radio buttons para la respuesta
        response_key = f"wpi_q_{i}"
        
        # Inicializar desde respuestas guardadas
        if response_key not in st.session_state and st.session_state.wpi_responses[i] is not None:
            st.session_state[response_key] = st.session_state.wpi_responses[i]
        
        col1, col2 = st.columns([4, 1])
        with col1:
            response = st.radio(
                f"Respuesta {i + 1}",
                options=[1, 2, 3, 4, 5],
                format_func=lambda x: {
                    1: "1 - Totalmente en desacuerdo",
                    2: "2 - En desacuerdo",
                    3: "3 - Neutral",
                    4: "4 - De acuerdo",
                    5: "5 - Totalmente de acuerdo"
                }[x],
                key=response_key,
                horizontal=False,
                index=None if response_key not in st.session_state or st.session_state[response_key] is None else st.session_state[response_key] - 1
            )
        
        with col2:
            st.markdown("<br>" * 2, unsafe_allow_html=True)
            if response is not None:
                st.success("✅")
                st.session_state.wpi_responses[i] = response
            else:
                st.warning("⚠️")
                all_answered = False

    # Navegación
    st.markdown("---")
    col_prev, col_space, col_next = st.columns([1, 4, 1])

    with col_prev:
        if page > 0:
            if st.button("⬅️ Anterior", key="wpi_prev"):
                st.session_state.wpi_page -= 1
                st.rerun()

    with col_next:
        is_last = q_end >= total
        btn_label = "✅ Finalizar Evaluación" if is_last else "Siguiente ➡️"
        if st.button(btn_label, key="wpi_next", disabled=not all_answered):
            # Verificar tiempo nuevamente
            remaining = db.check_session_time(db.get_session_by_id(session["id"]))
            if remaining == -1:
                st.error("⏰ El tiempo ha expirado.")
                return

            if is_last:
                # Verificar que todas las preguntas estén respondidas
                if None in st.session_state.wpi_responses:
                    st.warning("⚠️ Hay preguntas sin responder. Revisa las páginas anteriores.")
                else:
                    # Calcular resultados
                    responses = st.session_state.wpi_responses
                    raw, normalized, percentages = calculate_wpi_results(responses, questions)

                    # Guardar respuestas
                    answer_records = []
                    for i in range(total):
                        answer_records.append({
                            "question_index": i,
                            "question_text": questions[i]["question"],
                            "answer_value": responses[i],
                            "answer_b_value": None,  # No aplica para WPI
                        })
                    db.save_answers(session["id"], answer_records)

                    # Guardar resultados
                    results = {
                        "raw": raw,
                        "normalized": normalized,
                        "percentages": percentages
                    }
                    db.save_results(session["id"], results)
                    db.complete_test_session(session["id"])

                    # Limpiar session state
                    for key in ["wpi_questions", "wpi_responses", "wpi_page", "eri_questions", "eri_responses", "eri_page", "test_session"]:
                        st.session_state.pop(key, None)

                    nav("candidate_done")
                    st.rerun()
            else:
                st.session_state.wpi_page += 1
                st.rerun()


def page_eri_test():
    """
    Página del test ERI (Evaluación de Riesgo e Integridad) - 60 preguntas con escala Likert 1-5.
    """
    session = st.session_state.get("test_session")
    candidate = st.session_state.get("candidate")
    
    if not session or not candidate:
        nav("candidate_login")
        st.rerun()
        return

    session = db.get_session_by_id(session["id"])
    if not session or session["status"] not in ("in_progress",):
        if session and session["status"] == "expired":
            st.error("⏰ El tiempo de esta evaluación ha expirado.")
            if st.button("Volver"):
                nav("candidate_select_test")
                st.rerun()
            return
        nav("candidate_select_test")
        st.rerun()
        return

    # Verificar tiempo restante
    remaining = db.check_session_time(session)
    if remaining == -1:
        st.error("⏰ El tiempo de esta evaluación ha expirado.")
        if st.button("Volver"):
            nav("candidate_select_test")
            st.rerun()
        return

    # Mostrar timer
    deadline_ts = db.get_session_deadline_timestamp(session)
    if deadline_ts:
        render_timer(deadline_ts, session["id"])

    st.markdown(f"### 🔐 ERI - Evaluación de Riesgo e Integridad")
    st.caption(f"Candidato: {candidate['name']} | ID: {session['id']}")
    
    # Cargar preguntas si no están en session_state
    if "eri_questions" not in st.session_state:
        all_questions = load_eri_questions()
        # Mezclar preguntas de manera consistente por sesión
        rng = random.Random(session["id"])
        rng.shuffle(all_questions)
        st.session_state.eri_questions = all_questions
        db.update_session_questions(session["id"], all_questions)

    # Inicializar respuestas
    if "eri_responses" not in st.session_state:
        st.session_state.eri_responses = [None] * len(st.session_state.eri_questions)

    # Inicializar página
    if "eri_page" not in st.session_state:
        st.session_state.eri_page = 0

    questions = st.session_state.eri_questions
    total = len(questions)
    questions_per_page = 10  # 10 preguntas por página
    page = st.session_state.eri_page
    q_start = page * questions_per_page
    q_end = min(q_start + questions_per_page, total)

    # Barra de progreso
    progress = q_end / total
    st.progress(progress)
    st.markdown(f"**Preguntas {q_start + 1} - {q_end} de {total}**")

    # Instrucciones
    st.info("""
    **Instrucciones:** Responde con la máxima SINCERIDAD a cada afirmación. No hay respuestas correctas o incorrectas.
    
    Escala:
    - **1** = Totalmente de acuerdo
    - **2** = De acuerdo
    - **3** = Neutral / No estoy seguro
    - **4** = En desacuerdo
    - **5** = Totalmente en desacuerdo
    
    ⚠️ **IMPORTANTE:** Esta evaluación detecta patrones de respuesta poco sinceros. Por favor, responde honestamente.
    """)

    # Mostrar preguntas de la página actual
    all_answered = True
    
    for i in range(q_start, q_end):
        q = questions[i]
        q_text = q["question"]
        dim = q["dimension"]
        
        # Crear tarjeta visual para cada pregunta con colores de ERI
        st.markdown(
            f"""
            <div style="background: linear-gradient(135deg, #1e293b 0%, #334155 100%);
                        border-radius: 12px; padding: 20px; margin: 15px 0;
                        border-left: 4px solid {ERI_COLORS.get(dim, '#3b82f6')};">
                <div style="margin-bottom: 8px;">
                    <span style="background: {ERI_COLORS.get(dim, '#3b82f6')}; color: white; 
                                padding: 4px 12px; border-radius: 20px; 
                                font-size: 0.85em; font-weight: bold;">
                        Pregunta {i + 1} - {dim}
                    </span>
                </div>
                <p style="color: #e2e8f0; font-size: 1.1em; margin: 12px 0;">
                    {q_text}
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        
        # Radio buttons para la respuesta
        response_key = f"eri_q_{i}"
        
        # Inicializar desde respuestas guardadas
        if response_key not in st.session_state and st.session_state.eri_responses[i] is not None:
            st.session_state[response_key] = st.session_state.eri_responses[i]
        
        col1, col2 = st.columns([4, 1])
        with col1:
            response = st.radio(
                f"Respuesta {i + 1}",
                options=[1, 2, 3, 4, 5],
                format_func=lambda x: {
                    1: "1 - Totalmente de acuerdo",
                    2: "2 - De acuerdo",
                    3: "3 - Neutral",
                    4: "4 - En desacuerdo",
                    5: "5 - Totalmente en desacuerdo"
                }[x],
                key=response_key,
                horizontal=False,
                index=None if response_key not in st.session_state or st.session_state[response_key] is None else st.session_state[response_key] - 1
            )
        
        with col2:
            st.markdown("<br>" * 2, unsafe_allow_html=True)
            if response is not None:
                st.success("✅")
                st.session_state.eri_responses[i] = response
            else:
                st.warning("⚠️")
                all_answered = False

    # Navegación
    st.markdown("---")
    col_prev, col_space, col_next = st.columns([1, 4, 1])

    with col_prev:
        if page > 0:
            if st.button("⬅️ Anterior", key="eri_prev"):
                st.session_state.eri_page -= 1
                st.rerun()

    with col_next:
        is_last = q_end >= total
        btn_label = "✅ Finalizar Evaluación" if is_last else "Siguiente ➡️"
        if st.button(btn_label, key="eri_next", disabled=not all_answered):
            # Verificar tiempo nuevamente
            remaining = db.check_session_time(db.get_session_by_id(session["id"]))
            if remaining == -1:
                st.error("⏰ El tiempo ha expirado.")
                return

            if is_last:
                # Verificar que todas las preguntas estén respondidas
                if None in st.session_state.eri_responses:
                    st.warning("⚠️ Hay preguntas sin responder. Revisa las páginas anteriores.")
                else:
                    # Calcular resultados
                    responses = st.session_state.eri_responses
                    raw, normalized, percentages, validity_score, validity_flags = calculate_eri_results(responses, questions)

                    # Guardar respuestas
                    answer_records = []
                    for i in range(total):
                        answer_records.append({
                            "question_index": i,
                            "question_text": questions[i]["question"],
                            "answer_value": responses[i],
                            "answer_b_value": None,  # No aplica para ERI
                        })
                    db.save_answers(session["id"], answer_records)

                    # Guardar resultados
                    results = {
                        "raw": raw,
                        "normalized": normalized,
                        "percentages": percentages,
                        "validity_score": validity_score,
                        "validity_flags": validity_flags
                    }
                    db.save_results(session["id"], results)
                    db.complete_test_session(session["id"])

                    # Limpiar session state
                    for key in ["eri_questions", "eri_responses", "eri_page", "test_session"]:
                        st.session_state.pop(key, None)

                    nav("candidate_done")
                    st.rerun()
            else:
                st.session_state.eri_page += 1
                st.rerun()


# -------------------------------------------------------------------------
# CANDIDATE: DONE
# -------------------------------------------------------------------------
def page_candidate_done():
    candidate = st.session_state.get("candidate")
    name = candidate["name"] if candidate else "Candidato"

    st.markdown("<br><br>", unsafe_allow_html=True)
    st.markdown(f"""
    <div style="text-align:center; padding: 60px 20px; background: linear-gradient(135deg, #d1fae5 0%, #a7f3d0 100%); border-radius: 20px; margin: 20px 0;">
        <h1 style="color: #065f46; font-size: 2.5em;">✅ ¡Evaluación Completada!</h1>
        <p style="color: #047857; font-size: 1.3em;">Gracias, <strong>{name}</strong>.</p>
        <p style="color: #047857; font-size: 1.1em;">Tu evaluación ha sido registrada exitosamente.<br>
        Los resultados serán revisados por el equipo de Recursos Humanos.</p>
        <p style="color: #6b7280; margin-top: 30px;">Puedes cerrar esta ventana o continuar con otra evaluación pendiente.</p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    c1, c2, c3 = st.columns([1, 2, 1])
    with c2:
        if st.button("📋 Ver otras evaluaciones pendientes", use_container_width=True):
            nav("candidate_select_test")
            st.rerun()

        if st.button("🚪 Salir", use_container_width=True):
            for key in ["candidate", "pending_sessions", "test_session", 
                       "disc_questions", "disc_page", "disc_answers", 
                       "valanti_responses", "valanti_page",
                       "wpi_questions", "wpi_responses", "wpi_page",
                       "eri_questions", "eri_responses", "eri_page"]:
                st.session_state.pop(key, None)
            nav("home")
            st.rerun()


# =========================================================================
# MAIN ROUTING
# =========================================================================

_restore_admin_session()

if "page" not in st.session_state:
    st.session_state.page = "admin_dashboard" if st.session_state.get("admin") else "home"
else:
    if st.session_state.get("admin"):
        _touch_admin_session()

page = st.session_state.page

PAGE_MAP = {
    "home": page_home,
    "admin_login": page_admin_login,
    "admin_dashboard": page_admin_dashboard,
    "evaluador_login": page_evaluador_login,
    "evaluador_dashboard": page_evaluador_dashboard,
    "candidate_login": page_candidate_login,
    "candidate_select_test": page_candidate_select_test,
    "disc_test": page_disc_test,
    "valanti_test": page_valanti_test,
    "wpi_test": page_wpi_test,
    "eri_test": page_eri_test,
    "talent_map_test": page_talent_map_test,
    "desempeno_eval": page_desempeno_eval,
    "desempeno_lider_eval": page_desempeno_lider_eval,
    "desempeno_lider_employee_eval": page_desempeno_lider_employee_eval,
    "desempeno_lider_jefe_eval": page_desempeno_lider_jefe_eval,
    "periodo_prueba_eval": page_periodo_prueba_eval,
    "periodo_prueba_employee_eval": page_periodo_prueba_employee_eval,
    "periodo_prueba_jefe_eval": page_periodo_prueba_jefe_eval,
    "candidate_done": page_candidate_done,
}

if page in PAGE_MAP:
    PAGE_MAP[page]()
else:
    nav("home")
    st.rerun()

st.markdown("""
---
<div style="text-align:center; color: #888;">
    <small>Plataforma de Evaluaciones Psicométricas | Recursos Humanos</small>
</div>
""", unsafe_allow_html=True)
