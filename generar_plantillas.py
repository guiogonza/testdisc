"""
Genera plantillas Excel para cargue masivo de candidatos y pruebas.
Requiere: pip install openpyxl
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ─── Paleta de colores ─────────────────────────────────────────────────────────
COLOR_HEADER_BG   = "1F3864"   # Azul oscuro
COLOR_HEADER_FONT = "FFFFFF"   # Blanco
COLOR_REQUIRED_BG = "FFF2CC"   # Amarillo claro  (campos obligatorios)
COLOR_OPTIONAL_BG = "E2EFDA"   # Verde claro     (campos opcionales)
COLOR_EXAMPLE_BG  = "DDEEFF"   # Azul muy claro  (fila de ejemplo)
COLOR_NOTE_BG     = "FCE4D6"   # Naranja claro   (notas)
COLOR_BORDER      = "BFBFBF"


def thin_border():
    s = Side(border_style="thin", color=COLOR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def header_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    c.font = Font(color=COLOR_HEADER_FONT, bold=True, size=11)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def label_cell(ws, row, col, value, bg=COLOR_REQUIRED_BG, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, size=10)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = thin_border()
    return c


def data_cell(ws, row, col, value="", bg="FFFFFF", italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(italic=italic, size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = thin_border()
    return c


# ══════════════════════════════════════════════════════════════════════════════
# PLANTILLA 1: CARGUE MASIVO DE CANDIDATOS
# ══════════════════════════════════════════════════════════════════════════════
def crear_plantilla_candidatos(output_path):
    wb = openpyxl.Workbook()

    # ── Hoja de instrucciones ────────────────────────────────────────────────
    ws_inst = wb.active
    ws_inst.title = "Instrucciones"
    ws_inst.sheet_view.showGridLines = False
    ws_inst.column_dimensions["A"].width = 5
    ws_inst.column_dimensions["B"].width = 70

    ws_inst.row_dimensions[1].height = 40
    c = ws_inst.cell(row=1, column=2,
                     value="📋  PLANTILLA PARA CARGUE MASIVO DE CANDIDATOS")
    c.font = Font(bold=True, size=16, color=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")

    instrucciones = [
        ("", ""),
        ("OBJETIVO:", "Registrar múltiples candidatos/empleados de forma masiva en el sistema de evaluaciones RH."),
        ("", ""),
        ("HOJA DE DATOS:", "Diligencia los datos en la hoja «BD empleados». No modifiques los encabezados."),
        ("", ""),
        ("CAMPOS OBLIGATORIOS (fondo amarillo):", ""),
        ("  • CEDULA:", "Número de identificación único. Sin puntos ni comas. Ej: 1020304050"),
        ("  • APELLIDOS Y NOMBRES:", "Nombre completo del candidato. Ej: García López Juan Carlos"),
        ("  • EMPRESA:", "Código de empresa. Valores válidos: IRA / FALAB / PROLAB / SIPLAS / ANGEL"),
        ("", ""),
        ("CAMPOS OPCIONALES (fondo verde):", ""),
        ("  • REGIONAL:", "Ciudad o regional. Ej: Bogotá, Medellín, Cali"),
        ("  • CORREO:", "Email corporativo o personal. Ej: juan.garcia@empresa.com"),
        ("  • CARGO:", "Nombre del cargo. Ej: Analista de Calidad"),
        ("  • JEFE INMEDIATO:", "Nombre del jefe inmediato. Ej: María Rodríguez"),
        ("  • NIVEL DE CARGO:", "Operativo / Líder / Directivo"),
        ("  • INVITAR:", "SI o NO (por defecto SI). Indica si se enviará invitación de evaluación."),
        ("", ""),
        ("NOTAS IMPORTANTES:", ""),
        ("  ✓", "La fila 1 es el encabezado. No la modifiques."),
        ("  ✓", "La fila 2 es un ejemplo — bórrala antes de guardar si no aplica."),
        ("  ✓", "Si la cédula ya existe en el sistema, el registro se omitirá (no duplica)."),
        ("  ✓", "Guarda el archivo en formato .xlsx antes de importarlo."),
        ("  ✓", "Empresas válidas: IRA, FALAB, PROLAB, SIPLAS, ANGEL"),
    ]

    for i, (lbl, txt) in enumerate(instrucciones, start=3):
        if lbl.startswith("OBJETIVO") or lbl.startswith("HOJA") or \
           lbl.startswith("CAMPOS") or lbl.startswith("NOTAS"):
            c = ws_inst.cell(row=i, column=2, value=lbl + " " + txt)
            c.font = Font(bold=True, size=11, color=COLOR_HEADER_BG)
        elif lbl.startswith("  •") or lbl.startswith("  ✓"):
            ws_inst.cell(row=i, column=2, value=lbl + "  " + txt).font = Font(size=10)
        else:
            ws_inst.cell(row=i, column=2, value=txt).font = Font(italic=True, size=10, color="595959")
        ws_inst.row_dimensions[i].height = 18

    # ── Hoja de datos ────────────────────────────────────────────────────────
    ws = wb.create_sheet("BD empleados")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # Columnas: nombre, ancho, requerida
    COLUMNS = [
        ("CEDULA",             18, True),
        ("EMPRESA",            12, True),
        ("REGIONAL",           15, False),
        ("APELLIDOS Y NOMBRES",30, True),
        ("CORREO",             28, False),
        ("CARGO",              22, False),
        ("INVITAR",            10, False),
        ("JEFE INMEDIATO",     25, False),
        ("NIVEL DE CARGO",     18, False),
    ]

    ws.row_dimensions[1].height = 35

    for col_idx, (name, width, required) in enumerate(COLUMNS, start=1):
        header_cell(ws, 1, col_idx, name, width)
        # Color de fondo para cada columna de datos (filas 2 en adelante)
        bg = COLOR_REQUIRED_BG if required else COLOR_OPTIONAL_BG
        for r in range(2, 502):  # 500 filas de datos
            data_cell(ws, r, col_idx, bg=bg)

    # Fila de ejemplo (fila 2)
    ejemplos = [
        "1020304050", "IRA", "Bogotá", "García López Juan Carlos",
        "juan.garcia@ira.com", "Analista de Calidad",
        "SI", "María Rodríguez Pérez", "Operativo"
    ]
    for col_idx, val in enumerate(ejemplos, start=1):
        c = ws.cell(row=2, column=col_idx, value=val)
        c.fill = PatternFill("solid", fgColor=COLOR_EXAMPLE_BG)
        c.font = Font(italic=True, size=10, color="2F5496")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border()

    # Nota sobre la fila de ejemplo
    ws.cell(row=2, column=1).comment = None  # limpia por si acaso

    # ── Validaciones desplegables ────────────────────────────────────────────
    # EMPRESA (col 2)
    dv_empresa = DataValidation(
        type="list",
        formula1='"IRA,FALAB,PROLAB,SIPLAS,ANGEL"',
        allow_blank=True,
        showErrorMessage=True,
        error="Valor no válido. Use: IRA, FALAB, PROLAB, SIPLAS o ANGEL",
        errorTitle="Empresa inválida",
        showDropDown=False,
    )
    ws.add_data_validation(dv_empresa)
    dv_empresa.sqref = f"B2:B501"

    # INVITAR (col 7)
    dv_invitar = DataValidation(
        type="list",
        formula1='"SI,NO"',
        allow_blank=True,
        showErrorMessage=True,
        error="Use SI o NO",
        errorTitle="Valor inválido",
        showDropDown=False,
    )
    ws.add_data_validation(dv_invitar)
    dv_invitar.sqref = "G2:G501"

    # NIVEL DE CARGO (col 9)
    dv_nivel = DataValidation(
        type="list",
        formula1='"Operativo,Líder,Directivo"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv_nivel)
    dv_nivel.sqref = "I2:I501"

    # ── Tabla formal de Excel ────────────────────────────────────────────────
    table = Table(displayName="CandidatosMasivo", ref="A1:I501")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # ── Hoja de referencia de empresas ───────────────────────────────────────
    ws_ref = wb.create_sheet("Referencia")
    ws_ref.sheet_view.showGridLines = False
    header_cell(ws_ref, 1, 1, "CÓDIGO", 12)
    header_cell(ws_ref, 1, 2, "NOMBRE EMPRESA", 40)
    header_cell(ws_ref, 1, 3, "NIVELES DE CARGO VÁLIDOS", 30)

    empresas = [
        ("IRA",    "IRA - Inversiones y Representaciones"),
        ("FALAB",  "FALAB - Fabricación de Laboratorios"),
        ("PROLAB", "PROLAB - Productos de Laboratorio"),
        ("SIPLAS", "SIPLAS - Sistemas Plásticos"),
        ("ANGEL",  "ANGEL - Ángel Comercial"),
    ]
    for i, (cod, nom) in enumerate(empresas, start=2):
        label_cell(ws_ref, i, 1, cod, bold=True)
        label_cell(ws_ref, i, 2, nom)
        if i == 2:
            label_cell(ws_ref, i, 3, "Operativo / Líder / Directivo")

    wb.save(output_path)
    print(f"✓ Plantilla candidatos creada: {output_path}")


# ══════════════════════════════════════════════════════════════════════════════
# PLANTILLA 2: CARGUE MASIVO DE PRUEBAS
# ══════════════════════════════════════════════════════════════════════════════
def crear_plantilla_pruebas(output_path):
    wb = openpyxl.Workbook()

    # ── Hoja de instrucciones ────────────────────────────────────────────────
    ws_inst = wb.active
    ws_inst.title = "Instrucciones"
    ws_inst.sheet_view.showGridLines = False
    ws_inst.column_dimensions["A"].width = 5
    ws_inst.column_dimensions["B"].width = 75

    ws_inst.row_dimensions[1].height = 40
    c = ws_inst.cell(row=1, column=2,
                     value="🧪  PLANTILLA PARA CARGUE MASIVO DE PRUEBAS / EVALUACIONES")
    c.font = Font(bold=True, size=16, color=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")

    instrucciones = [
        ("", ""),
        ("OBJETIVO:", "Asignar múltiples pruebas o evaluaciones a candidatos ya registrados en el sistema."),
        ("", ""),
        ("HOJA DE DATOS:", "Diligencia los datos en la hoja «Pruebas». No modifiques los encabezados."),
        ("", ""),
        ("CAMPOS OBLIGATORIOS (fondo amarillo):", ""),
        ("  • CEDULA_CANDIDATO:", "Cédula del candidato al que se asigna la prueba. Debe existir en el sistema."),
        ("  • TIPO_PRUEBA:", "Tipo de evaluación. Selecciona del desplegable (ver tabla de referencia)."),
        ("", ""),
        ("CAMPOS OPCIONALES / CONDICIONALES (fondo verde):", ""),
        ("  • TIEMPO_LIMITE_MINUTOS:", "Tiempo máximo en minutos. Si se deja vacío se usa el predeterminado por prueba."),
        ("  • CEDULA_EVALUADOR:", "Solo para tipos desempeno_lider y periodo_prueba. Cédula del evaluador/jefe."),
        ("  • NOMBRE_EVALUADOR:", "Solo para tipos desempeno_lider y periodo_prueba. Nombre completo del evaluador."),
        ("", ""),
        ("TIPOS DE PRUEBA DISPONIBLES:", ""),
        ("  • disc", "Evaluación de personalidad DISC"),
        ("  • valanti", "Evaluación de valores VALANTI"),
        ("  • wpi", "Work Personality Index (WPI)"),
        ("  • eri", "Evaluación de Riesgo Integridad (ERI)"),
        ("  • talent_map", "Talent Map"),
        ("  • desempeno", "Evaluación de Desempeño Operativo"),
        ("  • desempeno_lider", "Evaluación de Desempeño Líderes (requiere evaluador)"),
        ("  • periodo_prueba", "Evaluación Período de Prueba (requiere evaluador)"),
        ("", ""),
        ("TIEMPOS PREDETERMINADOS POR TIPO (si no se especifica):", ""),
        ("  • disc / valanti:", "30 minutos"),
        ("  • wpi:", "30 minutos"),
        ("  • eri:", "20 minutos"),
        ("  • talent_map:", "25 minutos"),
        ("  • desempeno / desempeno_lider / periodo_prueba:", "60 minutos"),
        ("", ""),
        ("NOTAS IMPORTANTES:", ""),
        ("  ✓", "La fila 1 es el encabezado. No la modifiques."),
        ("  ✓", "La fila 2 es un ejemplo — bórrala antes de guardar si no aplica."),
        ("  ✓", "El candidato (CEDULA_CANDIDATO) debe estar registrado previamente en el sistema."),
        ("  ✓", "Para pruebas desempeno_lider y periodo_prueba los campos de evaluador son obligatorios."),
        ("  ✓", "Se puede asignar más de una prueba al mismo candidato en filas distintas."),
    ]

    for i, (lbl, txt) in enumerate(instrucciones, start=3):
        if lbl.startswith("OBJETIVO") or lbl.startswith("HOJA") or \
           lbl.startswith("CAMPOS") or lbl.startswith("TIPOS") or \
           lbl.startswith("TIEMPOS") or lbl.startswith("NOTAS"):
            c = ws_inst.cell(row=i, column=2, value=lbl + " " + txt)
            c.font = Font(bold=True, size=11, color=COLOR_HEADER_BG)
        elif lbl.startswith("  •") or lbl.startswith("  ✓"):
            ws_inst.cell(row=i, column=2, value=lbl + "  " + txt).font = Font(size=10)
        else:
            ws_inst.cell(row=i, column=2, value=txt).font = Font(italic=True, size=10, color="595959")
        ws_inst.row_dimensions[i].height = 18

    # ── Hoja de datos ────────────────────────────────────────────────────────
    ws = wb.create_sheet("Pruebas")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    COLUMNS = [
        ("CEDULA_CANDIDATO",       18, True),
        ("TIPO_PRUEBA",             22, True),
        ("TIEMPO_LIMITE_MINUTOS",   22, False),
        ("CEDULA_EVALUADOR",        18, False),
        ("NOMBRE_EVALUADOR",        30, False),
    ]

    ws.row_dimensions[1].height = 35

    for col_idx, (name, width, required) in enumerate(COLUMNS, start=1):
        header_cell(ws, 1, col_idx, name, width)
        bg = COLOR_REQUIRED_BG if required else COLOR_OPTIONAL_BG
        for r in range(2, 502):
            data_cell(ws, r, col_idx, bg=bg)

    # Fila de ejemplo
    ejemplos = [
        ("1020304050", "disc",              "30", "",           ""),
        ("1020304051", "wpi",               "30", "",           ""),
        ("1020304052", "desempeno_lider",   "60", "9876543210", "María Rodríguez Pérez"),
        ("1020304053", "periodo_prueba",    "60", "9876543210", "María Rodríguez Pérez"),
    ]
    for row_offset, vals in enumerate(ejemplos):
        r = 2 + row_offset
        for col_idx, val in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.fill = PatternFill("solid", fgColor=COLOR_EXAMPLE_BG)
            c.font = Font(italic=True, size=10, color="2F5496")
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = thin_border()

    # ── Validación: TIPO_PRUEBA (col 2) ─────────────────────────────────────
    tipos = "disc,valanti,wpi,eri,talent_map,desempeno,desempeno_lider,periodo_prueba"
    dv_tipo = DataValidation(
        type="list",
        formula1=f'"{tipos}"',
        allow_blank=False,
        showErrorMessage=True,
        error=f"Tipo no válido. Opciones: {tipos}",
        errorTitle="Tipo de prueba inválido",
        showDropDown=False,
    )
    ws.add_data_validation(dv_tipo)
    dv_tipo.sqref = "B2:B501"

    # ── Validación: TIEMPO_LIMITE_MINUTOS (col 3) — solo números ──────────
    dv_tiempo = DataValidation(
        type="whole",
        operator="between",
        formula1="5",
        formula2="180",
        allow_blank=True,
        showErrorMessage=True,
        error="Ingresa un número entre 5 y 180 (minutos)",
        errorTitle="Tiempo inválido",
    )
    ws.add_data_validation(dv_tiempo)
    dv_tiempo.sqref = "C2:C501"

    # ── Tabla formal ─────────────────────────────────────────────────────────
    table = Table(displayName="PruebasMasivo", ref="A1:E501")
    style = TableStyleInfo(
        name="TableStyleMedium7",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # ── Hoja de referencia ────────────────────────────────────────────────────
    ws_ref = wb.create_sheet("Referencia")
    ws_ref.sheet_view.showGridLines = False

    header_cell(ws_ref, 1, 1, "TIPO_PRUEBA",         22)
    header_cell(ws_ref, 1, 2, "NOMBRE COMPLETO",     35)
    header_cell(ws_ref, 1, 3, "TIEMPO SUGERIDO (min)", 22)
    header_cell(ws_ref, 1, 4, "¿REQUIERE EVALUADOR?", 22)

    tipos_ref = [
        ("disc",             "Evaluación de Personalidad DISC",          "30", "No"),
        ("valanti",          "Evaluación de Valores VALANTI",            "30", "No"),
        ("wpi",              "Work Personality Index",                   "30", "No"),
        ("eri",              "Evaluación de Riesgo de Integridad",       "20", "No"),
        ("talent_map",       "Talent Map",                               "25", "No"),
        ("desempeno",        "Desempeño Operativo",                      "60", "No"),
        ("desempeno_lider",  "Desempeño Líderes",                        "60", "Sí (obligatorio)"),
        ("periodo_prueba",   "Período de Prueba",                        "60", "Sí (obligatorio)"),
    ]
    for i, (tipo, nombre, tiempo, eval_req) in enumerate(tipos_ref, start=2):
        label_cell(ws_ref, i, 1, tipo, bold=True)
        label_cell(ws_ref, i, 2, nombre)
        label_cell(ws_ref, i, 3, tiempo)
        bg = COLOR_NOTE_BG if "obligatorio" in eval_req else COLOR_OPTIONAL_BG
        label_cell(ws_ref, i, 4, eval_req, bg=bg)

    wb.save(output_path)
    print(f"✓ Plantilla pruebas creada:     {output_path}")


# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    out_cand  = os.path.join(BASE_DIR, "plantilla_cargue_candidatos.xlsx")
    out_tests = os.path.join(BASE_DIR, "plantilla_cargue_pruebas.xlsx")

    crear_plantilla_candidatos(out_cand)
    crear_plantilla_pruebas(out_tests)

    print("\n✅ Listo. Archivos generados:")
    print(f"   {out_cand}")
    print(f"   {out_tests}")
