"""
Script para importar empleados desde el Excel FO-GH-40 Evaluación Desempeño - Operativo V.2.xlsm
"""
import openpyxl
import os
import sys
from database import create_empresa, create_empleado, get_all_empresas, get_empleados_by_empresa


def import_empresas():
    """Importar las 5 empresas del grupo."""
    empresas_data = [
        ("IRA", "IRA - Inversiones y Representaciones"),
        ("FALAB", "FALAB - Fabricación de Laboratorios"),
        ("PROLAB", "PROLAB - Productos de Laboratorio"),
        ("SIPLAS", "SIPLAS - Sistemas Plásticos"),
        ("ANGEL", "ANGEL - Ángel Comercial"),
    ]
    
    print("\n" + "="*70)
    print("IMPORTANDO EMPRESAS")
    print("="*70)
    
    for codigo, nombre in empresas_data:
        empresa = create_empresa(codigo, nombre, activa=1)
        if empresa:
            print(f"✓ Empresa creada/actualizada: {codigo} - {nombre} (ID: {empresa['id']})")
        else:
            print(f"✗ Error al crear empresa: {codigo}")
    
    print(f"\nTotal empresas activas: {len(get_all_empresas())}")


def import_empleados_from_excel(excel_path):
    """Importar empleados desde el Excel."""
    if not os.path.exists(excel_path):
        print(f"ERROR: No se encontró el archivo {excel_path}")
        return
    
    print("\n" + "="*70)
    print("CARGANDO ARCHIVO EXCEL")
    print("="*70)
    print(f"Archivo: {excel_path}")
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Leer la hoja "BD empleados"
    if "BD empleados" not in wb.sheetnames:
        print("ERROR: No se encontró la hoja 'BD empleados' en el Excel")
        wb.close()
        return
    
    sheet = wb["BD empleados"]
    print(f"Hoja: BD empleados ({sheet.max_row} filas x {sheet.max_column} columnas)")
    
    # Headers esperados en fila 1:
    # CEDULA | EMPRESA | REGIONAL | APELLIDOS Y NOMBRES | CORREO | CARGO | INVITAR | JEFE INMEDIATO | NIVEL DE CARGO
    
    print("\n" + "="*70)
    print("IMPORTANDO EMPLEADOS")
    print("="*70)
    
    imported = 0
    duplicados = 0
    errores = 0
    
    for row_idx in range(2, sheet.max_row + 1):  # Empezar en fila 2 (después del header)
        try:
            cedula = sheet.cell(row_idx, 1).value  # Columna A
            empresa_codigo = sheet.cell(row_idx, 2).value  # Columna B
            regional = sheet.cell(row_idx, 3).value  # Columna C
            nombre = sheet.cell(row_idx, 4).value  # Columna D
            correo = sheet.cell(row_idx, 5).value  # Columna E
            cargo = sheet.cell(row_idx, 6).value  # Columna F
            invitar = sheet.cell(row_idx, 7).value  # Columna G
            jefe_inmediato = sheet.cell(row_idx, 8).value  # Columna H
            nivel_cargo = sheet.cell(row_idx, 9).value  # Columna I
            
            # Validar datos obligatorios
            if not cedula or not nombre:
                continue
            
            # Convertir cedula a string
            cedula = str(int(cedula)) if isinstance(cedula, float) else str(cedula).strip()
            
            # Limpiar datos
            nombre = str(nombre).strip() if nombre else ""
            empresa_codigo = str(empresa_codigo).strip() if empresa_codigo else ""
            regional = str(regional).strip() if regional else ""
            correo = str(correo).strip() if correo else ""
            cargo = str(cargo).strip() if cargo else ""
            invitar = str(invitar).strip().upper() if invitar else "SI"
            jefe_inmediato = str(jefe_inmediato).strip() if jefe_inmediato else ""
            nivel_cargo = str(nivel_cargo).strip() if nivel_cargo else ""
            
            # Reemplazar valores #N/A
            if jefe_inmediato == "#N/A":
                jefe_inmediato = ""
            
            # Intentar crear el empleado
            empleado = create_empleado(
                cedula=cedula,
                name=nombre,
                empresa_codigo=empresa_codigo,
                regional=regional,
                correo=correo,
                position=cargo,
                jefe_inmediato=jefe_inmediato,
                nivel_cargo=nivel_cargo,
                invitar=invitar,
                age=None,
                sex=None,
                education=None
            )
            
            if empleado:
                imported += 1
                if imported <= 5 or imported % 20 == 0:  # Mostrar primeros 5 y cada 20
                    print(f"✓ {imported}. {cedula} - {nombre[:40]} ({empresa_codigo})")
            else:
                duplicados += 1
                if duplicados <= 3:
                    print(f"⚠ Duplicado: {cedula} - {nombre[:40]}")
        
        except Exception as e:
            errores += 1
            if errores <= 3:
                print(f"✗ Error en fila {row_idx}: {str(e)}")
    
    wb.close()
    
    print("\n" + "="*70)
    print("RESUMEN DE IMPORTACIÓN")
    print("="*70)
    print(f"✓ Importados exitosamente: {imported}")
    print(f"⚠ Duplicados (ya existían): {duplicados}")
    print(f"✗ Errores: {errores}")
    print(f"📊 Total procesados: {imported + duplicados + errores}")
    
    return imported


def show_statistics():
    """Mostrar estadísticas de empleados por empresa."""
    print("\n" + "="*70)
    print("ESTADÍSTICAS DE EMPLEADOS POR EMPRESA")
    print("="*70)
    
    empresas = get_all_empresas()
    total_empleados = 0
    
    for empresa in empresas:
        empleados = get_empleados_by_empresa(empresa["id"])
        count = len(empleados)
        total_empleados += count
        print(f"{empresa['codigo']:10} - {empresa['nombre']:35} : {count:3} empleados")
    
    # Empleados sin empresa asignada
    empleados_sin_empresa = get_empleados_by_empresa(empresa_id=None)
    sin_empresa = len([e for e in empleados_sin_empresa if e.get("empresa_id") is None])
    
    if sin_empresa > 0:
        print(f"{'SIN EMPRESA':10} - {'(No asignados)':35} : {sin_empresa:3} empleados")
        total_empleados += sin_empresa
    
    print("-" * 70)
    print(f"{'TOTAL':10}   {'':35}   {total_empleados:3} empleados")


def main():
    """Función principal de importación."""
    # Ruta al archivo Excel (en el directorio padre)
    base_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(os.path.dirname(base_dir), "FO-GH-40 Evaluación Desempeño - Operativo V.2.xlsm")
    
    print("\n" + "="*70)
    print("      IMPORTADOR DE EMPLEADOS Y EMPRESAS - SISTEMA RH")
    print("="*70)
    
    # 1. Importar empresas primero
    import_empresas()
    
    # 2. Importar empleados desde Excel
    imported = import_empleados_from_excel(excel_path)
    
    # 3. Mostrar estadísticas finales
    if imported and imported > 0:
        show_statistics()
    
    print("\n" + "="*70)
    print("IMPORTACIÓN COMPLETADA")
    print("="*70)
    print("\nAhora puedes ver los empleados en la aplicación:")
    print("  - Ir a 'Ver Candidatos/Empleados' en el menú lateral")
    print("  - Filtrar por empresa en el selector")
    print("\n")


if __name__ == "__main__":
    main()
